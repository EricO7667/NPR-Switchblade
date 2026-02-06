from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

# --- Your existing engine modules (keep using these) ---
from .data_loader import DataLoader
from .matching_engine import MatchingEngine
from .parsing_engine import parse_description
from .config_loader import load_config
from .data_models import MatchType
from openpyxl import Workbook, load_workbook
import os
from tkinter import messagebox

from .npr_workspace import (
    NPRWorkspace,
    NPRPrimaryNewItem,
    NPRSecondaryRow,
    NPRRowKind,
)

from .npr_export import (
    NPRExportMetadata,
    export_npr_from_rows,
)

# --- New domain model (Step 1) ---
from .data_models import DecisionNode, DecisionStatus
from .data_models import Alternate


# ----------------------------
# Helpers
# ----------------------------

def safe_get(obj: Any, *names: str, default: Any = "") -> Any:
    for n in names:
        if obj is None:
            return default
        if hasattr(obj, n):
            v = getattr(obj, n)
            if v is not None:
                return v
    return default


def clamp01(x: float) -> float:
    try:
        x = float(x)
    except Exception:
        return 0.0
    if x < 0.0:
        return 0.0
    if x > 1.0:
        return 1.0
    return x


# ----------------------------
# Controller
# ----------------------------

@dataclass
class ControllerConfig:
    components_yaml_path: str = "./config/components.yaml"
    npr_template_path: Optional[str] = None  # if None, controller will not export
    created_by: str = "NPR Tool"


class DecisionController:
    """
    Headless application controller.

    - Owns inventory, BOM/NPR list, match pairs, DecisionNodes
    - Enforces workflow rules (select/reject/ready/locked)
    - Owns external search caching (DigiKey etc.)
    - Builds NPRWorkspace and exports

    UI should ONLY talk to this class.
    """

    def __init__(
        self,
        cfg: Optional[ControllerConfig] = None,
        digikey_search_fn: Optional[Callable[[DecisionNode], List[Alternate]]] = None,
    ):
        self.cfg = cfg or ControllerConfig()
        self._digikey_search_fn = digikey_search_fn

        # Core data
        self.inventory: List[Any] = []
        self.npr_list: List[Any] = []
        self.match_pairs: List[Tuple[Any, Any]] = []  # (npr_part, match_result)

        # Decisions
        self.nodes: List[DecisionNode] = []

        # External cache: node_id -> alternates
        self.external_cache: Dict[str, List[Alternate]] = {}

        # Workspace
        self.workspace: Optional[NPRWorkspace] = None

    # ----------------------------
    # Loaders
    # ----------------------------

    def load_inventory(self, xlsx_path: str) -> int:
        self.inventory = DataLoader.load_inventory(xlsx_path)
        return len(self.inventory)

    def load_npr(self, xlsx_path: str) -> int:
        self.npr_list = DataLoader.load_npr(xlsx_path)
        return len(self.npr_list)

    # ----------------------------
    # Matching -> DecisionNodes
    # ----------------------------

    def run_matching(self) -> int:
        if not self.inventory or not self.npr_list:
            raise RuntimeError("Load inventory and NPR/BOM parts before matching.")

        cfg = load_config(self.cfg.components_yaml_path)

        # Parse both sides (same as ui_main)
        for inv in self.inventory:
            inv.parsed = parse_description(safe_get(inv, "description", default=""), cfg)
        for npr in self.npr_list:
            npr.parsed = parse_description(safe_get(npr, "description", default=""), cfg)

        engine = MatchingEngine(self.inventory, config=cfg)
        self.match_pairs = engine.match_npr_list(self.npr_list)

        # Convert to new DecisionNode model
        self.nodes = [self._pair_to_node(npr_part, match) for npr_part, match in self.match_pairs]
        return len(self.nodes)

    def _pair_to_node(self, npr_part: Any, match: Any) -> DecisionNode:
        """
        Convert your engine output (npr_part, match_result) into the new DecisionNode + Alternates.

        This is where we permanently escape the old UI and normalize the data.
        """
        match_type = safe_get(match, "match_type", default="")
        if hasattr(match_type, "value"):
            match_type = match_type.value
        match_type = str(match_type) if match_type else ""

        inv = safe_get(match, "inventory_part", default=None)

        # Determine base_type:
        # - NO_MATCH -> NEW
        # - else -> EXISTS if we have inventory_part
        base_type = "NEW" if match_type == MatchType.NO_MATCH else ("EXISTS" if inv is not None else "NEW")

        # BOM fields (use your known names; keep resilient)
        bom_uid = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default=""))
        bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default=""))
        description = str(safe_get(npr_part, "description", "desc", default=""))

        # Inventory base fields (EXISTS)
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default="")) if inv else ""
        inv_mpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="")) if inv else ""

        confidence = clamp01(safe_get(match, "confidence", default=0.0))

        node_id = bom_uid or (internal_pn or f"NODE-{len(self.nodes)+1}")

        node = DecisionNode(
            id=node_id,
            base_type=base_type,
            bom_uid=bom_uid,
            bom_mpn=bom_mpn,
            description=description,
            internal_part_number=internal_pn,
            inventory_mpn=inv_mpn,
            match_type=match_type,
            confidence=confidence,
            alternates=[],
            status=DecisionStatus.NEEDS_DECISION,
            locked=False,
            needs_approval=False,
            notes="",
            explain={},
        )

        # Internal candidate alternates:
        # If you already have ranked candidates in match, consume them here.
        # Otherwise: start with base inventory part as an "Exact" alternate if EXISTS.
        if inv is not None:
            node.alternates.append(self._inventory_part_to_alternate(inv, confidence=confidence, relationship="Base/Context"))

        # If match has a "candidates" list (or similar), convert them to alternates
        candidates = safe_get(match, "candidates", "candidate_parts", default=None)
        if candidates:
            for cand in candidates:
                node.alternates.append(self._inventory_part_to_alternate(cand, confidence=clamp01(safe_get(cand, "confidence", default=confidence)), relationship="Internal Candidate"))

        # Decide initial status heuristics
        self._recompute_node_flags(node)
        return node

    def _inventory_part_to_alternate(self, inv: Any, confidence: float, relationship: str) -> Alternate:
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default=""))
        mfgpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default=""))
        mfg = str(safe_get(inv, "manufacturer", "manufacturer_name", "mfgname", default=""))

        # stable alt id: prefer internal PN; fallback to MPN; else UUID
        alt_id = internal_pn or (f"MPN-{mfgpn}" if mfgpn else Alternate.new_id("INV"))

        return Alternate(
            id=alt_id,
            source="inventory",
            manufacturer=mfg,
            manufacturer_part_number=mfgpn,
            internal_part_number=internal_pn,
            description=str(safe_get(inv, "description", "desc", default="")),
            value=str(safe_get(inv, "value", default="")),
            package=str(safe_get(inv, "package", default="")),
            tolerance=str(safe_get(inv, "tolerance", default="")),
            voltage=str(safe_get(inv, "voltage", default="")),
            wattage=str(safe_get(inv, "wattage", default="")),
            stock=int(safe_get(inv, "stock", "qty", default=0) or 0),
            unit_cost=safe_get(inv, "unit_cost", "cost", default=None),
            supplier=str(safe_get(inv, "supplier", "vendor", default="")),
            confidence=clamp01(confidence),
            relationship=relationship,
            selected=False,
            rejected=False,
            raw=inv,
            meta={},
        )

    # ----------------------------
    # Query helpers for UI
    # ----------------------------

    def get_nodes(self) -> List[DecisionNode]:
        return self.nodes

    def get_node(self, node_id: str) -> DecisionNode:
        for n in self.nodes:
            if n.id == node_id:
                return n
        raise KeyError(f"No DecisionNode with id={node_id}")

    # ----------------------------
    # Mutations (UI actions)
    # ----------------------------

    def select_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.rejected = False
        alt.selected = True

        self._recompute_node_flags(node)

    def reject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False
        alt.rejected = True

        self._recompute_node_flags(node)

    def unselect_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False
        self._recompute_node_flags(node)

    def add_manual_alternate(self, node_id: str, manufacturer_part_number: str, description: str = "Manual alternate") -> Alternate:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = Alternate(
            id=Alternate.new_id("MAN"),
            source="manual",
            manufacturer="",
            manufacturer_part_number=manufacturer_part_number,
            internal_part_number="",
            description=description,
            confidence=0.0,
            relationship="Manual",
        )
        node.alternates.append(alt)
        self._recompute_node_flags(node)
        return alt

    def mark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        # Rule copied from ui_main: NEW requires at least one selected alt :contentReference[oaicite:1]{index=1}
        if node.base_type.upper() == "NEW" and not node.has_selection():
            raise ValueError("NEW nodes require at least one selected alternate before marking READY.")

        node.status = DecisionStatus.READY_FOR_EXPORT
        node.locked = True
        self._recompute_node_flags(node)

    # ----------------------------
    # External search (DigiKey etc.)
    # ----------------------------

    def search_digikey(self, node_id: str) -> List[Alternate]:
        """
        Cache is controller-owned (not UI). UI just calls this.
        """
        node = self.get_node(node_id)

        if node_id in self.external_cache:
            return self.external_cache[node_id]

        if not self._digikey_search_fn:
            # no client wired yet
            self.external_cache[node_id] = []
            return []

        results = self._digikey_search_fn(node) or []
        # Normalize + attach
        for alt in results:
            alt.source = alt.source or "digikey"
        self.external_cache[node_id] = results

        # Option: merge external alternates into node list (recommended)
        # so UI doesn’t need “internal vs external” stores.
        node.alternates.extend(results)
        self._recompute_node_flags(node)
        return results

    # ----------------------------
    # NPR Workspace (from DecisionNodes)
    # ----------------------------

    def build_npr_workspace_from_nodes(self) -> NPRWorkspace:
        """
        FINAL NPR WORKSPACE BUILD (DecisionNode authoritative)

        Page 1: all NEW nodes
        Page 2: approval rows:
            - Context exists rows (matched internal part)
            - Selected alternate proposal rows (internal/external/manual)
        """
        ws = NPRWorkspace()

        # -------------------------
        # Page 1 — NEW parts
        # -------------------------
        for node in self.nodes:
            if node.base_type.upper() != "NEW":
                continue

            primary = NPRPrimaryNewItem(
                bom_uid=node.bom_uid,
                bom_mpn=node.bom_mpn,
                description=node.description,
                component_type="",   # fill if you add it to DecisionNode
                populated=False,
                include_in_export=True,
                notes=node.notes or "",
            )

            ws.primary_new_items.append(primary)

        # -------------------------
        # Page 2 — Secondary rows
        # -------------------------
        for node in self.nodes:
            parent_uid = node.bom_uid
            parent_mpn = node.bom_mpn
            parent_desc = node.description

            has_selected_alts = bool(node.selected_alternates())
            # P2.1 Context EXISTS row
            if node.internal_part_number and (has_selected_alts or node.base_type.upper() == "NEW"):
                row_id = f"CTX-{parent_uid}-{node.internal_part_number}"
                ws.secondary_rows.append(
                    NPRSecondaryRow(
                        row_id=row_id,
                        kind=NPRRowKind.CONTEXT_EXISTS,
                        parent_bom_uid=parent_uid,
                        parent_bom_mpn=parent_mpn,
                        parent_description=parent_desc,
                        internal_part_number=node.internal_part_number,
                        exists_in_inventory=True,
                        include_in_export=True,
                        source="inventory",
                    )
                )

            # P2.2 Alternate proposals for every SELECTED alternate
            for alt in node.selected_alternates():
                exists = bool(alt.internal_part_number)

                row_id = f"ALT-{parent_uid}-{alt.id}"
                ws.secondary_rows.append(
                    NPRSecondaryRow(
                        row_id=row_id,
                        kind=NPRRowKind.ALTERNATE_PROPOSAL,
                        parent_bom_uid=parent_uid,
                        parent_bom_mpn=parent_mpn,
                        parent_description=parent_desc,
                        internal_part_number=alt.internal_part_number,
                        exists_in_inventory=exists,
                        include_in_export=True,
                        source=alt.source,
                        manufacturer_name=alt.manufacturer,
                        manufacturer_part_number=alt.manufacturer_part_number,
                        supplier=alt.supplier,
                        unit_cost=alt.unit_cost,
                    )
                )

        self.workspace = ws
        return ws


    def export_npr(self, output_path: str) -> None:
        """
        Exports NPR data to Excel safely on Windows.
        Automatically creates folders and handles locked files gracefully.
        """
        output_path = Path(output_path).expanduser().resolve()

        # ---- Validate workspace ----
        if not self.workspace or (not self.workspace.primary_new_items and not self.workspace.secondary_rows):
            raise RuntimeError("Workspace is empty. Build NPR workspace first.")

        # ---- Prepare export folder ----
        export_dir = output_path.parent
        if export_dir and not export_dir.exists():
            export_dir.mkdir(parents=True, exist_ok=True)

        # ---- Handle locked file (Excel open) ----
        if output_path.exists():
            try:
                os.rename(output_path, output_path)
            except OSError:
                raise PermissionError(
                    f"Cannot write to '{output_path}'.\nClose it in Excel and try again."
                )

        # ---- Load or create workbook ----
        template_path = self.cfg.npr_template_path
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
        else:
            wb = Workbook()

        ws1 = wb.active
        ws1.title = "NEW_PARTS"
        ws2 = wb.create_sheet("APPROVALS")

        # ---- Metadata ----
        meta = wb.create_sheet("META")
        meta["A1"], meta["B1"] = "Created By", self.cfg.created_by
        meta["A2"], meta["B2"] = "Created At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta["A3"], meta["B3"] = "Notes", "Exported from NPR Tool"

        # ---- Write Page 1: NEW PARTS ----
        ws1.append(["BOM UID", "BOM MPN", "Description", "Include", "Notes"])
        for p in self.workspace.primary_new_items:
            ws1.append([p.bom_uid, p.bom_mpn, p.description,
                        "Yes" if p.include_in_export else "No", p.notes])

        # ---- Write Page 2: APPROVALS ----
        ws2.append(["Row ID", "Type", "Parent UID", "MPN", "Description",
                    "Internal PN", "Exists", "Source", "Manufacturer",
                    "Mfg PN", "Supplier", "Cost"])
        for r in self.workspace.secondary_rows:
            ws2.append([
                r.row_id,
                getattr(r.kind, "value", str(r.kind)),
                r.parent_bom_uid,
                r.parent_bom_mpn,
                r.parent_description,
                r.internal_part_number,
                "Yes" if r.exists_in_inventory else "No",
                r.source,
                getattr(r, "manufacturer_name", ""),
                getattr(r, "manufacturer_part_number", ""),
                getattr(r, "supplier", ""),
                getattr(r, "unit_cost", "")
            ])

        # ---- Save Safely ----
        try:
            wb.save(output_path)
            wb.close()
        except PermissionError:
            raise PermissionError(f"Cannot save to '{output_path}'. File may be open in Excel.")
        except Exception as e:
            raise RuntimeError(f"Failed to save NPR file: {e}")

        print(f"✅ Exported NPR successfully to: {output_path}")

    # ----------------------------
    # Internal enforcement
    # ----------------------------

    def _find_alt(self, node: DecisionNode, alt_id: str) -> Alternate:
        for a in node.alternates:
            if a.id == alt_id:
                return a
        raise KeyError(f"No alternate with id={alt_id} on node={node.id}")

    def _ensure_unlocked(self, node: DecisionNode) -> None:
        if node.locked:
            raise PermissionError("Node is locked (READY_FOR_EXPORT). Unlocking not supported.")
        



    def _recompute_node_flags(self, node: DecisionNode) -> None:
        """
        Central place for status/approval heuristics.
        UI should just display these.
        """
        bt = node.base_type.upper()
        selected = len(node.selected_alternates())
        candidates = len(node.candidate_alternates())

        # Needs approval heuristic (mirrors your ui_main logic)
        if bt == "NEW":
            node.needs_approval = True
        elif bt == "EXISTS":
            node.needs_approval = selected > 0
        else:
            node.needs_approval = selected > 0

        # Status heuristic unless already READY/locked
        if node.locked:
            node.status = DecisionStatus.READY_FOR_EXPORT
            return

        if bt == "NEW" and selected == 0:
            node.status = DecisionStatus.NEEDS_ALTERNATE
            return

        if bt == "EXISTS" and candidates > 0 and selected == 0:
            node.status = DecisionStatus.NEEDS_DECISION
            return

        # default: auto-ish state (still editable)
        node.status = DecisionStatus.AUTO_MATCHED
