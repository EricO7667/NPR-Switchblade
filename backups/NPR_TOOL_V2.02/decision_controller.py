from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from collections import defaultdict, Counter
from .data_loader import DataLoader
from .matching_engine import MatchingEngine
from .parsing_engine import parse_description
from .config_loader import load_config
from .data_models import MatchType
from openpyxl import load_workbook
import os
import copy
from tkinter import messagebox
import re
import json
import threading
from .npr_workspace import (
    NPRWorkspace,
    NPRPrimaryNewItem,
    NPRSecondaryRow,
    NPRRowKind,
)
from .data_models import DecisionNode, DecisionStatus
from .data_models import Alternate
from rapidfuzz import fuzz
# ----------------------------
# Controller
# ----------------------------

@dataclass
class ControllerConfig:
    """Simplified configuration for NPR Controller."""

    # Base configuration
    components_yaml_path: str = "./config/components.yaml"

    # Clean, absolute template path (uses os + Path)
    npr_template_path: str = str(
        Path(os.path.dirname(os.path.abspath(__file__))) / "NPR_Master2023_v4_FormTEMPLATECOPY.xlsx"
    )

    # Metadata still needed for export
    created_by: str = "NPR Tool"

    PB_RE = re.compile(r"^\s*(\d{2})-(\d{5})\s*$")
    _PN_RE = re.compile(r"^(?P<prefix>\d{2})-(?P<body>\d{5})-(?P<suffix>[A-Za-z0-9]{4})$")

    # ----------------------------
    # Candidate limiting +  filters
    # ----------------------------
    MAX_INTERNAL_CANDIDATES = 10

    _IMPERIAL_PKGS = {"01005", "0201", "0402", "0603", "0805", "1206", "1210", "1812", "2010", "2512"}
    _METRIC_TO_IMPERIAL = {
        "0402": "01005",  # 0402 metric ~= 01005 imperial
        "0603": "0201",   # 0603 metric ~= 0201 imperial
        "1005": "0402",   # 1005 metric ~= 0402 imperial
        "1608": "0603",   # 1608 metric ~= 0603 imperial
        "2012": "0805",   # 2012 metric ~= 0805 imperial
        "3216": "1206",   # 3216 metric ~= 1206 imperial
        "3225": "1210",   # 3225 metric ~= 1210 imperial
        "4532": "1812",   # 4532 metric ~= 1812 imperial
        "5025": "2010",   # 5025 metric ~= 2010 imperial
        "6332": "2512",   # 6332 metric ~= 2512 imperial
    }

# ----------------------------
# Helpers
# ----------------------------

def safe_get(obj: Any, *names: str, default: Any = "") -> Any:
    for n in names:
        if obj is None:
            return default
        if hasattr(obj, n):
            v = getattr(obj, n)
            if v is not None:
                return v
    return default

def _split_company_pn(pn: str) -> tuple[str, str, str]:
    pn = (pn or "").strip()
    m = ControllerConfig._PN_RE.match(pn)
    if not m:
        return "", "", ""
    return m.group("prefix"), m.group("body"), m.group("suffix")

def clamp01(x: float) -> float:
    try:
        x = float(x)
    except Exception:
        return 0.0
    if x < 0.0:
        return 0.0
    if x > 1.0:
        return 1.0
    return x

class DecisionController:
    """
    Headless application controller.

    - Owns inventory, BOM/NPR list, match pairs, DecisionNodes
    - Enforces workflow rules (select/reject/ready/locked)
    - Owns external search caching (DigiKey etc.)
    - Builds NPRWorkspace and exports

    UI should ONLY talk to this class.
    """

    def __init__(
        self,
        cfg: Optional[ControllerConfig] = None,
        digikey_search_fn: Optional[Callable[[DecisionNode], List[Alternate]]] = None,
        stop_event: Optional[threading.Event] = None
    ):
        self.cfg = cfg or ControllerConfig()
        self._digikey_search_fn = digikey_search_fn

        # Core data
        self.inventory: List[Any] = []
        self.npr_list: List[Any] = []
        self.match_pairs: List[Tuple[Any, Any]] = []  # (npr_part, match_result)

        # Decisions
        self.nodes: List[DecisionNode] = []

        # External cache: node_id -> alternates
        self.external_cache: Dict[str, List[Alternate]] = {}
        self.stop_event = stop_event or threading.Event()

        # Workspace
        self.workspace: Optional[NPRWorkspace] = None

    # ----------------------------
    # Loaders
    # ----------------------------

    def load_inventory(self, xlsx_path: str) -> int:
        self.inventory = DataLoader.load_inventory(xlsx_path)
        return len(self.inventory)

    # I call it loading NPR but its really supposed to be LOAD BOM, but i messed up and made all the variable names NPR instead of BOM when making this
    def load_npr(self, xlsx_path: str) -> int:
        #self.npr_list = DataLoader.load_npr(xlsx_path)
        #return len(self.npr_list)

        # Accept “anything” BOMs:
        # - try full NPR/BOM parse
        # - fallback to simple parts list when headers are weird
        self.npr_list = DataLoader.load_bom_any(xlsx_path)
        return len(self.npr_list)
    
    def load_cns(self, xlsx_path: str) -> int:
        self.cns_records = DataLoader.load_cns_workbook(xlsx_path)
        return len(self.cns_records)
    
    def _should_stop(self) -> bool:
        try:
            return bool(self.stop_event and self.stop_event.is_set())
        except Exception:
            return False
    

    def load_cns_preview(self, xlsx_path: str, last_n_per_sheet: int = 1) -> dict:
        """
        Loads CNS and prints:
          - every sheet name discovered
          - the LAST prefix-body (or last N) discovered per sheet

        last_n_per_sheet:
          1 -> just last
          2 -> last two, etc.
        """
        records = DataLoader.load_cns_workbook(xlsx_path)
        self.cns_records = records

        # sheet -> ordered unique PBs (preserve discovery order)
        sheet_pbs = defaultdict(list)
        sheet_seen = defaultdict(set)

        for r in records:
            sheet = (getattr(r, "sheet_name", "") or "").strip()
            prefix = (getattr(r, "prefix", "") or "").strip()
            body = (getattr(r, "body", "") or "").strip()
            if not sheet or not prefix or not body:
                continue

            pb = f"{prefix}-{body}"
            if not ControllerConfig.PB_RE.match(pb):
                continue

            if pb not in sheet_seen[sheet]:
                sheet_seen[sheet].add(pb)
                sheet_pbs[sheet].append(pb)

        all_sheets = sorted(sheet_pbs.keys())

        #print("\n=== CNS SHEET LIST ===")
        #print(f"Sheets found with PBs: {len(all_sheets)}")
        #for s in all_sheets:
        #    print(f" - {s}")

        #print("\n=== CNS LAST PB PER SHEET ===")
        #for s in all_sheets:
        #    pbs = sheet_pbs[s]
        #    if not pbs:
        #        print(f"[{s}] unique PBs=0  last=[]")
        #        continue

        #    n = max(1, int(last_n_per_sheet))
        #    last_vals = pbs[-n:] if len(pbs) >= n else pbs[:]   # last N (or all if shorter)

        #    print(f"[{s}]  unique PBs={len(pbs)}  last={last_vals}")

        summary = {
            "total_records": len(records),
            "sheets_with_pbs": len(all_sheets),
            "unique_pb_total": sum(len(v) for v in sheet_pbs.values()),
            "sheets": all_sheets,
            "last_per_sheet": {
                s: (sheet_pbs[s][-max(1, int(last_n_per_sheet)):] if sheet_pbs[s] else [])
                for s in all_sheets
            },
            "unique_pb_per_sheet": {s: len(sheet_pbs[s]) for s in all_sheets},
        }
        return summary


    #def set_assigned_part_number(self, node_id: str, pn: str) -> None:
    #    node = self.get_node(node_id)
    #    self._ensure_unlocked(node)
    #
    #    pn = (pn or "").strip()
    #    node.assigned_part_number = pn
    #
    #    node.explain = dict(node.explain or {})
    #    node.explain["assigned_part_number"] = pn
    #
    #    self._recompute_node_flags(node)

    # decision_controller.py
    def set_assigned_part_number(self, node_id: str, pn: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
    
        pn = (pn or "").strip()
        node.assigned_part_number = pn
    
        node.explain = dict(node.explain or {})
        node.explain["assigned_part_number"] = pn
    
        self._recompute_node_flags(node)

    # ----------------------------
    # Matching -> DecisionNodes
    # ----------------------------

    def run_matching(self) -> int:
        """
        Run matching for all NPR/BOM parts.
    
        IMPORTANT: Customer-provided alternates are matched, but the winning attempt is selected
        deterministically by (tier_rank, tier_quality, stock, stable_order) — NOT by confidence.
        Confidence remains a UI/display artifact.
        """
        if not self.inventory or not self.npr_list:
            raise RuntimeError("Load inventory and NPR/BOM parts before matching.")
    
        cfg = load_config(self.cfg.components_yaml_path)
    
        # Parse both sides (robust attr names)
        for inv in self.inventory:
            if self._should_stop():
                return 0
            inv.parsed = parse_description(safe_get(inv, "description", "desc", default=""), cfg)
    
        # IMPORTANT: preserve loader-provided mpn_alts before overwriting parsed
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            existing_parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = []
            if isinstance(existing_parsed, dict):
                mpn_alts = existing_parsed.get("mpn_alts", []) or []
    
            new_parsed = parse_description(safe_get(npr, "description", "desc", default=""), cfg)
            if not isinstance(new_parsed, dict):
                new_parsed = {}
    
            if mpn_alts:
                new_parsed["mpn_alts"] = mpn_alts
    
            npr.parsed = new_parsed
    
        def uniq_keep_order(items):
            seen = set()
            out = []
            for x in items:
                x = (x or "").strip()
                if x and x not in seen:
                    seen.add(x)
                    out.append(x)
            return out
    
        def _match_type_key(match: Any) -> str:
            mt = safe_get(match, "match_type", default="")
            if hasattr(mt, "value"):
                mt = mt.value
            mt = str(mt or "")
            return mt.replace("MatchType.", "").strip().upper()
    
        def _tier_rank(mt_key: str) -> int:
            ranks = {
                "EXACT_MFG_PN": 600,
                "API_ASSISTED": 550,
                "SUBSTITUTE": 500,
                "PREFIX_FAMILY": 400,
                "PARTIAL_ITEMNUM": 300,
                "PARSED_MATCH": 200,
                "NO_MATCH": 0,
            }
            return int(ranks.get(mt_key, 0))
    
        def _tier_quality(match: Any, mt_key: str) -> float:
            if mt_key == "PARSED_MATCH":
                exp = safe_get(match, "explain", default=None) or {}
                if isinstance(exp, dict):
                    q = exp.get("best_ratio", None)
                    if q is None:
                        q = exp.get("match_ratio", None)
                    if q is None:
                        q = exp.get("ratio", None)
                    try:
                        return float(q or 0.0)
                    except Exception:
                        return 0.0
                return 0.0
    
            inv = safe_get(match, "inventory_part", default=None)
            return 1.0 if inv is not None and mt_key != "NO_MATCH" else 0.0
    
        def _inv_stock(inv: Any) -> float:
            if inv is None:
                return 0.0
            for k in ("stock", "qty", "onhand", "available"):
                v = safe_get(inv, k, default=None)
                if v is None:
                    continue
                try:
                    return float(v or 0.0)
                except Exception:
                    continue
            return 0.0
    
        # Grab the active Tk root (if UI is running). IMPORTANT: do this
        # before constructing the engine so we can pass ui_root correctly.
        try:
            import tkinter as _tk
            tk_root = getattr(_tk, "_default_root", None)
        except Exception:
            tk_root = None

        engine = MatchingEngine(
            self.inventory,
            config=cfg,
            ui_root=tk_root,
            stop_event=getattr(self, "stop_event", None),
        )

        phase_cb = getattr(tk_root, "loading_phase_callback", None) if tk_root else None
        progress_cb = getattr(tk_root, "loading_progress_callback", None) if tk_root else None

        # Phase: embeddings
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Building semantic cache...", True))
            except Exception:
                pass            

        # Build semantic cache (sync; run_matching is already in a worker thread)
        engine.ensure_embeddings_cache()

        # Phase: matching
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Matching parts...", True))
            except Exception:
                pass

        match_pairs = []
    
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            primary_mpn = safe_get(npr, "mfgpn", "bom_mpn", default="")
    
            mpn_options = uniq_keep_order([primary_mpn] + list(mpn_alts))
    
            attempts = []
            merged_candidates = []
    
            for idx, mpn in enumerate(mpn_options or [""]):
                if self._should_stop():
                    return 0
                tmp = copy.copy(npr)
                if hasattr(tmp, "mfgpn"):
                    tmp.mfgpn = mpn
                tmp.parsed = getattr(npr, "parsed", {}) or {}
    
                match = engine.match_single_part(tmp)
    
                mt_key = _match_type_key(match)
                tier = _tier_rank(mt_key)
                quality = _tier_quality(match, mt_key)
    
                inv = safe_get(match, "inventory_part", default=None)
                stock = _inv_stock(inv)
    
                cands = safe_get(match, "candidates", "candidate_parts", default=None) or []
                if inv is not None:
                    merged_candidates.append(inv)
                merged_candidates.extend(list(cands))
    
                inv_id = ""
                if inv is not None:
                    inv_id = str(
                        safe_get(inv, "itemnum", "internal_part_number", "vendoritem", default="") or ""
                    ).strip()
    
                attempts.append(
                    {
                        "customer_mpn": mpn,
                        "idx": idx,
                        "pair": (npr, match),  # keep ORIGINAL npr for node display
                        "match_type": mt_key,
                        "tier": tier,
                        "quality": quality,
                        "stock": stock,
                        "inv_id": inv_id,
                        "candidate_count": len(cands),
                    }
                )
    
            if not attempts:
                continue
            
            # Deterministic winner (no confidence)
            def _attempt_sort_key(a: dict) -> tuple:
                # prefer primary mpn on ties => earlier idx wins
                prefer_primary = -int(a.get("idx", 0))
                return (a["tier"], a["quality"], a["stock"], prefer_primary, a["inv_id"])
    
            winner = max(attempts, key=_attempt_sort_key)
            best_used_mpn = winner["customer_mpn"]
            npr_part, best_match = winner["pair"]
    
            # Stash customer MPN context and attempt audit for UI/debug
            try:
                if not isinstance(best_match.explain, dict):
                    best_match.explain = {}
                best_match.explain["customer_mpns"] = list(mpn_options)
                best_match.explain["winning_mpn"] = best_used_mpn
                best_match.explain["attempts"] = [
                    {
                        "customer_mpn": a["customer_mpn"],
                        "match_type": a["match_type"],
                        "tier": a["tier"],
                        "quality": a["quality"],
                        "stock": a["stock"],
                        "inv_id": a["inv_id"],
                        "candidate_count": a["candidate_count"],
                        "is_winner": a is winner,
                    }
                    for a in attempts
                ]
            except Exception:
                pass
            
            # Attach merged candidates back onto chosen match (best-effort)
            try:
                existing = safe_get(best_match, "candidates", "candidate_parts", default=None) or []
    
                seen = set()
                out = []
                for cand in (list(existing) + list(merged_candidates)):
                    key = safe_get(cand, "itemnum", "internal_part_number", "vendoritem", default=str(cand))
                    if key in seen:
                        continue
                    seen.add(key)
                    out.append(cand)
    
                # MatchingEngine assigns _pc_score and gates/reranks.
                out = sorted(out, key=lambda c: float(getattr(c, "_pc_score", 0.0) or 0.0), reverse=True)
                out = out[: ControllerConfig.MAX_INTERNAL_CANDIDATES]
    
                if hasattr(best_match, "candidates"):
                    best_match.candidates = out
                elif hasattr(best_match, "candidate_parts"):
                    best_match.candidate_parts = out
            except Exception:
                pass
            
            match_pairs.append((npr_part, best_match))

            # Update UI progress for matching phase
            if progress_cb and tk_root:
                ratio = (n_i + 1) / total_npr   
                try:
                    if tk_root and tk_root.winfo_exists():
                        tk_root.after(0, lambda r=ratio: progress_cb(r))
                except Exception:
                    pass

        # ---  sync winning_mpn and attempt flags ---
        try:
            if not hasattr(best_match, "explain") or not isinstance(best_match.explain, dict):
                best_match.explain = {}
            winning_mpn = best_match.explain.get("winning_mpn", "")
            attempts = best_match.explain.get("attempts", [])
            for a in attempts:
                a["is_winner"] = (a.get("customer_mpn") == winning_mpn)
            best_match.explain["attempts"] = attempts

            # NEW: mark the winning inventory part explicitly
            if hasattr(best_match, "inventory_part") and best_match.inventory_part:
                inv = best_match.inventory_part
                if hasattr(inv, "vendoritem"):
                    inv._is_winner = True
                elif hasattr(inv, "manufacturer_part_number"):
                    inv._is_winner = True
        except Exception as e:
            print(f"[CTRL PATCH] failed to sync winner info: {e}")


        # --- Save explain JSON for debugging ---
        debug_dir = Path(os.getenv("NPR_DEBUG_DIR", "debug_match"))
        debug_dir.mkdir(parents=True, exist_ok=True)

        debug_parse = str(os.getenv("NPR_DEBUG_PARSE", "")).strip().lower() not in ("", "0", "false", "no")
        debug_save  = str(os.getenv("NPR_DEBUG_SAVE", "")).strip().lower()  not in ("", "0", "false", "no")
        debug_jsonl_only = str(os.getenv("NPR_DEBUG_JSONL_ONLY", "")).strip().lower() in ("1", "true", "yes")

        if debug_parse or debug_save:
            try:
                jsonl_path = debug_dir / "explain_all.jsonl"

                def _inv_to_payload_dict(inv_obj: Any, *, is_winner: bool = False) -> dict:
                    if inv_obj is None:
                        return {
                            "inv_item": "",
                            "inv_desc": "",
                            "vendor_mpn": "",
                            "mfg": "",
                            "stock": 0,
                            "seed": 0.0,
                            "score": 0.0,
                            "is_winner": bool(is_winner),
                        }

                    inv_item = str(safe_get(inv_obj, "itemnum", "internal_part_number", default="") or "").strip()
                    inv_desc = str(safe_get(inv_obj, "desc", "description", default="") or "").strip()
                    vendor_mpn = str(safe_get(inv_obj, "vendoritem", "manufacturer_part_number", default="") or "").strip()
                    mfg = str(safe_get(inv_obj, "manufacturer", "manufacturer_name", "mfgname", default="") or "").strip()

                    stock = 0
                    for k in ("stock", "qty", "onhand", "available"):
                        v = safe_get(inv_obj, k, default=None)
                        if v is None:
                            continue
                        try:
                            stock = int(float(v or 0))
                            break
                        except Exception:
                            continue
                        
                    seed = safe_get(inv_obj, "_pc_seed", default=None)
                    if seed is None:
                        seed = safe_get(inv_obj, "confidence", default=0.0)
                    seed = clamp01(seed)

                    score = safe_get(inv_obj, "_pc_score", default=None)
                    if score is None:
                        score = seed
                    score = clamp01(score)

                    return {
                        "inv_item": inv_item,
                        "inv_desc": inv_desc,
                        "vendor_mpn": vendor_mpn,
                        "mfg": mfg,
                        "stock": stock,
                        "seed": float(seed),
                        "score": float(score),
                        "is_winner": bool(is_winner),
                    }

                for npr_part, match in match_pairs:
                    if self._should_stop():
                        return 0
                    exp = safe_get(match, "explain", default=None) or {}

                    npr_key = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default="NPR") or "NPR").strip()
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

                    bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default="") or "").strip()
                    npr_desc = str(safe_get(npr_part, "description", "desc", default="") or "")

                    winner_inv = safe_get(match, "inventory_part", default=None)
                    winner_item = str(safe_get(winner_inv, "itemnum", "internal_part_number", default="") or "").strip()

                    # Pull candidates from the chosen match (already merged+filtered+capped earlier)
                    cands = safe_get(match, "candidates", "candidate_parts", default=None) or []

                    # Ensure winner is present in candidates (training needs it)
                    # (Your merge usually does this, but this guarantees it.)
                    if winner_inv is not None:
                        found = False
                        for c in cands:
                            cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                            if cid and cid == winner_item:
                                found = True
                                break
                        if not found:
                            cands = [winner_inv] + list(cands)

                    cand_payloads = []
                    for c in cands:
                        cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                        cand_payloads.append(_inv_to_payload_dict(c, is_winner=(cid == winner_item)))

                    # Find winner rank in candidate list (0-based). If not found, -1.
                    winner_rank = -1
                    for i, cp in enumerate(cand_payloads):
                        if cp.get("is_winner"):
                            winner_rank = i
                            break
                        
                    payload = {
                        "ts": ts,

                        # identifiers
                        "npr_item": npr_key,
                        "bom_mpn": bom_mpn,

                        # main text fields
                        "npr_desc": npr_desc,

                        # deterministic winner choice context
                        "match_type": str(safe_get(match, "match_type", default="")),
                        "winning_mpn": exp.get("winning_mpn") if isinstance(exp, dict) else "",
                        "customer_mpns": exp.get("customer_mpns") if isinstance(exp, dict) else [],

                        # winner block
                        "winner": _inv_to_payload_dict(winner_inv, is_winner=True),
                        "winner_rank": winner_rank,

                        # THIS is the payload candidates you’re asking about
                        "candidates": cand_payloads,

                        # audit/debug (keep it for now; you can drop later for smaller JSONL)
                        "explain": exp,
                    }

                    with open(jsonl_path, "a", encoding="utf-8") as f:
                        f.write(json.dumps(payload, default=str) + "\n")

                    if not debug_jsonl_only:
                        inv_item = payload["winner"]["inv_item"] or "INV"
                        out_path = debug_dir / f"explain_{ts}_{npr_key}__{inv_item}.json"
                        with open(out_path, "w", encoding="utf-8") as f:
                            json.dump(payload, f, indent=2, default=str)
                        print(f"[MATCH DBG] wrote explain json -> {out_path}")

                print(f"[MATCH DBG] wrote JSONL -> {jsonl_path}")

            except Exception as e:
                print(f"[MATCH DBG] failed to write explain json/jsonl: {e}")
        
        self.match_pairs = match_pairs
    
        print("[CTRL DBG] building nodes...")
        self.nodes = [self._pair_to_node(npr_part, match) for npr_part, match in self.match_pairs]
        print(f"[CTRL DBG] nodes built: {len(self.nodes)}")
    
        print("[CTRL DBG] building type->prefix map...")
        self._type_prefix_map = self._build_type_prefix_map()
        print("[CTRL DBG] type->prefix map built OK")
    
        print("[CTRL DBG] applying CNS suggestions to nodes...")
        for i, node in enumerate(self.nodes, start=1):
            if self._should_stop():
                return 0
            #print(f"[CTRL DBG]  apply CNS {i}/{len(self.nodes)} node.id={getattr(node,'id',None)}")
            self._apply_cns_suggestion_to_node(node)
        print("[CTRL DBG] applied CNS suggestions OK")
    
        return len(self.nodes)
    

    ##must move back into the matching engine
    #def _filter_rank_cap_candidates(self, npr_part: Any, candidates: list, max_n: int = ControllerConfig.MAX_INTERNAL_CANDIDATES) -> list:
    #    """
    #    Hard-filter + re-rank candidates using deterministic rules.
    #    Goal: prevent fuzzy-only garbage and cap to max_n.
    #    """
    #    npr_parsed = _get_parsed(npr_part)
    #    n_type = _ptype(npr_parsed)
#
    #    # Pull normalized anchors (when available)
    #    n_pkg = _norm_pkg(str(npr_parsed.get("package") or safe_get(npr_part, "package", default="")))
    #    n_val = str(npr_parsed.get("value") or safe_get(npr_part, "value", default="")).strip()
#
    #    n_res = _norm_res_value(n_val) if n_type == "RESISTOR" else ""
    #    n_cap = _norm_cap_value(n_val) if n_type == "CAPACITOR" else ""
#
    #    filtered = []
    #    for cand in (candidates or []):
    #        c_parsed = _get_parsed(cand)
    #        c_type = _ptype(c_parsed)
#
    #        # ---------- HARD TYPE GATE ----------
    #        # If parser says both are clear types and they disagree, drop.
    #        if n_type != "OTHER" and c_type != "OTHER" and n_type != c_type:
    #            continue
#
    #        # ---------- HARD PACKAGE GATE (passives especially) ----------
    #        c_pkg_raw = str(c_parsed.get("package") or safe_get(cand, "package", default=""))
    #        c_pkg = _norm_pkg(c_pkg_raw)
#
    #        if n_type in {"RESISTOR", "CAPACITOR", "INDUCTOR"}:
    #            # If both have package, they must match after normalization
    #            if n_pkg and c_pkg and (n_pkg != c_pkg):
    #                continue
#
    #        # ---------- HARD VALUE GATE (only when both parse cleanly) ----------
    #        c_val = str(c_parsed.get("value") or safe_get(cand, "value", default="")).strip()
#
    #        if n_type == "RESISTOR":
    #            c_res = _norm_res_value(c_val)
    #            if n_res and c_res and (n_res != c_res):
    #                continue
#
    #        if n_type == "CAPACITOR":
    #            c_cap = _norm_cap_value(c_val)
    #            if n_cap and c_cap and (n_cap != c_cap):
    #                continue
#
    #        # ---------- SCORING ----------
    #        # Prefer tier-seeded confidence if present, otherwise fall back to object confidence.
    #        base_conf = safe_get(cand, "_pc_seed", default=None)
    #        if base_conf is None:
    #            base_conf = safe_get(cand, "confidence", default=0.0)
#
    #        base_conf = clamp01(base_conf)
#
    #        score = base_conf
#
    #        # Fallback: if we still have basically nothing, use description similarity
    #        if score <= 0.0001:
    #            try:
    #                
    #                n_desc = str(safe_get(npr_part, "desc", "description", default="") or "")
    #                c_desc = str(safe_get(cand, "desc", "description", default="") or "")
    #                if n_desc and c_desc:
    #                    ratio = fuzz.token_set_ratio(n_desc.upper(), c_desc.upper()) / 100.0
    #                    score = max(score, 0.15 + 0.55 * ratio)  # maps into ~[0.15..0.70]
    #            except Exception:
    #                pass
#
    #        # Strong bonuses for exact anchors (these dominate fuzzy noise)
    #        if n_pkg and c_pkg and n_pkg == c_pkg:
    #            score += 0.35
#
    #        if n_type == "RESISTOR" and n_res and _norm_res_value(c_val) == n_res:
    #            score += 0.45
    #        if n_type == "CAPACITOR" and n_cap and _norm_cap_value(c_val) == n_cap:
    #            score += 0.45
#
    #        # Minor bonuses for secondary fields if present
    #        n_tol = str(npr_parsed.get("tolerance") or safe_get(npr_part, "tolerance", default="")).strip().upper()
    #        c_tol = str(c_parsed.get("tolerance") or safe_get(cand, "tolerance", default="")).strip().upper()
    #        if n_tol and c_tol and n_tol == c_tol:
    #            score += 0.08
#
    #        n_v = str(npr_parsed.get("voltage") or safe_get(npr_part, "voltage", default="")).strip().upper()
    #        c_v = str(c_parsed.get("voltage") or safe_get(cand, "voltage", default="")).strip().upper()
    #        if n_v and c_v and n_v == c_v:
    #            score += 0.06
#
    #        s = clamp01(score)
#
    #        # Attach per-candidate score so UI cards can display it (CLAMPED)
    #        setattr(cand, "_pc_score", s)
#
    #        filtered.append((s, cand))
#
#
    #    # Sort best-first
    #    filtered.sort(key=lambda t: t[0], reverse=True)
#
    #    # Return only objects, capped
    #    return [c for _, c in filtered[:max_n]]
#

    #=====================================================
    # CNS matching. this is very experimental and optional. as of now ALOT of it is hard coded and being wired in.
    #=======================================================
    def _build_type_prefix_map(self) -> dict[str, str]:
        print("[CTRL DBG] ENTER _build_type_prefix_map")
        """
        Learn which CNS prefix is most common for each parsed component type,
        using existing inventory parts that already have a company PN.
        """
        counts = defaultdict(Counter)

        for inv in (self.inventory or []):
            pn = getattr(inv, "itemnum", "") or ""
            prefix, body, _ = _split_company_pn(pn)
            if not prefix or not body:
                continue

            p = getattr(inv, "parsed", {}) or {}
            ptype = str(getattr(inv, "_pc_ptype", "") or p.get("type") or "OTHER")
            ptype = ptype.upper().strip()
            counts[ptype][prefix] += 1

        out = {}
        for ptype, ctr in counts.items():
            out[ptype] = ctr.most_common(1)[0][0]

        #print("[CTRL DBG] EXIT _build_type_prefix_map")
        return out
    
    def _suggest_new_body_for_prefix(self, prefix: str) -> str:
        """
        Suggest a new 5-digit body:
          - find max(body) among rows with a non-empty description
          - propose next, skipping over bodies that already exist WITH a description
          - allowing bodies that exist with blank descriptions
        """
        prefix = (prefix or "").strip()
        if not prefix:
            return ""
    
        def field(obj, name: str, default=""):
            # supports dataclass/objects AND dicts
            if isinstance(obj, dict):
                return obj.get(name, default)
            return getattr(obj, name, default)
    
        bodies = {}  # body_int -> has_desc(bool)
    
        for r in (getattr(self, "cns_records", None) or []):
            rp = str(field(r, "prefix", "")).strip()
            if rp != prefix:
                continue
            
            rb = str(field(r, "body", "")).strip()
            if not rb.isdigit():
                continue
            
            desc = str(field(r, "description", "")).strip()
            bodies[int(rb)] = bool(desc)
    
        if not bodies:
            return "10000"
    
        max_desc_body = max((b for b, has_desc in bodies.items() if has_desc), default=max(bodies))
        cand = max_desc_body + 1
    
        while cand in bodies and bodies[cand] is True:
            cand += 1
    
        return f"{cand:05d}"
    
    def _apply_cns_suggestion_to_node(self, node) -> None:
        print(f"[CTRL DBG] ENTER _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)}")
        # 1) If EXISTS: derive from internal PN
        if getattr(node, "internal_part_number", ""):
            prefix, body, _ = _split_company_pn(node.internal_part_number)
            if prefix and body:
                node.suggested_prefix = prefix
                node.suggested_body = body
                node.suggested_pb = f"{prefix}-{body}"
                node.suggested_reason = "from existing inventory PN"

                print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

                return

        # 2) NEW: use parsed type → prefix mapping
        ptype = ""
        # Try to infer from the parsed NPRPart via the node’s description (or store parsed on node later)
        # If you have the NPRPart available during node creation, best: store node.parsed_type directly.
        # For now, we can recover from npr_list by bom_uid if needed; simplest v1: set it during _pair_to_node.
        ptype = getattr(node, "parsed_type", "") or "OTHER"

        prefix = ""
        if hasattr(self, "_type_prefix_map"):
            prefix = self._type_prefix_map.get(ptype, "")

        if not prefix:
            node.suggested_reason = "no prefix suggestion"
            print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

            return

        node.suggested_prefix = prefix

        # 3) NEW body suggestion (since we don’t have a confident body match yet)
        body = self._suggest_new_body_for_prefix(prefix)
        if body:
            node.suggested_body = body
            node.suggested_pb = f"{prefix}-{body}"
            node.suggested_reason = f"from type mapping ({ptype})"


    def _pair_to_node(self, npr_part: Any, match: Any) -> DecisionNode:
        match_type = safe_get(match, "match_type", default="")
        if hasattr(match_type, "value"):
            match_type = match_type.value
        match_type = str(match_type) if match_type else ""

        inv = safe_get(match, "inventory_part", default=None)

        # Determine base_type:
        # - NO_MATCH -> NEW
        # - else -> EXISTS if we have inventory_part
        base_type = "NEW" if match_type == MatchType.NO_MATCH else ("EXISTS" if inv is not None else "NEW")

        # BOM fields
        bom_uid = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default=""))
        bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default=""))
        description = str(safe_get(npr_part, "description", "desc", default=""))

        # Inventory base fields (EXISTS)
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default="")) if inv else ""
        inv_mpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="")) if inv else ""

        confidence = clamp01(safe_get(match, "confidence", default=0.0))

        node_id = bom_uid or (internal_pn or f"NODE-{len(self.nodes)+1}")

        # compute parsed type BEFORE node creation
        parsed = getattr(npr_part, "parsed", {}) or {}
        ptype = str(getattr(npr_part, "_pc_ptype", "") or (parsed.get("type") if isinstance(parsed, dict) else "") or "OTHER")
        ptype = ptype.upper().strip()

        node = DecisionNode(
            id=node_id,
            base_type=base_type,
            bom_uid=bom_uid,
            bom_mpn=bom_mpn,
            description=description,
            internal_part_number=internal_pn,
            inventory_mpn=inv_mpn,
            match_type=match_type,
            confidence=confidence,
            alternates=[],
            status=DecisionStatus.NEEDS_DECISION,
            locked=False,
            needs_approval=False,
            notes="",
            explain={},
        )

        # set dynamic attributes
        node.parsed_type = ptype

        # ----------------------------------------------------
        # Internal Alternates (inventory)
        # Always show the WINNER inventory_part as a Base card, even if candidates exist.
        # ----------------------------------------------------
        candidates = safe_get(match, "candidates", "candidate_parts", default=None) or []

        seen_ids = set()

        if inv is not None:
            base_alt = self._inventory_part_to_alternate(inv, confidence=confidence, relationship="Base/Selected")
            node.alternates.append(base_alt)
            seen_ids.add(base_alt.id)

        for cand in candidates:
            # Prefer per-candidate score; otherwise fall back so candidates never show 0%.
            cand_score = safe_get(cand, "_pc_score", default=None)
            conf = cand_score if cand_score is not None else 0.25
            conf = clamp01(conf)

            via_mpn = (getattr(cand, "_best_source_customer_mpn", "") or "").strip()
            via_mt = (getattr(cand, "_best_source_match_type", "") or "").strip()

            rel = "Internal Candidate"
            if via_mpn and via_mpn != (bom_mpn or "").strip():
                rel = f"Internal Candidate (via {via_mpn})"
            if via_mt:
                rel = f"{rel} [{via_mt}]"

            alt_obj = self._inventory_part_to_alternate(
                cand,
                confidence=conf,
                relationship=rel,
            )
            if alt_obj.id in seen_ids:
                continue
            seen_ids.add(alt_obj.id)
            node.alternates.append(alt_obj)

        # ----------------------------------------------------
        # Customer-provided alternates (external)
        # Show them as external cards in the UI.
        # ----------------------------------------------------
        explain = safe_get(match, "explain", default=None) or {}
        customer_mpns = []
        winning_mpn = ""

        if isinstance(explain, dict):
            customer_mpns = explain.get("customer_mpns", []) or []
            winning_mpn = (explain.get("winning_mpn", "") or "").strip()

        # Fallback if explain wasn’t populated (still works “okay”)
        if not customer_mpns:
            parsed = safe_get(npr_part, "parsed", default=None) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            customer_mpns = [bom_mpn] + list(mpn_alts)

        if not winning_mpn:
            winning_mpn = (bom_mpn or "").strip()

        # --- PATCH: propagate winner explicitly into node.explain ---
        node.explain = dict(safe_get(match, "explain", default={}) or {})
        node.explain["winning_mpn"] = winning_mpn
        for a in node.explain.get("attempts", []):
            a["is_winner"] = (a.get("customer_mpn") == winning_mpn)
        
        # for debugging
        print(f"[CTRL PATCH] node={node.id} winning_mpn={winning_mpn}")


        # Deduplicate against any alternates already present (inventory candidates etc.)
        seen_mpns = set()
        for a in node.alternates:
            m = (getattr(a, "manufacturer_part_number", "") or "").strip()
            if m:
                seen_mpns.add(m)

        for mpn in customer_mpns:
            mpn = (mpn or "").strip()
            if not mpn or mpn == winning_mpn or mpn in seen_mpns:
                continue
            seen_mpns.add(mpn)

            node.alternates.append(
                Alternate(
                    id=f"CUST-{node.id}-{mpn}",
                    source="customer_bom",
                    manufacturer=str(safe_get(npr_part, "mfgname", default="")),
                    manufacturer_part_number=mpn,
                    internal_part_number="",
                    description=description,
                    confidence=0.0,
                    relationship="Customer Provided",
                    selected=False,
                    rejected=False,
                    raw=None,
                    meta={"customer_provided": True},
                )
            )

        self._recompute_node_flags(node)
        return node


    def _inventory_part_to_alternate(self, inv: Any, confidence: float, relationship: str) -> Alternate:
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default=""))
        mfgpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default=""))
        mfg = str(safe_get(inv, "manufacturer", "manufacturer_name", "mfgname", default=""))

        # stable alt id: prefer internal PN; fallback to MPN; else UUID
        alt_id = internal_pn or (f"MPN-{mfgpn}" if mfgpn else Alternate.new_id("INV"))

        return Alternate(
            id=alt_id,
            source="inventory",
            manufacturer=mfg,
            manufacturer_part_number=mfgpn,
            internal_part_number=internal_pn,
            description=str(safe_get(inv, "description", "desc", default="")),
            value=str(safe_get(inv, "value", default="")),
            package=str(safe_get(inv, "package", default="")),
            tolerance=str(safe_get(inv, "tolerance", default="")),
            voltage=str(safe_get(inv, "voltage", default="")),
            wattage=str(safe_get(inv, "wattage", default="")),
            stock=int(safe_get(inv, "stock", "qty", default=0) or 0),
            unit_cost=safe_get(inv, "unit_cost", "cost", default=None),
            supplier=str(safe_get(inv, "supplier", "vendor", default="")),
            confidence=clamp01(confidence),
            relationship=relationship,
            selected=False,
            rejected=False,
            raw=inv,
            meta={},
        )

    # ----------------------------
    # Query helpers for UI
    # ----------------------------

    def get_nodes(self) -> List[DecisionNode]:
        return self.nodes

    def get_node(self, node_id: str) -> DecisionNode:
        for n in self.nodes:
            if n.id == node_id:
                return n
        raise KeyError(f"No DecisionNode with id={node_id}")

    # ----------------------------
    # Mutations (UI actions)
    # ----------------------------

    def select_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.rejected = False
        alt.selected = True

        # ELEVATE: if the selected alternate is inventory-backed, it becomes the anchor
        if getattr(alt, "source", "") == "inventory" and getattr(alt, "internal_part_number", ""):
            node.internal_part_number = alt.internal_part_number
            node.base_type = "EXISTS"
            # optional: keep a convenience copy of the inventory MPN
            node.inventory_mpn = getattr(alt, "manufacturer_part_number", "") or node.inventory_mpn

        self._recompute_node_flags(node)

    def reject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False
        alt.rejected = True

        # If this rejected inventory alternate is currently anchoring the node, clear the anchor
        if (
            getattr(alt, "source", "") == "inventory"
            and getattr(alt, "internal_part_number", "")
            and alt.internal_part_number == (node.internal_part_number or "")
            and float(getattr(node, "confidence", 0.0) or 0.0) < 0.999
        ):
            node.internal_part_number = ""
            node.inventory_mpn = ""

        # ---- DEMOTION RULE ----
        # If there are NO remaining inventory candidates (non-rejected, non-selected),
        # and we have no selected alternates, treat as NEW.
        remaining_inventory_candidates = [
            a for a in node.alternates
            if getattr(a, "source", "") == "inventory"
            and (not getattr(a, "rejected", False))
            and (not getattr(a, "selected", False))
        ]
        
        node.explain = dict(node.explain or {})
        node.explain["pre_demote"] = {
            "base_type": node.base_type,
            "confidence": node.confidence,
        }

        if (len(remaining_inventory_candidates) == 0) and (not node.has_selection()):


            node.base_type = "NEW"
            node.internal_part_number = ""
            node.inventory_mpn = ""
            node.confidence = 0.0
            node.explain = dict(node.explain or {})
            node.explain["demoted_to_new"] = "All inventory candidates were rejected."

        self._recompute_node_flags(node)


    def unreject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)

        # Undo reject
        alt.rejected = False
        alt.selected = False  # keep it simple: unreject ≠ auto-select

        # If we previously demoted to NEW because "all inventory candidates were rejected",
        # restore the pre-demote base_type/confidence when at least one inventory candidate exists again.
        if isinstance(getattr(node, "explain", None), dict):
            pre = node.explain.get("pre_demote")
            if pre and node.explain.get("demoted_to_new"):
                remaining_inventory_candidates = [
                    a for a in node.alternates
                    if getattr(a, "source", "") == "inventory"
                    and (not getattr(a, "rejected", False))
                ]
                if remaining_inventory_candidates:
                    node.base_type = pre.get("base_type", node.base_type)
                    node.confidence = pre.get("confidence", node.confidence)
                    node.explain.pop("demoted_to_new", None)
                    node.explain.pop("pre_demote", None)

        self._recompute_node_flags(node)



    def unselect_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False
        self._recompute_node_flags(node)

    def add_manual_alternate(self, node_id: str, manufacturer_part_number: str, description: str = "Manual alternate") -> Alternate:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = Alternate(
            id=Alternate.new_id("MAN"),
            source="manual",
            manufacturer="",
            manufacturer_part_number=manufacturer_part_number,
            internal_part_number="",
            description=description,
            confidence=0.0,
            relationship="Manual",
        )
        node.alternates.append(alt)
        self._recompute_node_flags(node)
        return alt

    def mark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        node.status = DecisionStatus.READY_FOR_EXPORT
        node.locked = True
        self._recompute_node_flags(node)

    # ----------------------------
    # External search (DigiKey etc.)
    # ----------------------------

    def search_digikey(self, node_id: str) -> List[Alternate]:
        """
        Cache is controller-owned (not UI). UI just calls this.
        """
        node = self.get_node(node_id)

        if node_id in self.external_cache:
            return self.external_cache[node_id]

        if not self._digikey_search_fn:
            # no client wired yet
            self.external_cache[node_id] = []
            return []

        results = self._digikey_search_fn(node) or []
        # Normalize + attach
        for alt in results:
            alt.source = alt.source or "digikey"
        self.external_cache[node_id] = results

        # Option: merge external alternates into node list (recommended)
        # so UI doesn’t need “internal vs external” stores.
        node.alternates.extend(results)
        self._recompute_node_flags(node)
        return results

    # ----------------------------
    # NPR Workspace (from DecisionNodes)
    # ----------------------------

    def build_npr_workspace_from_nodes(self) -> NPRWorkspace:
        ws = NPRWorkspace()

        # -------------------------
        # Page 1 — NEW parts
        # -------------------------
        for node in self.nodes:
            if node.status != DecisionStatus.READY_FOR_EXPORT:
                continue  # 🚫 Only export ready nodes

            if node.base_type.upper() != "NEW":
                continue

            primary = NPRPrimaryNewItem(
                bom_uid=node.bom_uid,
                bom_mpn=node.bom_mpn,
                description=node.description,
                component_type="",
                populated=False,
                include_in_export=True,
                notes=node.notes or "",
            )
            ws.primary_new_items.append(primary)

        # -------------------------
        # Page 2 — Secondary rows
        # -------------------------
        for node in self.nodes:
            if node.status != DecisionStatus.READY_FOR_EXPORT:
                continue  # 🚫 Skip non-ready nodes

            parent_uid = node.bom_uid
            parent_mpn = node.bom_mpn
            parent_desc = node.description

            if node.internal_part_number:
                row_id = f"CTX-{parent_uid}-{node.internal_part_number}"
                ws.secondary_rows.append(
                    NPRSecondaryRow(
                        row_id=row_id,
                        kind=NPRRowKind.CONTEXT_EXISTS,
                        parent_bom_uid=parent_uid,
                        parent_bom_mpn=parent_mpn,
                        parent_description=parent_desc,
                        internal_part_number=node.internal_part_number,
                        exists_in_inventory=True,
                        include_in_export=True,
                        source="inventory",
                    )
                )

            for alt in node.selected_alternates():
                exists = bool(alt.internal_part_number)
                row_id = f"ALT-{parent_uid}-{alt.id}"
                ws.secondary_rows.append(
                    NPRSecondaryRow(
                        row_id=row_id,
                        kind=NPRRowKind.ALTERNATE_PROPOSAL,
                        parent_bom_uid=parent_uid,
                        parent_bom_mpn=parent_mpn,
                        parent_description=parent_desc,
                        internal_part_number=alt.internal_part_number,
                        exists_in_inventory=exists,
                        include_in_export=True,
                        source=alt.source,
                        manufacturer_name=alt.manufacturer,
                        manufacturer_part_number=alt.manufacturer_part_number,
                        supplier=alt.supplier,
                        unit_cost=alt.unit_cost,
                    )
                )

        self.workspace = ws
        return ws

###############################################################################################################
    def export_npr(self, output_path: str = None):
        """Export NPR data to the official Engineering template."""

        # 1. Ensure template exists
        template_path = Path(self.cfg.npr_template_path)
        if not template_path.exists():
            messagebox.showerror("Export Error", f"NPR template not found: {template_path}")
            return

        # 2. Filter for only 'Ready for Export' nodes
        ready_nodes = [n for n in self.nodes if n.status == DecisionStatus.READY_FOR_EXPORT]
        if not ready_nodes:
            messagebox.showwarning("Nothing Ready", "No parts marked Ready for Export.")
            return

        # 3. Load workbook & select Engineering sheet
        wb = load_workbook(template_path)
        if "Engineering" not in wb.sheetnames:
            messagebox.showerror("Template Error", "Template missing 'Engineering' sheet.")
            return
        ws = wb["Engineering"]

        # 4. Fill NPR metadata (top of sheet)
        ws["B1"] = ""
        ws["B2"] = ""
        ws["B3"] = ""
        ws["B3"] = ""
        ws["B5"] = datetime.now().strftime("%Y-%m-%d")
        ws["B6"] = self.cfg.created_by

        # 5. Write parts data starting below the header block
        start_row = 10

        def uniq_keep_order(items):
            seen = set()
            out = []
            for x in items:
                if x and x not in seen:
                    seen.add(x)
                    out.append(x)
            return out

        def is_popcorn_pn(pn: str) -> bool:
            pn = (pn or "").strip()
            return pn.startswith(("06-", "21-", "24-"))

        def find_inventory_alt(node: DecisionNode, internal_pn: str) -> Optional[Alternate]:
            if not internal_pn:
                return None
            for a in node.alternates:
                if getattr(a, "source", "") == "inventory" and getattr(a, "internal_part_number", "") == internal_pn:
                    return a
            return None

        # Collect approval rows while exporting Engineering rows
        approval_rows: List[List[Any]] = []

        def is_full_match_anchor(alt_obj) -> bool:
            # Best-effort heuristics (same as your original intent)
            rel = (getattr(alt_obj, "relationship", "") or "").strip().lower()
            conf = float(getattr(alt_obj, "confidence", 0.0) or 0.0)
            return (rel == "exact") or (conf >= 0.999)

        out_row = start_row  # <-- important: since we may skip some rows now

        for node in ready_nodes:
            anchored_existing = bool(node.internal_part_number)

            # -------------------------
            # Detect primary (anchor) alt & whether anchor is full match
            # -------------------------
            primary_alt = None
            for alt in node.selected_alternates():
                if getattr(alt, "internal_part_number", "") and alt.internal_part_number == node.internal_part_number:
                    primary_alt = alt
                    break

            anchor_is_full_match = False
            if anchored_existing and primary_alt:
                anchor_is_full_match = is_full_match_anchor(primary_alt)
            elif anchored_existing and node.base_type == "EXISTS":
                # If it loaded as EXISTS and we don't have a selected-alt record,
                # treat as full match (conservative default).
                anchor_is_full_match = True

            # -------------------------
            # Compute "non-primary selected alternates"
            # (what you mean by "has alternates selected for it")
            # -------------------------
            non_primary_selected = []
            for alt in node.selected_alternates():
                is_primary = bool(getattr(alt, "internal_part_number", "")) and (
                    alt.internal_part_number == node.internal_part_number
                )
                if not is_primary:
                    non_primary_selected.append(alt)

            has_alternates_selected = len(non_primary_selected) > 0

            # -------------------------
            # RULE: If full match AND no alternates selected -> DO NOT write into Engineering
            # but DO write into Approvals sheet.
            # -------------------------
            skip_engineering_row = bool(anchored_existing and anchor_is_full_match and (not has_alternates_selected))

            # -------------------------
            # Always add anchor inventory part to approvals (unless popcorn)
            # - If it's a pure full match with no alternates: track it (FULL_MATCH)
            # - If it has alternates selected: treat like substitution universe (SUBSTITUTE)
            # -------------------------
            if anchored_existing and (not is_popcorn_pn(node.internal_part_number)):
                inv_alt = find_inventory_alt(node, node.internal_part_number)
                approval_rows.append(
                    [
                        node.bom_uid,
                        node.bom_mpn,
                        node.description or "",
                        node.internal_part_number,
                        (getattr(inv_alt, "manufacturer_part_number", "") if inv_alt else ""),
                        "inventory",
                        (getattr(inv_alt, "manufacturer", "") if inv_alt else ""),
                        (getattr(inv_alt, "supplier", "") if inv_alt else ""),
                        node.notes or "",
                    ]
                )


            # Also add any OTHER selected inventory alternates to approvals (unless popcorn)
            # Also add any OTHER selected inventory alternates to approvals (unless popcorn)
            for alt in non_primary_selected:
                if getattr(alt, "source", "") != "inventory":
                    continue
                
                ipn = (getattr(alt, "internal_part_number", "") or "").strip()
                if (not ipn) or is_popcorn_pn(ipn):
                    continue
                
                approval_rows.append(
                    [
                        node.bom_uid,
                        node.bom_mpn,
                        node.description or "",
                        ipn,
                        (getattr(alt, "manufacturer_part_number", "") or ""),
                        "inventory",
                        (getattr(alt, "manufacturer", "") or ""),
                        (getattr(alt, "supplier", "") or ""),
                        node.notes or "",
                    ]
                )


            if skip_engineering_row:
                # Full match with no alternates: don’t clutter Engineering NPR
                continue

            # -------------------------
            # NPR sheet row expansion:
            #   - If L is empty, fill it with the first available MPN.
            #   - If L already has something, emit a NEW ROW (duplicate A/B/C/R) for each additional alternate MPN.
            # -------------------------

            # Company PN (anchor)
            if anchored_existing:
                company_pn = node.internal_part_number
            else:
                company_pn = getattr(node, "assigned_part_number", "") or "NEW"

            # Collect NON-inventory alternates as separate MPN candidates (inventory ones are tracked in approvals sheet)
            alt_mpns: List[str] = []
            for alt in node.selected_alternates():
                if getattr(alt, "source", "") == "inventory":
                    continue
                mpn = (getattr(alt, "manufacturer_part_number", "") or "").strip()
                if mpn:
                    alt_mpns.append(mpn)

            # De-dupe while preserving order
            alt_mpns = uniq_keep_order(alt_mpns)

            # Decide the first row's L value:
            # - Prefer BOM MPN
            # - If BOM MPN empty, use first alternate MPN (if any)
            bom_mpn = (getattr(node, "bom_mpn", "") or "").strip()
            if bom_mpn:
                first_l = bom_mpn
            else:
                first_l = alt_mpns.pop(0) if alt_mpns else ""

            def write_row(row_idx: int, l_value: str):
                ws[f"A{row_idx}"] = company_pn
                ws[f"B{row_idx}"] = "EXISTING" if anchored_existing else "NEW"
                ws[f"C{row_idx}"] = node.description or ""
                ws[f"L{row_idx}"] = l_value
                ws[f"R{row_idx}"] = node.notes or ""

            # Write the primary row
            write_row(out_row, first_l)
            out_row += 1

            # Write one additional NPR row per remaining alternate MPN
            for mpn in alt_mpns:
                write_row(out_row, mpn)
                out_row += 1

        # -------------------------
        # 5b. Create/refresh Alternates (Approvals) sheet AFTER Engineering rows are written
        # -------------------------
        sheet_name = "Alternates for Approval"
        if sheet_name in wb.sheetnames:
            ws_alt = wb[sheet_name]
        else:
            ws_alt = wb.create_sheet(sheet_name)

        headers = [
            "BOM_UID",
            "BOM_MPN",
            "BOM_DESC",
            "INTERNAL_PN",
            "INVENTORY_MPN",
            "SOURCE",
            "MFG",
            "SUPPLIER",
            "NOTES",
        ]

        # Clear sheet (keep header row)
        if ws_alt.max_row > 1:
            ws_alt.delete_rows(2, ws_alt.max_row - 1)

        for col_idx, h in enumerate(headers, start=1):
            ws_alt.cell(row=1, column=col_idx, value=h)

        for r_idx, row in enumerate(approval_rows, start=2):
            for c_idx, v in enumerate(row, start=1):
                ws_alt.cell(row=r_idx, column=c_idx, value=v)

        # 6. Determine output path
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(template_path).parent / f"NPR_Export_{timestamp}.xlsx"

        # 7. Save workbook
        wb.save(output_path)
        messagebox.showinfo("NPR Export Complete", f"Exported NPR to:\n{output_path}")

        return output_path

    # ----------------------------
    # Internal enforcement
    # ----------------------------

    def _find_alt(self, node: DecisionNode, alt_id: str) -> Alternate:
        for a in node.alternates:
            if a.id == alt_id:
                return a
        raise KeyError(f"No alternate with id={alt_id} on node={node.id}")

    def _ensure_unlocked(self, node: DecisionNode) -> None:
        if node.locked:
            raise PermissionError("Node is locked (READY_FOR_EXPORT). Unlocking not supported.")
        

    def _recompute_node_flags(self, node: DecisionNode) -> None:
        """
        Central place for status/approval heuristics.
        UI should just display these.
        """
        bt = node.base_type.upper()
        selected = len(node.selected_alternates())
        candidates = len(node.candidate_alternates())

        # Needs approval heuristic
        if bt == "NEW":
            node.needs_approval = True
        elif bt == "EXISTS":
            node.needs_approval = selected > 0
        else:
            node.needs_approval = selected > 0

        # Status heuristic unless already READY/locked
        if node.locked:
            node.status = DecisionStatus.READY_FOR_EXPORT
            return
    

        if bt == "NEW" and selected == 0:
            node.status = DecisionStatus.NEEDS_ALTERNATE
            return

        if bt == "EXISTS" and candidates > 0 and selected == 0:
            node.status = DecisionStatus.NEEDS_DECISION
            return

        # default: auto-ish state (still editable)
        node.status = DecisionStatus.FULL_MATCH
