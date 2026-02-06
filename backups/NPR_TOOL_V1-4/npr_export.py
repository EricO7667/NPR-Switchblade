from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True)
class NPRExportMetadata:
    npr_number: str
    ecn_number: str
    project_number: str
    project_name: str
    author: str
    date: datetime


def _find_cell_with_text(ws, text: str) -> Optional[Tuple[int, int]]:
    target = (text or "").strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == target:
                return (cell.row, cell.column)
    return None


def _write_header_value_next_to_label(ws, label: str, value: Any) -> None:
    pos = _find_cell_with_text(ws, label)
    if not pos:
        raise ValueError(f"Header label '{label}' not found in template.")
    r, c = pos
    ws.cell(row=r, column=c + 1).value = value


def _find_table_header_row(ws, required_headers: Iterable[str]) -> int:
    required = {h.strip().lower() for h in required_headers}

    for r in range(1, ws.max_row + 1):
        found = set()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                found.add(v.strip().lower())
        if required.issubset(found):
            return r

    raise ValueError("Could not locate the parts table header row in template.")


def get_template_headers(template_path: str | Path, sheet_name: Optional[str] = None) -> List[str]:
    template_path = Path(template_path)
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = _find_table_header_row(ws, ["Part Number", "Item Description"])
    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            headers.append(v.strip())
    return headers


def export_npr_from_rows(
    template_path: str | Path,
    output_path: str | Path,
    metadata: Optional[NPRExportMetadata],
    excel_rows: list[dict],
    sheet_name: Optional[str] = None,
):

    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Header section
    if metadata:
        _write_header_value_next_to_label(ws, "NPR #", metadata.npr_number)
        _write_header_value_next_to_label(ws, "ECN #", metadata.ecn_number)
        _write_header_value_next_to_label(ws, "PROJECT #", metadata.project_number)
        _write_header_value_next_to_label(ws, "PROJECT NAME", metadata.project_name)
        _write_header_value_next_to_label(ws, "DATE", metadata.date.strftime("%Y-%m-%d"))
        _write_header_value_next_to_label(ws, "AUTHOR", metadata.author)

    # Table section
    header_row = _find_table_header_row(ws, ["Part Number", "Item Description"])
    data_start_row = header_row + 1

    # Build column index
    col_index: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            col_index[v.strip()] = c

    # Write values (do NOT touch formatting)
    for i, row_dict in enumerate(excel_rows):
        r = data_start_row + i

        # Clear only the values in that row (keep styles)
        for header, c in col_index.items():
            ws.cell(row=r, column=c).value = None

        # Write provided values
        for header, value in row_dict.items():
            if header not in col_index:
                continue
            ws.cell(row=r, column=col_index[header]).value = value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_excel_rows_from_workspace(
    workspace,
    template_headers: list[str],
) -> list[dict]:
    """
    Convert NPRWorkspace into Excel-ready row dictionaries.
    This function owns ALL template/header semantics.
    """
    rows: list[dict] = []

    # Page 1 — NEW parts
    for item in workspace.primary_new_items:
        if not item.include_in_export:
            continue

        d = {h: "" for h in template_headers}

        if "Part Status (new / exists)" in d:
            d["Part Status (new / exists)"] = "NEW"
        if "Item Description" in d:
            d["Item Description"] = item.description
        if "Manufacturer Part #" in d:
            d["Manufacturer Part #"] = item.bom_mpn

        rows.append(d)

    # Page 2 — Context + Alternates
    for row in workspace.get_export_selected_secondary_rows():
        d = {h: "" for h in template_headers}

        if "Part Number" in d:
            d["Part Number"] = (
                row.internal_part_number
                if row.kind.name == "CONTEXT_EXISTS"
                else ""
            )

        if "Part Status (new / exists)" in d:
            d["Part Status (new / exists)"] = (
                "EXISTS" if row.exists_in_inventory else "NEW"
            )

        if "Item Description" in d:
            d["Item Description"] = row.parent_description

        for flag in ("SMT", "TH", "Process", "Assembly", "PCB", "Mechanical"):
            if flag in d:
                d[flag] = "X" if row.flags.get(flag, False) else ""

        if row.kind.name == "ALTERNATE_PROPOSAL":
            if "Manufacturer Name" in d:
                d["Manufacturer Name"] = row.manufacturer_name
            if "Manufacturer Part #" in d:
                d["Manufacturer Part #"] = row.manufacturer_part_number
            if "Supplier" in d:
                d["Supplier"] = row.supplier
            if "Unit Cost" in d:
                d["Unit Cost"] = row.unit_cost or ""
            if "Stock Unit" in d:
                d["Stock Unit"] = row.stock_unit
            if "Lead Time (WKS)" in d:
                d["Lead Time (WKS)"] = row.lead_time_weeks or ""
            if "QC Required" in d:
                d["QC Required"] = (
                    "YES" if row.qc_required else "NO"
                    if row.qc_required is not None else ""
                )
            if "TARIFF CODE (HTSUS)" in d:
                d["TARIFF CODE (HTSUS)"] = row.tariff_code

        rows.append(d)

    return rows
