import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class PartCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PartChecker Pro v1.0 - Highlight Edition")
        self.root.geometry("800x600")
        self.root.resizable(False, False)

        self.db_path = None
        self.npr_path = None
        self.results_df = None

        self.setup_ui()

    def setup_ui(self):
        title = tk.Label(self.root, text="🔍 PartChecker Pro v1.0", font=("Segoe UI", 16, "bold"))
        title.pack(pady=10)

        self.db_label = tk.Label(self.root, text="Database Excel (ItemNumber): Not Loaded", fg="gray")
        self.db_label.pack()
        self.db_btn = tk.Button(self.root, text="📂 Load Database File", command=self.load_db)
        self.db_btn.pack(pady=5)

        self.npr_label = tk.Label(self.root, text="NPR Excel (Part Number): Not Loaded", fg="gray")
        self.npr_label.pack()
        self.npr_btn = tk.Button(self.root, text="📂 Load NPR File", command=self.load_npr)
        self.npr_btn.pack(pady=5)

        self.compare_btn = tk.Button(self.root, text="▶️ Compare Parts", command=self.compare_files, state="disabled")
        self.compare_btn.pack(pady=10)

        self.stats_label = tk.Label(self.root, text="✅ Found: 0   ❌ Missing: 0   Total: 0", font=("Segoe UI", 11))
        self.stats_label.pack(pady=10)

        # Table setup
        self.tree = ttk.Treeview(self.root, columns=("Part Number", "Exists"), show="headings", height=20)
        self.tree.heading("Part Number", text="Part Number")
        self.tree.heading("Exists", text="Exists")
        self.tree.column("Part Number", width=400)
        self.tree.column("Exists", width=150, anchor="center")

        style = ttk.Style()
        style.configure("Treeview", rowheight=22, font=("Segoe UI", 10))
        style.map('Treeview', background=[('selected', '#CCE5FF')])
        self.tree.pack(pady=10)

        self.export_btn = tk.Button(self.root, text="💾 Export Results to Excel", command=self.export_results, state="disabled")
        self.export_btn.pack(pady=10)

    def load_db(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.db_path = file_path
            self.db_label.config(text=f"Database File Loaded: {file_path.split('/')[-1]}", fg="green")
            self.check_ready()

    def load_npr(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.npr_path = file_path
            self.npr_label.config(text=f"NPR File Loaded: {file_path.split('/')[-1]}", fg="green")
            self.check_ready()

    def check_ready(self):
        if self.db_path and self.npr_path:
            self.compare_btn.config(state="normal")

    def compare_files(self):
        try:
            # === Load database and NPR ===
            db = pd.read_excel(self.db_path, dtype=str)
            npr_raw = pd.read_excel(self.npr_path, header=None, dtype=str)

            # === Find "Part Number" row in first column (first 20 rows only) ===
            part_start_row = None
            for i in range(min(20, len(npr_raw))):
                cell_value = str(npr_raw.iloc[i, 0]).strip().lower()
                if "part number" in cell_value:
                    part_start_row = i + 1  # next row after header
                    break

            if part_start_row is None:
                messagebox.showerror("Error", "Could not find 'Part Number' within first 20 rows of NPR file.")
                return

            # === Extract part numbers from below the header ===
            part_numbers = npr_raw.iloc[part_start_row:, 0].dropna().astype(str).str.strip().tolist()

            # === Compare against database ===
            if "ItemNumber" not in db.columns:
                messagebox.showerror("Error", "Database missing 'ItemNumber' column.")
                return

            db_filtered = db[db["ItemNumber"].astype(str).isin(part_numbers)]

            found_count = len(db_filtered)
            total = len(part_numbers)
            missing_count = total - found_count

            self.results_df = db_filtered

            self.stats_label.config(
                text=f"✅ Found: {found_count}   ❌ Missing: {missing_count}   Total Checked: {total}"
            )

            # === Clear table ===
            for row in self.tree.get_children():
                self.tree.delete(row)

            # === Populate table with found rows only ===
            for _, row in db_filtered.iterrows():
                self.tree.insert("", "end", values=(row["ItemNumber"], "Found"), tags=("found",))

            self.tree.tag_configure("found", background="#C6EFCE")
            self.export_btn.config(state="normal")

        except Exception as e:
            messagebox.showerror("Error", f"Error comparing files:\n{e}")

    def export_results(self):
        if self.results_df is None:
            messagebox.showerror("Error", "No results to export.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not save_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["Part Number", "Exists"])

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for _, row in self.results_df.iterrows():
                ws.append([row["Part Number"], "Found" if row["Exists"] else "Missing"])
                cell_fill = green_fill if row["Exists"] else red_fill
                ws.cell(ws.max_row, 1).fill = cell_fill
                ws.cell(ws.max_row, 2).fill = cell_fill

            wb.save(save_path)
            messagebox.showinfo("Export Complete", f"✅ Results exported successfully!\nFile saved at:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting results:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PartCheckerApp(root)
    root.mainloop()




# the general idea will be to make a general purpose NPR tool. the idea will be to have an NPR switchblade tool. The comapny has inventory of all their items.
# Each item in inventoy has an ItemNumber Description	PrimaryVendorNumber	VendorItem	MfgId	MfgName	MfgItemCount	LastCost	StdCost	AvgCost	Revision	ItemLeadTime	DefaultWhse	TotalQty
# when making a new parts order customers will provide to us their BOMs, which we MAY or MAYNOT have in stock. We have to check what we DO and DO NOT have in stock.
# this involves formatting their BOMS and then using our own NPR forming sheet. we will be ocusing on the NPR portion onot the BOM portion.
# The excel file that contains out inventry is called items.xlsx. our NPR form will contain a descrition, item partnumber, and the vendor number as well. 
# the problem is: cross checking our database with their needs is tedious and time consuming. I want to quickly check our database for if we have those parts or not
# this saves us time on having to create a new inventory part number for parts that already exist by accedent or by over looking. the descritions on alot of these parts are very simlar,
# but many of the descritpions are not. heres the logic. in our NPR we will check if we have a manufactuing part number in the npr.in the Items iventory. if we do then its an exct match.
# we will write into that line of our npr that the part number exists, then place the prt number into the partnumber slot, and replace the descrition in our NPR with the inventory part number description.
# IF we made an inventory part number for the itme int he NPR before checking the inventory for that part number, we will cross check that partnumber with the inventory part numbers to see IF that partnumber exsists.
# this does not necesarily mean that the part exists, just that the part number does, it could be a different part in totatlly from our NPR to the items list (unless the manufacturing part numbers are the same). in this case, we need to then parse
# and cross check the descriptions of the parts to see if they match. if the descritions match we wil have it marked off has passed the first two checks. but manual checking of the part is still needed in that case. 
# at the end of the day i guess we are only really speeding up the process of checking manufacturing part numbers in lists. what do you think about this idea any thoughts?
