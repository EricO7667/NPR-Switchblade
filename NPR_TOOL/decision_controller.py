
from __future__ import annotations

from dataclasses import dataclass, field, asdict, is_dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from collections import defaultdict, Counter
from .data_loader import DataLoader
from .matching_engine import MatchingEngine
from .parsing_engine import parse_description
from .config_loader import load_config
from .data_models import (
    MatchType, MatchResult, Alternate, DecisionNode, DecisionStatus,
    DecisionCard, CardDisplay, CardSection, CardState, NodeHeaderState, HeaderControlState,
    CardDetailState, CommittedExportState,
)
from types import SimpleNamespace
from copy import copy as _copy_style
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table

from .db import connect_db, init_db, DBConfig
from .repositories import WorkspaceRepo, BomRepo, CompanyPartRepo, InventoryImportRepo, DecisionNodeRepo, DecisionAltRepo

from openpyxl import load_workbook
import os
import copy
from tkinter import messagebox
import re
import json
import threading
from .npr_workspace import (
    NPRWorkspace,
    append_external_alternates,
    build_fake_external_alt_specs,
    external_alt_specs_from_node,
    restore_external_alternates,
)
import time


# ----------------------------
# Controller
# ----------------------------

@dataclass
class ControllerConfig:
    """Simplified configuration for NPR Controller."""

    # Base configuration
    components_yaml_path: str = "./config/components.yaml"

    # Clean, absolute template path (uses os + Path)
    npr_template_path: str = str(
        Path(os.path.dirname(os.path.abspath(__file__))) / "NPR_Master2023_v4_FormTEMPLATECOPY.xlsx"
    )

    # BOM template (used for BOM export sheet formatting/sections)
    bom_template_path: str = str(
        Path(os.path.dirname(os.path.abspath(__file__))) / "BOM_TEMPLATE.xlsx"
    )

    # Metadata still needed for export
    created_by: str = "NPR Tool"

    PB_RE = re.compile(r"^\s*(\d{2})-(\d{5})\s*$")
    _PN_RE = re.compile(r"^(?P<prefix>\d{2})-(?P<body>\d{5})-(?P<suffix>[A-Za-z0-9]{4})$")

    # ----------------------------
    # Candidate limiting +  filters
    # ----------------------------
    MAX_INTERNAL_CANDIDATES = 10

    _IMPERIAL_PKGS = {"01005", "0201", "0402", "0603", "0805", "1206", "1210", "1812", "2010", "2512"}
    _METRIC_TO_IMPERIAL = {
        "0402": "01005",  # 0402 metric ~= 01005 imperial
        "0603": "0201",   # 0603 metric ~= 0201 imperial
        "1005": "0402",   # 1005 metric ~= 0402 imperial
        "1608": "0603",   # 1608 metric ~= 0603 imperial
        "2012": "0805",   # 2012 metric ~= 0805 imperial
        "3216": "1206",   # 3216 metric ~= 1206 imperial
        "3225": "1210",   # 3225 metric ~= 1210 imperial
        "4532": "1812",   # 4532 metric ~= 1812 imperial
        "5025": "2010",   # 5025 metric ~= 2010 imperial
        "6332": "2512",   # 6332 metric ~= 2512 imperial
    }


@dataclass
class InventoryPart:
    """Lightweight controller-local inventory view used by matching/UI code."""

    itemnum: str
    desc: str = ""
    mfgid: str = ""
    mfgname: str = ""
    vendoritem: str = ""
    substitutes: List[str] = field(default_factory=list)
    raw_fields: Dict[str, Any] = field(default_factory=dict)
    parsed: Dict[str, Any] = field(default_factory=dict)
    api_data: Any = None


# ----------------------------
# Helpers
# ----------------------------

def safe_get(obj: Any, *names: str, default: Any = "") -> Any:
    for n in names:
        if obj is None:
            return default
        if hasattr(obj, n):
            v = getattr(obj, n)
            if v is not None:
                return v
    return default

def _split_company_pn(pn: str) -> tuple[str, str, str]:
    pn = (pn or "").strip()
    m = ControllerConfig._PN_RE.match(pn)
    if not m:
        return "", "", ""
    return m.group("prefix"), m.group("body"), m.group("suffix")

def clamp01(x: float) -> float:
    try:
        x = float(x)
    except Exception:
        return 0.0
    if x < 0.0:
        return 0.0
    if x > 1.0:
        return 1.0
    return x


class NPRWorkbookExporter:
    """Thin export service wrapper around the controller-owned export pipeline."""

    def __init__(self, controller: "DecisionController"):
        self.controller = controller

    def export(self, output_path: str = None):
        return self.controller._export_npr_impl(output_path)



class DecisionCardFactory:
    """Create explicit UI card objects from controller-owned DecisionNode/Alternate state."""

    STOCK_KEYS = ("totalqty", "total_qty", "qty_on_hand", "on_hand", "quantity")

    def build_cards(self, node: DecisionNode, focused_card_id: Optional[str] = None) -> List[DecisionCard]:
        cards: List[DecisionCard] = []
        for alt in list(getattr(node, "alternates", []) or []):
            cards.append(self.build_card(node=node, alt=alt, focused_card_id=focused_card_id))
        return cards

    def build_card(self, node: DecisionNode, alt: Alternate, focused_card_id: Optional[str] = None) -> DecisionCard:
        is_inventory = (getattr(alt, "source", "") == "inventory")
        is_rejected = bool(getattr(alt, "rejected", False))
        is_selected = bool(getattr(alt, "selected", False))
        card_id = f"CARD::{getattr(node, 'id', '')}::{getattr(alt, 'id', '')}"
        is_pinned = bool(focused_card_id and focused_card_id == card_id)

        section = self._section_for_alt(alt)
        confidence_ratio = self._confidence_for_alt(alt)
        title, subtitle = self._title_parts(alt)
        description = (getattr(alt, "description", "") or "").strip() or "(no description)"
        if len(description) > 92:
            description = description[:89].rstrip() + "..."

        border_role = "default"
        if is_selected:
            border_role = "selected"
        elif is_rejected:
            border_role = "rejected"
        elif is_pinned:
            border_role = "pinned"

        badges: List[str] = []
        if is_selected:
            badges.append("LOCKED IN")
        elif is_rejected:
            badges.append("REJECTED")
        elif is_pinned:
            badges.append("VIEWING")

        try:
            mfgpn_count = int(((getattr(alt, "meta", {}) or {}).get("company_pn_mfgpn_count", 0) or 0))
        except Exception:
            mfgpn_count = 0

        display = CardDisplay(
            title=title or "—",
            subtitle=subtitle,
            description=description,
            source_label=str(getattr(alt, "source", "") or "-"),
            stock_label=self._stock_label(alt),
            confidence_ratio=confidence_ratio,
            confidence_text=f"{int(confidence_ratio * 100)}%",
            border_role=border_role,
            badges=badges,
            mfgpn_count=mfgpn_count,
        )

        state = CardState(
            selected=is_selected,
            rejected=is_rejected,
            pinned=is_pinned,
            locked=bool(getattr(node, "locked", False)),
            visible=True,
            actionable=not bool(getattr(node, "locked", False)),
        )

        return DecisionCard(
            card_id=card_id,
            node_id=getattr(node, "id", ""),
            alt_id=getattr(alt, "id", ""),
            section=section,
            source=str(getattr(alt, "source", "") or ""),
            is_inventory=is_inventory,
            company_part_number=(getattr(alt, "internal_part_number", "") or "").strip(),
            manufacturer_part_number=(getattr(alt, "manufacturer_part_number", "") or "").strip(),
            manufacturer=(getattr(alt, "manufacturer", "") or "").strip(),
            relationship=(getattr(alt, "relationship", "") or "").strip(),
            state=state,
            display=display,
            alternate=alt,
            meta=dict(getattr(alt, "meta", {}) or {}),
        )

    def _section_for_alt(self, alt: Alternate) -> CardSection:
        if bool(getattr(alt, "rejected", False)):
            return CardSection.REJECTED
        if (getattr(alt, "source", "") or "") == "inventory":
            return CardSection.INTERNAL_MATCHES
        return CardSection.EXTERNAL_ALTERNATES

    def _confidence_for_alt(self, alt: Alternate) -> float:
        conf = float(getattr(alt, "confidence", 0.0) or 0.0)
        if (getattr(alt, "source", "") != "inventory") and conf == 0.0 and not bool(getattr(alt, "rejected", False)):
            conf = 1.0
        return clamp01(conf)

    def _stock_label(self, alt: Alternate) -> str:
        raw = getattr(alt, "raw", None)
        raw_fields = getattr(raw, "raw_fields", {}) or {}
        for key in self.STOCK_KEYS:
            value = raw_fields.get(key)
            if value not in (None, ""):
                return str(value)
        value = getattr(alt, "stock", None)
        return "-" if value in (None, "") else str(value)

    def _title_parts(self, alt: Alternate) -> tuple[str, str]:
        is_inventory = (getattr(alt, "source", "") == "inventory")
        company_pn = (getattr(alt, "internal_part_number", "") or "").strip()
        mfg_pn = (getattr(alt, "manufacturer_part_number", "") or "").strip()
        mfg_name = (getattr(alt, "manufacturer", "") or "").strip()

        rep_vendoritem = str(((getattr(alt, "meta", {}) or {}).get("company_pn_rep_vendoritem", "") or "")).strip()
        matched_ui = (getattr(alt, "_matched_mpn_ui", "") or "").strip()

        title = company_pn if is_inventory else (mfg_pn or company_pn or "(no part number)")
        subtitle = (matched_ui or rep_vendoritem or mfg_pn) if is_inventory else mfg_name
        return title, subtitle


class DecisionController:
    """
    Headless application controller.

    - Owns inventory, BOM/NPR list, match pairs, DecisionNodes
    - Enforces workflow rules (select/reject/ready/locked)
    - Owns external search caching (DigiKey etc.)
    - Builds NPRWorkspace and exports

    UI should ONLY talk to this class.
    """

    def __init__(
        self,
        cfg: Optional[ControllerConfig] = None,
        digikey_search_fn: Optional[Callable[[DecisionNode], List[Alternate]]] = None,
        stop_event: Optional[threading.Event] = None
    ):
        self.cfg = cfg or ControllerConfig()
        

        # Core data
        self.inventory: List[Any] = []
        self.npr_list: List[Any] = []
        self.match_pairs: List[Tuple[Any, Any]] = []  # (npr_part, match_result)

        # Pending inventory snapshot paths (may be loaded in any order)
        self._pending_inventory_master_path: Optional[str] = None
        self._pending_inventory_erp_path: Optional[str] = None
        
        # Pending inventory snapshot rows staged in-memory during imports
        self._pending_inventory_company_rows: list[Any] = []
        self._pending_master_rows: list[Any] = []
        self._pending_erp_rows: list[Any] = []

        # ----------------------------
        # DB persistence (workspace)
        # ----------------------------
        self.conn = connect_db(DBConfig())          # uses default ~/.npr_tool/npr.db
        init_db(self.conn)

        self.ws_repo = WorkspaceRepo(self.conn)
        self.bom_repo = BomRepo(self.conn)
        self.company_part_repo = CompanyPartRepo(self.conn)
        self.inventory_import_repo = InventoryImportRepo(self.conn)
        self.inv_company_repo = self.company_part_repo
        self.decision_node_repo = DecisionNodeRepo(self.conn)
        self.decision_alt_repo = DecisionAltRepo(self.conn)

        # Workspace identity (ONLY authoritative session state)
        self.workspace_id: Optional[str] = None

        # Map BOM uid -> bom_row_id (built during load_npr / load_bom)

        self._bom_row_by_uid: Dict[str, str] = {}


        # Decisions
        self.nodes: List[DecisionNode] = []

        # External cache: node_id -> alternates
        self.external_cache: Dict[str, List[Alternate]] = {}
        self.stop_event = stop_event or threading.Event()

        # Workspace
        self.workspace: Optional[NPRWorkspace] = None

        self._inv_by_itemnum: Dict[str, InventoryPart] = {}
        self._company_parts_by_cpn: Dict[str, Any] = {}
        self._alt_mpn_to_base: Dict[str, List[str]] = {}
        self._alt_loaded: bool = False

        self._erp_inventory: list[InventoryPart] = []
        self._erp_by_itemnum: dict[str, InventoryPart] = {}
        self._erp_loaded: bool = False

        # Derived / cached (NOT authoritative)
        self._inventory_cache = None
        self._views_cache: list[DecisionNode] = []

        self._digikey_search_fn = digikey_search_fn

        # Controller-owned builder for state-driven UI cards.
        self.card_factory = DecisionCardFactory()

        # Controller-owned header state snapshots for the upper panel.
        self._header_state_by_node: Dict[str, NodeHeaderState] = {}

        # Global inventory store workspace (single current master+ERP state, updated on each import)
        self._inventory_store_workspace_id = "__CURRENT_INVENTORY__"


    # ----------------------------
    # Loaders
    # ----------------------------


    def _ensure_inventory_store_workspace(self) -> str:
        """Ensure a hidden singleton workspace row exists for the current inventory store."""
        wsid = self._inventory_store_workspace_id
        try:
            row = self.ws_repo.get(wsid)
        except Exception:
            row = None
        if row:
            return wsid

        now = datetime.utcnow().isoformat(timespec="seconds")
        self.conn.execute(
            """
            INSERT OR IGNORE INTO workspace (
                workspace_id, name, status, created_at, updated_at,
                label, bom_source_path, inventory_master_path, inventory_erp_path, cns_path, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                wsid,
                "[SYSTEM] Current Inventory",
                "SYSTEM",
                now,
                now,
                "",
                "",
                "",
                "",
                "",
                "Do not open manually. Current master+ERP inventory store.",
            ),
        )
        self.conn.commit()
        return wsid

    def _dataclass_to_dict(self, obj: Any) -> Dict[str, Any]:
        if obj is None:
            return {}
        if isinstance(obj, dict):
            return dict(obj)
        if is_dataclass(obj):
            return asdict(obj)
        if hasattr(obj, "__dict__"):
            return dict(vars(obj))
        out: Dict[str, Any] = {}
        for name in dir(obj):
            if name.startswith("_"):
                continue
            try:
                value = getattr(obj, name)
            except Exception:
                continue
            if callable(value):
                continue
            out[name] = value
        return out

    def _sync_current_inventory_store(self, source_label: str = "INVENTORY") -> None:
        """Persist the current inventory snapshot using the active repository/schema layer."""
        wsid = self._ensure_inventory_store_workspace()
        now = datetime.utcnow().isoformat(timespec="seconds")
        self.conn.execute(
            "UPDATE workspace SET updated_at = ?, inventory_master_path = ?, inventory_erp_path = ? WHERE workspace_id = ?",
            (now, self._pending_inventory_master_path or "", self._pending_inventory_erp_path or "", wsid),
        )
        self.conn.commit()

        erp_rows = [self._dataclass_to_dict(r) for r in (self._pending_erp_rows or [])]
        master_rows = [self._dataclass_to_dict(r) for r in (self._pending_master_rows or [])]
        company_parts = list(self._pending_inventory_company_rows or [])

        stats = self.inventory_import_repo.replace_inventory_snapshot(
            wsid,
            erp_rows=erp_rows,
            master_rows=master_rows,
            company_parts=company_parts,
        )
        try:
            print(
                f"[DB][{(source_label or 'INVENTORY').upper()}] inventory sync:",
                f"company_parts={stats.get('company_parts', 0)}",
                f"manufacturer_parts={stats.get('manufacturer_parts', 0)}",
                f"erp_rows={stats.get('erp_rows', 0)}",
                f"master_rows={stats.get('master_rows', 0)}",
            )
        except Exception:
            pass

    def load_inventory(
        self,
        xlsx_path: str,
        progress_cb: Optional[Callable[[int, int, str], None]] = None,
        phase_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """Load the Master workbook, rebuild canonical company-part bundles, and sync the DB."""
        self._pending_inventory_master_path = xlsx_path

        if phase_cb:
            try:
                phase_cb("Reading master workbook...")
            except Exception:
                pass

        master_rows = DataLoader.load_alternate_master_rows(xlsx_path)
        self._pending_master_rows = list(master_rows or [])

        if phase_cb:
            try:
                phase_cb("Building canonical inventory records...")
                if progress_cb:
                    progress_cb(0, 1, "Building canonical inventory records")
            except Exception:
                pass

        company_parts = DataLoader.build_company_part_records(
            erp_rows=list(self._pending_erp_rows or []),
            master_rows=list(self._pending_master_rows or []),
        )
        self._pending_inventory_company_rows = list(company_parts or [])
        self._rebuild_company_part_maps(self._pending_inventory_company_rows)

        if phase_cb:
            try:
                phase_cb("Rebuilding inventory view...")
                if progress_cb:
                    progress_cb(0, 1, "Rebuilding inventory view")
            except Exception:
                pass

        self.inventory = self._company_rows_to_flat_inventory(self._pending_inventory_company_rows)
        self._inv_by_itemnum = {
            inv.itemnum: inv
            for inv in (self.inventory or [])
            if getattr(inv, "itemnum", "")
        }

        if phase_cb:
            try:
                phase_cb("Syncing inventory database...")
                if progress_cb:
                    progress_cb(0, 1, "Syncing inventory database")
            except Exception:
                pass
        self._sync_current_inventory_store(source_label="MASTER")

        try:
            if phase_cb:
                try:
                    phase_cb("Building alternate lookup map...")
                    if progress_cb:
                        progress_cb(0, 1, "Building alternate lookup map")
                except Exception:
                    pass
            self._alt_mpn_to_base = self._build_mpn_to_base_from_company_rows(self._pending_inventory_company_rows)
            self._alt_loaded = True
        except Exception:
            pass

        if self.workspace_id:
            try:
                self.ws_repo.update_meta(
                    self.workspace_id,
                    inventory_master_path=xlsx_path,
                    inventory_erp_path=(self._pending_inventory_erp_path or ""),
                )
            except Exception:
                pass

        if progress_cb:
            try:
                progress_cb(1, 1, "Done")
            except Exception:
                pass

        return len(self.inventory)

    def _dedupe_inventory_company_rows(self, rows: list[dict]) -> list[dict]:
        return list(rows or [])

    def _build_mpn_to_base_from_company_rows(self, company_rows: list[Any]) -> dict[str, list[str]]:
        """Build normalized MPN -> [company/base itemnum] from canonical company-part records."""
        out: dict[str, list[str]] = {}
        norm = getattr(DataLoader, "norm_mpn_key", None)
        if not callable(norm):
            norm = lambda s: str(s or "").strip().upper()

        for row in (company_rows or []):
            cpn = ""
            manufacturer_parts = []
            if isinstance(row, dict):
                cpn = str(row.get("cpn") or row.get("company_part_number") or "").strip()
                manufacturer_parts = list(row.get("manufacturer_parts") or row.get("alternates_json") or row.get("alternates") or [])
            else:
                cpn = str(getattr(row, "company_part_number", "") or "").strip()
                manufacturer_parts = list(getattr(row, "manufacturer_parts", []) or [])

            if not cpn:
                continue

            for part in manufacturer_parts:
                if isinstance(part, dict):
                    mpn = str(
                        part.get("manufacturer_part_number")
                        or part.get("mpn")
                        or part.get("mfgpn")
                        or part.get("vendoritem")
                        or ""
                    ).strip()
                else:
                    mpn = str(getattr(part, "manufacturer_part_number", "") or "").strip()

                key = norm(mpn)
                if not key:
                    continue
                out.setdefault(key, [])
                if cpn not in out[key]:
                    out[key].append(cpn)
        return out

    def _rebuild_company_part_maps(self, company_rows: list[Any]) -> None:
        """Rebuild the in-memory company-part lookup keyed by company part number."""
        out: dict[str, Any] = {}
        for row in (company_rows or []):
            cpn = ""
            if isinstance(row, dict):
                cpn = str(row.get("cpn") or row.get("company_part_number") or "").strip()
            else:
                cpn = str(getattr(row, "company_part_number", "") or "").strip()
            if cpn:
                out[cpn] = row
        self._company_parts_by_cpn = out

    def _get_company_part_row(self, company_part_number: str) -> Optional[Any]:
        cpn = str(company_part_number or "").strip()
        if not cpn:
            return None
        return (self._company_parts_by_cpn or {}).get(cpn)

    def _iter_company_part_manufacturer_parts(self, company_part_number: str) -> list[Any]:
        row = self._get_company_part_row(company_part_number)
        if row is None:
            return []
        if isinstance(row, dict):
            return list(row.get("manufacturer_parts") or row.get("alternates_json") or row.get("alternates") or [])
        return list(getattr(row, "manufacturer_parts", []) or [])

    def _manufacturer_part_mpn(self, part: Any) -> str:
        if isinstance(part, dict):
            return str(
                part.get("manufacturer_part_number")
                or part.get("mpn")
                or part.get("mfgpn")
                or part.get("vendoritem")
                or ""
            ).strip()
        return str(getattr(part, "manufacturer_part_number", "") or getattr(part, "mfgpn", "") or "").strip()

    def _manufacturer_part_to_spec_dict(self, company_part_number: str, part: Any, fallback_inv: Any = None) -> dict:
        def _pick(obj: Any, *attr_names: str) -> str:
            if obj is None:
                return ""
            raw = {}
            try:
                raw = dict(getattr(obj, "raw_fields", {}) or {})
            except Exception:
                raw = {}
            if isinstance(obj, dict):
                for name in attr_names:
                    value = obj.get(name)
                    if value not in (None, ""):
                        return str(value).strip()
                for name in attr_names:
                    value = raw.get(name)
                    if value not in (None, ""):
                        return str(value).strip()
                return ""
            for name in attr_names:
                value = getattr(obj, name, None)
                if value not in (None, ""):
                    return str(value).strip()
            for name in attr_names:
                value = raw.get(name)
                if value not in (None, ""):
                    return str(value).strip()
            return ""

        inv = fallback_inv
        if inv is None and company_part_number:
            inv = (self._inv_by_itemnum or {}).get(company_part_number)

        spec = {
            "ItemNumber": str(company_part_number or "").strip(),
            "VendorItem": self._manufacturer_part_mpn(part),
            "Description": _pick(part, "description", "desc"),
            "MfgName": _pick(part, "manufacturer_name", "manufacturer", "mfgname"),
            "MfgId": _pick(part, "manufacturer_id", "mfgid"),
            "PrimaryVendorNumber": "",
            "TotalQty": "",
            "LastCost": _pick(part, "last_cost", "lastcost"),
            "AvgCost": _pick(part, "average_cost", "avg_cost", "avgcost", "unit_cost", "price"),
            "StandardCost": _pick(part, "standard_cost", "standardcost"),
            "ItemLeadTime": _pick(part, "item_lead_time", "itemleadtime", "lead_time"),
            "DefaultWhse": "",
            "TariffCodeHTSUS": _pick(part, "tariff_code", "tariffcodehtsus", "htsus"),
            "TariffRate": _pick(part, "tariff_rate", "tariffrate"),
            "IsERPPrimary": _pick(part, "is_erp_primary"),
            "MasterSourceRowKey": _pick(part, "master_source_row_key"),
            "ERPSourceRowKey": _pick(part, "erp_source_row_key"),
        }

        if inv is not None:
            inv_raw = dict(getattr(inv, "raw_fields", {}) or {})
            spec["Description"] = spec["Description"] or str(getattr(inv, "desc", "") or "").strip()
            spec["PrimaryVendorNumber"] = str(inv_raw.get("primaryvendornumber") or inv_raw.get("supplier") or inv_raw.get("vendor") or "").strip()
            spec["TotalQty"] = str(inv_raw.get("totalqty") or inv_raw.get("total_qty") or inv_raw.get("qty_on_hand") or inv_raw.get("on_hand") or inv_raw.get("quantity") or "").strip()
            spec["DefaultWhse"] = str(inv_raw.get("defaultwhse") or inv_raw.get("default_whse") or inv_raw.get("warehouse") or "").strip()

        return spec

    def _get_company_part_mfgpn_options(self, company_part_number: str) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        row = self._get_company_part_row(company_part_number)
        inv = (self._inv_by_itemnum or {}).get(str(company_part_number or "").strip())

        def add(value: Any) -> None:
            m = str(value or "").strip()
            if not m:
                return
            k = m.lower()
            if k in seen:
                return
            seen.add(k)
            out.append(m)

        if inv is not None:
            add(getattr(inv, "vendoritem", "") or "")
            for sub in list(getattr(inv, "substitutes", None) or []):
                if isinstance(sub, str):
                    add(sub)
                else:
                    add(getattr(sub, "mfgpn", "") or getattr(sub, "manufacturer_part_number", "") or "")

        for part in self._iter_company_part_manufacturer_parts(company_part_number):
            add(self._manufacturer_part_mpn(part))

        if isinstance(row, dict):
            for m in list(row.get("company_pn_mfgpns") or []):
                add(m)

        return out

    def _resolve_company_part_manufacturer_part(self, company_part_number: str, mfgpn: str) -> Optional[Any]:
        target = str(mfgpn or "").strip()
        if not target:
            return None
        target_lower = target.lower()
        for part in self._iter_company_part_manufacturer_parts(company_part_number):
            mpn = self._manufacturer_part_mpn(part)
            if mpn and mpn.lower() == target_lower:
                return part
        return None

    def _company_rows_to_flat_inventory(self, company_rows: list[Any]) -> list[InventoryPart]:
        out: list[InventoryPart] = []
        for row in (company_rows or []):
            inv = self._company_row_to_inventory_part(row)
            if inv is not None:
                out.append(inv)
        return out

    def _company_row_to_inventory_part(self, row: Any) -> Optional[InventoryPart]:
        if row is None:
            return None

        if isinstance(row, dict):
            cpn = str(row.get("cpn") or row.get("company_part_number") or "").strip()
            desc = str(row.get("canonical_desc") or row.get("description") or "").strip()
            total_qty = row.get("stock_total")
            if total_qty is None:
                total_qty = row.get("total_qty")
            manufacturer_parts = list(row.get("manufacturer_parts") or row.get("alternates_json") or row.get("alternates") or [])
            raw_fields = dict(row.get("raw_fields") or {})
        else:
            cpn = str(getattr(row, "company_part_number", "") or "").strip()
            desc = str(getattr(row, "description", "") or "").strip()
            total_qty = getattr(row, "total_qty", None)
            manufacturer_parts = list(getattr(row, "manufacturer_parts", []) or [])
            raw_fields = dict(getattr(row, "raw_fields", {}) or {})

        if not cpn:
            return None

        try:
            stock_total = int(float(total_qty or 0))
        except Exception:
            stock_total = 0

        mpns: list[str] = []
        mfgname = ""
        mfgid = ""

        for part in manufacturer_parts:
            if isinstance(part, dict):
                mpn = str(
                    part.get("manufacturer_part_number")
                    or part.get("mpn")
                    or part.get("mfgpn")
                    or part.get("vendoritem")
                    or ""
                ).strip()
                part_mfgname = str(part.get("manufacturer_name") or part.get("mfgname") or "").strip()
                part_mfgid = str(part.get("manufacturer_id") or part.get("mfgid") or "").strip()
            else:
                mpn = str(getattr(part, "manufacturer_part_number", "") or "").strip()
                part_mfgname = str(getattr(part, "manufacturer_name", "") or "").strip()
                part_mfgid = str(getattr(part, "manufacturer_id", "") or "").strip()

            if not mpn:
                continue
            if not mfgname:
                mfgname = part_mfgname
            if not mfgid:
                mfgid = part_mfgid
            if mpn not in mpns:
                mpns.append(mpn)

        rep_mpn = mpns[0] if mpns else ""
        subs = mpns[1:] if len(mpns) > 1 else []

        inv = InventoryPart(
            itemnum=cpn,
            desc=desc,
            mfgid=mfgid,
            mfgname=mfgname,
            vendoritem=rep_mpn,
            substitutes=subs,
            raw_fields={"totalqty": stock_total, **raw_fields},
            parsed={},
            api_data=None,
        )
        inv._all_mpns = list(mpns)
        inv._manufacturer_parts = list(manufacturer_parts or [])
        return inv

    def load_alternates_db(self, xlsx_path: str) -> tuple[int, int]:
        count = self.load_inventory(xlsx_path)
        num_with_subs = sum(1 for inv in (self.inventory or []) if getattr(inv, "substitutes", []))
        return num_with_subs, count

    def load_items_inventory(
        self,
        xlsx_path: str,
        progress_cb: Optional[Callable[[int, int, str], None]] = None,
        phase_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """Load the ERP workbook, rebuild canonical company-part bundles, and sync the DB."""
        self._pending_inventory_erp_path = xlsx_path

        if phase_cb:
            try:
                phase_cb("Reading ERP workbook...")
            except Exception:
                pass

        erp_rows = DataLoader.load_erp_rows(xlsx_path)
        self._pending_erp_rows = list(erp_rows or [])

        if phase_cb:
            try:
                phase_cb("Building canonical inventory records...")
                if progress_cb:
                    progress_cb(0, 1, "Building canonical inventory records")
            except Exception:
                pass

        company_parts = DataLoader.build_company_part_records(
            erp_rows=list(self._pending_erp_rows or []),
            master_rows=list(self._pending_master_rows or []),
        )
        self._pending_inventory_company_rows = list(company_parts or [])
        self._rebuild_company_part_maps(self._pending_inventory_company_rows)

        try:
            if phase_cb:
                try:
                    phase_cb("Rebuilding inventory view...")
                    if progress_cb:
                        progress_cb(0, 1, "Rebuilding inventory view")
                except Exception:
                    pass
            self.inventory = self._company_rows_to_flat_inventory(self._pending_inventory_company_rows)
            self._inv_by_itemnum = {
                inv.itemnum: inv
                for inv in (self.inventory or [])
                if getattr(inv, "itemnum", "")
            }
        except Exception as e:
            print(f"[CTRL] failed to refresh in-memory inventory after ERP import: {e}")

        if phase_cb:
            try:
                phase_cb("Syncing inventory database...")
                if progress_cb:
                    progress_cb(0, 1, "Syncing inventory database")
            except Exception:
                pass
        self._sync_current_inventory_store(source_label="ERP")

        if self.workspace_id:
            try:
                self.ws_repo.update_meta(self.workspace_id, inventory_erp_path=xlsx_path)
            except Exception:
                pass

        if progress_cb:
            try:
                progress_cb(1, 1, "Done")
            except Exception:
                pass

        return len(self._pending_erp_rows)

    def _load_bom_rows_from_workbook(self, xlsx_path: str) -> List[Any]:
        """Load BOM rows only through the canonical DataLoader BOM pipeline."""
        rows = DataLoader.load_bom_any(xlsx_path)
        out: List[Any] = []
        for i, part in enumerate(rows, start=1):
            if not getattr(part, "partnum", ""):
                part.partnum = f"ROW-{i}"
            out.append(part)
        return out

    def load_npr(
        self,
        xlsx_path: str,
        progress_cb: Optional[Callable[[int, int, str], None]] = None,
        phase_cb: Optional[Callable[[str], None]] = None,
        ) -> int:
        """Load a BOM workbook, create a new workspace, and persist canonical BOM input rows."""
        self.npr_list = []
        self.match_pairs = []
        self.nodes = []
        self.external_cache = {}
        self.workspace = None
        self._bom_row_by_uid = {}

        if phase_cb:
            try:
                phase_cb("Reading BOM workbook...")
            except Exception:
                pass

        rows = self._load_bom_rows_from_workbook(xlsx_path)
        self.npr_list = rows

        if phase_cb:
            try:
                phase_cb("Creating workspace and saving BOM input...")
                if progress_cb:
                    progress_cb(0, 1, "Saving BOM input")
            except Exception:
                pass

        ws_name = Path(xlsx_path).stem
        self.workspace_id = self.ws_repo.create(
            name=ws_name,
            label="",
            bom_source_path=xlsx_path,
            inventory_master_path=(self._pending_inventory_master_path or ""),
            inventory_erp_path=(self._pending_inventory_erp_path or ""),
            cns_path="",
            notes="",
            status="ACTIVE",
        )

        inputs: List[Dict[str, Any]] = []
        for i, p in enumerate(self.npr_list or [], start=1):
            raw = p.to_dict() if hasattr(p, "to_dict") else {}
            qty = None
            try:
                q_raw = (getattr(p, "raw_fields", {}) or {}).get("quantity", "")
                qty = float(q_raw) if str(q_raw).strip() else None
            except Exception:
                qty = None

            inputs.append(
                {
                    "input_line_id": i,
                    "partnum": getattr(p, "partnum", f"ROW-{i}"),
                    "description": getattr(p, "desc", "") or getattr(p, "description", ""),
                    "qty": qty,
                    "refdes": (getattr(p, "raw_fields", {}) or {}).get("designator", ""),
                    "item_type": (getattr(p, "parsed", {}) or {}).get("type", ""),
                    "mfgname": getattr(p, "mfgname", ""),
                    "mfgpn": getattr(p, "mfgpn", ""),
                    "supplier": getattr(p, "supplier", ""),
                    "raw_json": raw,
                }
            )

        self.bom_repo.upsert_inputs(self.workspace_id, inputs)
        try:
            self.bom_repo.bootstrap_state_from_inputs(self.workspace_id, overwrite_existing=False)
        except Exception:
            pass

        if progress_cb:
            try:
                progress_cb(1, 1, "Done")
            except Exception:
                pass

        return len(self.npr_list)
    
    def load_cns(self, xlsx_path: str) -> int:
        self.cns_records = DataLoader.load_cns_workbook(xlsx_path)
        return len(self.cns_records)
    
    def _should_stop(self) -> bool:
        try:
            return bool(self.stop_event and self.stop_event.is_set())
        except Exception:
            return False

    def request_stop_matching(self) -> None:
        try:
            if self.stop_event:
                self.stop_event.set()
        except Exception:
            pass

    def reset_stop_matching(self) -> None:
        try:
            if self.stop_event:
                self.stop_event.clear()
        except Exception:
            pass

    def _db_enabled(self) -> bool:
        return self.workspace_id is not None


    def _node_line_id(self, node: DecisionNode) -> int:
        try:
            _ws, _line = (getattr(node, "id", "") or "").split(":", 1)
            return int(_line)
        except Exception:
            return 0

    def _sync_normalized_node_state(self, node: DecisionNode) -> DecisionNode:
        """Keep first-class committed node fields coherent with the persisted workflow state."""
        node.assigned_part_number = str(getattr(node, "assigned_part_number", "") or "").strip()
        node.preferred_inventory_mfgpn = str(getattr(node, "preferred_inventory_mfgpn", "") or "").strip()
        node.bom_section = str(getattr(node, "bom_section", "") or "SURFACE MOUNT").strip() or "SURFACE MOUNT"
        node.focused_alt_id = str(getattr(node, "focused_alt_id", "") or "").strip()
        node.exclude_customer_part_number_in_npr = bool(getattr(node, "exclude_customer_part_number_in_npr", False))
        return node

    def _focused_alt_id_for_node(self, node: DecisionNode) -> str:
        focused_alt_id = str(getattr(node, "focused_alt_id", "") or "").strip()
        if focused_alt_id:
            return focused_alt_id
        focused_card_id = str(getattr(node, "focused_card_id", "") or "").strip()
        if focused_card_id:
            try:
                card = node.get_card(focused_card_id) if hasattr(node, "get_card") else None
                if card is not None:
                    return str(getattr(card, "alt_id", "") or "").strip()
            except Exception:
                pass
        try:
            sels = node.selected_alternates() if hasattr(node, "selected_alternates") else []
            if sels:
                return str(getattr(sels[0], "id", "") or "").strip()
        except Exception:
            pass
        return ""

    def _node_to_db_row(self, node: DecisionNode) -> Dict[str, Any]:
        explain = dict(getattr(node, "explain", {}) or {})
        explain.pop("focused_alt_id", None)
        explain.pop("preferred_inventory_mfgpn", None)
        explain.pop("bom_section", None)
        explain.pop("assigned_part_number", None)
        focused_alt_id = self._focused_alt_id_for_node(node)
        node.focused_alt_id = focused_alt_id
        return {
            "node_id": str(getattr(node, "id", "") or ""),
            "line_id": self._node_line_id(node),
            "base_type": str(getattr(node, "base_type", "") or ""),
            "bom_uid": str(getattr(node, "bom_uid", "") or ""),
            "bom_mpn": str(getattr(node, "bom_mpn", "") or ""),
            "description": str(getattr(node, "description", "") or ""),
            "internal_part_number": str(getattr(node, "internal_part_number", "") or ""),
            "assigned_part_number": str(getattr(node, "assigned_part_number", "") or ""),
            "inventory_mpn": str(getattr(node, "inventory_mpn", "") or ""),
            "preferred_inventory_mfgpn": str(getattr(node, "preferred_inventory_mfgpn", "") or ""),
            "bom_section": str(getattr(node, "bom_section", "SURFACE MOUNT") or "SURFACE MOUNT"),
            "match_type": str(getattr(node, "match_type", "") or ""),
            "confidence": float(getattr(node, "confidence", 0.0) or 0.0),
            "status": str(getattr(getattr(node, "status", None), "value", getattr(node, "status", "")) or ""),
            "locked": 1 if bool(getattr(node, "locked", False)) else 0,
            "needs_approval": 1 if bool(getattr(node, "needs_approval", False)) else 0,
            "focused_alt_id": focused_alt_id,
            "exclude_customer_part_number_in_npr": 1 if bool(getattr(node, "exclude_customer_part_number_in_npr", False)) else 0,
            "notes": str(getattr(node, "notes", "") or ""),
            "explain_json": explain,
        }

    def _alt_to_db_row(self, alt: Alternate) -> Dict[str, Any]:
        return {
            "alt_id": str(getattr(alt, "id", "") or ""),
            "source": str(getattr(alt, "source", "") or ""),
            "manufacturer": str(getattr(alt, "manufacturer", "") or ""),
            "manufacturer_part_number": str(getattr(alt, "manufacturer_part_number", "") or ""),
            "internal_part_number": str(getattr(alt, "internal_part_number", "") or ""),
            "description": str(getattr(alt, "description", "") or ""),
            "value": str(getattr(alt, "value", "") or ""),
            "package": str(getattr(alt, "package", "") or ""),
            "tolerance": str(getattr(alt, "tolerance", "") or ""),
            "voltage": str(getattr(alt, "voltage", "") or ""),
            "wattage": str(getattr(alt, "wattage", "") or ""),
            "stock": int(getattr(alt, "stock", 0) or 0),
            "unit_cost": getattr(alt, "unit_cost", None),
            "supplier": str(getattr(alt, "supplier", "") or ""),
            "confidence": float(getattr(alt, "confidence", 0.0) or 0.0),
            "relationship": str(getattr(alt, "relationship", "") or ""),
            "matched_mpn": str(getattr(alt, "matched_mpn", "") or ""),
            "selected": 1 if bool(getattr(alt, "selected", False)) else 0,
            "rejected": 1 if bool(getattr(alt, "rejected", False)) else 0,
            "meta_json": dict(getattr(alt, "meta", {}) or {}),
            "raw_json": getattr(alt, "raw", {}) if isinstance(getattr(alt, "raw", None), (dict, list)) else {},
        }

    def _status_from_value(self, value: Any) -> DecisionStatus:
        raw = str(value or DecisionStatus.NEEDS_DECISION.value)
        try:
            return DecisionStatus(raw)
        except Exception:
            return DecisionStatus.NEEDS_DECISION

    def _restore_alternate_from_row(self, row: Dict[str, Any]) -> Alternate:
        return Alternate(
            id=str(row.get("alt_id", "") or ""),
            source=str(row.get("source", "") or ""),
            manufacturer=str(row.get("manufacturer", "") or ""),
            manufacturer_part_number=str(row.get("manufacturer_part_number", "") or ""),
            internal_part_number=str(row.get("internal_part_number", "") or ""),
            description=str(row.get("description", "") or ""),
            value=str(row.get("value", "") or ""),
            package=str(row.get("package", "") or ""),
            tolerance=str(row.get("tolerance", "") or ""),
            voltage=str(row.get("voltage", "") or ""),
            wattage=str(row.get("wattage", "") or ""),
            stock=int(row.get("stock", 0) or 0),
            unit_cost=row.get("unit_cost", None),
            supplier=str(row.get("supplier", "") or ""),
            confidence=float(row.get("confidence", 0.0) or 0.0),
            relationship=str(row.get("relationship", "") or ""),
            matched_mpn=str(row.get("matched_mpn", "") or ""),
            selected=bool(row.get("selected", 0) or 0),
            rejected=bool(row.get("rejected", 0) or 0),
            raw=row.get("raw_json", {}) or {},
            meta=dict(row.get("meta_json", {}) or {}),
        )

    def _restore_node_from_rows(self, node_row: Dict[str, Any], alt_rows: List[Dict[str, Any]]) -> DecisionNode:
        explain = dict(node_row.get("explain_json", {}) or {})
        focused_alt_id = str(node_row.get("focused_alt_id", "") or "").strip()
        node = DecisionNode(
            id=str(node_row.get("node_id", "") or ""),
            base_type=str(node_row.get("base_type", "") or ""),
            bom_uid=str(node_row.get("bom_uid", "") or ""),
            bom_mpn=str(node_row.get("bom_mpn", "") or ""),
            description=str(node_row.get("description", "") or ""),
            internal_part_number=str(node_row.get("internal_part_number", "") or ""),
            inventory_mpn=str(node_row.get("inventory_mpn", "") or ""),
            assigned_part_number=str(node_row.get("assigned_part_number", "") or ""),
            preferred_inventory_mfgpn=str(node_row.get("preferred_inventory_mfgpn", "") or ""),
            bom_section=str(node_row.get("bom_section", "") or "SURFACE MOUNT"),
            focused_alt_id=focused_alt_id,
            exclude_customer_part_number_in_npr=bool(node_row.get("exclude_customer_part_number_in_npr", 0) or 0),
            match_type=str(node_row.get("match_type", "") or ""),
            confidence=float(node_row.get("confidence", 0.0) or 0.0),
            alternates=[self._restore_alternate_from_row(r) for r in (alt_rows or [])],
            status=self._status_from_value(node_row.get("status")),
            locked=bool(node_row.get("locked", 0) or 0),
            needs_approval=bool(node_row.get("needs_approval", 0) or 0),
            notes=str(node_row.get("notes", "") or ""),
            explain=explain,
        )
        if focused_alt_id:
            node.focused_card_id = self._card_id_for_alt(node.id, focused_alt_id)
        return self._sync_normalized_node_state(node)

    def _rehydrate_npr_inputs(self, workspace_id: str) -> List[Any]:
        npr_list: List[Any] = []
        for i, part in enumerate(self.bom_repo.list_input_parts(workspace_id) or [], start=1):
            if not getattr(part, "partnum", ""):
                part.partnum = f"ROW-{i}"
            setattr(part, "bom_uid", getattr(part, "partnum", "") or str(i))
            setattr(part, "bom_mpn", getattr(part, "mfgpn", "") or "")
            npr_list.append(part)
        return npr_list

    def _card_id_for_alt(self, node_id: str, alt_id: str) -> str:
        return f"CARD::{node_id}::{alt_id}"

    def build_node_cards(self, node_id: str, focused_card_id: Optional[str] = None) -> List[DecisionCard]:
        """Materialize and attach UI card objects for the requested node."""
        node = self.get_node(node_id)

        if focused_card_id is not None:
            if hasattr(node, "set_focused_card"):
                card = None
                try:
                    card = node.get_card(focused_card_id) if hasattr(node, "get_card") else None
                except Exception:
                    card = None
                node.set_focused_card(focused_card_id, getattr(card, "alt_id", "") if card else "")
            else:
                node.focused_card_id = str(focused_card_id or "").strip()

        effective_focus = str(getattr(node, "focused_card_id", "") or "").strip() or None
        cards = self.card_factory.build_cards(node, focused_card_id=effective_focus)
        if hasattr(node, "set_cards"):
            node.set_cards(cards)
        else:
            node.cards = list(cards or [])

        valid_ids = {getattr(card, "card_id", "") for card in (node.cards or []) if getattr(card, "card_id", "")}
        if effective_focus and effective_focus not in valid_ids:
            if hasattr(node, "clear_focused_card"):
                node.clear_focused_card()
            else:
                node.focused_card_id = ""
                node.focused_alt_id = ""
            effective_focus = None
            cards = self.card_factory.build_cards(node, focused_card_id=None)
            if hasattr(node, "set_cards"):
                node.set_cards(cards)
            else:
                node.cards = list(cards or [])

        if not effective_focus:
            chosen = self._resolve_default_focus_card(node_id, persist=False)
            if chosen is not None:
                if hasattr(node, "set_focused_card"):
                    node.set_focused_card(getattr(chosen, "card_id", "") or "", getattr(chosen, "alt_id", "") or "")
                else:
                    node.focused_card_id = str(getattr(chosen, "card_id", "") or "")
                    node.focused_alt_id = str(getattr(chosen, "alt_id", "") or "")
                cards = self.card_factory.build_cards(node, focused_card_id=getattr(chosen, "card_id", "") or None)
                if hasattr(node, "set_cards"):
                    node.set_cards(cards)
                else:
                    node.cards = list(cards or [])
            else:
                node.focused_alt_id = ""

        return list(node.cards or [])

    def rebuild_all_node_cards(self, focused_by_node: Optional[Dict[str, Optional[str]]] = None) -> None:
        """Refresh card ownership snapshots for every node without touching workflow state."""
        focused_by_node = dict(focused_by_node or {})
        for node in list(self.nodes or []):
            try:
                self.build_node_cards(node.id, focused_card_id=focused_by_node.get(node.id))
            except Exception:
                continue

    def _find_card(self, node: DecisionNode, card_id: str, rebuild: bool = True) -> DecisionCard:
        if rebuild or not getattr(node, "cards", None):
            self.build_node_cards(node.id)
        card = node.get_card(card_id) if hasattr(node, "get_card") else None
        if card is None:
            raise ValueError(f"Card not found: {card_id}")
        return card

    def get_node_card(self, node_id: str, card_id: str, rebuild: bool = True) -> DecisionCard:
        node = self.get_node(node_id)
        return self._find_card(node, card_id, rebuild=rebuild)

    def get_card_for_alt(self, node_id: str, alt_id: str, rebuild: bool = True) -> Optional[DecisionCard]:
        node = self.get_node(node_id)
        if rebuild or not getattr(node, "cards", None):
            self.build_node_cards(node_id)
        return node.get_card_by_alt_id(alt_id) if hasattr(node, "get_card_by_alt_id") else None


    def _resolve_default_focus_card(self, node_id: str, persist: bool = False) -> Optional[DecisionCard]:
        node = self.get_node(node_id)
        cards = list(getattr(node, "cards", []) or [])
        if not cards:
            cards = self.card_factory.build_cards(node, focused_card_id=None)
            if hasattr(node, "set_cards"):
                node.set_cards(cards)
            else:
                node.cards = list(cards or [])
        if not cards:
            return None

        def _active(card: DecisionCard) -> bool:
            return not bool(getattr(getattr(card, "state", None), "rejected", False))

        focused_alt_id = str(getattr(node, "focused_alt_id", "") or "").strip()
        if focused_alt_id:
            card = node.get_card_by_alt_id(focused_alt_id) if hasattr(node, "get_card_by_alt_id") else None
            if card is not None and _active(card):
                if persist:
                    try:
                        self.focus_card(node_id, card.card_id)
                    except Exception:
                        pass
                return card

        try:
            for alt in (getattr(node, "alternates", []) or []):
                if bool(getattr(alt, "selected", False)) and not bool(getattr(alt, "rejected", False)):
                    card = node.get_card_by_alt_id(str(getattr(alt, "id", "") or "")) if hasattr(node, "get_card_by_alt_id") else None
                    if card is not None:
                        if persist:
                            try:
                                self.focus_card(node_id, card.card_id)
                            except Exception:
                                pass
                        return card
        except Exception:
            pass

        for card in cards:
            if _active(card):
                if persist:
                    try:
                        self.focus_card(node_id, card.card_id)
                    except Exception:
                        pass
                return card
        return None

    def focus_card(self, node_id: str, card_id: str, save: bool = True) -> DecisionCard:
        node = self.get_node(node_id)
        self._find_card(node, card_id, rebuild=True)
        card = self.get_node_card(node_id, card_id, rebuild=True)
        if hasattr(node, "set_focused_card"):
            node.set_focused_card(card_id, getattr(card, "alt_id", "") or "")
        else:
            node.focused_card_id = str(card_id or "").strip()
            node.focused_alt_id = str(getattr(card, "alt_id", "") or "")
        self.build_node_cards(node_id, focused_card_id=card_id)
        card = self.get_node_card(node_id, card_id, rebuild=False)
        if save:
            try:
                self._persist_node_and_alts(node)
            except Exception:
                pass
        return card

    def clear_focused_card(self, node_id: str, save: bool = True) -> None:
        node = self.get_node(node_id)
        if hasattr(node, "clear_focused_card"):
            node.clear_focused_card()
        else:
            node.focused_card_id = ""
        node.focused_alt_id = ""
        self.build_node_cards(node_id, focused_card_id="")
        if save:
            try:
                self._persist_node_and_alts(node)
            except Exception:
                pass

    def get_focused_card(self, node_id: str) -> Optional[DecisionCard]:
        node = self.get_node(node_id)
        focused_card_id = str(getattr(node, "focused_card_id", "") or "").strip()
        if focused_card_id:
            try:
                card = self.get_node_card(node_id, focused_card_id, rebuild=True)
                if not bool(getattr(getattr(card, "state", None), "rejected", False)):
                    return card
            except Exception:
                pass
        return self._resolve_default_focus_card(node_id, persist=False)

    def select_card(self, node_id: str, card_id: str) -> DecisionCard:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        self.select_alternate(node_id, card.alt_id)
        return self.focus_card(node_id, card_id, save=False)

    def unselect_card(self, node_id: str, card_id: str) -> DecisionCard:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        self.unselect_alternate(node_id, card.alt_id)
        return self.focus_card(node_id, card_id, save=False)

    def reject_card(self, node_id: str, card_id: str) -> DecisionCard:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        self.reject_alternate(node_id, card.alt_id)
        chosen = self._pick_detail_card(node_id, auto_focus=True)
        if chosen is not None:
            return chosen
        return self.get_node_card(node_id, card_id, rebuild=True)

    def unreject_card(self, node_id: str, card_id: str) -> DecisionCard:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        self.unreject_alternate(node_id, card.alt_id)
        return self.focus_card(node_id, card_id, save=False)

    def get_card_detail_payload(self, node_id: str, card_id: str) -> dict:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        return self.get_alt_detail_payload(node_id, card.alt_id)


    def set_preferred_inventory_mfgpn_for_card(self, node_id: str, card_id: str, mfgpn: str) -> DecisionCard:
        card = self.get_node_card(node_id, card_id, rebuild=True)
        if not getattr(card, "alt_id", ""):
            raise ValueError("Card has no backing alternate.")
        self.set_preferred_inventory_mfgpn(node_id, card.alt_id, mfgpn)
        return self.focus_card(node_id, card_id)

    def _pick_detail_card(self, node_id: str, auto_focus: bool = True) -> Optional[DecisionCard]:
        node = self.get_node(node_id)
        cards = list(self.build_node_cards(node_id) or [])
        if not cards:
            return None

        focused = self.get_focused_card(node_id)
        if focused is not None:
            return focused

        chosen = next(
            (c for c in cards if bool(getattr(getattr(c, "state", None), "selected", False))
             and not bool(getattr(getattr(c, "state", None), "rejected", False))),
            None,
        )
        if chosen is None:
            chosen = next(
                (c for c in cards if getattr(c, "source", "") == "inventory"
                 and not bool(getattr(getattr(c, "state", None), "rejected", False))),
                None,
            )
        if chosen is None:
            chosen = next(
                (c for c in cards if not bool(getattr(getattr(c, "state", None), "rejected", False))),
                None,
            )
        if chosen is None:
            return None

        if auto_focus:
            try:
                return self.focus_card(node_id, chosen.card_id)
            except Exception:
                return chosen
        return chosen

    def build_card_detail_state(
        self,
        node_id: str,
        card_id: Optional[str] = None,
        auto_focus: bool = True,
    ) -> CardDetailState:
        node = self.get_node(node_id)
        card: Optional[DecisionCard] = None

        if card_id:
            card = self.get_node_card(node_id, card_id, rebuild=True)
            if auto_focus:
                try:
                    card = self.focus_card(node_id, card.card_id)
                except Exception:
                    pass
        else:
            card = self._pick_detail_card(node_id, auto_focus=auto_focus)

        if card is None:
            return CardDetailState(
                node_id=node.id,
                title_text="Information",
                specs={},
                export_mfgpn_options=[],
                has_card=False,
                is_inventory=False,
            )

        payload = self.get_card_detail_payload(node_id, card.card_id) or {}
        specs = dict(payload.get("specs") or {})
        export_opts = [str(m).strip() for m in list(payload.get("export_mfgpn_options") or []) if str(m).strip()]
        selected_export_mfgpn = str(payload.get("selected_export_mfgpn", "") or specs.get("VendorItem", "") or "").strip()

        company_pn = (getattr(card, "company_part_number", "") or "").strip()
        shown_mfgpn = selected_export_mfgpn or (getattr(card, "manufacturer_part_number", "") or "").strip()
        if company_pn and shown_mfgpn:
            title_txt = f"Information • {company_pn} • {shown_mfgpn}"
        elif company_pn:
            title_txt = f"Information • {company_pn}"
        elif shown_mfgpn:
            title_txt = f"Information • {shown_mfgpn}"
        else:
            title_txt = "Information"

        return CardDetailState(
            node_id=node.id,
            card_id=getattr(card, "card_id", "") or "",
            alt_id=getattr(card, "alt_id", "") or "",
            title_text=title_txt,
            specs=specs,
            export_mfgpn_options=export_opts,
            selected_export_mfgpn=selected_export_mfgpn,
            has_card=True,
            is_inventory=bool(getattr(card, "is_inventory", False)),
        )

    def get_node_detail_state(
        self,
        node_id: str,
        card_id: Optional[str] = None,
        auto_focus: bool = True,
    ) -> CardDetailState:
        return self.build_card_detail_state(node_id=node_id, card_id=card_id, auto_focus=auto_focus)

    def _lookup_bom_row_id_for_node(self, node: DecisionNode) -> Optional[str]:
        """
        Best-effort: map node.bom_uid -> bom_row_id if Step B created the mapping.
        If not available yet, returns None (still persists node fine).
        """
        try:
            uid = (getattr(node, "bom_uid", "") or "").strip()
            if not uid:
                return None
            return (self._bom_row_by_uid or {}).get(uid)
        except Exception:
            return None
#------------------------------------#------------------------------------#------------------------------------
    def _bom_row_dict_to_part(self, r: dict):
        """
        Rehydrate a BOM row dict into a lightweight object that behaves like DataLoader rows.
        We only rely on attribute access in the rest of the pipeline.
        """
        # parsed/raw_fields are stored as JSON strings in DB
        try:
            import json
            raw_fields = json.loads(r.get("raw_fields_json", "") or "{}")
        except Exception:
            raw_fields = {}

        try:
            import json
            parsed = json.loads(r.get("parsed_json", "") or "{}")
        except Exception:
            parsed = {}

        return SimpleNamespace(
            partnum=r.get("partnum", "") or "",
            mfgpn=r.get("mfgpn", "") or "",
            mfgname=r.get("mfgname", "") or "",
            supplier=r.get("supplier", "") or "",
            description=r.get("description", "") or "",
            raw_fields=raw_fields,
            parsed=parsed,
        )
#------------------------------------#------------------------------------#------------------------------------#------------------------------------
    def _persist_inventory_resolved_for_itemnum(self, itemnum: str) -> None:
        # v2 schema: inventory is stored in inventory_company and is not updated per-item.
        return



    def _persist_node_and_alts(self, node: DecisionNode) -> None:
        if not self.workspace_id:
            try:
                self._sync_node_header_state(node)
            except Exception:
                pass
            return
        try:
            self._sync_normalized_node_state(node)
            explain = dict(getattr(node, "explain", {}) or {})
            explain["external_alternates"] = self._serialize_external_alternates_for_node(node)
            node.explain = explain
            self.decision_node_repo.save_node(self.workspace_id, self._node_to_db_row(node))
            alt_rows = [self._alt_to_db_row(alt) for alt in (getattr(node, "alternates", []) or [])]
            self.decision_alt_repo.save_node_alternates(self.workspace_id, node.id, alt_rows)
            try:
                self.ws_repo.touch(self.workspace_id)
            except Exception:
                pass
        except Exception as e:
            print(f"[DB] save node state failed for {getattr(node, 'id', '')}: {e}")
        try:
            self._sync_node_header_state(node)
        except Exception:
            pass

    def _serialize_external_alternates_for_node(self, node: DecisionNode) -> list[dict]:
        try:
            return external_alt_specs_from_node(node)
        except Exception:
            return []

    def _restore_external_alternates_for_node(self, node: DecisionNode) -> list[Alternate]:
        explain = dict(getattr(node, "explain", {}) or {})
        rows = explain.get("external_alternates") or []
        try:
            restored = restore_external_alternates(node, Alternate, rows)
        except Exception:
            restored = []
        if restored:
            node.explain = explain
        return restored


    def _persist_all_nodes(self) -> None:
        """Persist all in-memory DecisionNodes for the active workspace."""
        for n in (self.nodes or []):
            try:
                self._persist_node_and_alts(n)
            except Exception as e:
                print(f"[DB] save node failed: {e}")

    def save_workspace_state(self) -> int:
        """Save the current workspace state to the durable decision tables."""
        if not self.workspace_id:
            return 0
        self._persist_all_nodes()
        try:
            self.ws_repo.touch(self.workspace_id)
        except Exception:
            pass
        return len(self.nodes or [])

    def list_workspaces(self, *, status: str = "ACTIVE") -> List[Dict[str, Any]]:
        return self.ws_repo.list(status=status)

    def open_workspace(self, workspace_id: str) -> int:
        """Open an existing workspace and rebuild runtime state from persisted decision rows."""
        ws = self.ws_repo.get(workspace_id)
        if not ws:
            raise ValueError(f"Workspace not found: {workspace_id}")
        self.workspace_id = workspace_id
        self.reset_stop_matching()
        self.match_pairs = []
        self.external_cache = {}
        self.workspace = None
        self.nodes = []
        self._node_index = {}
        self._nodes_workspace_id = workspace_id
        self._inventory_cache = None
        self._views_cache = []
        self._header_state_by_node = {}
        self._pending_inventory_master_path = ws.get("inventory_master_path") or None
        self._pending_inventory_erp_path = ws.get("inventory_erp_path") or None
        try:
            self.npr_list = self._rehydrate_npr_inputs(workspace_id)
        except Exception as e:
            print(f"[DB] open_workspace: failed to rehydrate BOM inputs: {e}")
            self.npr_list = []
        views = self.build_decision_views()
        self.nodes = list(views or [])
        self._node_index = {n.id: n for n in (self.nodes or [])}
        self._nodes_workspace_id = workspace_id
        focused_by_node: Dict[str, Optional[str]] = {}
        for node in (self.nodes or []):
            focused_alt_id = str(getattr(node, "focused_alt_id", "") or "").strip()
            if focused_alt_id:
                focused_by_node[node.id] = self._card_id_for_alt(node.id, focused_alt_id)
        self.rebuild_all_node_cards(focused_by_node=focused_by_node)
        self.rebuild_all_node_header_states()
        return len(self.nodes)

    def build_decision_views(self) -> list[DecisionNode]:
        """Build UI-facing DecisionNode objects from persisted decision_node/decision_alt rows."""
        if not self.workspace_id:
            return []
        node_rows = self.decision_node_repo.list_nodes(self.workspace_id)
        if not node_rows:
            return []
        alt_rows = self.decision_alt_repo.list_workspace_alternates(self.workspace_id, include_rejected=True)
        alts_by_node: Dict[str, List[Dict[str, Any]]] = {}
        for row in (alt_rows or []):
            alts_by_node.setdefault(str(row.get("node_id", "") or ""), []).append(row)
        views: list[DecisionNode] = []
        for node_row in node_rows:
            node_id = str(node_row.get("node_id", "") or "")
            node = self._restore_node_from_rows(node_row, alts_by_node.get(node_id, []))
            views.append(node)
        return views

    def open_most_recent_workspace(self) -> int:
        wss = self.ws_repo.list(status="ACTIVE")
        if not wss:
            return 0
        return self.open_workspace(wss[0]["workspace_id"])
    
    def rematch_workspace_preserve_decisions(self) -> int:
        """
        Re-run matching for the CURRENT workspace using the current inventory + current npr_list,
        but preserve any existing decisions (status/locked/selected/rejected/notes).
        """
        if not self.workspace_id:
            raise RuntimeError("No workspace is open.")
        
        # Reload the authoritative decision state (including alternates + selected/rejected flags)
        # from the DB before re-running matching.
        wsid = self.workspace_id
        self.open_workspace(wsid)
    
        if not self.inventory:
            raise RuntimeError("Load inventory before re-running matching.")
    
        if not self.npr_list:
            raise RuntimeError("Workspace has no BOM loaded in memory; open workspace again or load BOM snapshot.")
    
        # Snapshot current decision-state nodes (authoritative)
        existing_nodes = list(self.nodes or [])
        existing_by_key = {self._node_key(n): n for n in existing_nodes}
    
        fresh_nodes = self._compute_fresh_nodes_via_run_matching()
    
        # Merge fresh suggestions into existing decisions
        merged: list[DecisionNode] = []
        for fresh in fresh_nodes:
            k = self._node_key(fresh)
            if k in existing_by_key:
                ex = existing_by_key[k]
                self._merge_suggestions_into_existing(ex, fresh)
                merged.append(ex)
            else:
                # New node (BOM row exists but no saved node yet)
                merged.append(fresh)
    
        self.nodes = merged
        return len(self.nodes)
    
    
    def _compute_fresh_nodes_via_run_matching(self) -> list:
        """
        Uses existing run_matching() to compute a fresh set of nodes,
        but returns them instead of leaving them as the authoritative self.nodes.
        """
        # Preserve current nodes while we run matching
        prev_nodes = self.nodes
        try:
            self.run_matching()           # this rebuilds self.nodes
            fresh_nodes = list(self.nodes or [])
            return fresh_nodes
        finally:
            # Restore; rematch_workspace_preserve_decisions() will set merged nodes after
            self.nodes = prev_nodes
    
    
    @staticmethod
    def _node_key(node) -> str:
        # DecisionNode has bom_uid (ROW-#). 
        k = (getattr(node, "bom_uid", "") or "").strip()
        if k:
            return k
        return (getattr(node, "id", "") or "").strip()
    
    
    @staticmethod
    def _alt_key(alt) -> tuple[str, str, str]:
        # Stable identity across reruns (ids will differ)
        src = (getattr(alt, "source", "") or "").strip().upper()
        ipn = (getattr(alt, "internal_part_number", "") or "").strip().upper()
        mpn = (getattr(alt, "manufacturer_part_number", "") or "").strip().upper()
        return (src, ipn, mpn)
    
    
    def _is_decided(self, node) -> bool:
        # Locked always means “do not touch”
        if bool(getattr(node, "locked", False)):
            return True
    
        st = getattr(node, "status", None)
        # Enum-safe + string-safe
        try:
            s = (getattr(st, "value", "") or getattr(st, "name", "") or str(st or "")).upper()
        except Exception:
            s = str(st or "").upper()
    
        return s not in ("", "NEEDS_DECISION", "OPEN")
    
 #------------------------------------#------------------------------------#------------------------------------#------------------------------------   
    def _merge_suggestions_into_existing(self, existing, fresh) -> None:
        """
        Refresh suggestion fields while preserving user decisions.
        """
    
        # Always refresh “suggestion/meta” fields
        existing.match_type = getattr(fresh, "match_type", existing.match_type)
        existing.confidence = float(getattr(fresh, "confidence", existing.confidence) or 0.0)
    
        fresh_explain = getattr(fresh, "explain", None)
        if isinstance(fresh_explain, dict):
            existing.explain = fresh_explain
    
        # Alternates: refresh list, preserve selected/rejected flags
        fresh_alts = list(getattr(fresh, "alternates", []) or [])
        old_alts = list(getattr(existing, "alternates", []) or [])
        old_by_key = {self._alt_key(a): a for a in old_alts}
    
        merged_alts = []
        for a in fresh_alts:
            k = self._alt_key(a)
            if k in old_by_key:
                old = old_by_key[k]
                a.selected = bool(getattr(old, "selected", False))
                a.rejected = bool(getattr(old, "rejected", False))
            merged_alts.append(a)
        existing.alternates = merged_alts
    
        # If decided, do NOT override chosen fields/status/notes/etc
        if self._is_decided(existing):
            return
    
        # If not decided yet, allow fresh “best guess” fields to update
        existing.internal_part_number = getattr(fresh, "internal_part_number", existing.internal_part_number)
        existing.inventory_mpn = getattr(fresh, "inventory_mpn", existing.inventory_mpn)
    

#------------------------------------#------------------------------------#------------------------------------#------------------------------------        
    def load_cns_preview(self, xlsx_path: str, last_n_per_sheet: int = 1) -> dict:
        """
        Loads CNS and prints:
          - every sheet name discovered
          - the LAST prefix-body (or last N) discovered per sheet

        last_n_per_sheet:
          1 -> just last
          2 -> last two, etc.
        """
        records = DataLoader.load_cns_workbook(xlsx_path)
        self.cns_records = records

        # sheet -> ordered unique PBs (preserve discovery order)
        sheet_pbs = defaultdict(list)
        sheet_seen = defaultdict(set)

        for r in records:
            sheet = (getattr(r, "sheet_name", "") or "").strip()
            prefix = (getattr(r, "prefix", "") or "").strip()
            body = (getattr(r, "body", "") or "").strip()
            if not sheet or not prefix or not body:
                continue

            pb = f"{prefix}-{body}"
            if not ControllerConfig.PB_RE.match(pb):
                continue

            if pb not in sheet_seen[sheet]:
                sheet_seen[sheet].add(pb)
                sheet_pbs[sheet].append(pb)

        all_sheets = sorted(sheet_pbs.keys())

        summary = {
            "total_records": len(records),
            "sheets_with_pbs": len(all_sheets),
            "unique_pb_total": sum(len(v) for v in sheet_pbs.values()),
            "sheets": all_sheets,
            "last_per_sheet": {
                s: (sheet_pbs[s][-max(1, int(last_n_per_sheet)):] if sheet_pbs[s] else [])
                for s in all_sheets
            },
            "unique_pb_per_sheet": {s: len(sheet_pbs[s]) for s in all_sheets},
        }
        return summary
 #------------------------------------#------------------------------------#------------------------------------#------------------------------------   
    def _merge_master_with_erp(self, master_inv: InventoryPart) -> InventoryPart:
        """
        Return an InventoryPart that keeps master identity/substitutes, but overlays
        ERP raw_fields for stock/cost display.
        """
        if not master_inv:
            return master_inv

        itemnum = (getattr(master_inv, "itemnum", "") or "").strip()
        erp = self._erp_by_itemnum.get(itemnum)

        if not erp:
            return master_inv  # no overlay available

        # Overlay: keep master object, but copy ERP raw_fields into it
        master_inv.raw_fields = dict(getattr(erp, "raw_fields", {}) or {})
        # If ERP has a better desc/mfg fields, optionally overlay those too:
        # master_inv.desc = master_inv.desc or erp.desc
        # master_inv.mfgname = master_inv.mfgname or erp.mfgname
        # master_inv.mfgid = master_inv.mfgid or erp.mfgid
        return master_inv

    def set_assigned_part_number(self, node_id: str, pn: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
    
        pn = (pn or "").strip()
        node.assigned_part_number = pn
    
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    # ----------------------------
    # Matching -> DecisionNodes
    # ----------------------------

    # ----------------------------
    # BOM section + Description overrides (UI-driven)
    # ----------------------------


    def _selected_committed_alt(self, node: DecisionNode) -> Optional[Alternate]:
        """Return the active selected alternate used by header/export state."""
        try:
            for alt in (getattr(node, "alternates", []) or []):
                print("this is inside of selcted_commited_alt")
                print(alt)
                if bool(getattr(alt, "selected", False)) and not bool(getattr(alt, "rejected", False)):
                    return alt
        except Exception:
            pass
        return None

    def _all_alternates_rejected(self, node: DecisionNode) -> bool:
        try:
            alts = list(getattr(node, "alternates", []) or [])
            return bool(alts) and all(bool(getattr(a, "rejected", False)) for a in alts)
        except Exception:
            return False

    def _header_manual_editable(self, node: DecisionNode, selected_alt: Optional[Alternate] = None) -> bool:
        if bool(getattr(node, "locked", False)):
            return False
        selected_alt = selected_alt if selected_alt is not None else self._selected_committed_alt(node)
        #print("_selected_committed_alt: inside of jeader_manuel_editable")
        if selected_alt is None:
            return True
        if self._all_alternates_rejected(node):
            return True
        return (str(getattr(selected_alt, "source", "") or "").strip().lower() != "inventory")

    def _compute_node_header_state(self, node: DecisionNode) -> NodeHeaderState:
        """Derive a controller-owned header snapshot from committed node state."""
        alts = list(getattr(node, "alternates", []) or [])
        selected_alt = self._selected_committed_alt(node)
        #print("_selected_committed_alt inside of _compute_node_header_state")

        all_rejected = self._all_alternates_rejected(node)

        locked = bool(getattr(node, "locked", False))
        status_text = str(getattr(getattr(node, "status", None), "value", getattr(node, "status", "")))
        is_ready = (status_text == "READY_FOR_EXPORT")

        if selected_alt is not None:
            pn_val = (getattr(selected_alt, "internal_part_number", "") or "").strip() or \
                     (getattr(node, "internal_part_number", "") or "").strip()
            desc_val = (getattr(selected_alt, "description", "") or "").strip() or \
                       (getattr(node, "description", "") or "").strip()
        else:
            pn_val = (getattr(node, "assigned_part_number", "") or "").strip()
            desc_val = (getattr(node, "description", "") or "").strip()

        suggested_val = (getattr(node, "suggested_pb", "") or "").strip()
        if pn_val:
            suggested_val = pn_val

        selected_is_internal = bool(
            selected_alt is not None and (getattr(selected_alt, "source", "") or "").strip().lower() == "inventory"
        )
        manual_unlock = self._header_manual_editable(node, selected_alt=selected_alt)
        pn_unlock = manual_unlock
        desc_unlock = manual_unlock

        controls = HeaderControlState(
            company_pn_editable=pn_unlock,
            apply_pn_enabled=pn_unlock,
            description_editable=desc_unlock,
            bom_section_editable=(not locked),
            approval_editable=(not locked),
            load_external_enabled=(not locked),
            mark_ready_enabled=(not is_ready),
            unmark_ready_enabled=is_ready,
            auto_reject_enabled=(not locked),
        )

        return NodeHeaderState(
            node_id=str(getattr(node, "id", "") or ""),
            title_text=f"Company PN: {pn_val or '—'}",
            subtitle_text=f"BOM MPN: {getattr(node, 'bom_mpn', '') or '—'}",
            status_text=status_text,
            company_part_number=pn_val,
            suggested_company_part_number=suggested_val,
            description_text=desc_val,
            bom_mpn=(getattr(node, "bom_mpn", "") or "").strip(),
            bom_section=str(getattr(node, "bom_section", "SURFACE MOUNT") or "SURFACE MOUNT"),
            include_approval=bool(getattr(node, "needs_approval", False)),
            committed_company_part_number=pn_val,
            committed_description_text=desc_val,
            committed_bom_section=str(getattr(node, "bom_section", "SURFACE MOUNT") or "SURFACE MOUNT"),
            committed_include_approval=bool(getattr(node, "needs_approval", False)),
            selected_alt_id=str(getattr(selected_alt, "id", "") or ""),
            selected_card_id=str(getattr(node, "focused_card_id", "") or ""),
            has_selected_alt=bool(selected_alt is not None),
            selected_is_internal=selected_is_internal,
            all_rejected=all_rejected,
            locked=locked,
            is_ready=is_ready,
            controls=controls,
        )

    def _update_header_apply_state(self, header: NodeHeaderState) -> NodeHeaderState:
        """Refresh header control toggles that depend on staged draft values."""
        controls = getattr(header, "controls", None)
        if controls is None:
            controls = HeaderControlState()
            header.controls = controls

        pn_val = (getattr(header, "company_part_number", "") or "").strip()
        pn_editable = bool(getattr(controls, "company_pn_editable", False))
        controls.apply_pn_enabled = pn_editable and bool(pn_val)
        return header

    def _merge_header_drafts(self, previous: Optional[NodeHeaderState], current: NodeHeaderState) -> NodeHeaderState:
        """Preserve staged upper-panel edits across node/card refreshes until they are applied."""
        if previous is None:
            return self._update_header_apply_state(current)

        if bool(getattr(previous, "dirty_company_part_number", False)):
            current.company_part_number = getattr(previous, "company_part_number", "") or ""
            current.dirty_company_part_number = True

        if bool(getattr(previous, "dirty_description", False)):
            current.description_text = getattr(previous, "description_text", "") or ""
            current.dirty_description = True

        if bool(getattr(previous, "dirty_bom_section", False)):
            current.bom_section = getattr(previous, "bom_section", "") or current.bom_section
            current.dirty_bom_section = True

        if bool(getattr(previous, "dirty_approval", False)):
            current.include_approval = bool(getattr(previous, "include_approval", False))
            current.dirty_approval = True

        return self._update_header_apply_state(current)

    def _ensure_header_state(self, node_id: str) -> NodeHeaderState:
        """Return a header state object for the node, creating it when needed."""
        return self.get_node_header_state(node_id, rebuild=False)

    def _sync_node_header_state(self, node: DecisionNode) -> NodeHeaderState:
        """Refresh the cached header state for a node and return it."""
        prev = self._header_state_by_node.get(str(getattr(node, "id", "") or ""))
        state = self._compute_node_header_state(node)
        state = self._merge_header_drafts(prev, state)
        self._header_state_by_node[state.node_id] = state
        return state

    def build_node_header_state(self, node_id: str) -> NodeHeaderState:
        """Materialize and cache the controller-owned header state for a node."""
        node = self.get_node(node_id)
        return self._sync_node_header_state(node)

    def get_node_header_state(self, node_id: str, rebuild: bool = False) -> NodeHeaderState:
        """Return the cached header state, rebuilding it when requested or missing."""
        if rebuild or node_id not in self._header_state_by_node:
            return self.build_node_header_state(node_id)
        return self._header_state_by_node[node_id]

    def rebuild_all_node_header_states(self) -> None:
        """Refresh upper-panel state for every node."""
        for node in list(self.nodes or []):
            try:
                self._sync_node_header_state(node)
            except Exception:
                continue

    def stage_header_company_pn(self, node_id: str, value: str) -> NodeHeaderState:
        """Stage a Company PN edit in the controller-owned header object without committing it yet."""
        header = self._ensure_header_state(node_id)
        header.company_part_number = (value or "").strip()
        header.dirty_company_part_number = (
            header.company_part_number != (getattr(header, "committed_company_part_number", "") or "").strip()
        )
        return self._update_header_apply_state(header)

    def apply_header_company_pn(self, node_id: str) -> NodeHeaderState:
        """Commit the staged Company PN from the header object into the node state."""
        header = self._ensure_header_state(node_id)
        self._update_header_apply_state(header)

        if not bool(getattr(header.controls, "company_pn_editable", False)):
            raise ValueError("Company Part Number cannot be edited for the current node state.")

        pn = (getattr(header, "company_part_number", "") or getattr(header, "suggested_company_part_number", "") or "").strip()
        if not pn:
            raise ValueError("Enter a Company Part Number first.")

        self.set_assigned_part_number(node_id, pn)
        fresh = self.get_node_header_state(node_id, rebuild=True)
        fresh.company_part_number = pn
        fresh.committed_company_part_number = pn
        fresh.dirty_company_part_number = False
        return self._update_header_apply_state(fresh)

    def stage_header_description(self, node_id: str, value: str) -> NodeHeaderState:
        """Stage a description edit in the controller-owned header object."""
        header = self._ensure_header_state(node_id)
        header.description_text = (value or "").strip()
        header.dirty_description = (
            header.description_text != (getattr(header, "committed_description_text", "") or "").strip()
        )
        return header

    def apply_header_description(self, node_id: str, value: Optional[str] = None) -> NodeHeaderState:
        """Commit the staged description into the node/export state and return the refreshed header object."""
        if value is not None:
            self.stage_header_description(node_id, value)

        header = self._ensure_header_state(node_id)
        if not bool(getattr(header.controls, "description_editable", False)):
            raise ValueError("Description cannot be edited for the current node state.")

        self.set_node_description(node_id, getattr(header, "description_text", "") or "")
        fresh = self.get_node_header_state(node_id, rebuild=True)
        fresh.description_text = getattr(fresh, "committed_description_text", getattr(fresh, "description_text", "") or "")
        fresh.dirty_description = False
        return fresh

    def stage_header_bom_section(self, node_id: str, value: str) -> NodeHeaderState:
        """Stage a BOM section selection in the header object."""
        header = self._ensure_header_state(node_id)
        header.bom_section = (value or "").strip() or "SURFACE MOUNT"
        header.dirty_bom_section = (
            header.bom_section != (getattr(header, "committed_bom_section", "") or "SURFACE MOUNT").strip()
        )
        return header

    def apply_header_bom_section(self, node_id: str, value: Optional[str] = None) -> NodeHeaderState:
        """Commit the staged BOM section into node explain/export state."""
        if value is not None:
            self.stage_header_bom_section(node_id, value)

        header = self._ensure_header_state(node_id)
        if not bool(getattr(header.controls, "bom_section_editable", False)):
            raise ValueError("BOM section cannot be edited for the current node state.")

        section = self.set_node_bom_section(node_id, getattr(header, "bom_section", "") or "SURFACE MOUNT")
        # set_node_bom_section does not currently persist node state, so refresh header state explicitly here.
        node = self.get_node(node_id)
        self._sync_node_header_state(node)
        fresh = self.get_node_header_state(node_id, rebuild=False)
        fresh.bom_section = section
        fresh.committed_bom_section = section
        fresh.dirty_bom_section = False
        return fresh

    def stage_header_approval(self, node_id: str, include_on_approval_sheet: bool) -> NodeHeaderState:
        """Stage approval-sheet inclusion in the header object."""
        header = self._ensure_header_state(node_id)
        header.include_approval = bool(include_on_approval_sheet)
        header.dirty_approval = (
            bool(header.include_approval) != bool(getattr(header, "committed_include_approval", False))
        )
        return header

    def apply_header_approval(self, node_id: str, include_on_approval_sheet: Optional[bool] = None) -> NodeHeaderState:
        """Commit approval-sheet inclusion into node state."""
        if include_on_approval_sheet is not None:
            self.stage_header_approval(node_id, include_on_approval_sheet)

        header = self._ensure_header_state(node_id)
        if not bool(getattr(header.controls, "approval_editable", False)):
            raise ValueError("Approval export state cannot be edited for the current node state.")

        self.set_node_approval_export(node_id, bool(getattr(header, "include_approval", False)))
        fresh = self.get_node_header_state(node_id, rebuild=True)
        fresh.include_approval = bool(getattr(fresh, "committed_include_approval", getattr(fresh, "include_approval", False)))
        fresh.dirty_approval = False
        return fresh

    def mark_ready_from_header(self, node_id: str) -> NodeHeaderState:
        """Commit Mark Ready from the upper-panel action layer and return the refreshed header object."""
        self.mark_ready(node_id)
        return self.get_node_header_state(node_id, rebuild=True)

    def unmark_ready_from_header(self, node_id: str) -> NodeHeaderState:
        """Commit Unmark Ready from the upper-panel action layer and return the refreshed header object."""
        self.unmark_ready(node_id)
        return self.get_node_header_state(node_id, rebuild=True)

    def auto_reject_all_from_header(self, node_id: str) -> NodeHeaderState:
        """Reject all active alternates for a node through the upper-panel action layer."""
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        for alt in list(getattr(node, "alternates", []) or []):
            if bool(getattr(alt, "rejected", False)):
                continue
            try:
                self.reject_alternate(node.id, alt.id)
            except Exception:
                continue

        return self.get_node_header_state(node_id, rebuild=True)

    def get_node_bom_section(self, node_id: str) -> str:
        """Return the export bucket/section for this node (SURFACE MOUNT, THROUGH-HOLE, AUXILIARY - ...)."""
        node = self.get_node(node_id)
        try:
            sec = str(getattr(node, "bom_section", "") or "").strip()
            if sec:
                return sec
        except Exception:
            pass
        try:
            parsed = getattr(node, "parsed", None) or {}
            sec = str((parsed or {}).get("type", "") or "").strip()
            if sec:
                return sec
        except Exception:
            pass
        return "SURFACE MOUNT"


    def set_node_bom_section(self, node_id: str, section: str) -> str:
        """Set the BOM export section for a node and persist it with the node decision state."""
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
        section = (section or "").strip() or "SURFACE MOUNT"
        node.bom_section = section
        self._persist_node_and_alts(node)
        return section

    def set_node_approval_export(self, node_id: str, include_on_approval_sheet: bool) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        node.needs_approval = bool(include_on_approval_sheet)
        self._persist_node_and_alts(node)

    def get_exclude_customer_part_number_in_npr(self, node_id: str) -> bool:
        node = self.get_node(node_id)
        return bool(getattr(node, "exclude_customer_part_number_in_npr", False))

    def set_exclude_customer_part_number_in_npr(self, node_id: str, exclude_value: bool) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
        node.exclude_customer_part_number_in_npr = bool(exclude_value)
        self._persist_node_and_alts(node)

    def get_internal_umbrella_count(self, node_id: str) -> int:
        node = self.get_node(node_id)
        cpn = (getattr(node, "internal_part_number", "") or "").strip()
        if not cpn:
            return 0

        inv_obj = self._inv_by_itemnum.get(cpn)
        if inv_obj is not None:
            all_mpns = list(getattr(inv_obj, "_all_mpns", []) or [])
            if all_mpns:
                return len([x for x in all_mpns if str(x or "").strip()])
            rep = str(getattr(inv_obj, "vendoritem", "") or "").strip()
            subs = [str(x or "").strip() for x in (getattr(inv_obj, "substitutes", []) or []) if str(x or "").strip()]
            vals = []
            if rep:
                vals.append(rep)
            vals.extend(subs)
            return len(vals)

        row_company = None
        try:
            row_company = self.inv_company_repo.get(self.workspace_id, cpn)
        except Exception:
            row_company = None
        if not row_company:
            try:
                row_company = self.inv_company_repo.get(self._inventory_store_workspace_id, cpn)
            except Exception:
                row_company = None
        if not row_company:
            return 0

        alts = row_company.get("alternates_json", []) or []
        if isinstance(alts, str):
            try:
                alts = json.loads(alts)
            except Exception:
                alts = []
        return len([
            a for a in (alts or [])
            if str((a or {}).get("mpn") or (a or {}).get("mfgpn") or (a or {}).get("manufacturer_part_number") or "").strip()
        ])

    #def set_node_description(self, node_id: str, description: str) -> None:
    #    """
    #    Persist the user-edited header description for export.
#
    #    Storage:
    #      - saved to node.explain["export_description_override"]
    #      - node.explain["export_description_override_applied"] = True
    #      - original BOM description preserved once in node.explain["input_description"]
#
    #    NOTE:
    #      Do NOT patch {"description": ...} into bom_line_state. That column doesn't exist.
    #      Persist via explain_json through _persist_node_and_alts().
    #    """
    #    node = self.get_node(node_id)
    #    self._ensure_unlocked(node)
#
    #    # Preserve original BOM description before overwriting
    #    prior_desc = (getattr(node, "description", "") or "").strip()
#
    #    new_desc = (description or "").strip()
#
    #    node.explain = dict(getattr(node, "explain", {}) or {})
    #    node.explain.setdefault("input_description", prior_desc)
#
    #    node.explain["export_description_override"] = new_desc
    #    node.explain["export_description_override_applied"] = True
#
    #    # Update live node description so UI reflects what's going to export
    #    node.description = new_desc
#
    #    self._recompute_node_flags(node)
    #    self._persist_node_and_alts(node)

    def set_node_description(self, node_id: str, description: str) -> None:
        """
        Persist the current header description text for this node.
        This is the export source of truth.
        """
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        new_desc = (description or "").strip()

        # Ensure explain exists so we can persist in bom_line_state.explain_json
        node.explain = dict(getattr(node, "explain", {}) or {})

        # Preserve original BOM description once (for your own reference/debug)
        node.explain.setdefault("input_description", (getattr(node, "description", "") or "").strip())

        # Store the current export description explicitly
        node.explain["export_description"] = new_desc

        # Keep node.description in sync with the header box
        node.description = new_desc

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def _selected_inventory_alt(self, node: DecisionNode) -> Optional[Alternate]:
        """Return the active inventory card for the node's anchored CPN, preferring an explicitly selected card."""
        ipn = (getattr(node, "internal_part_number", "") or "").strip()
        if not ipn:
            return None

        selected_inv = None
        try:
            for a in (node.selected_alternates() or []):
                if (getattr(a, "source", "") or "").lower() != "inventory":
                    continue
                if (getattr(a, "internal_part_number", "") or "").strip() != ipn:
                    continue
                if bool(getattr(a, "rejected", False)):
                    continue
                selected_inv = a
                break
        except Exception:
            selected_inv = None

        if selected_inv is not None:
            return selected_inv

        inv_mpn = (getattr(node, "inventory_mpn", "") or "").strip().lower()
        for a in (getattr(node, "alternates", []) or []):
            if (getattr(a, "source", "") or "").lower() != "inventory":
                continue
            if (getattr(a, "internal_part_number", "") or "").strip() != ipn:
                continue
            if bool(getattr(a, "rejected", False)):
                continue
            ampn = (getattr(a, "manufacturer_part_number", "") or "").strip().lower()
            if inv_mpn and ampn == inv_mpn:
                return a

        for a in (getattr(node, "alternates", []) or []):
            if (getattr(a, "source", "") or "").lower() == "inventory" and (getattr(a, "internal_part_number", "") or "").strip() == ipn:
                if not bool(getattr(a, "rejected", False)):
                    return a
        return None


    

    #def _export_description_for_node(self, node: DecisionNode) -> str:
    #    """
    #    Export description rules:
#
    #    1) If user edited header description (override applied), export that.
    #    2) Else if we can see a selected card (inventory or non-inventory), export that card's description.
    #    3) Else if selection exists only as persisted state (common after reload), resolve inventory desc by:
    #         - selected inventory CPN (node.internal_part_number or state cpn) OR
    #         - selected_mpn/selected_mfg
    #       and export the inventory description.
    #    4) Else fallback to node.description.
    #    """
    #    ex = getattr(node, "explain", {}) or {}
#
    #    # 1) Manual override wins when applied
    #    if bool(ex.get("export_description_override_applied")):
    #        d = str(ex.get("export_description_override") or "").strip()
    #        if d:
    #            return d
#
    #    # 2) In-memory selected card wins
    #    inv_alt = self._selected_inventory_alt(node)
    #    if inv_alt is not None:
    #        d = str(getattr(inv_alt, "description", "") or "").strip()
    #        if d:
    #            return d
#
    #    try:
    #        for a in (node.selected_alternates() or []):
    #            if bool(getattr(a, "rejected", False)):
    #                continue
    #            d = str(getattr(a, "description", "") or "").strip()
    #            if d:
    #                return d
    #    except Exception:
    #        pass
#
    #    # 3) Persisted-state resolution (important for export after reload)
    #    # 3a) Try by selected/anchored CPN
    #    cpn = (getattr(node, "internal_part_number", "") or "").strip()
    #    if cpn:
    #        try:
    #            inv = None
    #            if hasattr(self, "_inv_by_itemnum") and isinstance(self._inv_by_itemnum, dict):
    #                inv = self._inv_by_itemnum.get(cpn)
    #            if inv is not None:
    #                d = str(getattr(inv, "description", "") or "").strip()
    #                if d:
    #                    return d
    #                # Some InventoryPart variants store raw desc
    #                rf = getattr(inv, "raw_fields", {}) or {}
    #                d2 = str(rf.get("description") or rf.get("item_description") or "").strip()
    #                if d2:
    #                    return d2
    #        except Exception:
    #            pass
#
    #    # 3b) Try by selected MPN/MFG (if you persist these in bom_line_state)
    #    sel_mpn = (getattr(node, "selected_mpn", "") or "").strip()
    #    sel_mfg = (getattr(node, "selected_mfg", "") or "").strip()
    #    if sel_mpn:
    #        try:
    #            # Scan inventory for matching mpn/mfg if you don't have an index.
    #            # This is only used when the selected alt list isn't rebuilt.
    #            inv_list = getattr(self, "inventory", []) or []
    #            for inv in inv_list:
    #                mpn = (getattr(inv, "manufacturer_part_number", "") or "").strip()
    #                mfg = (getattr(inv, "manufacturer", "") or "").strip()
    #                if mpn and mpn == sel_mpn and (not sel_mfg or mfg == sel_mfg):
    #                    d = str(getattr(inv, "description", "") or "").strip()
    #                    if d:
    #                        return d
    #                    rf = getattr(inv, "raw_fields", {}) or {}
    #                    d2 = str(rf.get("description") or rf.get("item_description") or "").strip()
    #                    if d2:
    #                        return d2
    #                    break
    #        except Exception:
    #            pass
#
    #    # 4) Fallback
    #    return str(getattr(node, "description", "") or "").strip()


    def _export_description_for_node(self, node: DecisionNode) -> str:
        # Single source of truth: whatever is stored on the node.
        return str(getattr(node, "description", "") or "").strip()    
    
    def _resolve_inv(self, obj):
        """
        Ensure we always use the canonical InventoryPart from the master inventory
        (the one that has substitutes attached), even if the matching engine handed
        us a lightweight/copy candidate object.
        """
        if obj is None:
            return None
        try:
            itemnum = str(safe_get(obj, "itemnum", "internal_part_number", default="") or "").strip()
        except Exception:
            itemnum = ""
        if not itemnum:
            return obj
        return self._inv_by_itemnum.get(itemnum, obj)

    def run_matching(self) -> int:
        """
        Run matching for all NPR/BOM parts.
    
        IMPORTANT: Customer-provided alternates are matched, but the winning attempt is selected
        deterministically by (tier_rank, tier_quality, stock, stable_order) — NOT by confidence.
        Confidence remains a UI/display artifact.
        """
        self.reset_stop_matching()
        if not self.inventory or not self.npr_list:
            raise RuntimeError("Load inventory and NPR/BOM parts before matching.")
#------------------------------------    
        cfg = load_config(self.cfg.components_yaml_path)
    
        # Parse both sides (robust attr names)
        for inv in self.inventory:
            if self._should_stop():
                return 0
            inv.parsed = parse_description(safe_get(inv, "description", "desc", default=""), cfg)
    
        # IMPORTANT: preserve loader-provided mpn_alts before overwriting parsed
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            existing_parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = []
            if isinstance(existing_parsed, dict):
                mpn_alts = existing_parsed.get("mpn_alts", []) or []
    
            new_parsed = parse_description(safe_get(npr, "description", "desc", default=""), cfg)
            if not isinstance(new_parsed, dict):
                new_parsed = {}
    
            if mpn_alts:
                new_parsed["mpn_alts"] = mpn_alts
    
            npr.parsed = new_parsed
    
        def uniq_keep_order(items):
            seen = set()
            out = []
            for x in items:
                x = (x or "").strip()
                if x and x not in seen:
                    seen.add(x)
                    out.append(x)
            return out

            # Determine first empty row after header in NPR sheet
            def _npr_first_empty_row():
                r = npr_header_row + 1
                while r <= npr_ws.max_row:
                    if all(str(npr_ws.cell(row=r, column=cc).value or "").strip() == "" for cc in range(1, 6)):
                        return r
                    r += 1
                return npr_ws.max_row + 1

            npr_write_row = _npr_first_empty_row()
    
        def _match_type_key(match: Any) -> str:
            mt = safe_get(match, "match_type", default="")
            if hasattr(mt, "value"):
                mt = mt.value
            mt = str(mt or "")
            return mt.replace("MatchType.", "").strip().upper()
    
        def _tier_rank(mt_key: str) -> int:
            ranks = {
                "EXACT_MFG_PN": 600,
                "API_ASSISTED": 550,
                "SUBSTITUTE": 500,
                "PREFIX_FAMILY": 400,
                "PARTIAL_ITEMNUM": 300,
                "PARSED_MATCH": 200,
                "NO_MATCH": 0,
            }
            return int(ranks.get(mt_key, 0))
    
        def _tier_quality(match: Any, mt_key: str) -> float:
            if mt_key == "PARSED_MATCH":
                exp = safe_get(match, "explain", default=None) or {}
                if isinstance(exp, dict):
                    q = exp.get("best_ratio", None)
                    if q is None:
                        q = exp.get("match_ratio", None)
                    if q is None:
                        q = exp.get("ratio", None)
                    try:
                        return float(q or 0.0)
                    except Exception:
                        return 0.0
                return 0.0
    
            inv = safe_get(match, "inventory_part", default=None)
            return 1.0 if inv is not None and mt_key != "NO_MATCH" else 0.0
    
        def _inv_stock(inv: Any) -> float:
            if inv is None:
                return 0.0
            for k in ("stock", "qty", "onhand", "available"):
                v = safe_get(inv, k, default=None)
                if v is None:
                    continue
                try:
                    return float(v or 0.0)
                except Exception:
                    continue
            return 0.0
    
        # Grab the active Tk root (if UI is running). IMPORTANT: do this
        # before constructing the engine so we can pass ui_root correctly.
        try:
            import tkinter as _tk
            tk_root = getattr(_tk, "_default_root", None)
        except Exception:
            tk_root = None

        #================================
        # CALLING MATCHING ENGINE
        #=================================
        engine = MatchingEngine(
            self.inventory,
            config=cfg,
            ui_root=tk_root,
            stop_event=getattr(self, "stop_event", None),
        )

        phase_cb = getattr(tk_root, "loading_phase_callback", None) if tk_root else None
        progress_cb = getattr(tk_root, "loading_progress_callback", None) if tk_root else None

        # Phase: embeddings
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Building semantic cache...", True))
            except Exception:
                pass            

        # Build semantic cache (sync; run_matching is already in a worker thread)
        engine.ensure_embeddings_cache()

        # Phase: matching
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Matching Engine is matching parts...", True))
            except Exception:
                pass

        match_pairs = []
    
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            primary_mpn = safe_get(npr, "mfgpn", "bom_mpn", default="")
    
            mpn_options = uniq_keep_order([primary_mpn] + list(mpn_alts))
    
            attempts = []
            merged_candidates = []
    
            for idx, mpn in enumerate(mpn_options or [""]):
                if self._should_stop():
                    return 0
                tmp = copy.copy(npr)
                if hasattr(tmp, "mfgpn"):
                    tmp.mfgpn = mpn
                tmp.parsed = getattr(npr, "parsed", {}) or {}
    
                # --- alternates DB lookup (BOM MFGPN -> base internal item) ---
                match = None
                if getattr(self, "_alt_loaded", False) and getattr(self, "_alt_mpn_to_base", None):
                    key = DataLoader.norm_mpn_key(mpn)
                    bases = self._alt_mpn_to_base.get(key, []) if key else []

                    if bases:
                        inv_choices = [self._inv_by_itemnum.get(b) for b in bases]
                        inv_choices = [x for x in inv_choices if x is not None]

                        if inv_choices:
                            def _stock(x):
                                try:
                                    return float(getattr(x, "stock", 0) or 0)
                                except Exception:
                                    return 0.0

                            inv_choices.sort(key=_stock, reverse=True)
                            inv = inv_choices[0]
                            match = MatchResult(
                                match_type=MatchType.SUBSTITUTE,
                                confidence=1.0,
                                inventory_part=inv,
                                candidates=[inv],
                                notes="Resolved via master inventory (MPN->base)",
                                explain={
                                    "substitute_mpn": mpn,
                                    "conflict": (len(bases) > 1),
                                    "bases": bases,
                                    "base_itemnum": getattr(inv, "itemnum", None),
                                },
                            )

                # --- Fall back to normal ML matching if not resolved ---
                if match is None:
                    # Tell terminal + UI which part we're on (helps diagnose "stuck at 13%")
                    label = (getattr(tmp, "npr_item", "") or "").strip()
                    mpn = (getattr(tmp, "mfgpn", "") or "").strip()
                    desc = (getattr(tmp, "description", "") or "").strip()
                    short = (desc[:60] + "...") if len(desc) > 60 else desc

                    print(f"[MATCH] {n_i+1}/{total_npr} starting ML match: {label} | {mpn} | {short}")

                    if phase_cb and tk_root:
                        try:
                            if tk_root and tk_root.winfo_exists():
                                tk_root.after(
                                    0,
                                    lambda msg=f"Matching {n_i+1}/{total_npr}: {label or mpn or short}": phase_cb(msg)
                                )
                        except Exception:
                            pass

                    t_part = time.time()
                    match = engine.match_single_part(tmp)
                    dt = time.time() - t_part

                    mt = safe_get(match, "match_type", default=None)
                    print(f"[MATCH] {n_i+1}/{total_npr} done in {dt:.2f}s (match_type={mt})")

                mt_key = _match_type_key(match)
                tier = _tier_rank(mt_key)
                quality = _tier_quality(match, mt_key)
    
                inv = safe_get(match, "inventory_part", default=None)
                stock = _inv_stock(inv)
    
                cands = safe_get(match, "candidates", "candidate_parts", default=None) or []
                if inv is not None:
                    merged_candidates.append(inv)
                merged_candidates.extend(list(cands))
    
                inv_id = ""
                if inv is not None:
                    inv_id = str(
                        safe_get(inv, "itemnum", "internal_part_number", "vendoritem", default="") or ""
                    ).strip()
    
                attempts.append(
                    {
                        "customer_mpn": mpn,
                        "idx": idx,
                        "pair": (npr, match),  # keep ORIGINAL npr for node display
                        "match_type": mt_key,
                        "tier": tier,
                        "quality": quality,
                        "stock": stock,
                        "inv_id": inv_id,
                        "candidate_count": len(cands),
                    }
                )
    
            if not attempts:
                continue
            
            # Deterministic winner (no confidence)
            def _attempt_sort_key(a: dict) -> tuple:
                # prefer primary mpn on ties => earlier idx wins
                prefer_primary = -int(a.get("idx", 0))
                return (a["tier"], a["quality"], a["stock"], prefer_primary, a["inv_id"])
    
            winner = max(attempts, key=_attempt_sort_key)
            best_used_mpn = winner["customer_mpn"]
            npr_part, best_match = winner["pair"]
    
            # Stash customer MPN context and attempt audit for UI/debug
            try:
                if not isinstance(best_match.explain, dict):
                    best_match.explain = {}
                best_match.explain["customer_mpns"] = list(mpn_options)
                best_match.explain["winning_mpn"] = best_used_mpn
                best_match.explain["attempts"] = [
                    {
                        "customer_mpn": a["customer_mpn"],
                        "match_type": a["match_type"],
                        "tier": a["tier"],
                        "quality": a["quality"],
                        "stock": a["stock"],
                        "inv_id": a["inv_id"],
                        "candidate_count": a["candidate_count"],
                        "is_winner": a is winner,
                    }
                    for a in attempts
                ]
            except Exception:
                pass
            
            # Attach merged candidates back onto chosen match (best-effort)
            try:
                existing = safe_get(best_match, "candidates", "candidate_parts", default=None) or []

                # 1) raw de-dupe so we don't waste work.
                seen = set()
                raw = []
                for cand in (list(existing) + list(merged_candidates)):
                    key = safe_get(cand, "itemnum", "internal_part_number", "vendoritem", default=str(cand))
                    if key in seen:
                        continue
                    seen.add(key)
                    raw.append(cand)

                # 2) Sort by engine score and apply a loose raw cap (performance guard).
                raw = sorted(raw, key=lambda c: float(getattr(c, "_pc_score", 0.0) or 0.0), reverse=True)

                # Optional: raw cap BEFORE collapsing, to keep big BOMs snappy.
                RAW_CAP = max(int(getattr(ControllerConfig, "MAX_INTERNAL_CANDIDATES", 10) or 10) * 5, 50)
                raw = raw[:RAW_CAP]

                # 3) Collapse by company PN so UI shows one card per itemnum.
                collapsed, meta = self._collapse_candidates_by_itemnum(raw)

                # 4) Apply the real UI cap (top N itemnums).
                collapsed = collapsed[: ControllerConfig.MAX_INTERNAL_CANDIDATES]

                # ---------------------------------------------------------
                # stamp matched MPN onto collapsed inventory candidates
                # ---------------------------------------------------------
                winning_mpn = (best_match.explain.get("winning_mpn") or "").strip().lower()

                for inv in collapsed:
                    # InventoryPart → Alternate mapping happens later,
                    # but we pre-stamp the InventoryPart so the UI Alternate inherits it.
                    try:
                        # If this InventoryPart was resolved via master MPN mapping,
                        # mark it with the winning BOM MPN.
                        inv._matched_mpn = winning_mpn
                    except Exception:
                        pass

                # 5) Write back to match object + stash explain meta for UI/spec panels.
                if hasattr(best_match, "candidates"):
                    best_match.candidates = collapsed
                elif hasattr(best_match, "candidate_parts"):
                    best_match.candidate_parts = collapsed
                else:
                    # last resort: attach attribute
                    best_match.candidates = collapsed

                if not hasattr(best_match, "explain") or not isinstance(best_match.explain, dict):
                    best_match.explain = {}
                best_match.explain["collapsed_groups_by_itemnum"] = meta

            except Exception:
                pass

            match_pairs.append((npr_part, best_match))

            # Update UI progress for matching phase
            if progress_cb and tk_root:
                ratio = (n_i + 1) / total_npr   
                try:
                    if tk_root and tk_root.winfo_exists():
                        tk_root.after(0, lambda r=ratio: progress_cb(r))
                except Exception:
                    pass

        # ---  sync winning_mpn and attempt flags ---
        try:
            if not hasattr(best_match, "explain") or not isinstance(best_match.explain, dict):
                best_match.explain = {}
            winning_mpn = best_match.explain.get("winning_mpn", "")
            attempts = best_match.explain.get("attempts", [])
            for a in attempts:
                a["is_winner"] = (a.get("customer_mpn") == winning_mpn)
            best_match.explain["attempts"] = attempts

            # NEW: mark the winning inventory part explicitly
            if hasattr(best_match, "inventory_part") and best_match.inventory_part:
                inv = best_match.inventory_part
                if hasattr(inv, "vendoritem"):
                    inv._is_winner = True
                elif hasattr(inv, "manufacturer_part_number"):
                    inv._is_winner = True
        except Exception as e:
            print(f"[CTRL PATCH] failed to sync winner info: {e}")

        # --- Save explain JSON for debugging ---
        debug_dir = Path(os.getenv("NPR_DEBUG_DIR", "debug_match"))
        debug_dir.mkdir(parents=True, exist_ok=True)

        debug_parse = str(os.getenv("NPR_DEBUG_PARSE", "")).strip().lower() not in ("", "0", "false", "no")
        debug_save  = str(os.getenv("NPR_DEBUG_SAVE", "")).strip().lower()  not in ("", "0", "false", "no")
        debug_jsonl_only = str(os.getenv("NPR_DEBUG_JSONL_ONLY", "")).strip().lower() in ("1", "true", "yes")

        if debug_parse or debug_save:
            try:
                jsonl_path = debug_dir / "explain_all.jsonl"

                def _inv_to_payload_dict(inv_obj: Any, *, is_winner: bool = False) -> dict:
                    if inv_obj is None:
                        return {
                            "inv_item": "",
                            "inv_desc": "",
                            "vendor_mpn": "",
                            "mfg": "",
                            "stock": 0,
                            "seed": 0.0,
                            "score": 0.0,
                            "is_winner": bool(is_winner),
                        }

                    inv_item = str(safe_get(inv_obj, "itemnum", "internal_part_number", default="") or "").strip()
                    inv_desc = str(safe_get(inv_obj, "desc", "description", default="") or "").strip()
                    vendor_mpn = str(safe_get(inv_obj, "vendoritem", "manufacturer_part_number", default="") or "").strip()
                    mfg = str(safe_get(inv_obj, "manufacturer", "manufacturer_name", "mfgname", default="") or "").strip()

                    stock = 0
                    for k in ("stock", "qty", "onhand", "available"):
                        v = safe_get(inv_obj, k, default=None)
                        if v is None:
                            continue
                        try:
                            stock = int(float(v or 0))
                            break
                        except Exception:
                            continue
                        
                    seed = safe_get(inv_obj, "_pc_seed", default=None)
                    if seed is None:
                        seed = safe_get(inv_obj, "confidence", default=0.0)
                    seed = clamp01(seed)

                    score = safe_get(inv_obj, "_pc_score", default=None)
                    if score is None:
                        score = seed
                    score = clamp01(score)

                    return {
                        "inv_item": inv_item,
                        "inv_desc": inv_desc,
                        "vendor_mpn": vendor_mpn,
                        "mfg": mfg,
                        "stock": stock,
                        "seed": float(seed),
                        "score": float(score),
                        "is_winner": bool(is_winner),
                    }

                for npr_part, match in match_pairs:
                    if self._should_stop():
                        return 0
                    exp = safe_get(match, "explain", default=None) or {}

                    npr_key = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default="NPR") or "NPR").strip()
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

                    bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default="") or "").strip()
                    npr_desc = str(safe_get(npr_part, "description", "desc", default="") or "")

                    winner_inv = safe_get(match, "inventory_part", default=None)
                    winner_item = str(safe_get(winner_inv, "itemnum", "internal_part_number", default="") or "").strip()

                    # Pull candidates from the chosen match (already merged+filtered+capped earlier)
                    cands = safe_get(match, "candidates", "candidate_parts", default=None) or []

                    # Ensure winner is present in candidates (training needs it)
                    
                    if winner_inv is not None:
                        found = False
                        for c in cands:
                            cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                            if cid and cid == winner_item:
                                found = True
                                break
                        if not found:
                            cands = [winner_inv] + list(cands)

                    cand_payloads = []
                    for c in cands:
                        cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                        cand_payloads.append(_inv_to_payload_dict(c, is_winner=(cid == winner_item)))

                    # Find winner rank in candidate list (0-based). If not found, -1.
                    winner_rank = -1
                    for i, cp in enumerate(cand_payloads):
                        if cp.get("is_winner"):
                            winner_rank = i
                            break
                        
                    payload = {
                        "ts": ts,

                        # identifiers
                        "npr_item": npr_key,
                        "bom_mpn": bom_mpn,

                        # main text fields
                        "npr_desc": npr_desc,

                        # deterministic winner choice context
                        "match_type": str(safe_get(match, "match_type", default="")),
                        "winning_mpn": exp.get("winning_mpn") if isinstance(exp, dict) else "",
                        "customer_mpns": exp.get("customer_mpns") if isinstance(exp, dict) else [],

                        # winner block
                        "winner": _inv_to_payload_dict(winner_inv, is_winner=True),
                        "winner_rank": winner_rank,

                    
                        "candidates": cand_payloads,

                        "explain": exp,
                    }

                    with open(jsonl_path, "a", encoding="utf-8") as f:
                        f.write(json.dumps(payload, default=str) + "\n")

                    if not debug_jsonl_only:
                        inv_item = payload["winner"]["inv_item"] or "INV"
                        out_path = debug_dir / f"explain_{ts}_{npr_key}__{inv_item}.json"
                        with open(out_path, "w", encoding="utf-8") as f:
                            json.dump(payload, f, indent=2, default=str)
                        print(f"[MATCH DBG] wrote explain json -> {out_path}")

                print(f"[MATCH DBG] wrote JSONL -> {jsonl_path}")

            except Exception as e:
                print(f"[MATCH DBG] failed to write explain json/jsonl: {e}")
        
        self.match_pairs = match_pairs
    
        print("[CTRL DBG] building nodes...")
        self.nodes = [self._pair_to_node(npr_part, match, i) for i, (npr_part, match) in enumerate(self.match_pairs, start=1)]
        # Keep node index in sync for UI selection
        self._nodes_workspace_id = getattr(self, 'workspace_id', None)
        self._node_index = {n.id: n for n in (self.nodes or [])}
        print(f"[CTRL DBG] nodes built: {len(self.nodes)}")

        # ----------------------------
        # DB: persist nodes after matching
        # ----------------------------
        try:
            if self._db_enabled():
                self._persist_all_nodes()
        except Exception as e:
            print(f"[DB] failed to persist nodes after matching: {e}")

        return len(self.nodes)
    

    #=====================================================
    # CNS matching. this is very experimental and optional. as of now ALOT of it is hard coded and being wired in.
    #=======================================================
    def _build_type_prefix_map(self) -> dict[str, str]:
        print("[CTRL DBG] ENTER _build_type_prefix_map")
        """
        Learn which CNS prefix is most common for each parsed component type,
        using existing inventory parts that already have a company PN.
        """
        counts = defaultdict(Counter)

        for inv in (self.inventory or []):
            pn = getattr(inv, "itemnum", "") or ""
            prefix, body, _ = _split_company_pn(pn)
            if not prefix or not body:
                continue

            p = getattr(inv, "parsed", {}) or {}
            ptype = str(getattr(inv, "_pc_ptype", "") or p.get("type") or "OTHER")
            ptype = ptype.upper().strip()
            counts[ptype][prefix] += 1

        out = {}
        for ptype, ctr in counts.items():
            out[ptype] = ctr.most_common(1)[0][0]

        #print("[CTRL DBG] EXIT _build_type_prefix_map")
        return out
    
    def _suggest_new_body_for_prefix(self, prefix: str) -> str:
        """
        Suggest a new 5-digit body:
          - find max(body) among rows with a non-empty description
          - propose next, skipping over bodies that already exist WITH a description
          - allowing bodies that exist with blank descriptions
        """
        prefix = (prefix or "").strip()
        if not prefix:
            return ""
    
        def field(obj, name: str, default=""):
            # supports dataclass/objects AND dicts
            if isinstance(obj, dict):
                return obj.get(name, default)
            return getattr(obj, name, default)
    
        bodies = {}  # body_int -> has_desc(bool)
    
        for r in (getattr(self, "cns_records", None) or []):
            rp = str(field(r, "prefix", "")).strip()
            if rp != prefix:
                continue
            
            rb = str(field(r, "body", "")).strip()
            if not rb.isdigit():
                continue
            
            desc = str(field(r, "description", "")).strip()
            bodies[int(rb)] = bool(desc)
    
        if not bodies:
            return "10000"
    
        max_desc_body = max((b for b, has_desc in bodies.items() if has_desc), default=max(bodies))
        cand = max_desc_body + 1
    
        while cand in bodies and bodies[cand] is True:
            cand += 1
    
        return f"{cand:05d}"
    
    def _apply_cns_suggestion_to_node(self, node) -> None:
        print(f"[CTRL DBG] ENTER _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)}")
        # 1) If EXISTS: derive from internal PN
        if getattr(node, "internal_part_number", ""):
            prefix, body, _ = _split_company_pn(node.internal_part_number)
            if prefix and body:
                node.suggested_prefix = prefix
                node.suggested_body = body
                node.suggested_pb = f"{prefix}-{body}"
                node.suggested_reason = "from existing inventory PN"

                print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

                return

        # 2) NEW: use parsed type → prefix mapping
        ptype = ""
        # Try to infer from the parsed NPRPart via the node’s description (or store parsed on node later)
        ptype = getattr(node, "parsed_type", "") or "OTHER"

        prefix = ""
        if hasattr(self, "_type_prefix_map"):
            prefix = self._type_prefix_map.get(ptype, "")

        if not prefix:
            node.suggested_reason = "no prefix suggestion"
            print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

            return

        node.suggested_prefix = prefix

        # 3) NEW body suggestion (since we don’t have a confident body match yet)
        body = self._suggest_new_body_for_prefix(prefix)
        if body:
            node.suggested_body = body
            node.suggested_pb = f"{prefix}-{body}"
            node.suggested_reason = f"from type mapping ({ptype})"


    def _pair_to_node(self, npr_part: Any, match: Any, input_line_id: int) -> DecisionNode:
        """Convert a (BOM input, MatchResult) pair into a UI-facing DecisionNode.

        IMPORTANT INVARIANT:
          - node.id is always "{workspace_id}:{input_line_id}" so DB patching + UI selection are stable.
        """
        match_type = safe_get(match, "match_type", default="")
        if hasattr(match_type, "value"):
            match_type = match_type.value
        match_type = str(match_type) if match_type else ""

        inv = self._resolve_inv(safe_get(match, "inventory_part", default=None))

        # Determine base_type:
        # - NO_MATCH -> NEW
        # - else -> EXISTS if we have inventory_part
        base_type = "NEW" if match_type == MatchType.NO_MATCH else ("EXISTS" if inv is not None else "NEW")

        # BOM fields
        bom_uid = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default="")).strip()
        bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default=""))
        description = str(safe_get(npr_part, "description", "desc", default=""))

        # Inventory base fields (EXISTS)
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default="")) if inv else ""
        inv_mpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="")) if inv else ""

        confidence = clamp01(safe_get(match, "confidence", default=0.0))

        # Stable node id for DB + UI
        node_id = f"{self.workspace_id}:{int(input_line_id)}" if self.workspace_id else f"ROW-{int(input_line_id)}"

        # compute parsed type BEFORE node creation
        parsed = getattr(npr_part, "parsed", {}) or {}
        ptype = str(getattr(npr_part, "_pc_ptype", "") or (parsed.get("type") if isinstance(parsed, dict) else "") or "OTHER")
        ptype = ptype.upper().strip()

        node = DecisionNode(
            id=node_id,
            base_type=base_type,
            bom_uid=bom_uid,
            bom_mpn=bom_mpn,
            description=description,
            internal_part_number="",
            inventory_mpn=inv_mpn,
            match_type=match_type,
            confidence=confidence,
            alternates=[],
            status=DecisionStatus.NEEDS_DECISION,
            locked=False,
            needs_approval=False,
            notes="",
            explain={},
        )

        # set dynamic attributes (used by CNS preview tooling)
        node.parsed_type = ptype

        # Suggested (not committed) Company PN derived from best inventory match, if any.
        # IMPORTANT: This must NOT anchor/export until the user clicks Add/Apply.
        if internal_pn:
            try:
                prefix, body, _ = _split_company_pn(internal_pn)
                if prefix and body:
                    node.suggested_prefix = prefix
                    node.suggested_body = body
                    node.suggested_pb = f"{prefix}-{body}"
                else:
                    node.suggested_prefix = ""
                    node.suggested_body = ""
                    node.suggested_pb = internal_pn
            except Exception:
                node.suggested_prefix = ""
                node.suggested_body = ""
                node.suggested_pb = internal_pn
        else:
            node.suggested_prefix = ""
            node.suggested_body = ""
            node.suggested_pb = ""

        # ----------------------------------------------------
        # Determine the MPN that caused the match (for display)
        # ----------------------------------------------------
        match_explain = safe_get(match, "explain", default=None) or {}
        if not isinstance(match_explain, dict):
            match_explain = {}

        # Prefer substitute_mpn (explicit alias) → winning_mpn → bom_mpn fallback
        winning_mpn_for_display = (
            (match_explain.get("substitute_mpn") or "").strip()
            or (match_explain.get("winning_mpn") or "").strip()
            or (bom_mpn or "").strip()
        )

        # ----------------------------------------------------
        # Internal Alternates (inventory)
        # Always show the WINNER inventory_part as a Base card, even if candidates exist.
        # ----------------------------------------------------
        candidates = safe_get(match, "candidates", "candidate_parts", default=None) or []
        seen_itemnums = set()
        seen_ids = set()

        if inv is not None:
            itemnum = (getattr(inv, "itemnum", "") or "").strip()
            if itemnum and itemnum not in seen_itemnums:
                base_alt = self._inventory_part_to_alternate(
                    inv,
                    confidence=confidence,
                    relationship="Base/Selected",
                )

                # Stamp grouped MFGPN metadata for the base card too (same as candidates),
                # and ONLY show a matched MFGPN that actually belongs to this CPN group.
                try:
                    base_subs = getattr(inv, "substitutes", None) or []
                    rep_vendoritem = (getattr(inv, "vendoritem", "") or "").strip()
                    seen_grp = set()
                    grp_mpns = []
                    def _push_grp(m):
                        m = (m or "").strip()
                        if not m:
                            return
                        k = m.lower()
                        if k in seen_grp:
                            return
                        seen_grp.add(k)
                        grp_mpns.append(m)
                    _push_grp(rep_vendoritem)
                    for _s in base_subs:
                        if isinstance(_s, str):
                            _push_grp(_s)
                        else:
                            _push_grp(getattr(_s, "mfgpn", "") or getattr(_s, "manufacturer_part_number", "") or "")

                    base_alt.meta = dict(getattr(base_alt, "meta", {}) or {})
                    base_alt.meta["company_pn_mfgpn_count"] = len(grp_mpns)
                    # keep aliases only (exclude rep vendoritem) for consistency with candidate stamping
                    base_alt.meta["company_pn_mfgpns"] = grp_mpns[1:] if len(grp_mpns) > 1 else []
                    base_alt.meta["company_pn_rep_vendoritem"] = rep_vendoritem

                    win = (winning_mpn_for_display or "").strip()
                    if win and win.lower() in {m.lower() for m in grp_mpns}:
                        base_alt._matched_mpn_ui = win
                    else:
                        # BOM MPN may have won matching, but if it is not a true alternate under this CPN,
                        # display the representative/real grouped MFGPN on the card.
                        base_alt._matched_mpn_ui = rep_vendoritem or (grp_mpns[0] if grp_mpns else "")
                except Exception:
                    if winning_mpn_for_display:
                        base_alt._matched_mpn_ui = winning_mpn_for_display

                node.alternates.append(base_alt)
                seen_itemnums.add(itemnum)

        for cand in candidates:
            # Prefer per-candidate score; otherwise fall back so candidates never show 0%.
            cand_score = safe_get(cand, "_pc_score", default=None)
            conf = cand_score if cand_score is not None else 0.25
            conf = clamp01(conf)

            via_mpn = (getattr(cand, "_best_source_customer_mpn", "") or "").strip()
            via_mt = (getattr(cand, "_best_source_match_type", "") or "").strip()

            rel = "Internal Candidate"
            if via_mpn and via_mpn != (bom_mpn or "").strip():
                rel = f"Internal Candidate (via {via_mpn})"
            if via_mt:
                rel = f"{rel} [{via_mt}]"

            # show collapsed alias count (if stamped)
            grp_count = int(getattr(cand, "_ui_group_count", 1) or 1)
            grp_mpns = list(getattr(cand, "_ui_group_mpns", []) or [])
            if grp_count > 1:
                rel = f"{rel} (+{grp_count - 1} MFGPNs)"

            # Rebind candidate -> master inventory by itemnum so EVERY card has substitutes
            cand_item = str(safe_get(cand, "itemnum", "internal_part_number", default="") or "").strip()
            master_inv = self._inv_by_itemnum.get(cand_item) if cand_item else None
            inv_for_ui = master_inv or cand

            cand_res = self._resolve_inv(cand)

            # preserve any temporary UI/grouping attributes computed on the candidate
            for attr in (
                "_pc_score",
                "_ui_group_count",
                "_ui_group_mpns",
                "_ui_group_best_vendoritem",
                "_best_source_customer_mpn",
                "_best_source_match_type",
            ):
                if hasattr(cand, attr) and not hasattr(cand_res, attr):
                    try:
                        setattr(cand_res, attr, getattr(cand, attr))
                    except Exception:
                        pass

            alt_obj = self._inventory_part_to_alternate(cand_res, confidence=conf, relationship=rel)

            # keep meta stamping (if group mpns exist)
            alt_obj.meta["company_pn_mfgpn_count"] = grp_count
            alt_obj.meta["company_pn_mfgpns"] = grp_mpns
            alt_obj.meta["company_pn_rep_vendoritem"] = str(getattr(cand, "_ui_group_best_vendoritem", "") or "")

            itemnum = (getattr(alt_obj, "internal_part_number", "") or "").strip()
            if itemnum and itemnum in seen_itemnums:
                continue
            seen_itemnums.add(itemnum)
            seen_ids.add(alt_obj.id)
            node.alternates.append(alt_obj)

        # ----------------------------------------------------
        # Customer-provided alternates (external)
        # Show them as external cards in the UI.
        # ----------------------------------------------------
        explain = safe_get(match, "explain", default=None) or {}
        customer_mpns = []
        winning_mpn = ""

        if isinstance(explain, dict):
            customer_mpns = explain.get("customer_mpns", []) or []
            winning_mpn = (explain.get("winning_mpn", "") or "").strip()

        if not customer_mpns:
            parsed = safe_get(npr_part, "parsed", default=None) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            customer_mpns = [bom_mpn] + list(mpn_alts)

        if not winning_mpn:
            winning_mpn = (bom_mpn or "").strip()

        node.explain = dict(safe_get(match, "explain", default={}) or {})
        node.explain["winning_mpn"] = winning_mpn

        # Deduplicate against alternates already present
        seen_mpns = set()
        for a in node.alternates:
            m = (getattr(a, "manufacturer_part_number", "") or "").strip()
            if m:
                seen_mpns.add(m)

        for mpn in customer_mpns:
            mpn = (mpn or "").strip()
            if not mpn or mpn == winning_mpn or mpn in seen_mpns:
                continue
            seen_mpns.add(mpn)

            node.alternates.append(
                Alternate(
                    id=f"CUST-{node.id}-{mpn}",
                    source="customer_bom",
                    manufacturer=str(safe_get(npr_part, "mfgname", default="")),
                    manufacturer_part_number=mpn,
                    internal_part_number="",
                    description=description,
                    confidence=0.0,
                    relationship="Customer Provided",
                    selected=False,
                    rejected=False,
                    raw=None,
                    meta={"customer_provided": True},
                )
            )

        # Ensure the surviving inventory card shows the resolved MPN
        for alt in node.alternates:
            if (
                alt.source == "inventory"
                and alt.internal_part_number == internal_pn
                and winning_mpn_for_display
            ):
                try:
                    inv._matched_mpn = winning_mpn_for_display
                except Exception:
                    pass

        self._recompute_node_flags(node)
        return node

    def _inventory_part_to_alternate(self, inv: InventoryPart, confidence: float, relationship: str = "Inventory") -> Alternate:
        """
        Convert an InventoryPart into a UI-facing Alternate card.

        IMPORTANT:
        - Stock/cost come from inv.raw_fields (ITEMS sheet), not from the master MPN sheet.
        - MFGPN-count badge comes from master sheet substitutes attached to InventoryPart.substitutes.
        """

        def _raw_get(*keys: str, default=""):
            raw = getattr(inv, "raw_fields", {}) or {}
            for k in keys:
                if k in raw:
                    v = raw.get(k)
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s != "":
                        return s
            return default

        def _to_int(x, default=0):
            try:
                if x is None:
                    return default
                s = str(x).strip()
                if s == "":
                    return default
                # handle "1,234"
                s = s.replace(",", "")
                return int(float(s))
            except Exception:
                return default

        def _to_float(x, default=None):
            try:
                if x is None:
                    return default
                s = str(x).strip()
                if s == "":
                    return default
                s = s.replace(",", "")
                return float(s)
            except Exception:
                return default

        # ---- STOCK / COST from ITEMS-sheet raw_fields (or fallback keys) ----
        qty_raw = _raw_get("totalqty", "total_qty", "qty_on_hand", "on_hand", "quantity", default="0")
        stock_qty = _to_int(qty_raw, default=0)

        # prefer avgcost, then lastcost, then stdcost
        cost_raw = _raw_get("avgcost", "avg_cost", "average_cost", "lastcost", "last_cost", "stdcost", "std_cost", default="")
        unit_cost = _to_float(cost_raw, default=None)

        # ---- MFGPN badge count from MASTER-sheet substitutes ----
        subs = getattr(inv, "substitutes", None) or []
        rep_vendoritem = (getattr(inv, "vendoritem", "") or "").strip()
        mfgpn_count = 0
        if rep_vendoritem:
            mfgpn_count += 1
        # count unique substitute mpns
        seen = set()
        for s in subs:
            if isinstance(s, str):
                mpn = s.strip()
            else:
                mpn = (getattr(s, "mfgpn", "") or "").strip()
            k = mpn.lower()
            if mpn and k not in seen:
                seen.add(k)
                mfgpn_count += 1


        # -------------------------------------------------
        # DISPLAY MPN SELECTION (CRITICAL FIX)
        # -------------------------------------------------
        display_mpn = rep_vendoritem

        # If the matching pipeline stamped a resolved MPN for UI,
        # ALWAYS prefer it for display (do NOT mutate inventory)
        ui_mpn = getattr(inv, "_matched_mpn", None)
        if ui_mpn:
            display_mpn = str(ui_mpn).strip()

        alt = Alternate(
            id=Alternate.new_id("INV"),
            source="inventory",
            manufacturer=(getattr(inv, "mfgname", "") or "").strip(),
            manufacturer_part_number=display_mpn,
            internal_part_number=(getattr(inv, "itemnum", "") or "").strip(),
            description=(getattr(inv, "desc", "") or "").strip(),
            value=str(getattr(inv, "parsed", {}).get("value", "") or ""),
            package=str(getattr(inv, "parsed", {}).get("package", "") or ""),
            tolerance=str(getattr(inv, "parsed", {}).get("tolerance", "") or ""),
            voltage=str(getattr(inv, "parsed", {}).get("voltage", "") or ""),
            wattage=str(getattr(inv, "parsed", {}).get("wattage", "") or ""),
            stock=int(stock_qty or 0),
            unit_cost=unit_cost,
            supplier=str(_raw_get("primaryvendornumber", "primary_vendor_number", "supplier", "vendor", default="")),
            confidence=clamp01(confidence),
            relationship=relationship,
            selected=False,
            rejected=False,
            raw=inv,
            meta={
                # used by UI badge + display logic
                "company_pn_mfgpn_count": int(mfgpn_count or 0),
                "company_pn_rep_vendoritem": rep_vendoritem,
            },
        )
        return alt
#------------------------------------
    

    def _collapse_candidates_by_itemnum(self, candidates):
        """
        Collapse a list of InventoryPart candidates so UI shows ONE card per company PN (itemnum).

        Returns:
            collapsed_list: list of representative InventoryPart (best _pc_score per itemnum)
            meta_by_itemnum: dict[itemnum] -> {"count": int, "mpns": [..], "best_vendoritem": str}
        
        Side effects:
            Stamps each representative candidate with:
              - _ui_group_count (int)
              - _ui_group_mpns (list[str])
              - _ui_group_best_vendoritem (str)
        """
        groups = {}   # itemnum -> {"count":int, "mpns":[...], "best_vendoritem":str}
        best = {}     # itemnum -> (inv, score)

        def _key(inv):
            return str(safe_get(inv, "itemnum", "internal_part_number", default="") or "").strip()

        def _mpn(inv):
            return str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="") or "").strip()

        def _score(inv):
            try:
                return float(getattr(inv, "_pc_score", 0.0) or 0.0)
            except Exception:
                return 0.0

        for inv in (candidates or []):
            item = _key(inv)
            if not item:
                # No stable grouping key -> skip (or treat as singleton elsewhere)
                continue

            mpn = _mpn(inv)
            sc = _score(inv)

            g = groups.setdefault(item, {"count": 0, "mpns": [], "best_vendoritem": ""})
            g["count"] += 1
            if mpn and mpn not in g["mpns"]:
                g["mpns"].append(mpn)

            if (item not in best) or (sc > best[item][1]):
                best[item] = (inv, sc)
                g["best_vendoritem"] = mpn

        collapsed = [t[0] for t in best.values()]
        collapsed.sort(key=lambda inv: _score(inv), reverse=True)

        # Stamp UI metadata onto representative rows
        for inv in collapsed:
            item = _key(inv)
            meta = groups.get(item, {})
            inv._ui_group_count = int(meta.get("count", 1) or 1)
            inv._ui_group_mpns = list(meta.get("mpns", []) or [])
            inv._ui_group_best_vendoritem = str(meta.get("best_vendoritem", "") or "")

        return collapsed, groups


    # ----------------------------
    # Query helpers for UI
    # ----------------------------

    def get_nodes(self) -> list[DecisionNode]:
        """UI-facing list of DecisionNode objects for the current workspace.

        IMPORTANT: The UI must be able to select a node id it sees in the table and
        then immediately fetch the same node from the controller. Therefore:
          - get_nodes() and get_node() MUST reference the same in-memory node set.
          - We only rebuild nodes from DB when needed (open_workspace / lazy load).
        """
        if not getattr(self, "workspace_id", None):
            return []

        # Lazy load if we haven't materialized nodes for this workspace yet
        if not getattr(self, "nodes", None) or getattr(self, "_nodes_workspace_id", None) != self.workspace_id:
            views = self.build_decision_views()
            self.nodes = list(views or [])
            self._node_index = {n.id: n for n in (self.nodes or [])}
            self._nodes_workspace_id = self.workspace_id

        return list(self.nodes or [])

    def get_node(self, node_id: str) -> DecisionNode:
        # Fast path: index lookup
        idx = getattr(self, "_node_index", None)
        if isinstance(idx, dict) and node_id in idx:
            return idx[node_id]

        # Slow path: scan current nodes, then backfill the index
        for n in (self.nodes or []):
            if n.id == node_id:
                if not isinstance(idx, dict):
                    self._node_index = {}
                self._node_index[n.id] = n
                return n

        raise KeyError(f"No DecisionNode with id={node_id}")

    # ----------------------------
    # Mutations (UI actions)
    # ----------------------------

    def select_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        # Cache the original BOM description once so we can restore it if the user unselects everything.
        node.explain = dict(getattr(node, "explain", {}) or {})
        if "input_description" not in node.explain:
            node.explain["input_description"] = (getattr(node, "description", "") or "")

        alt = self._find_alt(node, alt_id)
        alt.rejected = False
        alt.selected = True

        # ELEVATE: if the selected alternate has a Company PN (CNS), it becomes the anchor.
        alt_cpn = (getattr(alt, "internal_part_number", "") or "").strip()
        if alt_cpn:
            for other in (getattr(node, "alternates", []) or []):
                if other is alt:
                    continue
                if (getattr(other, "source", "") or "").lower() == "inventory":
                    other.selected = False
            node.internal_part_number = alt_cpn
            node.base_type = "EXISTS"
            node.explain = dict(getattr(node, "explain", {}) or {})
            preferred_mpn = (getattr(node, "preferred_inventory_mfgpn", "") or "").strip()
            matched_ui_mpn = (getattr(alt, "_matched_mpn_ui", "") or "").strip()
            chosen_inventory_mpn = preferred_mpn or matched_ui_mpn or (getattr(alt, "manufacturer_part_number", "") or "").strip()
            if chosen_inventory_mpn:
                alt.manufacturer_part_number = chosen_inventory_mpn
                node.inventory_mpn = chosen_inventory_mpn
                node.preferred_inventory_mfgpn = chosen_inventory_mpn
            else:
                node.inventory_mpn = getattr(alt, "manufacturer_part_number", "") or node.inventory_mpn

        ## Always update the export description to the selected card's description (header should reflect this).
        #chosen_desc = (getattr(alt, "description", "") or "").strip()
        #if chosen_desc:
        #    node.description = chosen_desc


        # After selection is committed
        chosen_desc = (getattr(alt, "description", "") or "").strip()
        if chosen_desc:
            self.set_node_description(node.id, chosen_desc)
        else:
            # still persist whatever node.description is currently
            self.set_node_description(node.id, getattr(node, "description", "") or "")


        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def reject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        if bool(getattr(alt, "selected", False)):
            raise ValueError("Cannot reject a locked in card. Unlock it first.")
        alt.selected = False
        alt.rejected = True

        # If this rejected inventory alternate is currently anchoring the node, clear the anchor
        if (
            getattr(alt, "source", "") == "inventory"
            and getattr(alt, "internal_part_number", "")
            and alt.internal_part_number == (node.internal_part_number or "")
            and float(getattr(node, "confidence", 0.0) or 0.0) < 0.999
        ):
            node.internal_part_number = ""
            node.inventory_mpn = ""

        # ---- DEMOTION RULE ----
        # If there are NO remaining inventory candidates (non-rejected, non-selected),
        # and we have no selected alternates, treat as NEW.
        remaining_inventory_candidates = [
            a for a in node.alternates
            if getattr(a, "source", "") == "inventory"
            and (not getattr(a, "rejected", False))
            and (not getattr(a, "selected", False))
        ]
        
        node.explain = dict(node.explain or {})
        node.explain["pre_demote"] = {
            "base_type": node.base_type,
            "confidence": node.confidence,
        }

        if (len(remaining_inventory_candidates) == 0) and (not node.has_selection()):


            node.base_type = "NEW"
            node.internal_part_number = ""
            node.inventory_mpn = ""
            node.confidence = 0.0
            node.explain = dict(node.explain or {})
            node.explain["demoted_to_new"] = "All inventory candidates were rejected."

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)



    def unreject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)

        # Undo reject
        alt.rejected = False
        alt.selected = False  # keep it simple: unreject ≠ auto-select

        # If we previously demoted to NEW because "all inventory candidates were rejected",
        # restore the pre-demote base_type/confidence when at least one inventory candidate exists again.
        if isinstance(getattr(node, "explain", None), dict):
            pre = node.explain.get("pre_demote")
            if pre and node.explain.get("demoted_to_new"):
                remaining_inventory_candidates = [
                    a for a in node.alternates
                    if getattr(a, "source", "") == "inventory"
                    and (not getattr(a, "rejected", False))
                ]
                if remaining_inventory_candidates:
                    node.base_type = pre.get("base_type", node.base_type)
                    node.confidence = pre.get("confidence", node.confidence)
                    node.explain.pop("demoted_to_new", None)
                    node.explain.pop("pre_demote", None)

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

    def unselect_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False

        # If nothing is selected anymore, clear anchor + restore original BOM description
        # (unless the user has explicitly applied/confirmed another description).
        if hasattr(node, "has_selection") and (not node.has_selection()):
            node.internal_part_number = ""
            node.base_type = "NEW"
            node.inventory_mpn = ""

            ex = dict(getattr(node, "explain", {}) or {})
            confirmed = bool(ex.get("export_description_confirmed"))
            override_applied = bool(ex.get("export_description_override_applied"))
            if (not confirmed) and (not override_applied):
                orig = (ex.get("input_description") or "").strip()
                if orig:
                    node.description = orig

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def add_manual_alternate(self, node_id: str, manufacturer_part_number: str, description: str = "Manual alternate") -> Alternate:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = Alternate(
            id=Alternate.new_id("MAN"),
            source="manual",
            manufacturer="",
            manufacturer_part_number=manufacturer_part_number,
            internal_part_number="",
            description=description,
            confidence=0.0,
            relationship="Manual",
        )
        node.alternates.append(alt)
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

        return alt

    def _has_selected_external_alt(self, node: DecisionNode) -> bool:
        try:
            for a in (node.selected_alternates() or []):
                if (getattr(a, "source", "") or "").lower() == "inventory":
                    continue
                if bool(getattr(a, "rejected", False)):
                    continue
                return True
        except Exception:
            pass
        return False

    def mark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        has_external_selected = self._has_selected_external_alt(node)
        node.status = DecisionStatus.READY_FOR_EXPORT
        node.locked = not has_external_selected
        self._recompute_node_flags(node)
        node.status = DecisionStatus.READY_FOR_EXPORT
        self._persist_node_and_alts(node)


    def unmark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        node.locked = False
        # move back to a workable status without trying to fully infer everything
        if getattr(node, 'internal_part_number', '') or node.has_selection():
            node.status = DecisionStatus.NEEDS_DECISION
        else:
            node.status = DecisionStatus.NEEDS_DECISION
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

    def set_preferred_inventory_mfgpn(self, node_id: str, alt_id: str, mfgpn: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
        alt = self._find_alt(node, alt_id)
        if (getattr(alt, 'source', '') or '').strip().lower() != 'inventory':
            raise ValueError('Preferred MFG PN can only be set for inventory alternates.')
        mfgpn = (mfgpn or '').strip()
        if not mfgpn:
            raise ValueError('MFG PN is blank.')

        company_part_number = (getattr(alt, 'internal_part_number', '') or '').strip()
        part_row = self._resolve_company_part_manufacturer_part(company_part_number, mfgpn)

        alt.manufacturer_part_number = mfgpn
        alt.meta = dict(getattr(alt, 'meta', {}) or {})
        alt.meta['company_pn_rep_vendoritem'] = mfgpn

        if part_row is not None:
            part_specs = self._manufacturer_part_to_spec_dict(company_part_number, part_row)
            alt.description = str(part_specs.get('Description', '') or getattr(alt, 'description', '') or '').strip()
            alt.manufacturer = str(part_specs.get('MfgName', '') or getattr(alt, 'manufacturer', '') or '').strip()
            avg_cost_raw = str(part_specs.get('AvgCost', '') or '').strip()
            try:
                alt.unit_cost = float(avg_cost_raw) if avg_cost_raw else getattr(alt, 'unit_cost', None)
            except Exception:
                pass
            total_qty_raw = str(part_specs.get('TotalQty', '') or '').strip()
            try:
                if total_qty_raw:
                    alt.stock = int(float(total_qty_raw))
            except Exception:
                pass
            alt.raw = part_row
            alt.meta['selected_company_part_mfgpn'] = mfgpn
            alt.meta['selected_company_part_mfgpn_specs'] = part_specs

        node.preferred_inventory_mfgpn = mfgpn
        node.inventory_mpn = mfgpn
        self._persist_node_and_alts(node)



    def get_alt_detail_payload(self, node_id: str, alt_id: str) -> dict:
        """
        Controller-backed detail payload for the UI (DB/controller is the truth).
        Returns:
          { "specs": {...}, "export_mfgpn_options": [..] }
        """
        node = self.get_node(node_id)
        alt = self._find_alt(node, alt_id)

        def _raw_get(obj, *keys, default=""):
            if obj is None:
                return default
            raw = getattr(obj, "raw_fields", {}) or {}
            if isinstance(raw, dict):
                for k in keys:
                    v = raw.get(k)
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s:
                        return s
            # fallback: direct attrs
            for k in keys:
                v = getattr(obj, k, None)
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return default

        def _pick_sub_field(sub_obj, *keys):
            # Try common direct attrs first, then raw_fields aliases.
            for k in keys:
                v = getattr(sub_obj, k, None)
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return _raw_get(sub_obj, *keys, default="")

        specs = {}
        export_opts = []

        if (getattr(alt, "source", "") or "").strip().lower() == "inventory":
            itemnum = (getattr(alt, "internal_part_number", "") or "").strip()
            inv = self._inv_by_itemnum.get(itemnum) if itemnum else None

            export_opts = self._get_company_part_mfgpn_options(itemnum)
            current_selected = (getattr(alt, "manufacturer_part_number", "") or "").strip()
            if current_selected and current_selected.lower() in {m.lower() for m in export_opts}:
                selected_vendoritem = current_selected
            else:
                pref = str((getattr(node, "preferred_inventory_mfgpn", "") or "")).strip()
                if pref and pref.lower() in {m.lower() for m in export_opts}:
                    selected_vendoritem = pref
                elif export_opts:
                    selected_vendoritem = export_opts[0]
                else:
                    selected_vendoritem = current_selected

            part_row = self._resolve_company_part_manufacturer_part(itemnum, selected_vendoritem)
            if part_row is not None:
                specs = self._manufacturer_part_to_spec_dict(itemnum, part_row, fallback_inv=inv)
            else:
                specs = {
                    "ItemNumber": itemnum,
                    "VendorItem": selected_vendoritem,
                    "Description": (getattr(alt, "description", "") or "").strip(),
                    "MfgName": (getattr(alt, "manufacturer", "") or "").strip(),
                    "PrimaryVendorNumber": _raw_get(inv, "primaryvendornumber", "supplier", "vendor") if inv is not None else "",
                    "TotalQty": _raw_get(inv, "totalqty", "total_qty", "qty_on_hand", "on_hand", "quantity") if inv is not None else "",
                    "AvgCost": "" if getattr(alt, "unit_cost", None) in (None, "") else str(getattr(alt, "unit_cost", "")),
                    "DefaultWhse": _raw_get(inv, "defaultwhse", "default_whse", "warehouse") if inv is not None else "",
                }

            if export_opts:
                specs["AlternatesCount"] = str(len(export_opts))
                specs["AlternatesList"] = "\n".join(export_opts)
        else:
            # Non-inventory alt fallback details
            specs = {
                "ItemNumber": (getattr(alt, "internal_part_number", "") or "").strip(),
                "VendorItem": (getattr(alt, "manufacturer_part_number", "") or "").strip(),
                "Description": (getattr(alt, "description", "") or "").strip(),
                "MfgName": (getattr(alt, "manufacturer", "") or "").strip(),
                "PrimaryVendorNumber": (getattr(alt, "supplier", "") or "").strip(),
                "TotalQty": "" if getattr(alt, "stock", None) in (None, "") else str(getattr(alt, "stock", "")),
                "AvgCost": "" if getattr(alt, "unit_cost", None) in (None, "") else str(getattr(alt, "unit_cost", "")),
            }

        return {
            "specs": specs,
            "export_mfgpn_options": export_opts,
            "selected_export_mfgpn": str(specs.get("VendorItem", "") or "").strip(),
        }

    def seed_fake_external_alternates(self, node_id: str) -> List[Alternate]:
        """Seed deterministic external alternate cards for UI testing without touching export rules."""
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        rows = build_fake_external_alt_specs(node)
        created = append_external_alternates(node, Alternate, rows)
        if created:
            node.explain = dict(getattr(node, "explain", {}) or {})
            node.explain["external_alternates"] = self._serialize_external_alternates_for_node(node)
            self._recompute_node_flags(node)
            self._persist_node_and_alts(node)
        return created

    def search_digikey(self, node_id: str) -> List[Alternate]:
        """
        Cache is controller-owned (not UI). UI just calls this.
        """
        node = self.get_node(node_id)

        if node_id in self.external_cache:
            return self.external_cache[node_id]

        if not self._digikey_search_fn:
            # no client wired yet
            self.external_cache[node_id] = []
            return []

        results = self._digikey_search_fn(node) or []
        # Normalize + attach
        for alt in results:
            alt.source = alt.source or "digikey"
        self.external_cache[node_id] = results

        # Option: merge external alternates into node list (recommended)
        # so UI doesn’t need “internal vs external” stores.
        append_external_alternates(node, Alternate, [
            {
                "id": getattr(alt, "id", ""),
                "source": getattr(alt, "source", "") or "digikey",
                "manufacturer": getattr(alt, "manufacturer", ""),
                "manufacturer_part_number": getattr(alt, "manufacturer_part_number", ""),
                "internal_part_number": getattr(alt, "internal_part_number", ""),
                "description": getattr(alt, "description", ""),
                "confidence": getattr(alt, "confidence", 0.0),
                "relationship": getattr(alt, "relationship", "External Alternate"),
                "selected": getattr(alt, "selected", False),
                "rejected": getattr(alt, "rejected", False),
                "meta": dict(getattr(alt, "meta", {}) or {}),
            }
            for alt in results
        ])
        node.explain = dict(getattr(node, "explain", {}) or {})
        node.explain["external_alternates"] = self._serialize_external_alternates_for_node(node)
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

        return results

    #------------------------------------   # ----------------------------
    # NPR Workspace (from DecisionNodes)
    # ----------------------------
    def _section_bucket(self, sec_raw: str) -> tuple[str, str]:
        s = (sec_raw or "").strip().upper()
        if s.startswith("SURFACE"):
            return "SURFACE MOUNT", "SMT"
        if s.startswith("THROUGH") or s.startswith("THRU"):
            return "THROUGH-HOLE", "TH"
        if s.startswith("AUXILIARY") or s.startswith("AUX"):
            if "ASSEMBLY" in s:
                return "AUXILIARY", "assembly"
            if "MECH" in s:
                return "AUXILIARY", "mech"
            if "PRODUCTION" in s:
                return "AUXILIARY", "production"
            if "OTHER" in s:
                return "AUXILIARY", "other"
            return "AUXILIARY", "aux"
        return "SURFACE MOUNT", "SMT"

    def _export_company_part_number_for_node(self, node: DecisionNode) -> str:
        assigned = str(getattr(node, "assigned_part_number", "") or "").strip()
        internal = str(getattr(node, "internal_part_number", "") or "").strip()
        return (assigned or internal or "").strip() or "NEW"

    def _selected_external_alts_for_export(self, node: DecisionNode) -> list[Alternate]:
        out = []
        try:
            for a in (node.selected_alternates() or []):
                if (getattr(a, "source", "") or "").lower() == "inventory":
                    continue
                if bool(getattr(a, "rejected", False)):
                    continue
                out.append(a)
        except Exception:
            pass
        return out

    def _export_mfg_fields_for_node(self, node: DecisionNode) -> tuple[str, str]:
        sel = self._selected_committed_alt(node)
        print(f"_EXPORT_MFG_FIELDS this is sel: {sel}")

        # 1. Selected alt always wins
        if sel is not None and not bool(getattr(sel, "rejected", False)):
            return (
                str(getattr(sel, "manufacturer", "") or "").strip(),
                str(getattr(sel, "manufacturer_part_number", "") or "").strip(),
            )

        # 2. No selected alt: try anchored inventory state
        anchored_cpn = self._export_company_part_number_for_node(node)
        if anchored_cpn == "NEW":
            anchored_cpn = ""

        inv_alt = self._selected_inventory_alt(node)
        if inv_alt is not None or anchored_cpn:
            manufacturer = str(getattr(inv_alt, "manufacturer", "") or "").strip() if inv_alt is not None else ""
            mpn = (
                str(getattr(node, "preferred_inventory_mfgpn", "") or "").strip()
                or str(getattr(node, "inventory_mpn", "") or "").strip()
                or (str(getattr(inv_alt, "manufacturer_part_number", "") or "").strip() if inv_alt is not None else "")
                or (str(getattr(inv_alt, "_matched_mpn_ui", "") or "").strip() if inv_alt is not None else "")
            )
            if not mpn and anchored_cpn:
                inv_obj = self._inv_by_itemnum.get(anchored_cpn)
                if inv_obj is not None:
                    mpn = str(getattr(inv_obj, "vendoritem", "") or "").strip()
                    if not manufacturer:
                        manufacturer = str(getattr(inv_obj, "mfgname", "") or "").strip()
            if mpn:
                return (manufacturer, mpn)

        # 3. External/manual fallback
        for a in self._selected_external_alts_for_export(node):
            return (
                str(getattr(a, "manufacturer", "") or "").strip(),
                str(getattr(a, "manufacturer_part_number", "") or "").strip(),
            )

        # 4. Last fallback
        return ("", str(getattr(node, "bom_mpn", "") or "").strip())

    def build_committed_export_state(self, node: DecisionNode) -> CommittedExportState:
        selected_alt = self._selected_committed_alt(node)
        #print(f"this is the selected_alt in the build committed export state: {self.selected_alt}")
        mfg, mfgpn = self._export_mfg_fields_for_node(node)
        section = self.get_node_bom_section(node.id)
        bucket, type_value = self._section_bucket(section)
        return CommittedExportState(
            node_id=str(getattr(node, "id", "") or ""),
            line_id=self._node_line_id(node),
            company_part_number=self._export_company_part_number_for_node(node),
            description_text=self._export_description_for_node(node),
            bom_mpn=str(getattr(node, "bom_mpn", "") or "").strip(),
            bom_section=section,
            bucket=bucket,
            type_value=type_value,
            manufacturer_name=str(mfg or "").strip(),
            manufacturer_part_number=str(mfgpn or "").strip(),
            selected_alt_id=str(getattr(selected_alt, "id", "") or ""),
            selected_alt_source=str(getattr(selected_alt, "source", "") or "").strip(),
            include_approval=bool(getattr(node, "needs_approval", False)),
            has_internal=bool(str(getattr(node, "internal_part_number", "") or "").strip()),
            has_selected_external=bool(self._selected_external_alts_for_export(node)),
            preferred_inventory_mfgpn=str(getattr(node, "preferred_inventory_mfgpn", "") or "").strip(),
            exclude_customer_part_number_in_npr=bool(getattr(node, "exclude_customer_part_number_in_npr", False)),
            notes=str(getattr(node, "notes", "") or "").strip(),
        )

    def export_npr(self, output_path: str = None):
        return NPRWorkbookExporter(self).export(output_path)
    
    def flush_workspace_to_db(self) -> None:
        """Persist the current workspace decision state."""
        if not self.workspace_id:
            return
        try:
            self.save_workspace_state()
        except Exception as e:
            print(f"[DB] save_workspace_state failed: {e}")

    def _export_npr_impl(self, output_path: str = None):
            """Export a single Excel workbook containing BOM + NPR outputs in separate sheets.

            Sheets:
              - BOM (from BOM_TEMPLATE if available)
              - NPR
              - Alternates for Approval
              - Bomexistences

            Description export rules:
              - Default is BLANK (input BOM description is not exported unless user overrides).
              - If an inventory card is selected/anchored, export that card's description.
              - If user applied an override, export the override.
            """

            # Export always reads from the current in-memory committed node state.
            try:
                self._persist_all_nodes()
            except Exception:
                pass


            try:
                self.flush_workspace_to_db()
            except Exception as e:
                print(f"[DB] flush on export failed: {e}")

            if not self.workspace_id:
                raise RuntimeError("No active workspace. Load a BOM or open a workspace first.")

            ready_nodes = [n for n in (self.get_nodes() or []) if getattr(n, "status", None) == DecisionStatus.READY_FOR_EXPORT]
            if not ready_nodes:
                raise RuntimeError("No parts marked Ready for Export.")

            from openpyxl import Workbook

            bom_template_path = Path(self.cfg.bom_template_path)
            if bom_template_path.exists():
                wb = load_workbook(bom_template_path)
            else:
                wb = Workbook()
                try:
                    wb.remove(wb.active)
                except Exception:
                    pass

            if wb.sheetnames:
                bom_ws = wb[wb.sheetnames[0]]
                bom_ws.title = "BOM"
            else:
                bom_ws = wb.create_sheet("BOM")

            npr_ws = wb["NPR"] if "NPR" in wb.sheetnames else wb.create_sheet("NPR")
            ws_alt = wb["Alternates for Approval"] if "Alternates for Approval" in wb.sheetnames else wb.create_sheet("Alternates for Approval")
            ws_bomexist = wb["BOMNOTES"] if "BOMNOTES" in wb.sheetnames else wb.create_sheet("BOMNOTES")

            def _norm(v):
                return str(v or "").strip().lower()

            def find_bom_sections(ws):
                sections = {}
                for r in range(1, ws.max_row + 1):
                    b = _norm(ws.cell(row=r, column=2).value)
                    c = _norm(ws.cell(row=r, column=3).value)
                    d = _norm(ws.cell(row=r, column=4).value)
                    e = _norm(ws.cell(row=r, column=5).value)
                    if b == "elan part number" and c == "description" and d.startswith("qty") and e == "refdes":
                        label = str(ws.cell(row=r - 1, column=2).value or ws.cell(row=r - 1, column=1).value or "").strip()
                        label_u = label.upper()
                        if "SURFACE" in label_u:
                            sections["SURFACE MOUNT"] = r
                        elif "THROUGH" in label_u or "THRU" in label_u:
                            sections["THROUGH-HOLE"] = r
                        elif "AUX" in label_u:
                            sections["AUXILIARY"] = r
                        else:
                            sections[label or f"SECTION@{r}"] = r
                return sections

            sections = find_bom_sections(bom_ws)
            # --- Template table expansion ---
            # BOM_TEMPLATE uses multiple section tables with a fixed row footprint. Writing past the
            # footprint can overwrite adjacent sections or make Excel report corruption. We prevent
            # that by inserting rows inside the section and expanding the corresponding Table.ref.

            def _find_table_by_header_row(ws, header_row: int):
                try:
                    for t in getattr(ws, "tables", {}).values():
                        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                        if min_row == header_row:
                            return t
                except Exception:
                    pass
                return None

            def _shift_table_ref(t: Table, delta_rows: int, start_row_inclusive: int):
                # Shift a table's ref if rows are inserted at/above it.
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                    if min_row >= start_row_inclusive:
                        min_row += delta_rows
                        max_row += delta_rows
                    elif max_row >= start_row_inclusive:
                        max_row += delta_rows
                    t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                except Exception:
                    pass

            def _ensure_section_capacity(ws, header_row: int, required_rows: int, next_header_row: int = None):
                data_start = header_row + 1

                # Next section header row implies a label row one row above it; insert before that label
                if next_header_row:
                    insert_at = max(1, next_header_row - 1)
                    available = max(0, insert_at - data_start)
                else:
                    insert_at = ws.max_row + 1
                    available = max(0, (ws.max_row + 1) - data_start)

                if required_rows <= available:
                    return

                need = required_rows - available
                if need <= 0:
                    return

                tables = list(getattr(ws, "tables", {}).values()) if getattr(ws, "tables", None) else []
                ws.insert_rows(insert_at, amount=need)

                # Shift all table refs below the insertion point
                for t in tables:
                    _shift_table_ref(t, need, insert_at)

                # Expand the section table itself (if present)
                t = _find_table_by_header_row(ws, header_row)
                if t is not None:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(t.ref)  # need to outsource this funcitnality to the data_loader. this is testing for now for functinlaity building. 
                        if next_header_row:
                            new_max_row = next_header_row - 2  # last data row before next label row
                        else:
                            new_max_row = max_row + need
                        if new_max_row > max_row:
                            t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
                    except Exception:
                        pass

                # Copy styling from the first existing data row (or header if no data row exists)
                style_src_row = data_start if data_start <= ws.max_row else header_row
                for rr in range(insert_at, insert_at + need):
                    ws.row_dimensions[rr].height = ws.row_dimensions[style_src_row].height
                    for cc in range(1, ws.max_column + 1):
                        src = ws.cell(row=style_src_row, column=cc)
                        dst = ws.cell(row=rr, column=cc)
                        dst._style = _copy_style(src._style)
                        dst.number_format = src.number_format
                        dst.font = _copy_style(src.font)
                        dst.fill = _copy_style(src.fill)
                        dst.border = _copy_style(src.border)
                        dst.alignment = _copy_style(src.alignment)
                        dst.protection = _copy_style(src.protection)
                        dst.comment = None


            if not sections:
                bom_ws.delete_rows(1, bom_ws.max_row)
                bom_ws["A1"] = "BOM Export"
                base_headers = ["Item", "ELAN Part Number", "Description", "QTY.", "REFDES", "TYPE", "Manufacturer", "Manufacturer PN", "Price", "EXT Price"]
                start = 4
                for sec_name in ("SURFACE MOUNT", "THROUGH-HOLE", "AUXILIARY"):
                    bom_ws.cell(row=start, column=2, value=sec_name)
                    hdr = start + 1
                    for j, h in enumerate(base_headers, start=1):
                        bom_ws.cell(row=hdr, column=j, value=h)
                    sections[sec_name] = hdr
                    start = hdr + 3

            def first_empty_row(ws, header_row):
                r = header_row + 1
                while r <= ws.max_row:
                    if str(ws.cell(row=r, column=2).value or "").strip() == "" and str(ws.cell(row=r, column=3).value or "").strip() == "":
                        return r
                    r += 1
                return ws.max_row + 1

            write_row_by_section = {sec: first_empty_row(bom_ws, hdr) for sec, hdr in sections.items()}

            input_rows = self.bom_repo.list_inputs(self.workspace_id) or []
            by_line_id = {}
            for r in input_rows:
                try:
                    lid = int(r.get("input_line_id") or 0)
                    if not lid:
                        continue
                    raw = r.get("raw_json") or {}
                    by_line_id[lid] = {
                        "input_line_id": lid,
                        "input_qty": r.get("qty"),
                        "qty": r.get("qty"),
                        "input_refdes": r.get("refdes"),
                        "refdes": r.get("refdes"),
                        "input_description": r.get("description"),
                        "input_partnum": r.get("partnum"),
                        "input_mfgpn": r.get("mfgpn"),
                        "input_mfgname": r.get("mfgname"),
                        "input_supplier": r.get("supplier"),
                        "input_raw_json": raw,
                    }
                except Exception:
                    pass

            item_counters = {"SURFACE MOUNT": 1, "THROUGH-HOLE": 1, "AUXILIARY": 1}


            # Ensure the BOM sections have enough rows for this export before we start writing.
            try:
                from collections import Counter as _Counter
                counts = _Counter()
                for _n in ready_nodes:
                    _sec = self.get_node_bom_section(_n.id)
                    _bucket, _t = self._section_bucket(_sec)
                    counts[_bucket] += 1

                sec_items = [(sec_name, int(hdr_row)) for sec_name, hdr_row in (sections or {}).items() if isinstance(hdr_row, int)]
                sec_items.sort(key=lambda x: x[1])
                hdr_rows_sorted = [r for _, r in sec_items]
                next_by_hdr = {r: (hdr_rows_sorted[i + 1] if i + 1 < len(hdr_rows_sorted) else None) for i, r in enumerate(hdr_rows_sorted)}

                for sec_name, hdr_row in sec_items:
                    add_count = int(counts.get(sec_name, 0))
                    if add_count <= 0:
                        continue
                    used = int(write_row_by_section.get(sec_name, hdr_row + 1)) - (hdr_row + 1)
                    required = max(0, used) + add_count
                    _ensure_section_capacity(bom_ws, hdr_row, required, next_header_row=next_by_hdr.get(hdr_row))
            except Exception as _cap_e:
                print(f"[EXPORT] BOM section expansion skipped: {_cap_e}")


            for node in ready_nodes:
                export_state = self.build_committed_export_state(node)
                cpn = export_state.company_part_number
                bucket = export_state.bucket
                type_val = export_state.type_value

                line_id = 0
                try:
                    _ws, _line = (node.id or "").split(":", 1)
                    line_id = int(_line)
                except Exception:
                    line_id = 0
                jr = by_line_id.get(line_id, {}) if line_id else {}

                qty = ""
                for k in ("input_qty", "qty", "input_quantity", "quantity"):
                    if k in jr and jr.get(k) not in (None, ""):
                        qty = jr.get(k)
                        break
                if qty in (None, ""):
                    raw = jr.get("input_raw_json") or {}
                    if isinstance(raw, str):
                        try:
                            raw = json.loads(raw)
                        except Exception:
                            raw = {}
                    if isinstance(raw, dict):
                        qty = (raw.get("raw_fields") or {}).get("quantity", "") or raw.get("qty", "") or ""

                refdes = ""
                for k in ("input_refdes", "refdes", "input_designator", "designator"):
                    if k in jr and jr.get(k) not in (None, ""):
                        refdes = jr.get(k)
                        break
                if not refdes:
                    raw = jr.get("input_raw_json") or {}
                    if isinstance(raw, str):
                        try:
                            raw = json.loads(raw)
                        except Exception:
                            raw = {}
                    if isinstance(raw, dict):
                        refdes = (raw.get("raw_fields") or {}).get("designator", "") or ""

                desc = export_state.description_text
                mfg, mfgpn = export_state.manufacturer_name, export_state.manufacturer_part_number

                row_idx = write_row_by_section.get(bucket) or (bom_ws.max_row + 1)
                item = item_counters[bucket]
                item_counters[bucket] += 1

                bom_ws.cell(row=row_idx, column=1, value=item)
                bom_ws.cell(row=row_idx, column=2, value=cpn)
                bom_ws.cell(row=row_idx, column=3, value=desc)
                bom_ws.cell(row=row_idx, column=4, value=qty)
                bom_ws.cell(row=row_idx, column=5, value=refdes)
                bom_ws.cell(row=row_idx, column=6, value=type_val)
                bom_ws.cell(row=row_idx, column=7, value=mfg)
                bom_ws.cell(row=row_idx, column=8, value=mfgpn)

                write_row_by_section[bucket] = row_idx + 1

            # NPR + Approval templates
            try:
                npr_template_path = Path(self.cfg.npr_template_path)
            except Exception:
                npr_template_path = Path(str(getattr(self.cfg, "npr_template_path", "") or ""))

            alts_template_path = Path(os.path.dirname(os.path.abspath(__file__))) / "ALTS_template.xlsx"

            def _copy_template_sheet(template_path: Path, target_name: str, fallback_headers: list[str]):
                if target_name in wb.sheetnames:
                    wb.remove(wb[target_name])
                ws = wb.create_sheet(target_name)

                if template_path and template_path.exists():
                    try:
                        twb = load_workbook(template_path)
                        tws = twb[twb.sheetnames[0]]
                        from copy import copy as _cpy
                        for r in range(1, tws.max_row + 1):
                            ws.row_dimensions[r].height = tws.row_dimensions[r].height
                            for c in range(1, tws.max_column + 1):
                                s = tws.cell(row=r, column=c)
                                d = ws.cell(row=r, column=c, value=s.value)
                                d._style = _cpy(s._style)
                                d.number_format = s.number_format
                                d.font = _cpy(s.font)
                                d.fill = _cpy(s.fill)
                                d.border = _cpy(s.border)
                                d.alignment = _cpy(s.alignment)
                                d.protection = _cpy(s.protection)
                        try:
                            for col_letter, dim in tws.column_dimensions.items():
                                ws.column_dimensions[col_letter].width = dim.width
                        except Exception:
                            pass
                        return ws
                    except Exception as e:
                        print(f"[EXPORT] template copy failed for {target_name}: {e}")

                ws.delete_rows(1, ws.max_row)
                ws.append(fallback_headers)
                return ws

            npr_ws = _copy_template_sheet(
                npr_template_path,
                "NPR",
                [
                    "Part Number",
                    "Part Status (new / exists)",
                    "Item Description",
                    "Custom Fields- SMT, TH, Process, Assembly,PCB, Mechanical",
                    "Preferred Part",
                    "Unit Cost",
                    "Default Units of Measure",
                    "Stock Unit-EA, ml, ounce, gram",
                    "per",
                    "Purchase Unit",
                    "Manufacturer Name",
                    "Manufacturer Part #",
                    "Supplier",
                    "Lead Time (WKS)",
                    "QC Required",
                    "TARIFF CODE (HTSUS)",
                    "Quote Number (Attach a Copy)",
                ],
            )

            ws_alt = _copy_template_sheet(
                alts_template_path,
                "Alternates for Approval",
                ["ELAN Inventory PN", "MfgNs under this PN", "Manufacturer", "link", "DESC", "BOM MfgN subbing for", "Designator", "NOTES"],
            )

            if "BOMNOTES" in wb.sheetnames:
                ws_bomexist = wb["BOMNOTES"]
                ws_bomexist.delete_rows(1, ws_bomexist.max_row)
            else:
                ws_bomexist = wb.create_sheet("BOMNOTES")
            ws_bomexist.append(["BOM_UID", "BOM_MPN", "BOM_DESC", "COMPANY_PN", "STATUS", "NOTES"])

            def _find_row_by_header(ws, expected: list[str], max_scan: int = 200) -> int:
                norm_expected = [str(x or "").strip().lower() for x in expected]
                for r in range(1, min(ws.max_row, max_scan) + 1):
                    vals = [str(ws.cell(row=r, column=i + 1).value or "").strip().lower() for i in range(len(norm_expected))]
                    if vals == norm_expected:
                        return r
                return 1

            npr_header_row = _find_row_by_header(
                npr_ws,
                [
                    "Part Number",
                    "Part Status (new / exists)",
                    "Item Description",
                    "Custom Fields- SMT, TH, Process, Assembly,PCB, Mechanical",
                    "Preferred Part",
                    "Unit Cost",
                    "Default Units of Measure",
                    "Stock Unit-EA, ml, ounce, gram",
                    "per",
                    "Purchase Unit",
                    "Manufacturer Name",
                    "Manufacturer Part #",
                    "Supplier",
                    "Lead Time (WKS)",
                    "QC Required",
                    "TARIFF CODE (HTSUS)",
                    "Quote Number (Attach a Copy)",
                ],
            )
            npr_write_row = max(9, npr_header_row + 1)
            while npr_write_row <= npr_ws.max_row:
                if all(str(npr_ws.cell(row=npr_write_row, column=cc).value or "").strip() == "" for cc in range(1, 18)):
                    break
                npr_write_row += 1

            alts_header_row = _find_row_by_header(
                ws_alt,
                ["ELAN Inventory PN", "MfgNs under this PN", "Manufacturer", "link", "DESC", "BOM MfgN subbing for", "Designator", "NOTES"],
            )
            alts_write_row = max(5, alts_header_row + 1)
            while alts_write_row <= ws_alt.max_row:
                if all(str(ws_alt.cell(row=alts_write_row, column=cc).value or "").strip() == "" for cc in range(1, 9)):
                    break
                alts_write_row += 1

            def _normalize_text(v: Any) -> str:
                return str(v or "").strip()

            def _line_context_for_node(node: DecisionNode) -> dict:
                line_id = 0
                try:
                    _ws, _line = (node.id or "").split(":", 1)
                    line_id = int(_line)
                except Exception:
                    line_id = 0
                jr = by_line_id.get(line_id, {}) if line_id else {}

                qty = ""
                for k in ("input_qty", "qty", "input_quantity", "quantity"):
                    if k in jr and jr.get(k) not in (None, ""):
                        qty = jr.get(k)
                        break

                raw = jr.get("input_raw_json") or {}
                if isinstance(raw, str):
                    try:
                        raw = json.loads(raw)
                    except Exception:
                        raw = {}
                raw_fields = (raw.get("raw_fields") or {}) if isinstance(raw, dict) else {}

                refdes = ""
                for k in ("input_refdes", "refdes", "input_designator", "designator"):
                    if k in jr and jr.get(k) not in (None, ""):
                        refdes = jr.get(k)
                        break
                if not refdes:
                    refdes = raw_fields.get("designator", "") or ""

                mfgname = ""
                for k in ("input_mfgname", "mfgname", "manufacturer", "manufacturer_name"):
                    if k in jr and jr.get(k) not in (None, ""):
                        mfgname = jr.get(k)
                        break

                return {
                    "qty": qty,
                    "refdes": refdes,
                    "mfgname": mfgname,
                }

            def _selected_external_alts(node: DecisionNode) -> list[Alternate]:
                out = []
                try:
                    for a in (node.selected_alternates() or []):
                        if (getattr(a, "source", "") or "").lower() == "inventory":
                            continue
                        if bool(getattr(a, "rejected", False)):
                            continue
                        out.append(a)
                except Exception:
                    pass
                return out

            def _internal_selected_is_exact(node: DecisionNode, inv_alt: Optional[Alternate]) -> bool:
                if inv_alt is None:
                    return False
                bom_mpn = _normalize_text(getattr(node, "bom_mpn", ""))
                export_state = self.build_committed_export_state(node)
                _mfg, selected_mpn = export_state.manufacturer_name, export_state.manufacturer_part_number
                selected_mpn = _normalize_text(selected_mpn)
                if bom_mpn and selected_mpn and bom_mpn.lower() == selected_mpn.lower():
                    return True
                rel = _normalize_text(getattr(inv_alt, "relationship", ""))
                if "exact" in rel.lower():
                    return True
                match_type = _normalize_text(getattr(node, "match_type", ""))
                if "EXACT" in match_type.upper():
                    return True
                return False

            def _append_npr_row(
                part_number: str,
                part_status: str,
                item_description: str,
                custom_field: str,
                preferred_part: str = "",
                manufacturer_name: str = "",
                manufacturer_part_number: str = "",
                supplier: str = "",
                unit_cost: Any = "",
                default_uom: str = "",
                stock_unit: str = "",
                per_value: str = "",
                purchase_unit: str = "",
                lead_time: Any = "",
                qc_required: Any = "",
                tariff_code: str = "",
                quote_number: str = "",
            ) -> None:
                nonlocal npr_write_row
                row_vals = [
                    part_number,
                    part_status,
                    item_description,
                    custom_field,
                    preferred_part,
                    unit_cost,
                    default_uom,
                    stock_unit,
                    per_value,
                    purchase_unit,
                    manufacturer_name,
                    manufacturer_part_number,
                    supplier,
                    lead_time,
                    qc_required,
                    tariff_code,
                    quote_number,
                ]
                for idx, val in enumerate(row_vals, start=1):
                    npr_ws.cell(row=npr_write_row, column=idx, value=val)
                npr_write_row += 1

            def _alts_optional_approved_col(ws) -> int | None:
                try:
                    for c in range(1, ws.max_column + 1):
                        hdr = _normalize_text(ws.cell(row=alts_header_row, column=c).value)
                        if hdr.lower().startswith("approved"):
                            return c
                except Exception:
                    pass
                return None

            def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
                from copy import copy as _cpy
                try:
                    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
                except Exception:
                    pass
                for cc in range(1, max_col + 1):
                    s = ws.cell(row=src_row, column=cc)
                    d = ws.cell(row=dst_row, column=cc)
                    try:
                        d._style = _cpy(s._style)
                        d.number_format = s.number_format
                        d.font = _cpy(s.font)
                        d.fill = _cpy(s.fill)
                        d.border = _cpy(s.border)
                        d.alignment = _cpy(s.alignment)
                        d.protection = _cpy(s.protection)
                    except Exception:
                        pass

            def _unmerge_overlapping(ws, row_start: int, row_end: int) -> None:
                try:
                    ranges = list(ws.merged_cells.ranges)
                except Exception:
                    ranges = []
                for mr in ranges:
                    try:
                        if not (mr.max_row < row_start or mr.min_row > row_end):
                            ws.unmerge_cells(str(mr))
                    except Exception:
                        pass

            def _set_link_cell(cell, value: str) -> None:
                value = _normalize_text(value)
                cell.value = value
                if value and (value.startswith("http://") or value.startswith("https://")):
                    try:
                        cell.hyperlink = value
                    except Exception:
                        pass

            def _render_alts_approval_sheet(internal_groups: list[dict], external_rows: list[dict]) -> None:
                nonlocal alts_header_row

                inv_start = alts_header_row + 1
                ext_label_row = None
                for rr in range(inv_start, ws_alt.max_row + 1):
                    v = _normalize_text(ws_alt.cell(row=rr, column=2).value)
                    if v.upper().startswith("EXTERNAL ALTERNATES"):
                        ext_label_row = rr
                        break
                if ext_label_row is None:
                    ext_label_row = max(inv_start + 1, ws_alt.max_row + 2)
                    ws_alt.cell(row=ext_label_row, column=2, value="EXTERNAL ALTERNATES")
                    ws_alt.cell(row=ext_label_row + 1, column=2, value="List of parts that which may be fit as an alternate of designated BOM Manufacturer Part Number (MfgN) not within ELAN INVENTORY")
                    ws_alt.cell(row=ext_label_row + 2, column=2, value="MfgNs")
                    ws_alt.cell(row=ext_label_row + 2, column=3, value="Manufacturer")
                    ws_alt.cell(row=ext_label_row + 2, column=4, value="link")
                    ws_alt.cell(row=ext_label_row + 2, column=5, value="DESC")
                    ws_alt.cell(row=ext_label_row + 2, column=6, value="BOM MfgN Alternate For")
                    ws_alt.cell(row=ext_label_row + 2, column=7, value="Designator")
                    ws_alt.cell(row=ext_label_row + 2, column=8, value="NOTES")

                ext_header_row_local = ext_label_row + 2
                ext_start = ext_header_row_local + 1
                inv_capacity = max(0, ext_label_row - inv_start)
                internal_needed = sum(max(1, len(g.get("rows", []) or [])) for g in (internal_groups or []))
                if internal_needed > inv_capacity:
                    need = internal_needed - inv_capacity
                    ws_alt.insert_rows(ext_label_row, amount=need)
                    for rr in range(ext_label_row, ext_label_row + need):
                        _copy_row_style(ws_alt, max(inv_start, ext_label_row - 1), rr, max(8, ws_alt.max_column))
                    ext_label_row += need
                    ext_header_row_local += need
                    ext_start += need

                ext_capacity = max(0, ws_alt.max_row - ext_start + 1)
                external_needed = len(external_rows or [])
                if external_needed > ext_capacity:
                    need = external_needed - ext_capacity
                    ws_alt.insert_rows(ws_alt.max_row + 1, amount=need)
                    for rr in range(ws_alt.max_row - need + 1, ws_alt.max_row + 1):
                        _copy_row_style(ws_alt, ext_start, rr, max(8, ws_alt.max_column))

                _unmerge_overlapping(ws_alt, inv_start, ws_alt.max_row)

                for rr in range(inv_start, ws_alt.max_row + 1):
                    for cc in range(1, max(8, ws_alt.max_column) + 1):
                        ws_alt.cell(row=rr, column=cc).value = None
                        try:
                            ws_alt.cell(row=rr, column=cc).hyperlink = None
                        except Exception:
                            pass

                # Always rewrite the external section headers after clearing the data region.
                ws_alt.cell(row=ext_label_row, column=2, value="EXTERNAL ALTERNATES")
                ws_alt.cell(row=ext_label_row + 1, column=2, value="List of parts that which may be fit as an alternate of designated BOM Manufacturer Part Number (MfgN) not within ELAN INVENTORY")
                ws_alt.cell(row=ext_header_row_local, column=2, value="MfgNs")
                ws_alt.cell(row=ext_header_row_local, column=3, value="Manufacturer")
                ws_alt.cell(row=ext_header_row_local, column=4, value="link")
                ws_alt.cell(row=ext_header_row_local, column=5, value="DESC")
                ws_alt.cell(row=ext_header_row_local, column=6, value="BOM MfgN Alternate For")
                ws_alt.cell(row=ext_header_row_local, column=7, value="Designator")
                ws_alt.cell(row=ext_header_row_local, column=8, value="NOTES")

                approved_col = _alts_optional_approved_col(ws_alt)
                cur = inv_start
                for group in (internal_groups or []):
                    rows = list(group.get("rows", []) or [])
                    span = max(1, len(rows))
                    for off in range(span):
                        _copy_row_style(ws_alt, inv_start, cur + off, max(8, ws_alt.max_column))
                    if span > 1:
                        for col in (1, 5, 6, 7, 8):
                            ws_alt.merge_cells(start_row=cur, start_column=col, end_row=cur + span - 1, end_column=col)
                        if approved_col:
                            ws_alt.merge_cells(start_row=cur, start_column=approved_col, end_row=cur + span - 1, end_column=approved_col)
                    ws_alt.cell(row=cur, column=1, value=group.get("elan_pn", ""))
                    ws_alt.cell(row=cur, column=5, value=group.get("desc", ""))
                    ws_alt.cell(row=cur, column=6, value=group.get("bom_mpn", ""))
                    ws_alt.cell(row=cur, column=7, value=group.get("designator", ""))
                    ws_alt.cell(row=cur, column=8, value=group.get("notes", "") or "---")
                    if approved_col:
                        ws_alt.cell(row=cur, column=approved_col, value=group.get("approved", "Yes"))
                    for off, row in enumerate(rows):
                        rr = cur + off
                        ws_alt.cell(row=rr, column=2, value=row.get("mfgpn", ""))
                        ws_alt.cell(row=rr, column=3, value=row.get("manufacturer", ""))
                        _set_link_cell(ws_alt.cell(row=rr, column=4), row.get("link", ""))
                    cur += span

                cur = ext_start
                for row in (external_rows or []):
                    _copy_row_style(ws_alt, ext_start, cur, max(8, ws_alt.max_column))
                    ws_alt.cell(row=cur, column=2, value=row.get("mfgpn", ""))
                    ws_alt.cell(row=cur, column=3, value=row.get("manufacturer", ""))
                    _set_link_cell(ws_alt.cell(row=cur, column=4), row.get("link", ""))
                    ws_alt.cell(row=cur, column=5, value=row.get("desc", ""))
                    ws_alt.cell(row=cur, column=6, value=row.get("bom_mpn", ""))
                    ws_alt.cell(row=cur, column=7, value=row.get("designator", ""))
                    ws_alt.cell(row=cur, column=8, value=row.get("notes", "") or "---")
                    if approved_col:
                        ws_alt.cell(row=cur, column=approved_col, value=row.get("approved", "Yes"))
                    cur += 1

            def _iter_company_mpn_rows(company_pn: str) -> list[tuple[str, str, str]]:
                company_pn = _normalize_text(company_pn)
                if not company_pn:
                    return []
                inv_obj = self._inv_by_itemnum.get(company_pn)
                row_company = None
                try:
                    row_company = self.inv_company_repo.get(self.workspace_id, company_pn)
                except Exception:
                    row_company = None
                if not row_company:
                    try:
                        row_company = self.inv_company_repo.get(self._inventory_store_workspace_id, company_pn)
                    except Exception:
                        row_company = None

                out = []
                if row_company:
                    alts = row_company.get("alternates_json", []) or []
                    if isinstance(alts, str):
                        try:
                            alts = json.loads(alts)
                        except Exception:
                            alts = []
                    seen = set()
                    for a in (alts or []):
                        mpn = _normalize_text((a or {}).get("mpn") or (a or {}).get("mfgpn") or (a or {}).get("manufacturer_part_number"))
                        mfg = _normalize_text((a or {}).get("mfgname") or (a or {}).get("manufacturer"))
                        link = _normalize_text((a or {}).get("link") or (a or {}).get("url") or (a or {}).get("purchase_url") or (a or {}).get("datasheet") or "")
                        if not mpn:
                            continue
                        key = (mpn.lower(), mfg.lower(), link.lower())
                        if key in seen:
                            continue
                        seen.add(key)
                        out.append((mpn, mfg, link))
                elif inv_obj is not None:
                    mpns = []
                    rep = _normalize_text(getattr(inv_obj, "vendoritem", ""))
                    if rep:
                        mpns.append(rep)
                    mpns.extend([_normalize_text(x) for x in (getattr(inv_obj, "substitutes", []) or []) if _normalize_text(x)])
                    seen = set()
                    for mpn in mpns:
                        k = mpn.lower()
                        if k in seen:
                            continue
                        seen.add(k)
                        out.append((mpn, _normalize_text(getattr(inv_obj, "mfgname", "")), ""))
                return out

            internal_approval_groups = []
            external_approval_rows = []

            for node in ready_nodes:
                ctx = _line_context_for_node(node)
                export_state = self.build_committed_export_state(node)
                desc = export_state.description_text

                # need to change this to be for the NODE.mfgpn not the BOM mfgpn
                bom_mpn = _normalize_text(export_state.bom_mpn)


                company_pn = _normalize_text(export_state.company_part_number)
                type_val = export_state.type_value
                note_text = _normalize_text(export_state.notes)
                npr_company_pn = "" if (
                    bool(export_state.has_selected_external)
                    and not bool(export_state.has_internal)
                    and bool(getattr(export_state, "exclude_customer_part_number_in_npr", False))
                ) else company_pn

                inv_alt = self._selected_inventory_alt(node)
                external_alts = self._selected_external_alts_for_export(node)
                has_internal = bool(export_state.has_internal)
                include_approval = bool(export_state.include_approval)

                ws_bomexist.append([
                    _normalize_text(getattr(node, "bom_uid", "")),
                    bom_mpn,
                    desc,
                    company_pn,
                    "EXISTING" if has_internal else "NEW",
                    note_text,
                ])

                if has_internal and not external_alts:
                    if include_approval and inv_alt is not None and not _internal_selected_is_exact(node, inv_alt):
                        internal_rows = []
                        for mpn, mfg, link in _iter_company_mpn_rows(company_pn):
                            internal_rows.append({"mfgpn": mpn, "manufacturer": mfg, "link": link})
                        if internal_rows:
                            internal_approval_groups.append({
                                "elan_pn": company_pn,
                                "rows": internal_rows,
                                "desc": desc,
                                "bom_mpn": bom_mpn,
                                "designator": _normalize_text(ctx.get("refdes")),
                                "notes": note_text,
                                "approved": "Yes",
                            })
                    continue

                if has_internal:
                    if include_approval and inv_alt is not None and (external_alts or not _internal_selected_is_exact(node, inv_alt)):
                        internal_rows = []
                        for mpn, mfg, link in _iter_company_mpn_rows(company_pn):
                            internal_rows.append({"mfgpn": mpn, "manufacturer": mfg, "link": link})
                        if internal_rows:
                            internal_approval_groups.append({
                                "elan_pn": company_pn,
                                "rows": internal_rows,
                                "desc": desc,
                                "bom_mpn": bom_mpn,
                                "designator": _normalize_text(ctx.get("refdes")),
                                "notes": note_text,
                                "approved": "Yes",
                            })
                else:
                    input_mfgname = _normalize_text(ctx.get("mfgname"))
                    _append_npr_row(
                        npr_company_pn,
                        "NEW",
                        desc,
                        type_val,
                        preferred_part="",
                        manufacturer_name=input_mfgname,
                        manufacturer_part_number=bom_mpn,
                    )

                for alt in external_alts:
                    # IMPORTANT:
                    # External NPR rows should stay correlated to the committed header values
                    # for this BOM line. The alternate contributes the external sourcing fields
                    # (MFG / MFGPN / supplier / link / cost), but the exported ELAN-side identity
                    # stays on the node: company_pn + header description.
                    alt_desc = desc
                    alt_supplier = _normalize_text(getattr(alt, "supplier", "") or getattr(alt, "source", ""))
                    alt_link = _normalize_text(
                        getattr(alt, "link", "")
                        or getattr(alt, "url", "")
                        or getattr(alt, "purchase_url", "")
                        or getattr(alt, "datasheet", "")
                        or alt_supplier
                    )
                    alt_mfg = _normalize_text(getattr(alt, "manufacturer", ""))
                    alt_mpn = _normalize_text(getattr(alt, "manufacturer_part_number", ""))
                    _append_npr_row(
                        npr_company_pn,
                        "NEW",
                        alt_desc,
                        type_val,
                        preferred_part="",
                        manufacturer_name=alt_mfg,
                        manufacturer_part_number=alt_mpn,
                        supplier=alt_supplier,
                        unit_cost=getattr(alt, "unit_cost", ""),
                        stock_unit=_normalize_text(getattr(alt, "stock_unit", "")),
                        lead_time=getattr(alt, "lead_time_weeks", ""),
                        qc_required=getattr(alt, "qc_required", ""),
                        tariff_code=_normalize_text(getattr(alt, "tariff_code", "")),
                    )
                    if include_approval:
                        external_approval_rows.append({
                            "mfgpn": alt_mpn,
                            "manufacturer": alt_mfg,
                            "link": alt_link,
                            "desc": alt_desc,
                            "bom_mpn": bom_mpn,
                            "designator": _normalize_text(ctx.get("refdes")),
                            "notes": note_text,
                            "approved": "Yes",
                        })

            _render_alts_approval_sheet(internal_approval_groups, external_approval_rows)
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = str(Path.cwd() / f"EXPORT_{timestamp}.xlsx")

            try:
                wb.save(output_path)
            except PermissionError:
                raise PermissionError(f"Cannot save. Close the Excel file first: {output_path}")
            except Exception as e:
                raise RuntimeError(f"Export failed to save workbook: {e}")

            return output_path

    # ----------------------------
    # Internal enforcement
    # ----------------------------

    def _find_alt(self, node: DecisionNode, alt_id: str) -> Alternate:
        for a in node.alternates:
            if a.id == alt_id:
                return a
        raise KeyError(f"No alternate with id={alt_id} on node={node.id}")

    def _ensure_unlocked(self, node: DecisionNode) -> None:
        if node.locked:
            raise PermissionError("Node is locked (READY_FOR_EXPORT). Unlocking not supported.")
        

    def _recompute_node_flags(self, node: DecisionNode) -> None:
        """
        Central place for status/approval heuristics.
        UI should just display these.
        """
        bt = node.base_type.upper()
        selected = len(node.selected_alternates())
        candidates = len(node.candidate_alternates())

        node.needs_approval = bool(getattr(node, "needs_approval", False))

        # Status heuristic unless already READY/locked
        if node.locked:
            node.status = DecisionStatus.READY_FOR_EXPORT
            return

        if getattr(node, "status", None) == DecisionStatus.READY_FOR_EXPORT and self._has_selected_external_alt(node):
            # External-alternate flow stays editable after Mark Ready so description / assigned PN
            # can still be adjusted without losing Ready-for-export state.
            node.status = DecisionStatus.READY_FOR_EXPORT
            return

        if bt == "NEW" and selected == 0:
            node.status = DecisionStatus.NEEDS_ALTERNATE
            return

        if bt == "EXISTS" and candidates > 0 and selected == 0:
            node.status = DecisionStatus.NEEDS_DECISION
            return

        # default: auto-ish state (still editable)
        node.status = DecisionStatus.FULL_MATCH



        ### input in a set_node_description which uses the passed in description to change the description in the Exported BOM and NPR.
        ### 