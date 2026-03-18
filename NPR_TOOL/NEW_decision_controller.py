
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple
from collections import defaultdict, Counter
from .data_loader import DataLoader
from .matching_engine import MatchingEngine
from .parsing_engine import parse_description
from .config_loader import load_config
from .data_models import MatchType, MatchResult, InventoryPart, Alternate, DecisionNode, DecisionStatus
from types import SimpleNamespace

from .db import connect_db, init_db, DBConfig
from .repositories import WorkspaceRepo, BomRepo, InventoryCompanyRepo, ExportLogRepo

from openpyxl import load_workbook
import os
import copy
from tkinter import messagebox
import re
import json
import threading
from .npr_workspace import (
    NPRWorkspace,
    NPRPrimaryNewItem,
    NPRSecondaryRow,
    NPRRowKind,
    append_external_alternates,
    build_fake_external_alt_specs,
    external_alt_specs_from_node,
    restore_external_alternates,
)
import time


# ----------------------------
# Controller
# ----------------------------

@dataclass
class ControllerConfig:
    """Simplified configuration for NPR Controller."""

    # Base configuration
    components_yaml_path: str = "./config/components.yaml"

    # Clean, absolute template path (uses os + Path)
    npr_template_path: str = str(
        Path(os.path.dirname(os.path.abspath(__file__))) / "NPR_Master2023_v4_FormTEMPLATECOPY.xlsx"
    )

    # BOM template (used for BOM export sheet formatting/sections)
    bom_template_path: str = str(
        Path(os.path.dirname(os.path.abspath(__file__))) / "BOM_TEMPLATE.xlsx"
    )

    # Metadata still needed for export
    created_by: str = "NPR Tool"

    PB_RE = re.compile(r"^\s*(\d{2})-(\d{5})\s*$")
    _PN_RE = re.compile(r"^(?P<prefix>\d{2})-(?P<body>\d{5})-(?P<suffix>[A-Za-z0-9]{4})$")

    # ----------------------------
    # Candidate limiting +  filters
    # ----------------------------
    MAX_INTERNAL_CANDIDATES = 10

    _IMPERIAL_PKGS = {"01005", "0201", "0402", "0603", "0805", "1206", "1210", "1812", "2010", "2512"}
    _METRIC_TO_IMPERIAL = {
        "0402": "01005",  # 0402 metric ~= 01005 imperial
        "0603": "0201",   # 0603 metric ~= 0201 imperial
        "1005": "0402",   # 1005 metric ~= 0402 imperial
        "1608": "0603",   # 1608 metric ~= 0603 imperial
        "2012": "0805",   # 2012 metric ~= 0805 imperial
        "3216": "1206",   # 3216 metric ~= 1206 imperial
        "3225": "1210",   # 3225 metric ~= 1210 imperial
        "4532": "1812",   # 4532 metric ~= 1812 imperial
        "5025": "2010",   # 5025 metric ~= 2010 imperial
        "6332": "2512",   # 6332 metric ~= 2512 imperial
    }

# ----------------------------
# Helpers
# ----------------------------

def safe_get(obj: Any, *names: str, default: Any = "") -> Any:
    for n in names:
        if obj is None:
            return default
        if hasattr(obj, n):
            v = getattr(obj, n)
            if v is not None:
                return v
    return default

def _split_company_pn(pn: str) -> tuple[str, str, str]:
    pn = (pn or "").strip()
    m = ControllerConfig._PN_RE.match(pn)
    if not m:
        return "", "", ""
    return m.group("prefix"), m.group("body"), m.group("suffix")

def clamp01(x: float) -> float:
    try:
        x = float(x)
    except Exception:
        return 0.0
    if x < 0.0:
        return 0.0
    if x > 1.0:
        return 1.0
    return x



class DecisionController:
    """
    Headless application controller.

    - Owns inventory, BOM/NPR list, match pairs, DecisionNodes
    - Enforces workflow rules (select/reject/ready/locked)
    - Owns external search caching (DigiKey etc.)
    - Builds NPRWorkspace and exports

    UI should ONLY talk to this class.
    """

    def __init__(
        self,
        cfg: Optional[ControllerConfig] = None,
        digikey_search_fn: Optional[Callable[[DecisionNode], List[Alternate]]] = None,
        stop_event: Optional[threading.Event] = None
    ):
        self.cfg = cfg or ControllerConfig()
        

        # Core data
        self.inventory: List[Any] = []
        self.npr_list: List[Any] = []
        self.match_pairs: List[Tuple[Any, Any]] = []  # (npr_part, match_result)

        # Pending inventory snapshot paths (may be loaded in any order)
        self._pending_inventory_master_path: Optional[str] = None
        self._pending_inventory_erp_path: Optional[str] = None
        
        # Pending inventory_company rows staged before workspace exists
        self._pending_inventory_company_rows: list[dict] = []

        # ----------------------------
        # DB persistence (workspace)
        # ----------------------------
        self.conn = connect_db(DBConfig())          # uses default ~/.npr_tool/npr.db
        init_db(self.conn)

        self.ws_repo  = WorkspaceRepo(self.conn)
        self.bom_repo = BomRepo(self.conn)
        self.inv_company_repo = InventoryCompanyRepo(self.conn)
        self.export_repo = ExportLogRepo(self.conn)

        # Workspace identity (ONLY authoritative session state)
        self.workspace_id: Optional[str] = None

        # Map BOM uid -> bom_row_id (built during load_npr / load_bom)

        self._bom_row_by_uid: Dict[str, str] = {}


        # Decisions
        self.nodes: List[DecisionNode] = []

        # External cache: node_id -> alternates
        self.external_cache: Dict[str, List[Alternate]] = {}
        self.stop_event = stop_event or threading.Event()

        # Workspace
        self.workspace: Optional[NPRWorkspace] = None

        self._inv_by_itemnum: Dict[str, InventoryPart] = {}
        self._alt_mpn_to_base: Dict[str, List[str]] = {}
        self._alt_loaded: bool = False

        self._erp_inventory: list[InventoryPart] = []
        self._erp_by_itemnum: dict[str, InventoryPart] = {}
        self._erp_loaded: bool = False

        # Derived / cached (NOT authoritative)
        self._inventory_cache = None
        self._views_cache: list[DecisionNode] = []

        self._digikey_search_fn = digikey_search_fn

        # Global inventory store workspace (single current master+ERP state, updated on each import)
        self._inventory_store_workspace_id = "__CURRENT_INVENTORY__"


    # ----------------------------
    # Loaders
    # ----------------------------

    def _ensure_inventory_store_workspace(self) -> str:
        """Ensure a hidden singleton workspace row exists for the current inventory store."""
        wsid = self._inventory_store_workspace_id
        try:
            row = self.ws_repo.get(wsid)
        except Exception:
            row = None
        if row:
            return wsid
        # Create directly to avoid repo assumptions about generated IDs.
        now = datetime.utcnow().isoformat(timespec="seconds")
        self.conn.execute(
            """
            INSERT OR IGNORE INTO workspace (
                workspace_id, name, status, created_at, updated_at,
                label, bom_source_path, inventory_master_path, inventory_erp_path, cns_path, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                wsid,
                "[SYSTEM] Current Inventory",
                "SYSTEM",
                now,
                now,
                "",
                "",
                "",
                "",
                "",
                "Do not open manually. Current master+ERP inventory store.",
            ),
        )
        self.conn.commit()
        return wsid

    def _upsert_current_inventory_store(self, company_rows: list[dict], source_label: str = "INVENTORY") -> None:
        """Incrementally sync the single current inventory store (no delete+rewrite)."""
        if company_rows is None:
            return
        wsid = self._ensure_inventory_store_workspace()
        now = datetime.utcnow().isoformat(timespec="seconds")
        self.conn.execute(
            "UPDATE workspace SET updated_at = ?, inventory_master_path = ?, inventory_erp_path = ? WHERE workspace_id = ?",
            (now, self._pending_inventory_master_path or "", self._pending_inventory_erp_path or "", wsid),
        )
        self.conn.commit()

        # Incremental DB sync:
        # - upsert only NEW/CHANGED CPN rows
        # - prune CPNs missing from the latest master snapshot
        # - rebuild inventory_company_item only for changed CPNs
        if hasattr(self.inv_company_repo, "sync_company_parts_incremental"):
            stats = self.inv_company_repo.sync_company_parts_incremental(wsid, company_rows)
            try:
                print(
                    f"[DB][{(source_label or 'INVENTORY').upper()}] inventory sync:",
                    f"new_cpn={stats.get('new',0)}",
                    f"changed_cpn={stats.get('changed',0)}",
                    f"unchanged_cpn={stats.get('unchanged',0)}",
                    f"pruned_cpn={stats.get('pruned',0)}",
                )
            except Exception:
                pass
        else:
            # Fallback for older repo versions
            self.inv_company_repo.upsert_company_parts(wsid, company_rows)

    def load_inventory(
        self,
        xlsx_path: str,
        progress_cb: Optional[Callable[[int, int, str], None]] = None,
        phase_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """Load **master inventory** (CPN-level + alternates) and stage it for DB persistence.

        - Builds inventory_company rows (CPN pooled stock + alternates JSON)
        - Builds a flat InventoryPart list for MatchingEngine compatibility
          (CRITICAL: ensure InventoryPart.substitutes includes ALL MPNs under the CPN)
        """
        if phase_cb:
            try:
                phase_cb("Reading workbook and parsing descriptions...")
            except Exception:
                pass

        company_rows, flat_inventory = DataLoader.build_inventory_company_parts(
            xlsx_path,
            erp_inventory_path=None,
            progress_cb=progress_cb,
        )

        # IMPORTANT: do not dedupe here; preserve alternates_json payload exactly as built by DataLoader.
        # If duplicate rows exist in source data, we want to inspect/fix the loader mapping, not collapse rows here.

        # ---------------------------------------------------------
        # IMPORTANT FIX:
        # Do NOT trust flat_inventory to carry the CPN->MPN list correctly.
        # Rebuild a CPN-level InventoryPart list where:
        #   vendoritem = representative MPN
        #   substitutes = all other MPNs under that CPN
        # This is what the UI bottom-right panel expects.
        # ---------------------------------------------------------
        if phase_cb:
            try:
                phase_cb("Rebuilding inventory view...")
                if progress_cb:
                    progress_cb(0, 1, "Rebuilding inventory view")
            except Exception:
                pass

        rebuilt_flat = self._company_rows_to_flat_inventory(company_rows)

        self.inventory = rebuilt_flat
        self._inv_by_itemnum = {
            inv.itemnum: inv
            for inv in (self.inventory or [])
            if getattr(inv, "itemnum", "")
        }

        # Stage for DB (workspace may not exist yet)
        self._pending_inventory_company_rows = company_rows
        self._pending_inventory_master_path = xlsx_path

        # Persist current inventory store (single current master+ERP state)
        # IMPORTANT: MASTER should NOT prune ERP-only rows. Merge MASTER data onto current DB rows.
        try:
            wsid = self._ensure_inventory_store_workspace()
            existing_rows = []
            try:
                existing_rows = self.inv_company_repo.list(wsid) or []
            except Exception:
                existing_rows = []

            merged_by_cpn = {}
            for r in (existing_rows or []):
                cpn_key = str((r or {}).get("cpn", "") or "").strip()
                if cpn_key:
                    merged_by_cpn[cpn_key] = dict(r)

            for r in (company_rows or []):
                cpn_key = str((r or {}).get("cpn", "") or "").strip()
                if not cpn_key:
                    continue
                ex = merged_by_cpn.get(cpn_key, {})
                merged_by_cpn[cpn_key] = {
                    "cpn": cpn_key,
                    "canonical_desc": str((r or {}).get("canonical_desc", "") or ex.get("canonical_desc", "") or ""),
                    "stock_total": int((ex or {}).get("stock_total", 0) or (r or {}).get("stock_total", 0) or 0),
                    "alternates_json": (r or {}).get("alternates_json") or (r or {}).get("alternates") or [],
                }

            merged_rows = list(merged_by_cpn.values())
            self._pending_inventory_company_rows = merged_rows
            if phase_cb:
                try:
                    phase_cb("Syncing inventory database...")
                    if progress_cb:
                        progress_cb(0, 1, "Syncing inventory database")
                except Exception:
                    pass
            self._upsert_current_inventory_store(merged_rows, source_label="MASTER")
            if self.workspace_id:
                self.ws_repo.update_meta(
                    self.workspace_id,
                    inventory_master_path=xlsx_path,
                    inventory_erp_path=(self._pending_inventory_erp_path or ""),
                )
        except Exception as e:
            print(f"[DB] current inventory store update failed: {e}")

        # Treat master inventory as alternates DB too (MPN -> [base itemnum]) using the
        # already-built company_rows payload. Do NOT reread the master workbook here; that
        # duplicates the heaviest load/parsing path on the UI thread and can freeze the UI.
        try:
            if phase_cb:
                try:
                    phase_cb("Building alternate lookup map...")
                    if progress_cb:
                        progress_cb(0, 1, "Building alternate lookup map")
                except Exception:
                    pass
            self._alt_mpn_to_base = self._build_mpn_to_base_from_company_rows(company_rows)
            self._alt_loaded = True
        except Exception:
            pass

        if progress_cb:
            try:
                progress_cb(1, 1, "Done")
            except Exception:
                pass

        return len(self.inventory)


    def _dedupe_inventory_company_rows(self, rows: list[dict]) -> list[dict]:
        """Legacy compatibility shim (no-op).

        We intentionally do NOT dedupe inventory_company rows here because it can
        collapse/strip alternate payloads and hide loader issues. Inventory rows
        should be preserved exactly as produced by DataLoader, then persisted.
        """
        return list(rows or [])

    def _build_mpn_to_base_from_company_rows(self, company_rows: list[dict]) -> dict[str, list[str]]:
        """Build normalized MPN -> [company/base itemnum] from inventory_company rows.

        This replaces the previous controller behavior of calling
        DataLoader.load_master_inventory(xlsx_path) a second time after the master had already
        been loaded into company_rows. Re-reading the workbook here doubled the synchronous
        work on the UI thread and caused freezes/regressions on large master files.
        """
        out: dict[str, list[str]] = {}
        norm = getattr(DataLoader, 'norm_mpn_key', None)
        if not callable(norm):
            return out

        for row in (company_rows or []):
            if not isinstance(row, dict):
                continue
            base = str(row.get('cpn', '')).strip()
            if not base:
                continue

            alts = row.get('alternates_json') or row.get('alternates') or []
            if isinstance(alts, str):
                try:
                    alts = json.loads(alts)
                except Exception:
                    alts = []

            for alt in (alts or []):
                if not isinstance(alt, dict):
                    continue
                mpn = str(alt.get('mpn', '') or alt.get('mfgpn', '') or alt.get('manufacturer_part_number', '')).strip()
                key = norm(mpn)
                if not key:
                    continue
                out.setdefault(key, [])
                if base not in out[key]:
                    out[key].append(base)
        return out

    def _company_rows_to_flat_inventory(self, company_rows: list[dict]) -> list[InventoryPart]:
        """
        Convert inventory_company-style rows into a MatchingEngine/UI friendly
        InventoryPart list where each CPN becomes:

          InventoryPart.itemnum   = CPN
          InventoryPart.vendoritem = representative MPN (shown as the main MPN)
          InventoryPart.substitutes = all other MPNs under that CPN

        This is REQUIRED for:
          - bottom-left cards badge counts
          - bottom-right “all MFGPNs under company PN” panel
          - ensuring exact match MPN still shows up (as vendoritem)
        """
        out: list[InventoryPart] = []

        for r in (company_rows or []):
            inv = self._company_row_to_inventory_part(r)
            if inv is not None:
                out.append(inv)

        return out

    def _company_row_to_inventory_part(self, row: dict) -> Optional[InventoryPart]:
        """
        Build a synthetic InventoryPart from an inventory_company row.

        IMPORTANT:
        - Do not re-run description parsing here.
        - The loader is responsible for parsing once and attaching row["parsed"].
        - Re-parsing during rebuild caused the post-progress-bar hang/crash.
        """

        if not isinstance(row, dict):
            return None

        cpn = str(row.get("cpn", "")).strip()
        if not cpn:
            return None

        desc = str(row.get("canonical_desc", "")).strip()
        stock_total = int(row.get("stock_total", 0) or 0)

        parsed_raw = row.get("parsed")
        parsed = dict(parsed_raw) if isinstance(parsed_raw, dict) else {}

        alts = row.get("alternates_json") or row.get("alternates") or []
        if isinstance(alts, str):
            try:
                alts = json.loads(alts)
            except Exception:
                alts = []
        if not isinstance(alts, list):
            alts = []

        mpns: list[str] = []
        mfgname = ""
        mfgid = ""

        for a in alts:
            if not isinstance(a, dict):
                continue

            mpn = str(
                a.get("mpn")
                or a.get("mfgpn")
                or a.get("vendoritem")
                or ""
            ).strip()
            if not mpn:
                continue

            if not mfgname:
                mfgname = str(a.get("mfgname", "") or "").strip()
            if not mfgid:
                mfgid = str(a.get("mfgid", "") or "").strip()

            if mpn not in mpns:
                mpns.append(mpn)

        rep_mpn = mpns[0] if mpns else ""
        subs = mpns[1:] if len(mpns) > 1 else []

        inv = InventoryPart(
            itemnum=cpn,
            desc=desc,
            mfgid=mfgid,
            mfgname=mfgname,
            vendoritem=rep_mpn,
            substitutes=subs,
            raw_fields={
                "totalqty": stock_total,
            },
            parsed=parsed,
            api_data=None,
        )

        inv._all_mpns = list(mpns)
        return inv
    


    def load_alternates_db(self, xlsx_path: str) -> tuple[int, int]:
        """Compatibility shim.

        The master file now *is* the inventory + alternates source. This call simply reloads the
        master inventory and returns counts like the old function did.
        """
        inventory_parts, subs_by_base, mpn_to_base = DataLoader.load_master_inventory(xlsx_path)

        self.inventory = inventory_parts
        self._inv_by_itemnum = {
            inv.itemnum: inv
            for inv in (self.inventory or [])
            if getattr(inv, "itemnum", "")
        }

        self._alt_mpn_to_base = mpn_to_base or {}
        self._alt_loaded = True

        num_with_subs = sum(1 for _, subs in (subs_by_base or {}).items() if subs)
        return num_with_subs, len(self._alt_mpn_to_base)

    def load_items_inventory(
        self,
        xlsx_path: str,
        progress_cb: Optional[Callable[[int, int, str], None]] = None,
        phase_cb: Optional[Callable[[str], None]] = None,
    ) -> int:
        """Load the ERP inventory extract and attach stock totals at the CPN level.

        In the v2 model, stock is ONLY stored on the company part number record (inventory_company).
        """
        self._pending_inventory_erp_path = xlsx_path

        if phase_cb:
            try:
                phase_cb("Reading ERP workbook...")
            except Exception:
                pass

        try:
            stock_totals = DataLoader.load_erp_stock_totals(xlsx_path)
        except Exception as e:
            raise ValueError(f"Failed to load ERP inventory stock totals: {e}")

        merged = 0

        # Build working set from current DB inventory store + any staged rows.
        wsid = self._ensure_inventory_store_workspace()
        db_rows = []
        try:
            db_rows = self.inv_company_repo.list(wsid) or []
        except Exception:
            db_rows = []

        by_cpn = {str(r.get("cpn", "")).strip(): dict(r) for r in (db_rows or []) if str(r.get("cpn", "")).strip()}
        for r in (self._pending_inventory_company_rows or []):
            cpn_key = str((r or {}).get("cpn", "")).strip()
            if cpn_key:
                # staged rows (e.g., just-loaded MASTER) should override DB metadata
                by_cpn[cpn_key] = dict(r)

        # DEBUG: show sample ERP rows that are not found in the currently loaded master CPN set.
        master_cpns = {str((r or {}).get("cpn", "")).strip() for r in (self._pending_inventory_company_rows or []) if str((r or {}).get("cpn", "")).strip()}
        erp_only = []
        for cpn, qty in stock_totals.items():
            cpn_key = str(cpn or "").strip()
            if not cpn_key:
                continue
            if master_cpns and cpn_key not in master_cpns:
                erp_only.append((cpn_key, int(qty or 0)))

        if erp_only:
            print(f"[DB][DEBUG] ERP rows not found in MASTER: count={len(erp_only)} (showing up to 5)")
            for cpn_key, qty in erp_only[:5]:
                print(f"[DB][DEBUG] ERP-only sample: cpn={cpn_key!r} qty={qty}")

        total_stock_rows = max(1, len(stock_totals))
        if phase_cb:
            try:
                phase_cb("Merging ERP stock totals...")
            except Exception:
                pass

        for idx, (cpn, qty) in enumerate(stock_totals.items(), start=1):
            cpn_key = str(cpn or "").strip()
            if not cpn_key:
                continue
            row = by_cpn.get(cpn_key)
            if row is None:
                row = {
                    "cpn": cpn_key,
                    "canonical_desc": "",
                    "stock_total": 0,
                    "alternates_json": [],
                }
                by_cpn[cpn_key] = row
            row["stock_total"] = int(qty or 0)
            merged += 1
            if progress_cb and (idx % 250 == 0 or idx == total_stock_rows):
                try:
                    progress_cb(idx, total_stock_rows, f"Merging ERP stock totals {idx}/{total_stock_rows}")
                except Exception:
                    pass

        self._pending_inventory_company_rows = list(by_cpn.values())


        # Refresh the live in-memory inventory snapshot too.
        # Otherwise UI cards keep using stale InventoryPart.raw_fields["totalqty"].
        try:
            if phase_cb:
                try:
                    phase_cb("Rebuilding inventory view...")
                    if progress_cb:
                        progress_cb(0, 1, "Rebuilding inventory view")
                except Exception:
                    pass
            rebuilt_flat = self._company_rows_to_flat_inventory(self._pending_inventory_company_rows)
            self.inventory = rebuilt_flat
            self._inv_by_itemnum = {
                inv.itemnum: inv
                for inv in (self.inventory or [])
                if getattr(inv, "itemnum", "")
            }
        except Exception as e:
            print(f"[CTRL] failed to refresh in-memory inventory after ERP merge: {e}")

        # IMPORTANT: no dedupe pass here. Preserve the staged rows/alternates exactly as loaded.

        # Persist current inventory store immediately
        try:
            if phase_cb:
                try:
                    phase_cb("Syncing inventory database...")
                    if progress_cb:
                        progress_cb(0, 1, "Syncing inventory database")
                except Exception:
                    pass
            self._upsert_current_inventory_store(self._pending_inventory_company_rows, source_label="ERP")
            if self.workspace_id:
                self.ws_repo.update_meta(self.workspace_id, inventory_erp_path=xlsx_path)
        except Exception as e:
            print(f"[DB] current inventory store update (ERP) failed: {e}")

        if progress_cb:
            try:
                progress_cb(1, 1, "Done")
            except Exception:
                pass

        return merged

    def load_npr(self, xlsx_path: str) -> int:
        """
        Load a BOM/NPR file and ALWAYS create a NEW workspace for it.

        Rationale:
          - Workspaces represent independent jobs/projects.
          - BOMs should never be mixed into the same workspace.
          - Re-opening an existing workspace is handled via open_workspace().
        """
        # Hard reset "session state" that must NOT bleed across workspaces
        self.npr_list = []
        self.match_pairs = []
        self.nodes = []
        self.external_cache = {}
        self.workspace = None
        self._bom_row_by_uid = {}

        rows = DataLoader.load_bom_any(xlsx_path)
        self.npr_list = rows

        # --- DB persistence layer ---
        with open(xlsx_path, "rb") as f:
            file_bytes = f.read()

        # Create a new BOM artifact every time we import
                # ----------------------------
        # DB: create workspace + persist canonical BOM input + bootstrap mutable state
        # ----------------------------
        ws_name = Path(xlsx_path).stem
        self.workspace_id = self.ws_repo.create(
            name=ws_name,
            label="",
            bom_source_path=xlsx_path,
            inventory_master_path=(self._pending_inventory_master_path or ""),
            inventory_erp_path=(self._pending_inventory_erp_path or ""),
            cns_path="",
            notes="",
            status="ACTIVE",
        )

        # Persist BOM input (canonical)
        inputs: List[Dict[str, Any]] = []
        for i, p in enumerate(self.npr_list or [], start=1):
            raw = p.to_dict() if hasattr(p, "to_dict") else {}
            raw.setdefault("raw_fields", getattr(p, "raw_fields", {}))
            raw.setdefault("parsed", getattr(p, "parsed", {}))

            qty = None
            try:
                q_raw = (getattr(p, "raw_fields", {}) or {}).get("quantity", "")
                qty = float(q_raw) if str(q_raw).strip() else None
            except Exception:
                qty = None

            inputs.append(
                {
                    "input_line_id": i,
                    "partnum": getattr(p, "partnum", f"ROW-{i}"),
                    "description": getattr(p, "desc", "") or getattr(p, "description", ""),
                    "qty": qty,
                    "refdes": (getattr(p, "raw_fields", {}) or {}).get("designator", ""),
                    "item_type": (getattr(p, "parsed", {}) or {}).get("type", ""),
                    "mfgname": getattr(p, "mfgname", ""),
                    "mfgpn": getattr(p, "mfgpn", ""),
                    "supplier": getattr(p, "supplier", ""),
                    "raw_json": raw,
                }
            )

        self.bom_repo.upsert_inputs(self.workspace_id, inputs)
        self.bom_repo.bootstrap_state_from_inputs(self.workspace_id, overwrite_existing=True)

        # Do NOT persist inventory per-workspace (prevents DB bloat).
        # Inventory is stored once in the current inventory store.

        return len(self.npr_list)

    def flush_workspace_to_db(self) -> None:
        """Persist current decisions to bom_line_state (v2).

        Safe no-op if no workspace is active.
        """
        if not self.workspace_id:
            return
        try:
            self._persist_all_nodes()
        except Exception as e:
            print(f"[DB] flush_workspace_to_db failed: {e}")

    def cleanup_unused_tables(self) -> None:
        """Optional maintenance: clear currently-unused match/item tables and reclaim space later with VACUUM."""
        try:
            self.conn.execute("DELETE FROM match_alt")
            self.conn.execute("DELETE FROM match_node")
            self.conn.execute("DELETE FROM match_run")
            self.conn.execute("DELETE FROM inventory_company_item")
            self.conn.commit()
        except Exception as e:
            print(f"[DB] cleanup_unused_tables failed: {e}")

    def load_cns(self, xlsx_path: str) -> int:
        self.cns_records = DataLoader.load_cns_workbook(xlsx_path)
        return len(self.cns_records)
    
    def _should_stop(self) -> bool:
        try:
            return bool(self.stop_event and self.stop_event.is_set())
        except Exception:
            return False

    def _db_enabled(self) -> bool:
        return self.workspace_id is not None


    def _lookup_bom_row_id_for_node(self, node: DecisionNode) -> Optional[str]:
        """
        Best-effort: map node.bom_uid -> bom_row_id if Step B created the mapping.
        If not available yet, returns None (still persists node fine).
        """
        try:
            uid = (getattr(node, "bom_uid", "") or "").strip()
            if not uid:
                return None
            return (self._bom_row_by_uid or {}).get(uid)
        except Exception:
            return None
#------------------------------------#------------------------------------#------------------------------------
    def _bom_row_dict_to_part(self, r: dict):
        """
        Rehydrate a BOM row dict into a lightweight object that behaves like DataLoader rows.
        We only rely on attribute access in the rest of the pipeline.
        """
        # parsed/raw_fields are stored as JSON strings in DB
        try:
            import json
            raw_fields = json.loads(r.get("raw_fields_json", "") or "{}")
        except Exception:
            raw_fields = {}

        try:
            import json
            parsed = json.loads(r.get("parsed_json", "") or "{}")
        except Exception:
            parsed = {}

        return SimpleNamespace(
            partnum=r.get("partnum", "") or "",
            mfgpn=r.get("mfgpn", "") or "",
            mfgname=r.get("mfgname", "") or "",
            supplier=r.get("supplier", "") or "",
            description=r.get("description", "") or "",
            raw_fields=raw_fields,
            parsed=parsed,
        )
#------------------------------------#------------------------------------#------------------------------------#------------------------------------
    def _persist_inventory_resolved_for_itemnum(self, itemnum: str) -> None:
        # v2 schema: inventory is stored in inventory_company and is not updated per-item.
        return


    def _persist_node_and_alts(self, node: DecisionNode) -> None:
        if not self.workspace_id:
            return

        try:
            _ws, _line = (node.id or "").split(":", 1)
            line_id = int(_line)
        except Exception:
            return

        sel = None
        try:
            sels = node.selected_alternates() if hasattr(node, "selected_alternates") else []
            sel = sels[0] if sels else None
        except Exception:
            sel = None

        anchored_cpn = (getattr(node, "internal_part_number", "") or "").strip()
        needs_new_cpn = False

        if sel and (getattr(sel, "source", "") or "").lower() == "inventory":
            anchored_cpn = (getattr(sel, "internal_part_number", "") or "").strip()
            needs_new_cpn = False
        elif anchored_cpn:
            needs_new_cpn = False
        else:
            anchored_cpn = ""
            needs_new_cpn = False

        ex = dict(getattr(node, "explain", {}) or {})
        ex["external_alternates"] = self._serialize_external_alternates_for_node(node)
        node.explain = ex
        ex_json = json.dumps(ex, ensure_ascii=False, default=str)

        patch: Dict[str, Any] = {
            "cpn": anchored_cpn,
            "needs_new_cpn": bool(needs_new_cpn),
            "notes": getattr(node, "notes", "") or "",
            "locked": 1 if bool(getattr(node, "locked", False)) else 0,
            "needs_approval": 1 if bool(getattr(node, "needs_approval", False)) else 0,
            "match_type": str(getattr(node, "match_type", "") or ""),
            "confidence": float(getattr(node, "confidence", 0.0) or 0.0),
            "explain_json": ex_json,
        }

        selected_mfg = ""
        selected_mpn = ""

        if anchored_cpn:
            selected_mpn = (
                str(ex.get("preferred_inventory_mfgpn") or "").strip()
                or str(getattr(node, "inventory_mpn", "") or "").strip()
            )

        if sel and (getattr(sel, "source", "") or "").lower() == "inventory":
            if not selected_mpn:
                selected_mpn = str(getattr(sel, "manufacturer_part_number", "") or "").strip()
            selected_mfg = str(getattr(sel, "manufacturer", "") or "").strip()
        else:
            inv_alt = None
            try:
                inv_alt = self._selected_inventory_alt(node)
            except Exception:
                inv_alt = None
            if inv_alt is not None:
                if not selected_mpn:
                    selected_mpn = str(getattr(inv_alt, "manufacturer_part_number", "") or "").strip()
                selected_mfg = str(getattr(inv_alt, "manufacturer", "") or "").strip()

        patch["selected_mfg"] = selected_mfg
        patch["selected_mpn"] = selected_mpn

        try:
            self.bom_repo.patch_state(self.workspace_id, line_id, patch)
        except Exception as e:
            print(f"[DB] bom_line_state patch failed for line {line_id}: {e}")

    def _serialize_external_alternates_for_node(self, node: DecisionNode) -> list[dict]:
        try:
            return external_alt_specs_from_node(node)
        except Exception:
            return []

    def _restore_external_alternates_for_node(self, node: DecisionNode) -> list[Alternate]:
        explain = dict(getattr(node, "explain", {}) or {})
        rows = explain.get("external_alternates") or []
        try:
            restored = restore_external_alternates(node, Alternate, rows)
        except Exception:
            restored = []
        if restored:
            node.explain = explain
        return restored

    def _persist_all_nodes(self) -> None:
        """Persist all in-memory DecisionNodes to DB (best effort)."""
        for n in (self.nodes or []):
            try:
                self._persist_node_and_alts(n)
            except Exception as e:
                print(f"[DB] persist node failed: {e}")

    def list_workspaces(self, *, status: str = "ACTIVE") -> List[Dict[str, Any]]:
        return self.ws_repo.list(status=status)

    def open_workspace(self, workspace_id: str) -> int:
        """Open an existing workspace and materialize UI nodes from authoritative DB state.

        This does **not** infer or mutate anything. It only:
          - Loads canonical BOM inputs from DB into self.npr_list (for matching/export workflows)
          - Builds DecisionNode views from DB state (bom_line_input + bom_line_state)
          - Resets per-workspace caches so UI interactions are consistent

        Node ids are always:  "{workspace_id}:{input_line_id}"
        """
        ws = self.ws_repo.get(workspace_id)
        if not ws:
            raise ValueError(f"Workspace not found: {workspace_id}")

        # --- HARD RESET: anything session-only must not bleed across workspaces ---
        self.workspace_id = workspace_id
        self.match_pairs = []
        self.external_cache = {}
        self.workspace = None

        # Clear caches / indices
        self.nodes = []
        self._node_index = {}
        self._nodes_workspace_id = workspace_id
        self._inventory_cache = None
        self._views_cache = []

        # Rehydrate BOM inputs (so run_matching can operate on an opened workspace)
        self.npr_list = []
        try:
            joined = self.bom_repo.load_joined_lines(workspace_id)
            for r in (joined or []):
                raw = r.get("input_raw_json") or {}
                if isinstance(raw, str):
                    try:
                        raw = json.loads(raw)
                    except Exception:
                        raw = {}
                raw_fields = {}
                parsed = {}
                if isinstance(raw, dict):
                    raw_fields = raw.get("raw_fields") or {}
                    parsed = raw.get("parsed") or {}

                self.npr_list.append(
                    SimpleNamespace(
                        partnum=r.get("input_partnum", "") or f"ROW-{r.get('input_line_id')}",
                        bom_uid=r.get("input_partnum", "") or str(r.get("input_line_id") or ""),
                        mfgpn=r.get("input_mfgpn", "") or "",
                        bom_mpn=r.get("input_mfgpn", "") or "",
                        mfgname=r.get("input_mfgname", "") or "",
                        supplier=r.get("input_supplier", "") or "",
                        description=r.get("input_description", "") or "",
                        raw_fields=raw_fields if isinstance(raw_fields, dict) else {},
                        parsed=parsed if isinstance(parsed, dict) else {},
                    )
                )
        except Exception as e:
            print(f"[DB] open_workspace: failed to rehydrate BOM inputs: {e}")

        # Build fresh decision views for THIS workspace
        views = self.build_decision_views()
        self.nodes = list(views or [])
        self._node_index = {n.id: n for n in (self.nodes or [])}

        return len(self.nodes)



    def build_decision_views(self) -> list[DecisionNode]:
        """
        Build UI-facing DecisionNode objects from authoritative DB state.
        No mutation. No inference. No persistence.

        CRITICAL:
          - When a line has a CPN, ensure node.alternates contains at least the
            internal inventory card (so bottom panels can populate).
          - Ensure the InventoryPart behind that card has vendoritem + substitutes
            so the UI shows ALL MFGPNs under the CPN (including exact).
        """
        if not self.workspace_id:
            return []

        rows = self.bom_repo.load_joined_lines(self.workspace_id)
        views: list[DecisionNode] = []

        for r in rows:
            cpn = str(r.get("cpn", "") or "").strip()
            needs_new_cpn = bool(r.get("needs_new_cpn", 0) or 0)

            base_type = "EXISTS" if cpn else "NEW"
            if needs_new_cpn and not cpn:
                base_type = "NEW"

            if needs_new_cpn:
                status = DecisionStatus.NEEDS_ALTERNATE
            elif cpn:
                status = DecisionStatus.EXISTS
            else:
                status = DecisionStatus.NEEDS_DECISION

            node = DecisionNode(
                id=f"{self.workspace_id}:{r['input_line_id']}",
                base_type=base_type,

                bom_uid=r.get("input_partnum") or str(r["input_line_id"]),
                bom_mpn=r.get("input_mfgpn") or "",
                description=(
                    str((r.get("explain_json") or {}).get("export_description") or "").strip()
                    or str((r.get("explain_json") or {}).get("export_description_override") or "").strip()
                    or r.get("input_description")
                    or ""
                ),

                internal_part_number=cpn,
                inventory_mpn=str(r.get("selected_mpn", "") or "").strip(),

                match_type=str(r.get("match_type", "") or "").strip(),
                confidence=float(r.get("confidence", 0.0) or 0.0),

                alternates=[],

                status=status,
                locked=bool(r.get("locked", 0) or 0),

                notes=str(r.get("notes", "") or ""),
                explain=dict(r.get("explain_json") or {}),
                
            )
            try:
                ex = dict(getattr(node, "explain", {}) or {})
                saved = (ex.get("export_description") or "").strip()
                if saved:
                    node.description = saved
                assigned = (ex.get("assigned_part_number") or "").strip()
                if assigned:
                    node.assigned_part_number = assigned
                if "approval_export_selected" in ex:
                    node.needs_approval = bool(ex.get("approval_export_selected"))
                else:
                    node.needs_approval = bool(r.get("needs_approval", 0) or 0)
            except Exception:
                pass
            # ---------------------------------------------------------
            # HYDRATE the internal inventory card when CPN exists
            # ---------------------------------------------------------
            if cpn:
                inv_obj = self._inv_by_itemnum.get(cpn)

                # If inventory wasn't loaded this session, reconstruct from DB
                if inv_obj is None:
                    try:
                        row_company = self.inv_company_repo.get(self.workspace_id, cpn)
                    except Exception:
                        row_company = None
                    if not row_company:
                        try:
                            row_company = self.inv_company_repo.get(self._inventory_store_workspace_id, cpn)
                        except Exception:
                            row_company = None

                    if row_company:
                        # repo returns alternates_json; normalize into the same shape our builder expects
                        company_row = {
                            "cpn": row_company.get("cpn", cpn),
                            "canonical_desc": row_company.get("canonical_desc", "") or "",
                            "stock_total": int(row_company.get("stock_total", 0) or 0),
                            "alternates": row_company.get("alternates_json", []) or [],
                        }
                        inv_obj = self._company_row_to_inventory_part(company_row)
                        if inv_obj is not None:
                            self._inv_by_itemnum[cpn] = inv_obj

                if inv_obj is not None:
                    # Rehydrate the anchored internal inventory card exactly as selected for this BOM line.
                    sel_mpn = str(r.get("selected_mpn", "") or "").strip()
                    bom_mpn = str(r.get("input_mfgpn", "") or "").strip()
                    conf = float(r.get("confidence", 0.0) or 0.0)
                    rel = "Base/Selected"

                    if sel_mpn and bom_mpn and sel_mpn.lower() == bom_mpn.lower():
                        conf = 1.0
                        rel = "Base/Selected [Exact MFGPN]"

                    alt = self._inventory_part_to_alternate(inv_obj, confidence=conf, relationship=rel)

                    # CRITICAL: the rebuilt inventory card must carry the line's persisted selected MPN,
                    # otherwise export falls back to the BOM input MPN after reopen/rehydration.
                    if sel_mpn:
                        alt.manufacturer_part_number = sel_mpn
                        alt.selected = True
                        node.inventory_mpn = sel_mpn
                        node.explain = dict(getattr(node, "explain", {}) or {})
                        node.explain["preferred_inventory_mfgpn"] = sel_mpn
                    else:
                        node.inventory_mpn = (node.inventory_mpn or getattr(inv_obj, "vendoritem", "") or "").strip()

                    node.alternates.append(alt)

            self._restore_external_alternates_for_node(node)
            views.append(node)

        self._views_cache = views
        return views

    def open_most_recent_workspace(self) -> int:
        wss = self.ws_repo.list(status="ACTIVE")
        if not wss:
            return 0
        return self.open_workspace(wss[0]["workspace_id"])
    
    def rematch_workspace_preserve_decisions(self) -> int:
        """
        Re-run matching for the CURRENT workspace using the current inventory + current npr_list,
        but preserve any existing decisions (status/locked/selected/rejected/notes).
        """
        if not self.workspace_id:
            raise RuntimeError("No workspace is open.")
        
        # Reload the authoritative decision state (including alternates + selected/rejected flags)
        # from the DB before re-running matching.
        wsid = self.workspace_id
        self.open_workspace(wsid)
    
        if not self.inventory:
            raise RuntimeError("Load inventory before re-running matching.")
    
        if not self.npr_list:
            raise RuntimeError("Workspace has no BOM loaded in memory; open workspace again or load BOM snapshot.")
    
        # Snapshot current decision-state nodes (authoritative)
        existing_nodes = list(self.nodes or [])
        existing_by_key = {self._node_key(n): n for n in existing_nodes}
    
        fresh_nodes = self._compute_fresh_nodes_via_run_matching()
    
        # Merge fresh suggestions into existing decisions
        merged: list[DecisionNode] = []
        for fresh in fresh_nodes:
            k = self._node_key(fresh)
            if k in existing_by_key:
                ex = existing_by_key[k]
                self._merge_suggestions_into_existing(ex, fresh)
                merged.append(ex)
            else:
                # New node (BOM row exists but no saved node yet)
                merged.append(fresh)
    
        self.nodes = merged
        return len(self.nodes)
    
    
    def _compute_fresh_nodes_via_run_matching(self) -> list:
        """
        Uses existing run_matching() to compute a fresh set of nodes,
        but returns them instead of leaving them as the authoritative self.nodes.
        """
        # Preserve current nodes while we run matching
        prev_nodes = self.nodes
        try:
            self.run_matching()           # this rebuilds self.nodes
            fresh_nodes = list(self.nodes or [])
            return fresh_nodes
        finally:
            # Restore; rematch_workspace_preserve_decisions() will set merged nodes after
            self.nodes = prev_nodes
    
    
    @staticmethod
    def _node_key(node) -> str:
        # DecisionNode has bom_uid (ROW-#). 
        k = (getattr(node, "bom_uid", "") or "").strip()
        if k:
            return k
        return (getattr(node, "id", "") or "").strip()
    
    
    @staticmethod
    def _alt_key(alt) -> tuple[str, str, str]:
        # Stable identity across reruns (ids will differ)
        src = (getattr(alt, "source", "") or "").strip().upper()
        ipn = (getattr(alt, "internal_part_number", "") or "").strip().upper()
        mpn = (getattr(alt, "manufacturer_part_number", "") or "").strip().upper()
        return (src, ipn, mpn)
    
    
    def _is_decided(self, node) -> bool:
        # Locked always means “do not touch”
        if bool(getattr(node, "locked", False)):
            return True
    
        st = getattr(node, "status", None)
        # Enum-safe + string-safe
        try:
            s = (getattr(st, "value", "") or getattr(st, "name", "") or str(st or "")).upper()
        except Exception:
            s = str(st or "").upper()
    
        return s not in ("", "NEEDS_DECISION", "OPEN")
    
 #------------------------------------#------------------------------------#------------------------------------#------------------------------------   
    def _merge_suggestions_into_existing(self, existing, fresh) -> None:
        """
        Refresh suggestion fields while preserving user decisions.
        """
    
        # Always refresh “suggestion/meta” fields
        existing.match_type = getattr(fresh, "match_type", existing.match_type)
        existing.confidence = float(getattr(fresh, "confidence", existing.confidence) or 0.0)
    
        fresh_explain = getattr(fresh, "explain", None)
        if isinstance(fresh_explain, dict):
            existing.explain = fresh_explain
    
        # Alternates: refresh list, preserve selected/rejected flags
        fresh_alts = list(getattr(fresh, "alternates", []) or [])
        old_alts = list(getattr(existing, "alternates", []) or [])
        old_by_key = {self._alt_key(a): a for a in old_alts}
    
        merged_alts = []
        for a in fresh_alts:
            k = self._alt_key(a)
            if k in old_by_key:
                old = old_by_key[k]
                a.selected = bool(getattr(old, "selected", False))
                a.rejected = bool(getattr(old, "rejected", False))
            merged_alts.append(a)
        existing.alternates = merged_alts
    
        # If decided, do NOT override chosen fields/status/notes/etc
        if self._is_decided(existing):
            return
    
        # If not decided yet, allow fresh “best guess” fields to update
        existing.internal_part_number = getattr(fresh, "internal_part_number", existing.internal_part_number)
        existing.inventory_mpn = getattr(fresh, "inventory_mpn", existing.inventory_mpn)
    

#------------------------------------#------------------------------------#------------------------------------#------------------------------------        
    def load_cns_preview(self, xlsx_path: str, last_n_per_sheet: int = 1) -> dict:
        """
        Loads CNS and prints:
          - every sheet name discovered
          - the LAST prefix-body (or last N) discovered per sheet

        last_n_per_sheet:
          1 -> just last
          2 -> last two, etc.
        """
        records = DataLoader.load_cns_workbook(xlsx_path)
        self.cns_records = records

        # sheet -> ordered unique PBs (preserve discovery order)
        sheet_pbs = defaultdict(list)
        sheet_seen = defaultdict(set)

        for r in records:
            sheet = (getattr(r, "sheet_name", "") or "").strip()
            prefix = (getattr(r, "prefix", "") or "").strip()
            body = (getattr(r, "body", "") or "").strip()
            if not sheet or not prefix or not body:
                continue

            pb = f"{prefix}-{body}"
            if not ControllerConfig.PB_RE.match(pb):
                continue

            if pb not in sheet_seen[sheet]:
                sheet_seen[sheet].add(pb)
                sheet_pbs[sheet].append(pb)

        all_sheets = sorted(sheet_pbs.keys())

        summary = {
            "total_records": len(records),
            "sheets_with_pbs": len(all_sheets),
            "unique_pb_total": sum(len(v) for v in sheet_pbs.values()),
            "sheets": all_sheets,
            "last_per_sheet": {
                s: (sheet_pbs[s][-max(1, int(last_n_per_sheet)):] if sheet_pbs[s] else [])
                for s in all_sheets
            },
            "unique_pb_per_sheet": {s: len(sheet_pbs[s]) for s in all_sheets},
        }
        return summary
 #------------------------------------#------------------------------------#------------------------------------#------------------------------------   
    def _merge_master_with_erp(self, master_inv: InventoryPart) -> InventoryPart:
        """
        Return an InventoryPart that keeps master identity/substitutes, but overlays
        ERP raw_fields for stock/cost display.
        """
        if not master_inv:
            return master_inv

        itemnum = (getattr(master_inv, "itemnum", "") or "").strip()
        erp = self._erp_by_itemnum.get(itemnum)

        if not erp:
            return master_inv  # no overlay available

        # Overlay: keep master object, but copy ERP raw_fields into it
        master_inv.raw_fields = dict(getattr(erp, "raw_fields", {}) or {})
        # If ERP has a better desc/mfg fields, optionally overlay those too:
        # master_inv.desc = master_inv.desc or erp.desc
        # master_inv.mfgname = master_inv.mfgname or erp.mfgname
        # master_inv.mfgid = master_inv.mfgid or erp.mfgid
        return master_inv

    def set_assigned_part_number(self, node_id: str, pn: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
    
        pn = (pn or "").strip()
        node.assigned_part_number = pn
    
        node.explain = dict(node.explain or {})
        node.explain["assigned_part_number"] = pn
    
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    # ----------------------------
    # Matching -> DecisionNodes
    # ----------------------------

    # ----------------------------
    # BOM section + Description overrides (UI-driven)
    # ----------------------------

    def get_node_bom_section(self, node_id: str) -> str:
        """Return the export bucket/section for this node (SURFACE MOUNT, THROUGH-HOLE, AUXILIARY - ...)."""
        node = self.get_node(node_id)
        try:
            sec = str((getattr(node, "explain", {}) or {}).get("bom_section", "") or "").strip()
            if sec:
                return sec
        except Exception:
            pass
        try:
            parsed = getattr(node, "parsed", None) or {}
            sec = str((parsed or {}).get("type", "") or "").strip()
            if sec:
                return sec
        except Exception:
            pass
        return "SURFACE MOUNT"

    def set_node_bom_section(self, node_id: str, section: str) -> str:
        """Set the BOM export section for a node. Best-effort persists to DB if schema supports it."""
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        section = (section or "").strip() or "SURFACE MOUNT"
        node.explain = dict(getattr(node, "explain", {}) or {})
        node.explain["bom_section"] = section

        if self.workspace_id:
            try:
                _ws, _line = (node.id or "").split(":", 1)
                line_id = int(_line)
                self.bom_repo.patch_state(self.workspace_id, line_id, {"bom_section": section})
            except Exception:
                pass

        return section


    def set_node_approval_export(self, node_id: str, include_on_approval_sheet: bool) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        node.explain = dict(getattr(node, "explain", {}) or {})
        node.explain["approval_export_selected"] = bool(include_on_approval_sheet)
        node.needs_approval = bool(include_on_approval_sheet)
        self._persist_node_and_alts(node)

    def get_internal_umbrella_count(self, node_id: str) -> int:
        node = self.get_node(node_id)
        cpn = (getattr(node, "internal_part_number", "") or "").strip()
        if not cpn:
            return 0

        inv_obj = self._inv_by_itemnum.get(cpn)
        if inv_obj is not None:
            all_mpns = list(getattr(inv_obj, "_all_mpns", []) or [])
            if all_mpns:
                return len([x for x in all_mpns if str(x or "").strip()])
            rep = str(getattr(inv_obj, "vendoritem", "") or "").strip()
            subs = [str(x or "").strip() for x in (getattr(inv_obj, "substitutes", []) or []) if str(x or "").strip()]
            vals = []
            if rep:
                vals.append(rep)
            vals.extend(subs)
            return len(vals)

        row_company = None
        try:
            row_company = self.inv_company_repo.get(self.workspace_id, cpn)
        except Exception:
            row_company = None
        if not row_company:
            try:
                row_company = self.inv_company_repo.get(self._inventory_store_workspace_id, cpn)
            except Exception:
                row_company = None
        if not row_company:
            return 0

        alts = row_company.get("alternates_json", []) or []
        if isinstance(alts, str):
            try:
                alts = json.loads(alts)
            except Exception:
                alts = []
        return len([
            a for a in (alts or [])
            if str((a or {}).get("mpn") or (a or {}).get("mfgpn") or (a or {}).get("manufacturer_part_number") or "").strip()
        ])

    #def set_node_description(self, node_id: str, description: str) -> None:
    #    """
    #    Persist the user-edited header description for export.
#
    #    Storage:
    #      - saved to node.explain["export_description_override"]
    #      - node.explain["export_description_override_applied"] = True
    #      - original BOM description preserved once in node.explain["input_description"]
#
    #    NOTE:
    #      Do NOT patch {"description": ...} into bom_line_state. That column doesn't exist.
    #      Persist via explain_json through _persist_node_and_alts().
    #    """
    #    node = self.get_node(node_id)
    #    self._ensure_unlocked(node)
#
    #    # Preserve original BOM description before overwriting
    #    prior_desc = (getattr(node, "description", "") or "").strip()
#
    #    new_desc = (description or "").strip()
#
    #    node.explain = dict(getattr(node, "explain", {}) or {})
    #    node.explain.setdefault("input_description", prior_desc)
#
    #    node.explain["export_description_override"] = new_desc
    #    node.explain["export_description_override_applied"] = True
#
    #    # Update live node description so UI reflects what's going to export
    #    node.description = new_desc
#
    #    self._recompute_node_flags(node)
    #    self._persist_node_and_alts(node)

    def set_node_description(self, node_id: str, description: str) -> None:
        """
        Persist the current header description text for this node.
        This is the export source of truth.
        """
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        new_desc = (description or "").strip()

        # Ensure explain exists so we can persist in bom_line_state.explain_json
        node.explain = dict(getattr(node, "explain", {}) or {})

        # Preserve original BOM description once (for your own reference/debug)
        node.explain.setdefault("input_description", (getattr(node, "description", "") or "").strip())

        # Store the current export description explicitly
        node.explain["export_description"] = new_desc

        # Keep node.description in sync with the header box
        node.description = new_desc

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def _selected_inventory_alt(self, node: DecisionNode) -> Optional[Alternate]:
        """Return the active inventory card for the node's anchored CPN, preferring an explicitly selected card."""
        ipn = (getattr(node, "internal_part_number", "") or "").strip()
        if not ipn:
            return None

        selected_inv = None
        try:
            for a in (node.selected_alternates() or []):
                if (getattr(a, "source", "") or "").lower() != "inventory":
                    continue
                if (getattr(a, "internal_part_number", "") or "").strip() != ipn:
                    continue
                if bool(getattr(a, "rejected", False)):
                    continue
                selected_inv = a
                break
        except Exception:
            selected_inv = None

        if selected_inv is not None:
            return selected_inv

        inv_mpn = (getattr(node, "inventory_mpn", "") or "").strip().lower()
        for a in (getattr(node, "alternates", []) or []):
            if (getattr(a, "source", "") or "").lower() != "inventory":
                continue
            if (getattr(a, "internal_part_number", "") or "").strip() != ipn:
                continue
            if bool(getattr(a, "rejected", False)):
                continue
            ampn = (getattr(a, "manufacturer_part_number", "") or "").strip().lower()
            if inv_mpn and ampn == inv_mpn:
                return a

        for a in (getattr(node, "alternates", []) or []):
            if (getattr(a, "source", "") or "").lower() == "inventory" and (getattr(a, "internal_part_number", "") or "").strip() == ipn:
                if not bool(getattr(a, "rejected", False)):
                    return a
        return None


    

    #def _export_description_for_node(self, node: DecisionNode) -> str:
    #    """
    #    Export description rules:
#
    #    1) If user edited header description (override applied), export that.
    #    2) Else if we can see a selected card (inventory or non-inventory), export that card's description.
    #    3) Else if selection exists only as persisted state (common after reload), resolve inventory desc by:
    #         - selected inventory CPN (node.internal_part_number or state cpn) OR
    #         - selected_mpn/selected_mfg
    #       and export the inventory description.
    #    4) Else fallback to node.description.
    #    """
    #    ex = getattr(node, "explain", {}) or {}
#
    #    # 1) Manual override wins when applied
    #    if bool(ex.get("export_description_override_applied")):
    #        d = str(ex.get("export_description_override") or "").strip()
    #        if d:
    #            return d
#
    #    # 2) In-memory selected card wins
    #    inv_alt = self._selected_inventory_alt(node)
    #    if inv_alt is not None:
    #        d = str(getattr(inv_alt, "description", "") or "").strip()
    #        if d:
    #            return d
#
    #    try:
    #        for a in (node.selected_alternates() or []):
    #            if bool(getattr(a, "rejected", False)):
    #                continue
    #            d = str(getattr(a, "description", "") or "").strip()
    #            if d:
    #                return d
    #    except Exception:
    #        pass
#
    #    # 3) Persisted-state resolution (important for export after reload)
    #    # 3a) Try by selected/anchored CPN
    #    cpn = (getattr(node, "internal_part_number", "") or "").strip()
    #    if cpn:
    #        try:
    #            inv = None
    #            if hasattr(self, "_inv_by_itemnum") and isinstance(self._inv_by_itemnum, dict):
    #                inv = self._inv_by_itemnum.get(cpn)
    #            if inv is not None:
    #                d = str(getattr(inv, "description", "") or "").strip()
    #                if d:
    #                    return d
    #                # Some InventoryPart variants store raw desc
    #                rf = getattr(inv, "raw_fields", {}) or {}
    #                d2 = str(rf.get("description") or rf.get("item_description") or "").strip()
    #                if d2:
    #                    return d2
    #        except Exception:
    #            pass
#
    #    # 3b) Try by selected MPN/MFG (if you persist these in bom_line_state)
    #    sel_mpn = (getattr(node, "selected_mpn", "") or "").strip()
    #    sel_mfg = (getattr(node, "selected_mfg", "") or "").strip()
    #    if sel_mpn:
    #        try:
    #            # Scan inventory for matching mpn/mfg if you don't have an index.
    #            # This is only used when the selected alt list isn't rebuilt.
    #            inv_list = getattr(self, "inventory", []) or []
    #            for inv in inv_list:
    #                mpn = (getattr(inv, "manufacturer_part_number", "") or "").strip()
    #                mfg = (getattr(inv, "manufacturer", "") or "").strip()
    #                if mpn and mpn == sel_mpn and (not sel_mfg or mfg == sel_mfg):
    #                    d = str(getattr(inv, "description", "") or "").strip()
    #                    if d:
    #                        return d
    #                    rf = getattr(inv, "raw_fields", {}) or {}
    #                    d2 = str(rf.get("description") or rf.get("item_description") or "").strip()
    #                    if d2:
    #                        return d2
    #                    break
    #        except Exception:
    #            pass
#
    #    # 4) Fallback
    #    return str(getattr(node, "description", "") or "").strip()


    def _export_description_for_node(self, node: DecisionNode) -> str:
        # Single source of truth: whatever is stored on the node.
        return str(getattr(node, "description", "") or "").strip()    
    
    def _resolve_inv(self, obj):
        """
        Ensure we always use the canonical InventoryPart from the master inventory
        (the one that has substitutes attached), even if the matching engine handed
        us a lightweight/copy candidate object.
        """
        if obj is None:
            return None
        try:
            itemnum = str(safe_get(obj, "itemnum", "internal_part_number", default="") or "").strip()
        except Exception:
            itemnum = ""
        if not itemnum:
            return obj
        return self._inv_by_itemnum.get(itemnum, obj)

    def run_matching(self) -> int:
        """
        Run matching for all NPR/BOM parts.
    
        IMPORTANT: Customer-provided alternates are matched, but the winning attempt is selected
        deterministically by (tier_rank, tier_quality, stock, stable_order) — NOT by confidence.
        Confidence remains a UI/display artifact.
        """
        if not self.inventory or not self.npr_list:
            raise RuntimeError("Load inventory and NPR/BOM parts before matching.")
#------------------------------------    
        cfg = load_config(self.cfg.components_yaml_path)
    
        # Parse both sides (robust attr names)
        for inv in self.inventory:
            if self._should_stop():
                return 0
            inv.parsed = parse_description(safe_get(inv, "description", "desc", default=""), cfg)
    
        # IMPORTANT: preserve loader-provided mpn_alts before overwriting parsed
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            existing_parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = []
            if isinstance(existing_parsed, dict):
                mpn_alts = existing_parsed.get("mpn_alts", []) or []
    
            new_parsed = parse_description(safe_get(npr, "description", "desc", default=""), cfg)
            if not isinstance(new_parsed, dict):
                new_parsed = {}
    
            if mpn_alts:
                new_parsed["mpn_alts"] = mpn_alts
    
            npr.parsed = new_parsed
    
        def uniq_keep_order(items):
            seen = set()
            out = []
            for x in items:
                x = (x or "").strip()
                if x and x not in seen:
                    seen.add(x)
                    out.append(x)
            return out

            # Determine first empty row after header in NPR sheet
            def _npr_first_empty_row():
                r = npr_header_row + 1
                while r <= npr_ws.max_row:
                    if all(str(npr_ws.cell(row=r, column=cc).value or "").strip() == "" for cc in range(1, 6)):
                        return r
                    r += 1
                return npr_ws.max_row + 1

            npr_write_row = _npr_first_empty_row()
    
        def _match_type_key(match: Any) -> str:
            mt = safe_get(match, "match_type", default="")
            if hasattr(mt, "value"):
                mt = mt.value
            mt = str(mt or "")
            return mt.replace("MatchType.", "").strip().upper()
    
        def _tier_rank(mt_key: str) -> int:
            ranks = {
                "EXACT_MFG_PN": 600,
                "API_ASSISTED": 550,
                "SUBSTITUTE": 500,
                "PREFIX_FAMILY": 400,
                "PARTIAL_ITEMNUM": 300,
                "PARSED_MATCH": 200,
                "NO_MATCH": 0,
            }
            return int(ranks.get(mt_key, 0))
    
        def _tier_quality(match: Any, mt_key: str) -> float:
            if mt_key == "PARSED_MATCH":
                exp = safe_get(match, "explain", default=None) or {}
                if isinstance(exp, dict):
                    q = exp.get("best_ratio", None)
                    if q is None:
                        q = exp.get("match_ratio", None)
                    if q is None:
                        q = exp.get("ratio", None)
                    try:
                        return float(q or 0.0)
                    except Exception:
                        return 0.0
                return 0.0
    
            inv = safe_get(match, "inventory_part", default=None)
            return 1.0 if inv is not None and mt_key != "NO_MATCH" else 0.0
    
        def _inv_stock(inv: Any) -> float:
            if inv is None:
                return 0.0
            for k in ("stock", "qty", "onhand", "available"):
                v = safe_get(inv, k, default=None)
                if v is None:
                    continue
                try:
                    return float(v or 0.0)
                except Exception:
                    continue
            return 0.0
    
        # Grab the active Tk root (if UI is running). IMPORTANT: do this
        # before constructing the engine so we can pass ui_root correctly.
        try:
            import tkinter as _tk
            tk_root = getattr(_tk, "_default_root", None)
        except Exception:
            tk_root = None

        #================================
        # CALLING MATCHING ENGINE
        #=================================
        engine = MatchingEngine(
            self.inventory,
            config=cfg,
            ui_root=tk_root,
            stop_event=getattr(self, "stop_event", None),
        )

        phase_cb = getattr(tk_root, "loading_phase_callback", None) if tk_root else None
        progress_cb = getattr(tk_root, "loading_progress_callback", None) if tk_root else None

        # Phase: embeddings
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Building semantic cache...", True))
            except Exception:
                pass            

        # Build semantic cache (sync; run_matching is already in a worker thread)
        engine.ensure_embeddings_cache()

        # Phase: matching
        if phase_cb and tk_root:
            try:
                if tk_root and tk_root.winfo_exists():
                    tk_root.after(0, lambda: phase_cb("Matching Engine is matching parts...", True))
            except Exception:
                pass

        match_pairs = []
    
        total_npr = max(1, len(self.npr_list))
        for n_i, npr in enumerate(self.npr_list):
            if self._should_stop():
                return 0
            parsed = getattr(npr, "parsed", {}) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            primary_mpn = safe_get(npr, "mfgpn", "bom_mpn", default="")
    
            mpn_options = uniq_keep_order([primary_mpn] + list(mpn_alts))
    
            attempts = []
            merged_candidates = []
    
            for idx, mpn in enumerate(mpn_options or [""]):
                if self._should_stop():
                    return 0
                tmp = copy.copy(npr)
                if hasattr(tmp, "mfgpn"):
                    tmp.mfgpn = mpn
                tmp.parsed = getattr(npr, "parsed", {}) or {}
    
                # --- alternates DB lookup (BOM MFGPN -> base internal item) ---
                match = None
                if getattr(self, "_alt_loaded", False) and getattr(self, "_alt_mpn_to_base", None):
                    key = DataLoader.norm_mpn_key(mpn)
                    bases = self._alt_mpn_to_base.get(key, []) if key else []

                    if bases:
                        inv_choices = [self._inv_by_itemnum.get(b) for b in bases]
                        inv_choices = [x for x in inv_choices if x is not None]

                        if inv_choices:
                            def _stock(x):
                                try:
                                    return float(getattr(x, "stock", 0) or 0)
                                except Exception:
                                    return 0.0

                            inv_choices.sort(key=_stock, reverse=True)
                            inv = inv_choices[0]
                            match = MatchResult(
                                match_type=MatchType.SUBSTITUTE,
                                confidence=1.0,
                                inventory_part=inv,
                                candidates=[inv],
                                notes="Resolved via master inventory (MPN->base)",
                                explain={
                                    "substitute_mpn": mpn,
                                    "conflict": (len(bases) > 1),
                                    "bases": bases,
                                    "base_itemnum": getattr(inv, "itemnum", None),
                                },
                            )

                # --- Fall back to normal ML matching if not resolved ---
                if match is None:
                    # Tell terminal + UI which part we're on (helps diagnose "stuck at 13%")
                    label = (getattr(tmp, "npr_item", "") or "").strip()
                    mpn = (getattr(tmp, "mfgpn", "") or "").strip()
                    desc = (getattr(tmp, "description", "") or "").strip()
                    short = (desc[:60] + "...") if len(desc) > 60 else desc

                    print(f"[MATCH] {n_i+1}/{total_npr} starting ML match: {label} | {mpn} | {short}")

                    if phase_cb and tk_root:
                        try:
                            if tk_root and tk_root.winfo_exists():
                                tk_root.after(
                                    0,
                                    lambda msg=f"Matching {n_i+1}/{total_npr}: {label or mpn or short}": phase_cb(msg)
                                )
                        except Exception:
                            pass

                    t_part = time.time()
                    match = engine.match_single_part(tmp)
                    dt = time.time() - t_part

                    mt = safe_get(match, "match_type", default=None)
                    print(f"[MATCH] {n_i+1}/{total_npr} done in {dt:.2f}s (match_type={mt})")

                mt_key = _match_type_key(match)
                tier = _tier_rank(mt_key)
                quality = _tier_quality(match, mt_key)
    
                inv = safe_get(match, "inventory_part", default=None)
                stock = _inv_stock(inv)
    
                cands = safe_get(match, "candidates", "candidate_parts", default=None) or []
                if inv is not None:
                    merged_candidates.append(inv)
                merged_candidates.extend(list(cands))
    
                inv_id = ""
                if inv is not None:
                    inv_id = str(
                        safe_get(inv, "itemnum", "internal_part_number", "vendoritem", default="") or ""
                    ).strip()
    
                attempts.append(
                    {
                        "customer_mpn": mpn,
                        "idx": idx,
                        "pair": (npr, match),  # keep ORIGINAL npr for node display
                        "match_type": mt_key,
                        "tier": tier,
                        "quality": quality,
                        "stock": stock,
                        "inv_id": inv_id,
                        "candidate_count": len(cands),
                    }
                )
    
            if not attempts:
                continue
            
            # Deterministic winner (no confidence)
            def _attempt_sort_key(a: dict) -> tuple:
                # prefer primary mpn on ties => earlier idx wins
                prefer_primary = -int(a.get("idx", 0))
                return (a["tier"], a["quality"], a["stock"], prefer_primary, a["inv_id"])
    
            winner = max(attempts, key=_attempt_sort_key)
            best_used_mpn = winner["customer_mpn"]
            npr_part, best_match = winner["pair"]
    
            # Stash customer MPN context and attempt audit for UI/debug
            try:
                if not isinstance(best_match.explain, dict):
                    best_match.explain = {}
                best_match.explain["customer_mpns"] = list(mpn_options)
                best_match.explain["winning_mpn"] = best_used_mpn
                best_match.explain["attempts"] = [
                    {
                        "customer_mpn": a["customer_mpn"],
                        "match_type": a["match_type"],
                        "tier": a["tier"],
                        "quality": a["quality"],
                        "stock": a["stock"],
                        "inv_id": a["inv_id"],
                        "candidate_count": a["candidate_count"],
                        "is_winner": a is winner,
                    }
                    for a in attempts
                ]
            except Exception:
                pass
            
            # Attach merged candidates back onto chosen match (best-effort)
            try:
                existing = safe_get(best_match, "candidates", "candidate_parts", default=None) or []

                # 1) raw de-dupe so we don't waste work.
                seen = set()
                raw = []
                for cand in (list(existing) + list(merged_candidates)):
                    key = safe_get(cand, "itemnum", "internal_part_number", "vendoritem", default=str(cand))
                    if key in seen:
                        continue
                    seen.add(key)
                    raw.append(cand)

                # 2) Sort by engine score and apply a loose raw cap (performance guard).
                raw = sorted(raw, key=lambda c: float(getattr(c, "_pc_score", 0.0) or 0.0), reverse=True)

                # Optional: raw cap BEFORE collapsing, to keep big BOMs snappy.
                RAW_CAP = max(int(getattr(ControllerConfig, "MAX_INTERNAL_CANDIDATES", 10) or 10) * 5, 50)
                raw = raw[:RAW_CAP]

                # 3) Collapse by company PN so UI shows one card per itemnum.
                collapsed, meta = self._collapse_candidates_by_itemnum(raw)

                # 4) Apply the real UI cap (top N itemnums).
                collapsed = collapsed[: ControllerConfig.MAX_INTERNAL_CANDIDATES]

                # ---------------------------------------------------------
                # stamp matched MPN onto collapsed inventory candidates
                # ---------------------------------------------------------
                winning_mpn = (best_match.explain.get("winning_mpn") or "").strip().lower()

                for inv in collapsed:
                    # InventoryPart → Alternate mapping happens later,
                    # but we pre-stamp the InventoryPart so the UI Alternate inherits it.
                    try:
                        # If this InventoryPart was resolved via master MPN mapping,
                        # mark it with the winning BOM MPN.
                        inv._matched_mpn = winning_mpn
                    except Exception:
                        pass

                # 5) Write back to match object + stash explain meta for UI/spec panels.
                if hasattr(best_match, "candidates"):
                    best_match.candidates = collapsed
                elif hasattr(best_match, "candidate_parts"):
                    best_match.candidate_parts = collapsed
                else:
                    # last resort: attach attribute
                    best_match.candidates = collapsed

                if not hasattr(best_match, "explain") or not isinstance(best_match.explain, dict):
                    best_match.explain = {}
                best_match.explain["collapsed_groups_by_itemnum"] = meta

            except Exception:
                pass

            match_pairs.append((npr_part, best_match))

            # Update UI progress for matching phase
            if progress_cb and tk_root:
                ratio = (n_i + 1) / total_npr   
                try:
                    if tk_root and tk_root.winfo_exists():
                        tk_root.after(0, lambda r=ratio: progress_cb(r))
                except Exception:
                    pass

        # ---  sync winning_mpn and attempt flags ---
        try:
            if not hasattr(best_match, "explain") or not isinstance(best_match.explain, dict):
                best_match.explain = {}
            winning_mpn = best_match.explain.get("winning_mpn", "")
            attempts = best_match.explain.get("attempts", [])
            for a in attempts:
                a["is_winner"] = (a.get("customer_mpn") == winning_mpn)
            best_match.explain["attempts"] = attempts

            # NEW: mark the winning inventory part explicitly
            if hasattr(best_match, "inventory_part") and best_match.inventory_part:
                inv = best_match.inventory_part
                if hasattr(inv, "vendoritem"):
                    inv._is_winner = True
                elif hasattr(inv, "manufacturer_part_number"):
                    inv._is_winner = True
        except Exception as e:
            print(f"[CTRL PATCH] failed to sync winner info: {e}")

        # --- Save explain JSON for debugging ---
        debug_dir = Path(os.getenv("NPR_DEBUG_DIR", "debug_match"))
        debug_dir.mkdir(parents=True, exist_ok=True)

        debug_parse = str(os.getenv("NPR_DEBUG_PARSE", "")).strip().lower() not in ("", "0", "false", "no")
        debug_save  = str(os.getenv("NPR_DEBUG_SAVE", "")).strip().lower()  not in ("", "0", "false", "no")
        debug_jsonl_only = str(os.getenv("NPR_DEBUG_JSONL_ONLY", "")).strip().lower() in ("1", "true", "yes")

        if debug_parse or debug_save:
            try:
                jsonl_path = debug_dir / "explain_all.jsonl"

                def _inv_to_payload_dict(inv_obj: Any, *, is_winner: bool = False) -> dict:
                    if inv_obj is None:
                        return {
                            "inv_item": "",
                            "inv_desc": "",
                            "vendor_mpn": "",
                            "mfg": "",
                            "stock": 0,
                            "seed": 0.0,
                            "score": 0.0,
                            "is_winner": bool(is_winner),
                        }

                    inv_item = str(safe_get(inv_obj, "itemnum", "internal_part_number", default="") or "").strip()
                    inv_desc = str(safe_get(inv_obj, "desc", "description", default="") or "").strip()
                    vendor_mpn = str(safe_get(inv_obj, "vendoritem", "manufacturer_part_number", default="") or "").strip()
                    mfg = str(safe_get(inv_obj, "manufacturer", "manufacturer_name", "mfgname", default="") or "").strip()

                    stock = 0
                    for k in ("stock", "qty", "onhand", "available"):
                        v = safe_get(inv_obj, k, default=None)
                        if v is None:
                            continue
                        try:
                            stock = int(float(v or 0))
                            break
                        except Exception:
                            continue
                        
                    seed = safe_get(inv_obj, "_pc_seed", default=None)
                    if seed is None:
                        seed = safe_get(inv_obj, "confidence", default=0.0)
                    seed = clamp01(seed)

                    score = safe_get(inv_obj, "_pc_score", default=None)
                    if score is None:
                        score = seed
                    score = clamp01(score)

                    return {
                        "inv_item": inv_item,
                        "inv_desc": inv_desc,
                        "vendor_mpn": vendor_mpn,
                        "mfg": mfg,
                        "stock": stock,
                        "seed": float(seed),
                        "score": float(score),
                        "is_winner": bool(is_winner),
                    }

                for npr_part, match in match_pairs:
                    if self._should_stop():
                        return 0
                    exp = safe_get(match, "explain", default=None) or {}

                    npr_key = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default="NPR") or "NPR").strip()
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

                    bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default="") or "").strip()
                    npr_desc = str(safe_get(npr_part, "description", "desc", default="") or "")

                    winner_inv = safe_get(match, "inventory_part", default=None)
                    winner_item = str(safe_get(winner_inv, "itemnum", "internal_part_number", default="") or "").strip()

                    # Pull candidates from the chosen match (already merged+filtered+capped earlier)
                    cands = safe_get(match, "candidates", "candidate_parts", default=None) or []

                    # Ensure winner is present in candidates (training needs it)
                    
                    if winner_inv is not None:
                        found = False
                        for c in cands:
                            cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                            if cid and cid == winner_item:
                                found = True
                                break
                        if not found:
                            cands = [winner_inv] + list(cands)

                    cand_payloads = []
                    for c in cands:
                        cid = str(safe_get(c, "itemnum", "internal_part_number", default="") or "").strip()
                        cand_payloads.append(_inv_to_payload_dict(c, is_winner=(cid == winner_item)))

                    # Find winner rank in candidate list (0-based). If not found, -1.
                    winner_rank = -1
                    for i, cp in enumerate(cand_payloads):
                        if cp.get("is_winner"):
                            winner_rank = i
                            break
                        
                    payload = {
                        "ts": ts,

                        # identifiers
                        "npr_item": npr_key,
                        "bom_mpn": bom_mpn,

                        # main text fields
                        "npr_desc": npr_desc,

                        # deterministic winner choice context
                        "match_type": str(safe_get(match, "match_type", default="")),
                        "winning_mpn": exp.get("winning_mpn") if isinstance(exp, dict) else "",
                        "customer_mpns": exp.get("customer_mpns") if isinstance(exp, dict) else [],

                        # winner block
                        "winner": _inv_to_payload_dict(winner_inv, is_winner=True),
                        "winner_rank": winner_rank,

                    
                        "candidates": cand_payloads,

                        "explain": exp,
                    }

                    with open(jsonl_path, "a", encoding="utf-8") as f:
                        f.write(json.dumps(payload, default=str) + "\n")

                    if not debug_jsonl_only:
                        inv_item = payload["winner"]["inv_item"] or "INV"
                        out_path = debug_dir / f"explain_{ts}_{npr_key}__{inv_item}.json"
                        with open(out_path, "w", encoding="utf-8") as f:
                            json.dump(payload, f, indent=2, default=str)
                        print(f"[MATCH DBG] wrote explain json -> {out_path}")

                print(f"[MATCH DBG] wrote JSONL -> {jsonl_path}")

            except Exception as e:
                print(f"[MATCH DBG] failed to write explain json/jsonl: {e}")
        
        self.match_pairs = match_pairs
    
        print("[CTRL DBG] building nodes...")
        self.nodes = [self._pair_to_node(npr_part, match, i) for i, (npr_part, match) in enumerate(self.match_pairs, start=1)]
        # Keep node index in sync for UI selection
        self._nodes_workspace_id = getattr(self, 'workspace_id', None)
        self._node_index = {n.id: n for n in (self.nodes or [])}
        print(f"[CTRL DBG] nodes built: {len(self.nodes)}")

        # ----------------------------
        # DB: persist nodes after matching
        # ----------------------------
        try:
            if self._db_enabled():
                self._persist_all_nodes()
        except Exception as e:
            print(f"[DB] failed to persist nodes after matching: {e}")

        return len(self.nodes)
    

    #=====================================================
    # CNS matching. this is very experimental and optional. as of now ALOT of it is hard coded and being wired in.
    #=======================================================
    def _build_type_prefix_map(self) -> dict[str, str]:
        print("[CTRL DBG] ENTER _build_type_prefix_map")
        """
        Learn which CNS prefix is most common for each parsed component type,
        using existing inventory parts that already have a company PN.
        """
        counts = defaultdict(Counter)

        for inv in (self.inventory or []):
            pn = getattr(inv, "itemnum", "") or ""
            prefix, body, _ = _split_company_pn(pn)
            if not prefix or not body:
                continue

            p = getattr(inv, "parsed", {}) or {}
            ptype = str(getattr(inv, "_pc_ptype", "") or p.get("type") or "OTHER")
            ptype = ptype.upper().strip()
            counts[ptype][prefix] += 1

        out = {}
        for ptype, ctr in counts.items():
            out[ptype] = ctr.most_common(1)[0][0]

        #print("[CTRL DBG] EXIT _build_type_prefix_map")
        return out
    
    def _suggest_new_body_for_prefix(self, prefix: str) -> str:
        """
        Suggest a new 5-digit body:
          - find max(body) among rows with a non-empty description
          - propose next, skipping over bodies that already exist WITH a description
          - allowing bodies that exist with blank descriptions
        """
        prefix = (prefix or "").strip()
        if not prefix:
            return ""
    
        def field(obj, name: str, default=""):
            # supports dataclass/objects AND dicts
            if isinstance(obj, dict):
                return obj.get(name, default)
            return getattr(obj, name, default)
    
        bodies = {}  # body_int -> has_desc(bool)
    
        for r in (getattr(self, "cns_records", None) or []):
            rp = str(field(r, "prefix", "")).strip()
            if rp != prefix:
                continue
            
            rb = str(field(r, "body", "")).strip()
            if not rb.isdigit():
                continue
            
            desc = str(field(r, "description", "")).strip()
            bodies[int(rb)] = bool(desc)
    
        if not bodies:
            return "10000"
    
        max_desc_body = max((b for b, has_desc in bodies.items() if has_desc), default=max(bodies))
        cand = max_desc_body + 1
    
        while cand in bodies and bodies[cand] is True:
            cand += 1
    
        return f"{cand:05d}"
    
    def _apply_cns_suggestion_to_node(self, node) -> None:
        print(f"[CTRL DBG] ENTER _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)}")
        # 1) If EXISTS: derive from internal PN
        if getattr(node, "internal_part_number", ""):
            prefix, body, _ = _split_company_pn(node.internal_part_number)
            if prefix and body:
                node.suggested_prefix = prefix
                node.suggested_body = body
                node.suggested_pb = f"{prefix}-{body}"
                node.suggested_reason = "from existing inventory PN"

                print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

                return

        # 2) NEW: use parsed type → prefix mapping
        ptype = ""
        # Try to infer from the parsed NPRPart via the node’s description (or store parsed on node later)
        ptype = getattr(node, "parsed_type", "") or "OTHER"

        prefix = ""
        if hasattr(self, "_type_prefix_map"):
            prefix = self._type_prefix_map.get(ptype, "")

        if not prefix:
            node.suggested_reason = "no prefix suggestion"
            print(f"[CTRL DBG] EXIT _apply_cns_suggestion_to_node node.id={getattr(node,'id',None)} suggested={getattr(node,'suggested_pb',None)}")

            return

        node.suggested_prefix = prefix

        # 3) NEW body suggestion (since we don’t have a confident body match yet)
        body = self._suggest_new_body_for_prefix(prefix)
        if body:
            node.suggested_body = body
            node.suggested_pb = f"{prefix}-{body}"
            node.suggested_reason = f"from type mapping ({ptype})"

    def _pair_to_node(self, npr_part: Any, match: Any, input_line_id: int) -> DecisionNode:
        """Convert a (BOM input, MatchResult) pair into a UI-facing DecisionNode.

        IMPORTANT INVARIANT:
          - node.id is always "{workspace_id}:{input_line_id}" so DB patching + UI selection are stable.
        """
        match_type = safe_get(match, "match_type", default="")
        if hasattr(match_type, "value"):
            match_type = match_type.value
        match_type = str(match_type) if match_type else ""

        inv = self._resolve_inv(safe_get(match, "inventory_part", default=None))

        # Determine base_type:
        # - NO_MATCH -> NEW
        # - else -> EXISTS if we have inventory_part
        base_type = "NEW" if match_type == MatchType.NO_MATCH else ("EXISTS" if inv is not None else "NEW")

        # BOM fields
        bom_uid = str(safe_get(npr_part, "bom_uid", "partnum", "itemnum", default="")).strip()
        bom_mpn = str(safe_get(npr_part, "bom_mpn", "mfgpn", "mpn", default=""))
        description = str(safe_get(npr_part, "description", "desc", default=""))

        # Inventory base fields (EXISTS)
        internal_pn = str(safe_get(inv, "itemnum", "internal_part_number", default="")) if inv else ""
        inv_mpn = str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="")) if inv else ""

        confidence = clamp01(safe_get(match, "confidence", default=0.0))

        # Stable node id for DB + UI
        node_id = f"{self.workspace_id}:{int(input_line_id)}" if self.workspace_id else f"ROW-{int(input_line_id)}"

        # compute parsed type BEFORE node creation
        parsed = getattr(npr_part, "parsed", {}) or {}
        ptype = str(getattr(npr_part, "_pc_ptype", "") or (parsed.get("type") if isinstance(parsed, dict) else "") or "OTHER")
        ptype = ptype.upper().strip()

        node = DecisionNode(
            id=node_id,
            base_type=base_type,
            bom_uid=bom_uid,
            bom_mpn=bom_mpn,
            description=description,
            internal_part_number="",
            inventory_mpn=inv_mpn,
            match_type=match_type,
            confidence=confidence,
            alternates=[],
            status=DecisionStatus.NEEDS_DECISION,
            locked=False,
            needs_approval=False,
            notes="",
            explain={},
        )

        # set dynamic attributes (used by CNS preview tooling)
        node.parsed_type = ptype

        # Suggested (not committed) Company PN derived from best inventory match, if any.
        # IMPORTANT: This must NOT anchor/export until the user clicks Add/Apply.
        if internal_pn:
            try:
                prefix, body, _ = _split_company_pn(internal_pn)
                if prefix and body:
                    node.suggested_prefix = prefix
                    node.suggested_body = body
                    node.suggested_pb = f"{prefix}-{body}"
                else:
                    node.suggested_prefix = ""
                    node.suggested_body = ""
                    node.suggested_pb = internal_pn
            except Exception:
                node.suggested_prefix = ""
                node.suggested_body = ""
                node.suggested_pb = internal_pn
        else:
            node.suggested_prefix = ""
            node.suggested_body = ""
            node.suggested_pb = ""

        # ----------------------------------------------------
        # Determine the MPN that caused the match (for display)
        # ----------------------------------------------------
        match_explain = safe_get(match, "explain", default=None) or {}
        if not isinstance(match_explain, dict):
            match_explain = {}

        # Prefer substitute_mpn (explicit alias) → winning_mpn → bom_mpn fallback
        winning_mpn_for_display = (
            (match_explain.get("substitute_mpn") or "").strip()
            or (match_explain.get("winning_mpn") or "").strip()
            or (bom_mpn or "").strip()
        )

        # ----------------------------------------------------
        # Internal Alternates (inventory)
        # Always show the WINNER inventory_part as a Base card, even if candidates exist.
        # ----------------------------------------------------
        candidates = safe_get(match, "candidates", "candidate_parts", default=None) or []
        seen_itemnums = set()
        seen_ids = set()

        if inv is not None:
            itemnum = (getattr(inv, "itemnum", "") or "").strip()
            if itemnum and itemnum not in seen_itemnums:
                base_alt = self._inventory_part_to_alternate(
                    inv,
                    confidence=confidence,
                    relationship="Base/Selected",
                )

                # Stamp grouped MFGPN metadata for the base card too (same as candidates),
                # and ONLY show a matched MFGPN that actually belongs to this CPN group.
                try:
                    base_subs = getattr(inv, "substitutes", None) or []
                    rep_vendoritem = (getattr(inv, "vendoritem", "") or "").strip()
                    seen_grp = set()
                    grp_mpns = []
                    def _push_grp(m):
                        m = (m or "").strip()
                        if not m:
                            return
                        k = m.lower()
                        if k in seen_grp:
                            return
                        seen_grp.add(k)
                        grp_mpns.append(m)
                    _push_grp(rep_vendoritem)
                    for _s in base_subs:
                        if isinstance(_s, str):
                            _push_grp(_s)
                        else:
                            _push_grp(getattr(_s, "mfgpn", "") or getattr(_s, "manufacturer_part_number", "") or "")

                    base_alt.meta = dict(getattr(base_alt, "meta", {}) or {})
                    base_alt.meta["company_pn_mfgpn_count"] = len(grp_mpns)
                    # keep aliases only (exclude rep vendoritem) for consistency with candidate stamping
                    base_alt.meta["company_pn_mfgpns"] = grp_mpns[1:] if len(grp_mpns) > 1 else []
                    base_alt.meta["company_pn_rep_vendoritem"] = rep_vendoritem

                    win = (winning_mpn_for_display or "").strip()
                    if win and win.lower() in {m.lower() for m in grp_mpns}:
                        base_alt._matched_mpn_ui = win
                    else:
                        # BOM MPN may have won matching, but if it is not a true alternate under this CPN,
                        # display the representative/real grouped MFGPN on the card.
                        base_alt._matched_mpn_ui = rep_vendoritem or (grp_mpns[0] if grp_mpns else "")
                except Exception:
                    if winning_mpn_for_display:
                        base_alt._matched_mpn_ui = winning_mpn_for_display

                node.alternates.append(base_alt)
                seen_itemnums.add(itemnum)

        for cand in candidates:
            # Prefer per-candidate score; otherwise fall back so candidates never show 0%.
            cand_score = safe_get(cand, "_pc_score", default=None)
            conf = cand_score if cand_score is not None else 0.25
            conf = clamp01(conf)

            via_mpn = (getattr(cand, "_best_source_customer_mpn", "") or "").strip()
            via_mt = (getattr(cand, "_best_source_match_type", "") or "").strip()

            rel = "Internal Candidate"
            if via_mpn and via_mpn != (bom_mpn or "").strip():
                rel = f"Internal Candidate (via {via_mpn})"
            if via_mt:
                rel = f"{rel} [{via_mt}]"

            # show collapsed alias count (if stamped)
            grp_count = int(getattr(cand, "_ui_group_count", 1) or 1)
            grp_mpns = list(getattr(cand, "_ui_group_mpns", []) or [])
            if grp_count > 1:
                rel = f"{rel} (+{grp_count - 1} MFGPNs)"

            # Rebind candidate -> master inventory by itemnum so EVERY card has substitutes
            cand_item = str(safe_get(cand, "itemnum", "internal_part_number", default="") or "").strip()
            master_inv = self._inv_by_itemnum.get(cand_item) if cand_item else None
            inv_for_ui = master_inv or cand

            cand_res = self._resolve_inv(cand)

            # preserve any temporary UI/grouping attributes computed on the candidate
            for attr in (
                "_pc_score",
                "_ui_group_count",
                "_ui_group_mpns",
                "_ui_group_best_vendoritem",
                "_best_source_customer_mpn",
                "_best_source_match_type",
            ):
                if hasattr(cand, attr) and not hasattr(cand_res, attr):
                    try:
                        setattr(cand_res, attr, getattr(cand, attr))
                    except Exception:
                        pass

            alt_obj = self._inventory_part_to_alternate(cand_res, confidence=conf, relationship=rel)

            # keep meta stamping (if group mpns exist)
            alt_obj.meta["company_pn_mfgpn_count"] = grp_count
            alt_obj.meta["company_pn_mfgpns"] = grp_mpns
            alt_obj.meta["company_pn_rep_vendoritem"] = str(getattr(cand, "_ui_group_best_vendoritem", "") or "")

            itemnum = (getattr(alt_obj, "internal_part_number", "") or "").strip()
            if itemnum and itemnum in seen_itemnums:
                continue
            seen_itemnums.add(itemnum)
            seen_ids.add(alt_obj.id)
            node.alternates.append(alt_obj)

        # ----------------------------------------------------
        # Customer-provided alternates (external)
        # Show them as external cards in the UI.
        # ----------------------------------------------------
        explain = safe_get(match, "explain", default=None) or {}
        customer_mpns = []
        winning_mpn = ""

        if isinstance(explain, dict):
            customer_mpns = explain.get("customer_mpns", []) or []
            winning_mpn = (explain.get("winning_mpn", "") or "").strip()

        if not customer_mpns:
            parsed = safe_get(npr_part, "parsed", default=None) or {}
            mpn_alts = parsed.get("mpn_alts", []) if isinstance(parsed, dict) else []
            customer_mpns = [bom_mpn] + list(mpn_alts)

        if not winning_mpn:
            winning_mpn = (bom_mpn or "").strip()

        node.explain = dict(safe_get(match, "explain", default={}) or {})
        node.explain["winning_mpn"] = winning_mpn

        # Deduplicate against alternates already present
        seen_mpns = set()
        for a in node.alternates:
            m = (getattr(a, "manufacturer_part_number", "") or "").strip()
            if m:
                seen_mpns.add(m)

        for mpn in customer_mpns:
            mpn = (mpn or "").strip()
            if not mpn or mpn == winning_mpn or mpn in seen_mpns:
                continue
            seen_mpns.add(mpn)

            node.alternates.append(
                Alternate(
                    id=f"CUST-{node.id}-{mpn}",
                    source="customer_bom",
                    manufacturer=str(safe_get(npr_part, "mfgname", default="")),
                    manufacturer_part_number=mpn,
                    internal_part_number="",
                    description=description,
                    confidence=0.0,
                    relationship="Customer Provided",
                    selected=False,
                    rejected=False,
                    raw=None,
                    meta={"customer_provided": True},
                )
            )

        # Ensure the surviving inventory card shows the resolved MPN
        for alt in node.alternates:
            if (
                alt.source == "inventory"
                and alt.internal_part_number == internal_pn
                and winning_mpn_for_display
            ):
                try:
                    inv._matched_mpn = winning_mpn_for_display
                except Exception:
                    pass

        self._recompute_node_flags(node)
        return node

    def _inventory_part_to_alternate(self, inv: InventoryPart, confidence: float, relationship: str = "Inventory") -> Alternate:
        """
        Convert an InventoryPart into a UI-facing Alternate card.

        IMPORTANT:
        - Stock/cost come from inv.raw_fields (ITEMS sheet), not from the master MPN sheet.
        - MFGPN-count badge comes from master sheet substitutes attached to InventoryPart.substitutes.
        """

        def _raw_get(*keys: str, default=""):
            raw = getattr(inv, "raw_fields", {}) or {}
            for k in keys:
                if k in raw:
                    v = raw.get(k)
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s != "":
                        return s
            return default

        def _to_int(x, default=0):
            try:
                if x is None:
                    return default
                s = str(x).strip()
                if s == "":
                    return default
                # handle "1,234"
                s = s.replace(",", "")
                return int(float(s))
            except Exception:
                return default

        def _to_float(x, default=None):
            try:
                if x is None:
                    return default
                s = str(x).strip()
                if s == "":
                    return default
                s = s.replace(",", "")
                return float(s)
            except Exception:
                return default

        # ---- STOCK / COST from ITEMS-sheet raw_fields (or fallback keys) ----
        qty_raw = _raw_get("totalqty", "total_qty", "qty_on_hand", "on_hand", "quantity", default="0")
        stock_qty = _to_int(qty_raw, default=0)

        # prefer avgcost, then lastcost, then stdcost
        cost_raw = _raw_get("avgcost", "avg_cost", "average_cost", "lastcost", "last_cost", "stdcost", "std_cost", default="")
        unit_cost = _to_float(cost_raw, default=None)

        # ---- MFGPN badge count from MASTER-sheet substitutes ----
        subs = getattr(inv, "substitutes", None) or []
        rep_vendoritem = (getattr(inv, "vendoritem", "") or "").strip()
        mfgpn_count = 0
        if rep_vendoritem:
            mfgpn_count += 1
        # count unique substitute mpns
        seen = set()
        for s in subs:
            if isinstance(s, str):
                mpn = s.strip()
            else:
                mpn = (getattr(s, "mfgpn", "") or "").strip()
            k = mpn.lower()
            if mpn and k not in seen:
                seen.add(k)
                mfgpn_count += 1


        # -------------------------------------------------
        # DISPLAY MPN SELECTION (CRITICAL FIX)
        # -------------------------------------------------
        display_mpn = rep_vendoritem

        # If the matching pipeline stamped a resolved MPN for UI,
        # ALWAYS prefer it for display (do NOT mutate inventory)
        ui_mpn = getattr(inv, "_matched_mpn", None)
        if ui_mpn:
            display_mpn = str(ui_mpn).strip()

        alt = Alternate(
            id=Alternate.new_id("INV"),
            source="inventory",
            manufacturer=(getattr(inv, "mfgname", "") or "").strip(),
            manufacturer_part_number=display_mpn,
            internal_part_number=(getattr(inv, "itemnum", "") or "").strip(),
            description=(getattr(inv, "desc", "") or "").strip(),
            value=str(getattr(inv, "parsed", {}).get("value", "") or ""),
            package=str(getattr(inv, "parsed", {}).get("package", "") or ""),
            tolerance=str(getattr(inv, "parsed", {}).get("tolerance", "") or ""),
            voltage=str(getattr(inv, "parsed", {}).get("voltage", "") or ""),
            wattage=str(getattr(inv, "parsed", {}).get("wattage", "") or ""),
            stock=int(stock_qty or 0),
            unit_cost=unit_cost,
            supplier=str(_raw_get("primaryvendornumber", "primary_vendor_number", "supplier", "vendor", default="")),
            confidence=clamp01(confidence),
            relationship=relationship,
            selected=False,
            rejected=False,
            raw=inv,
            meta={
                # used by UI badge + display logic
                "company_pn_mfgpn_count": int(mfgpn_count or 0),
                "company_pn_rep_vendoritem": rep_vendoritem,
            },
        )
        return alt
#------------------------------------
    

    def _collapse_candidates_by_itemnum(self, candidates):
        """
        Collapse a list of InventoryPart candidates so UI shows ONE card per company PN (itemnum).

        Returns:
            collapsed_list: list of representative InventoryPart (best _pc_score per itemnum)
            meta_by_itemnum: dict[itemnum] -> {"count": int, "mpns": [..], "best_vendoritem": str}
        
        Side effects:
            Stamps each representative candidate with:
              - _ui_group_count (int)
              - _ui_group_mpns (list[str])
              - _ui_group_best_vendoritem (str)
        """
        groups = {}   # itemnum -> {"count":int, "mpns":[...], "best_vendoritem":str}
        best = {}     # itemnum -> (inv, score)

        def _key(inv):
            return str(safe_get(inv, "itemnum", "internal_part_number", default="") or "").strip()

        def _mpn(inv):
            return str(safe_get(inv, "vendoritem", "manufacturer_part_number", default="") or "").strip()

        def _score(inv):
            try:
                return float(getattr(inv, "_pc_score", 0.0) or 0.0)
            except Exception:
                return 0.0

        for inv in (candidates or []):
            item = _key(inv)
            if not item:
                # No stable grouping key -> skip (or treat as singleton elsewhere)
                continue

            mpn = _mpn(inv)
            sc = _score(inv)

            g = groups.setdefault(item, {"count": 0, "mpns": [], "best_vendoritem": ""})
            g["count"] += 1
            if mpn and mpn not in g["mpns"]:
                g["mpns"].append(mpn)

            if (item not in best) or (sc > best[item][1]):
                best[item] = (inv, sc)
                g["best_vendoritem"] = mpn

        collapsed = [t[0] for t in best.values()]
        collapsed.sort(key=lambda inv: _score(inv), reverse=True)

        # Stamp UI metadata onto representative rows
        for inv in collapsed:
            item = _key(inv)
            meta = groups.get(item, {})
            inv._ui_group_count = int(meta.get("count", 1) or 1)
            inv._ui_group_mpns = list(meta.get("mpns", []) or [])
            inv._ui_group_best_vendoritem = str(meta.get("best_vendoritem", "") or "")

        return collapsed, groups


    # ----------------------------
    # Query helpers for UI
    # ----------------------------

    def get_nodes(self) -> list[DecisionNode]:
        """UI-facing list of DecisionNode objects for the current workspace.

        IMPORTANT: The UI must be able to select a node id it sees in the table and
        then immediately fetch the same node from the controller. Therefore:
          - get_nodes() and get_node() MUST reference the same in-memory node set.
          - We only rebuild nodes from DB when needed (open_workspace / lazy load).
        """
        if not getattr(self, "workspace_id", None):
            return []

        # Lazy load if we haven't materialized nodes for this workspace yet
        if not getattr(self, "nodes", None) or getattr(self, "_nodes_workspace_id", None) != self.workspace_id:
            views = self.build_decision_views()
            self.nodes = list(views or [])
            self._node_index = {n.id: n for n in (self.nodes or [])}
            self._nodes_workspace_id = self.workspace_id

        return list(self.nodes or [])

    def get_node(self, node_id: str) -> DecisionNode:
        # Fast path: index lookup
        idx = getattr(self, "_node_index", None)
        if isinstance(idx, dict) and node_id in idx:
            return idx[node_id]

        # Slow path: scan current nodes, then backfill the index
        for n in (self.nodes or []):
            if n.id == node_id:
                if not isinstance(idx, dict):
                    self._node_index = {}
                self._node_index[n.id] = n
                return n

        raise KeyError(f"No DecisionNode with id={node_id}")

    # ----------------------------
    # Mutations (UI actions)
    # ----------------------------

    def select_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        # Cache the original BOM description once so we can restore it if the user unselects everything.
        node.explain = dict(getattr(node, "explain", {}) or {})
        if "input_description" not in node.explain:
            node.explain["input_description"] = (getattr(node, "description", "") or "")

        alt = self._find_alt(node, alt_id)
        alt.rejected = False
        alt.selected = True

        # ELEVATE: if the selected alternate has a Company PN (CNS), it becomes the anchor.
        alt_cpn = (getattr(alt, "internal_part_number", "") or "").strip()
        if alt_cpn:
            for other in (getattr(node, "alternates", []) or []):
                if other is alt:
                    continue
                if (getattr(other, "source", "") or "").lower() == "inventory":
                    other.selected = False
            node.internal_part_number = alt_cpn
            node.base_type = "EXISTS"
            node.explain = dict(getattr(node, "explain", {}) or {})
            preferred_mpn = (node.explain.get("preferred_inventory_mfgpn") or "").strip()
            matched_ui_mpn = (getattr(alt, "_matched_mpn_ui", "") or "").strip()
            chosen_inventory_mpn = preferred_mpn or matched_ui_mpn or (getattr(alt, "manufacturer_part_number", "") or "").strip()
            if chosen_inventory_mpn:
                alt.manufacturer_part_number = chosen_inventory_mpn
                node.inventory_mpn = chosen_inventory_mpn
                node.explain["preferred_inventory_mfgpn"] = chosen_inventory_mpn
            else:
                node.inventory_mpn = getattr(alt, "manufacturer_part_number", "") or node.inventory_mpn

        ## Always update the export description to the selected card's description (header should reflect this).
        #chosen_desc = (getattr(alt, "description", "") or "").strip()
        #if chosen_desc:
        #    node.description = chosen_desc


        # After selection is committed
        chosen_desc = (getattr(alt, "description", "") or "").strip()
        if chosen_desc:
            self.set_node_description(node.id, chosen_desc)
        else:
            # still persist whatever node.description is currently
            self.set_node_description(node.id, getattr(node, "description", "") or "")


        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def reject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False
        alt.rejected = True

        # If this rejected inventory alternate is currently anchoring the node, clear the anchor
        if (
            getattr(alt, "source", "") == "inventory"
            and getattr(alt, "internal_part_number", "")
            and alt.internal_part_number == (node.internal_part_number or "")
            and float(getattr(node, "confidence", 0.0) or 0.0) < 0.999
        ):
            node.internal_part_number = ""
            node.inventory_mpn = ""

        # ---- DEMOTION RULE ----
        # If there are NO remaining inventory candidates (non-rejected, non-selected),
        # and we have no selected alternates, treat as NEW.
        remaining_inventory_candidates = [
            a for a in node.alternates
            if getattr(a, "source", "") == "inventory"
            and (not getattr(a, "rejected", False))
            and (not getattr(a, "selected", False))
        ]
        
        node.explain = dict(node.explain or {})
        node.explain["pre_demote"] = {
            "base_type": node.base_type,
            "confidence": node.confidence,
        }

        if (len(remaining_inventory_candidates) == 0) and (not node.has_selection()):


            node.base_type = "NEW"
            node.internal_part_number = ""
            node.inventory_mpn = ""
            node.confidence = 0.0
            node.explain = dict(node.explain or {})
            node.explain["demoted_to_new"] = "All inventory candidates were rejected."

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)



    def unreject_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)

        # Undo reject
        alt.rejected = False
        alt.selected = False  # keep it simple: unreject ≠ auto-select

        # If we previously demoted to NEW because "all inventory candidates were rejected",
        # restore the pre-demote base_type/confidence when at least one inventory candidate exists again.
        if isinstance(getattr(node, "explain", None), dict):
            pre = node.explain.get("pre_demote")
            if pre and node.explain.get("demoted_to_new"):
                remaining_inventory_candidates = [
                    a for a in node.alternates
                    if getattr(a, "source", "") == "inventory"
                    and (not getattr(a, "rejected", False))
                ]
                if remaining_inventory_candidates:
                    node.base_type = pre.get("base_type", node.base_type)
                    node.confidence = pre.get("confidence", node.confidence)
                    node.explain.pop("demoted_to_new", None)
                    node.explain.pop("pre_demote", None)

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

    def unselect_alternate(self, node_id: str, alt_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = self._find_alt(node, alt_id)
        alt.selected = False

        # If nothing is selected anymore, clear anchor + restore original BOM description
        # (unless the user has explicitly applied/confirmed another description).
        if hasattr(node, "has_selection") and (not node.has_selection()):
            node.internal_part_number = ""
            node.base_type = "NEW"
            node.inventory_mpn = ""

            ex = dict(getattr(node, "explain", {}) or {})
            confirmed = bool(ex.get("export_description_confirmed"))
            override_applied = bool(ex.get("export_description_override_applied"))
            if (not confirmed) and (not override_applied):
                orig = (ex.get("input_description") or "").strip()
                if orig:
                    node.description = orig

        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)


    def add_manual_alternate(self, node_id: str, manufacturer_part_number: str, description: str = "Manual alternate") -> Alternate:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        alt = Alternate(
            id=Alternate.new_id("MAN"),
            source="manual",
            manufacturer="",
            manufacturer_part_number=manufacturer_part_number,
            internal_part_number="",
            description=description,
            confidence=0.0,
            relationship="Manual",
        )
        node.alternates.append(alt)
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

        return alt

    def _has_selected_external_alt(self, node: DecisionNode) -> bool:
        try:
            for a in (node.selected_alternates() or []):
                if (getattr(a, "source", "") or "").lower() == "inventory":
                    continue
                if bool(getattr(a, "rejected", False)):
                    continue
                return True
        except Exception:
            pass
        return False

    def mark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        has_external_selected = self._has_selected_external_alt(node)
        node.status = DecisionStatus.READY_FOR_EXPORT
        node.locked = not has_external_selected
        self._recompute_node_flags(node)
        node.status = DecisionStatus.READY_FOR_EXPORT
        self._persist_node_and_alts(node)


    def unmark_ready(self, node_id: str) -> None:
        node = self.get_node(node_id)
        node.locked = False
        # move back to a workable status without trying to fully infer everything
        if getattr(node, 'internal_part_number', '') or node.has_selection():
            node.status = DecisionStatus.NEEDS_DECISION
        else:
            node.status = DecisionStatus.NEEDS_DECISION
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

    def set_preferred_inventory_mfgpn(self, node_id: str, alt_id: str, mfgpn: str) -> None:
        node = self.get_node(node_id)
        self._ensure_unlocked(node)
        alt = self._find_alt(node, alt_id)
        if (getattr(alt, 'source', '') or '').strip().lower() != 'inventory':
            raise ValueError('Preferred MFG PN can only be set for inventory alternates.')
        mfgpn = (mfgpn or '').strip()
        if not mfgpn:
            raise ValueError('MFG PN is blank.')
        alt.manufacturer_part_number = mfgpn
        alt.meta = dict(getattr(alt, 'meta', {}) or {})
        alt.meta['company_pn_rep_vendoritem'] = mfgpn
        node.explain = dict(getattr(node, 'explain', {}) or {})
        node.explain['preferred_inventory_mfgpn'] = mfgpn
        self._persist_node_and_alts(node)



    def get_alt_detail_payload(self, node_id: str, alt_id: str) -> dict:
        """
        Controller-backed detail payload for the UI (DB/controller is the truth).
        Returns:
          { "specs": {...}, "export_mfgpn_options": [..] }
        """
        node = self.get_node(node_id)
        alt = self._find_alt(node, alt_id)

        def _raw_get(obj, *keys, default=""):
            if obj is None:
                return default
            raw = getattr(obj, "raw_fields", {}) or {}
            if isinstance(raw, dict):
                for k in keys:
                    v = raw.get(k)
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s:
                        return s
            # fallback: direct attrs
            for k in keys:
                v = getattr(obj, k, None)
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return default

        def _pick_sub_field(sub_obj, *keys):
            # Try common direct attrs first, then raw_fields aliases.
            for k in keys:
                v = getattr(sub_obj, k, None)
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return _raw_get(sub_obj, *keys, default="")

        specs = {}
        export_opts = []

        if (getattr(alt, "source", "") or "").strip().lower() == "inventory":
            itemnum = (getattr(alt, "internal_part_number", "") or "").strip()
            inv = None
            if itemnum:
                inv = self._inv_by_itemnum.get(itemnum)
            inv = inv or getattr(alt, "raw", None)

            if inv is not None:
                seen = set()

                def add_mpn(m):
                    m = (m or "").strip()
                    if not m:
                        return
                    k = m.lower()
                    if k in seen:
                        return
                    seen.add(k)
                    export_opts.append(m)

                # Build export options strictly from company-part grouping truth.
                base_vendoritem = (getattr(inv, "vendoritem", "") or "").strip() or _raw_get(inv, "vendoritem", "vendor_item", "mfgpn")
                add_mpn(base_vendoritem)

                # grouped aliases stamped during candidate collapsing
                for m in list((getattr(alt, "meta", {}) or {}).get("company_pn_mfgpns", []) or []):
                    add_mpn(m)

                # inventory substitutes
                subs = list(getattr(inv, "substitutes", None) or [])
                for s in subs:
                    if isinstance(s, str):
                        add_mpn(s)
                    else:
                        add_mpn(getattr(s, "mfgpn", "") or getattr(s, "manufacturer_part_number", "") or _pick_sub_field(s, "mfgpn", "vendoritem", "vendor_item"))

                # selected export MFG PN (must be valid under this CPN)
                current_selected = (getattr(alt, "manufacturer_part_number", "") or "").strip()
                if current_selected and current_selected.lower() in seen:
                    selected_vendoritem = current_selected
                else:
                    pref = str(((getattr(node, "explain", {}) or {}).get("preferred_inventory_mfgpn") or "")).strip()
                    if pref and pref.lower() in seen:
                        selected_vendoritem = pref
                    else:
                        selected_vendoritem = base_vendoritem or (export_opts[0] if export_opts else "")

                # If the selected vendor item corresponds to a substitute object, surface THAT
                # substitute's details in the panel (description/manufacturer/etc.).
                selected_sub = None
                if selected_vendoritem:
                    sv = selected_vendoritem.strip().lower()
                    for s in subs:
                        if isinstance(s, str):
                            continue
                        smpn = (getattr(s, "mfgpn", "") or getattr(s, "manufacturer_part_number", "") or _pick_sub_field(s, "mfgpn", "vendoritem", "vendor_item")).strip()
                        if smpn and smpn.lower() == sv:
                            selected_sub = s
                            break

                # Base values from the inventory/company item row
                desc_val = (getattr(inv, "desc", "") or "").strip() or _raw_get(inv, "desc", "description")
                mfg_name_val = (getattr(inv, "mfgname", "") or "").strip() or _raw_get(inv, "mfgname", "manufacturer_name")
                mfg_id_val = (getattr(inv, "mfgid", "") or "").strip() or _raw_get(inv, "mfgid", "manufacturer_id")
                supplier_val = _raw_get(inv, "primaryvendornumber", "supplier", "vendor")
                last_cost_val = _raw_get(inv, "lastcost", "last_cost")
                avg_cost_val = _raw_get(inv, "avgcost", "avg_cost", "average_cost")
                lead_val = _raw_get(inv, "itemleadtime", "item_lead_time", "lead_time")

                # Override with substitute details when available for the selected alternate MPN
                if selected_sub is not None:
                    desc_val = _pick_sub_field(selected_sub, "description", "desc") or desc_val
                    mfg_name_val = _pick_sub_field(selected_sub, "manufacturer", "mfgname", "manufacturer_name") or mfg_name_val
                    mfg_id_val = _pick_sub_field(selected_sub, "mfgid", "manufacturer_id") or mfg_id_val
                    supplier_val = _pick_sub_field(selected_sub, "supplier", "vendor", "primaryvendornumber") or supplier_val
                    # cost fields vary by source naming
                    avg_cost_val = (
                        _pick_sub_field(selected_sub, "unit_cost", "avgcost", "avg_cost", "price")
                        or avg_cost_val
                    )
                    last_cost_val = (
                        _pick_sub_field(selected_sub, "lastcost", "last_cost", "unit_cost", "price")
                        or last_cost_val
                    )
                    lead_val = _pick_sub_field(selected_sub, "lead_time", "itemleadtime", "item_lead_time") or lead_val

                specs = {
                    "ItemNumber": (getattr(inv, "itemnum", "") or "").strip() or _raw_get(inv, "itemnum", "item_number", "itemnumber"),
                    "VendorItem": selected_vendoritem,
                    "Description": desc_val,
                    "MfgName": mfg_name_val,
                    "MfgId": mfg_id_val,
                    "PrimaryVendorNumber": supplier_val,
                    "TotalQty": _raw_get(inv, "totalqty", "total_qty", "qty_on_hand", "on_hand", "quantity"),
                    "LastCost": last_cost_val,
                    "AvgCost": avg_cost_val,
                    "ItemLeadTime": lead_val,
                    "DefaultWhse": _raw_get(inv, "defaultwhse", "default_whse", "warehouse"),
                    "TariffCodeHTSUS": _raw_get(inv, "tariffcodehtsus", "htsus", "tariff_code"),
                }

                if export_opts:
                    specs["AlternatesCount"] = str(len(export_opts))
                    specs["AlternatesList"] = "\n".join(export_opts)

        else:
            # Non-inventory alt fallback details
            specs = {
                "ItemNumber": (getattr(alt, "internal_part_number", "") or "").strip(),
                "VendorItem": (getattr(alt, "manufacturer_part_number", "") or "").strip(),
                "Description": (getattr(alt, "description", "") or "").strip(),
                "MfgName": (getattr(alt, "manufacturer", "") or "").strip(),
                "PrimaryVendorNumber": (getattr(alt, "supplier", "") or "").strip(),
                "TotalQty": "" if getattr(alt, "stock", None) in (None, "") else str(getattr(alt, "stock", "")),
                "AvgCost": "" if getattr(alt, "unit_cost", None) in (None, "") else str(getattr(alt, "unit_cost", "")),
            }

        return {"specs": specs, "export_mfgpn_options": export_opts}

    def seed_fake_external_alternates(self, node_id: str) -> List[Alternate]:
        """Seed deterministic external alternate cards for UI testing without touching export rules."""
        node = self.get_node(node_id)
        self._ensure_unlocked(node)

        rows = build_fake_external_alt_specs(node)
        created = append_external_alternates(node, Alternate, rows)
        if created:
            node.explain = dict(getattr(node, "explain", {}) or {})
            node.explain["external_alternates"] = self._serialize_external_alternates_for_node(node)
            self._recompute_node_flags(node)
            self._persist_node_and_alts(node)
        return created

    def search_digikey(self, node_id: str) -> List[Alternate]:
        """
        Cache is controller-owned (not UI). UI just calls this.
        """
        node = self.get_node(node_id)

        if node_id in self.external_cache:
            return self.external_cache[node_id]

        if not self._digikey_search_fn:
            # no client wired yet
            self.external_cache[node_id] = []
            return []

        results = self._digikey_search_fn(node) or []
        # Normalize + attach
        for alt in results:
            alt.source = alt.source or "digikey"
        self.external_cache[node_id] = results

        # Option: merge external alternates into node list (recommended)
        # so UI doesn’t need “internal vs external” stores.
        append_external_alternates(node, Alternate, [
            {
                "id": getattr(alt, "id", ""),
                "source": getattr(alt, "source", "") or "digikey",
                "manufacturer": getattr(alt, "manufacturer", ""),
                "manufacturer_part_number": getattr(alt, "manufacturer_part_number", ""),
                "internal_part_number": getattr(alt, "internal_part_number", ""),
                "description": getattr(alt, "description", ""),
                "confidence": getattr(alt, "confidence", 0.0),
                "relationship": getattr(alt, "relationship", "External Alternate"),
                "selected": getattr(alt, "selected", False),
                "rejected": getattr(alt, "rejected", False),
                "meta": dict(getattr(alt, "meta", {}) or {}),
            }
            for alt in results
        ])
        node.explain = dict(getattr(node, "explain", {}) or {})
        node.explain["external_alternates"] = self._serialize_external_alternates_for_node(node)
        self._recompute_node_flags(node)
        self._persist_node_and_alts(node)

        return results

 #------------------------------------   # ----------------------------
    # NPR Workspace (from DecisionNodes)
    # ----------------------------
    def export_npr(self, output_path: str = None):
            """Export a single Excel workbook containing BOM + NPR outputs in separate sheets.

            Sheets:
              - BOM (from BOM_TEMPLATE if available)
              - NPR
              - Alternates for Approval
              - Bomexistences

            Description export rules:
              - Default is BLANK (input BOM description is not exported unless user overrides).
              - If an inventory card is selected/anchored, export that card's description.
              - If user applied an override, export the override.
            """

            try:
                self.flush_workspace_to_db()
            except Exception as e:
                print(f"[DB] flush on export failed: {e}")

            if not self.workspace_id:
                raise RuntimeError("No active workspace. Load a BOM or open a workspace first.")

            ready_nodes = [n for n in (self.get_nodes() or []) if getattr(n, "status", None) == DecisionStatus.READY_FOR_EXPORT]
            if not ready_nodes:
                raise RuntimeError("No parts marked Ready for Export.")

            from openpyxl import Workbook

            bom_template_path = Path(self.cfg.bom_template_path)
            if bom_template_path.exists():
                wb = load_workbook(bom_template_path)
            else:
                wb = Workbook()
                try:
                    wb.remove(wb.active)
                except Exception:
                    pass

            if wb.sheetnames:
                bom_ws = wb[wb.sheetnames[0]]
                bom_ws.title = "BOM"
            else:
                bom_ws = wb.create_sheet("BOM")

            npr_ws = wb["NPR"] if "NPR" in wb.sheetnames else wb.create_sheet("NPR")
            ws_alt = wb["Alternates for Approval"] if "Alternates for Approval" in wb.sheetnames else wb.create_sheet("Alternates for Approval")
            ws_bomexist = wb["BOMNOTES"] if "BOMNOTES" in wb.sheetnames else wb.create_sheet("BOMNOTES")

            def _norm(v):
                return str(v or "").strip().lower()

            def find_bom_sections(ws):
                sections = {}
                for r in range(1, ws.max_row + 1):
                    b = _norm(ws.cell(row=r, column=2).value)
                    c = _norm(ws.cell(row=r, column=3).value)
                    d = _norm(ws.cell(row=r, column=4).value)
                    e = _norm(ws.cell(row=r, column=5).value)
                    if b == "elan part number" and c == "description" and d.startswith("qty") and e == "refdes":
                        label = str(ws.cell(row=r - 1, column=2).value or ws.cell(row=r - 1, column=1).value or "").strip()
                        label_u = label.upper()
                        if "SURFACE" in label_u:
                            sections["SURFACE MOUNT"] = r
                        elif "THROUGH" in label_u or "THRU" in label_u:
                            sections["THROUGH-HOLE"] = r
                        elif "AUX" in label_u:
                            sections["AUXILIARY"] = r
                        else:
                            sections[label or f"SECTION@{r}"] = r
                return sections

            sections = find_bom_sections(bom_ws)
            # --- Template table expansion ---
            # BOM_TEMPLATE uses multiple section tables with a fixed row footprint. Writing past the
            # footprint can overwrite adjacent sections or make Excel report corruption. We prevent
            # that by inserting rows inside the section and expanding the corresponding Table.ref.

            from copy import copy as _copy_style
            from openpyxl.utils.cell import range_boundaries, get_column_letter
            from openpyxl.worksheet.table import Table

            def _find_table_by_header_row(ws, header_row: int):
                try:
                    for t in getattr(ws, "tables", {}).values():
                        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                        if min_row == header_row:
                            return t
                except Exception:
                    pass
                return None

            def _shift_table_ref(t: Table, delta_rows: int, start_row_inclusive: int):
                # Shift a table's ref if rows are inserted at/above it.
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                    if min_row >= start_row_inclusive:
                        min_row += delta_rows
                        max_row += delta_rows
                    elif max_row >= start_row_inclusive:
                        max_row += delta_rows
                    t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                except Exception:
                    pass

            def _ensure_section_capacity(ws, header_row: int, required_rows: int, next_header_row: int = None):
                data_start = header_row + 1

                # Next section header row implies a label row one row above it; insert before that label
                if next_header_row:
                    insert_at = max(1, next_header_row - 1)
                    available = max(0, insert_at - data_start)
                else:
                    insert_at = ws.max_row + 1
                    available = max(0, (ws.max_row + 1) - data_start)

                if required_rows <= available:
                    return

                need = required_rows - available
                if need <= 0:
                    return

                tables = list(getattr(ws, "tables", {}).values()) if getattr(ws, "tables", None) else []
                ws.insert_rows(insert_at, amount=need)

                # Shift all table refs below the insertion point
                for t in tables:
                    _shift_table_ref(t, need, insert_at)

                # Expand the section table itself (if present)
                t = _find_table_by_header_row(ws, header_row)
                if t is not None:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
                        if next_header_row:
                            new_max_row = next_header_row - 2  # last data row before next label row
                        else:
                            new_max_row = max_row + need
                        if new_max_row > max_row:
                            t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
                    except Exception:
                        pass

                # Copy styling from the first existing data row (or header if no data row exists)
                style_src_row = data_start if data_start <= ws.max_row else header_row
                for rr in range(insert_at, insert_at + need):
                    ws.row_dimensions[rr].height = ws.row_dimensions[style_src_row].height
                    for cc in range(1, ws.max_column + 1):
                        src = ws.cell(row=style_src_row, column=cc)
                        dst = ws.cell(row=rr, column=cc)
                        dst._style = _copy_style(src._style)
                        dst.number_format = src.number_format
                        dst.font = _copy_style(src.font)
                        dst.fill = _copy_style(src.fill)
                        dst.border = _copy_style(src.border)
                        dst.alignment = _copy_style(src.alignment)
                        dst.protection = _copy_style(src.protection)
                        dst.comment = None


            if not sections:
                bom_ws.delete_rows(1, bom_ws.max_row)
                bom_ws["A1"] = "BOM Export"
                base_headers = ["Item", "ELAN Part Number", "Description", "QTY.", "REFDES", "TYPE", "Manufacturer", "Manufacturer PN", "Price", "EXT Price"]
                start = 4
                for sec_name in ("SURFACE MOUNT", "THROUGH-HOLE", "AUXILIARY"):
                    bom_ws.cell(row=start, column=2, value=sec_name)
                    hdr = start + 1
                    for j, h in enumerate(base_headers, start=1):
                        bom_ws.cell(row=hdr, column=j, value=h)
                    sections[sec_name] = hdr
                    start = hdr + 3

            def first_empty_row(ws, header_row):
                r = header_row + 1
                while r <= ws.max_row:
                    if str(ws.cell(row=r, column=2).value or "").strip() == "" and str(ws.cell(row=r, column=3).value or "").strip() == "":
                        return r
                    r += 1
                return ws.max_row + 1

            write_row_by_section = {sec: first_empty_row(bom_ws, hdr) for sec, hdr in sections.items()}

            joined = self.bom_repo.load_joined_lines(self.workspace_id) or []
            by_line_id = {}
            for r in joined:
                try:
                    lid = int(r.get("input_line_id") or 0)
                    if lid:
                        by_line_id[lid] = r
                except Exception:
                    pass

            item_counters = {"SURFACE MOUNT": 1, "THROUGH-HOLE": 1, "AUXILIARY": 1}

            def section_bucket(sec_raw: str) -> tuple[str, str]:
                s = (sec_raw or "").strip().upper()
                if s.startswith("SURFACE"):
                    return "SURFACE MOUNT", "SMT"
                if s.startswith("THROUGH") or s.startswith("THRU"):
                    return "THROUGH-HOLE", "TH"
                if s.startswith("AUXILIARY") or s.startswith("AUX"):
                    if "ASSEMBLY" in s:
                        return "AUXILIARY", "assembly"
                    if "MECH" in s:
                        return "AUXILIARY", "mech"
                    if "PRODUCTION" in s:
                        return "AUXILIARY", "production"
                    if "OTHER" in s:
                        return "AUXILIARY", "other"
                    return "AUXILIARY", "aux"
                return "SURFACE MOUNT", "SMT"

            def export_mfg_fields(node: DecisionNode) -> tuple[str, str]:
                ex = dict(getattr(node, "explain", {}) or {})
                preferred_inventory_mfgpn = (ex.get("preferred_inventory_mfgpn") or "").strip()
                anchored_cpn = (getattr(node, "internal_part_number", "") or getattr(node, "assigned_part_number", "") or "").strip()
                inv_alt = self._selected_inventory_alt(node)

                if anchored_cpn:
                    manufacturer = str(getattr(inv_alt, "manufacturer", "") or "").strip() if inv_alt is not None else ""
                    mpn = (
                        preferred_inventory_mfgpn
                        or str(getattr(node, "inventory_mpn", "") or "").strip()
                        or (str(getattr(inv_alt, "manufacturer_part_number", "") or "").strip() if inv_alt is not None else "")
                        or (str(getattr(inv_alt, "_matched_mpn_ui", "") or "").strip() if inv_alt is not None else "")
                    )
                    if not mpn:
                        inv_obj = self._inv_by_itemnum.get(anchored_cpn)
                        if inv_obj is not None:
                            mpn = str(getattr(inv_obj, "vendoritem", "") or "").strip()
                            if not manufacturer:
                                manufacturer = str(getattr(inv_obj, "mfgname", "") or "").strip()
                    return (manufacturer, mpn)

                try:
                    for a in (node.selected_alternates() or []):
                        if (getattr(a, "source", "") or "").lower() == "inventory":
                            continue
                        if bool(getattr(a, "rejected", False)):
                            continue
                        return (str(getattr(a, "manufacturer", "") or "").strip(), str(getattr(a, "manufacturer_part_number", "") or "").strip())
                except Exception:
                    pass
                return ("", str(getattr(node, "bom_mpn", "") or "").strip())


            # Ensure the BOM sections have enough rows for this export before we start writing.
            try:
                from collections import Counter as _Counter
                counts = _Counter()
                for _n in ready_nodes:
                    _sec = self.get_node_bom_section(_n.id)
                    _bucket, _t = section_bucket(_sec)
                    counts[_bucket] += 1

                sec_items = [(sec_name, int(hdr_row)) for sec_name, hdr_row in (sections or {}).items() if isinstance(hdr_row, int)]
                sec_items.sort(key=lambda x: x[1])
                hdr_rows_sorted = [r for _, r in sec_items]
                next_by_hdr = {r: (hdr_rows_sorted[i + 1] if i + 1 < len(hdr_rows_sorted) else None) for i, r in enumerate(hdr_rows_sorted)}

                for sec_name, hdr_row in sec_items:
                    add_count = int(counts.get(sec_name, 0))
                    if add_count <= 0:
                        continue
                    used = int(write_row_by_section.get(sec_name, hdr_row + 1)) - (hdr_row + 1)
                    required = max(0, used) + add_count
                    _ensure_section_capacity(bom_ws, hdr_row, required, next_header_row=next_by_hdr.get(hdr_row))
            except Exception as _cap_e:
                print(f"[EXPORT] BOM section expansion skipped: {_cap_e}")


            for node in ready_nodes:
                cpn = (getattr(node, "internal_part_number", "") or getattr(node, "assigned_part_number", "") or "").strip() or "NEW"
                sec = self.get_node_bom_section(node.id)
                bucket, type_val = section_bucket(sec)

                line_id = 0
                try:
                    _ws, _line = (node.id or "").split(":", 1)
                    line_id = int(_line)
                except Exception:
                    line_id = 0
                jr = by_line_id.get(line_id, {}) if line_id else {}

                qty = ""
                for k in ("input_qty", "qty", "input_quantity", "quantity"):
                    if k in jr and jr.get(k) not in (None, ""):
                        qty = jr.get(k)
                        break
                if qty in (None, ""):
                    raw = jr.get("input_raw_json") or {}
                    if isinstance(raw, str):
                        try:
                            raw = json.loads(raw)
                        except Exception:
                            raw = {}
                    if isinstance(raw, dict):
                        qty = (raw.get("raw_fields") or {}).get("quantity", "") or raw.get("qty", "") or ""

                refdes = ""
                for k in ("input_refdes", "refdes", "input_designator", "designator"):
                    if k in jr and jr.get(k) not in (None, ""):
                        refdes = jr.get(k)
                        break
                if not refdes:
                    raw = jr.get("input_raw_json") or {}
                    if isinstance(raw, str):
                        try:
                            raw = json.loads(raw)
                        except Exception:
                            raw = {}
                    if isinstance(raw, dict):
                        refdes = (raw.get("raw_fields") or {}).get("designator", "") or ""

                desc = self._export_description_for_node(node)
                mfg, mfgpn = export_mfg_fields(node)

                row_idx = write_row_by_section.get(bucket) or (bom_ws.max_row + 1)
                item = item_counters[bucket]
                item_counters[bucket] += 1

                bom_ws.cell(row=row_idx, column=1, value=item)
                bom_ws.cell(row=row_idx, column=2, value=cpn)
                bom_ws.cell(row=row_idx, column=3, value=desc)
                bom_ws.cell(row=row_idx, column=4, value=qty)
                bom_ws.cell(row=row_idx, column=5, value=refdes)
                bom_ws.cell(row=row_idx, column=6, value=type_val)
                bom_ws.cell(row=row_idx, column=7, value=mfg)
                bom_ws.cell(row=row_idx, column=8, value=mfgpn)

                write_row_by_section[bucket] = row_idx + 1

            # NPR + Approval templates
            try:
                npr_template_path = Path(self.cfg.npr_template_path)
            except Exception:
                npr_template_path = Path(str(getattr(self.cfg, "npr_template_path", "") or ""))

            alts_template_path = Path(os.path.dirname(os.path.abspath(__file__))) / "ALTS_template.xlsx"

            def _copy_template_sheet(template_path: Path, target_name: str, fallback_headers: list[str]):
                if target_name in wb.sheetnames:
                    wb.remove(wb[target_name])
                ws = wb.create_sheet(target_name)

                if template_path and template_path.exists():
                    try:
                        twb = load_workbook(template_path)
                        tws = twb[twb.sheetnames[0]]
                        from copy import copy as _cpy
                        for r in range(1, tws.max_row + 1):
                            ws.row_dimensions[r].height = tws.row_dimensions[r].height
                            for c in range(1, tws.max_column + 1):
                                s = tws.cell(row=r, column=c)
                                d = ws.cell(row=r, column=c, value=s.value)
                                d._style = _cpy(s._style)
                                d.number_format = s.number_format
                                d.font = _cpy(s.font)
                                d.fill = _cpy(s.fill)
                                d.border = _cpy(s.border)
                                d.alignment = _cpy(s.alignment)
                                d.protection = _cpy(s.protection)
                        try:
                            for col_letter, dim in tws.column_dimensions.items():
                                ws.column_dimensions[col_letter].width = dim.width
                        except Exception:
                            pass
                        return ws
                    except Exception as e:
                        print(f"[EXPORT] template copy failed for {target_name}: {e}")

                ws.delete_rows(1, ws.max_row)
                ws.append(fallback_headers)
                return ws

            npr_ws = _copy_template_sheet(
                npr_template_path,
                "NPR",
                [
                    "Part Number",
                    "Part Status (new / exists)",
                    "Item Description",
                    "Custom Fields- SMT, TH, Process, Assembly,PCB, Mechanical",
                    "Preferred Part",
                    "Unit Cost",
                    "Default Units of Measure",
                    "Stock Unit-EA, ml, ounce, gram",
                    "per",
                    "Purchase Unit",
                    "Manufacturer Name",
                    "Manufacturer Part #",
                    "Supplier",
                    "Lead Time (WKS)",
                    "QC Required",
                    "TARIFF CODE (HTSUS)",
                    "Quote Number (Attach a Copy)",
                ],
            )

            ws_alt = _copy_template_sheet(
                alts_template_path,
                "Alternates for Approval",
                ["ELAN Inventory PN", "MfgNs under this PN", "Manufacturer", "link", "DESC", "BOM MfgN subbing for", "Designator", "NOTES"],
            )

            if "BOMNOTES" in wb.sheetnames:
                ws_bomexist = wb["BOMNOTES"]
                ws_bomexist.delete_rows(1, ws_bomexist.max_row)
            else:
                ws_bomexist = wb.create_sheet("BOMNOTES")
            ws_bomexist.append(["BOM_UID", "BOM_MPN", "BOM_DESC", "COMPANY_PN", "STATUS", "NOTES"])

            def _find_row_by_header(ws, expected: list[str], max_scan: int = 200) -> int:
                norm_expected = [str(x or "").strip().lower() for x in expected]
                for r in range(1, min(ws.max_row, max_scan) + 1):
                    vals = [str(ws.cell(row=r, column=i + 1).value or "").strip().lower() for i in range(len(norm_expected))]
                    if vals == norm_expected:
                        return r
                return 1

            npr_header_row = _find_row_by_header(
                npr_ws,
                [
                    "Part Number",
                    "Part Status (new / exists)",
                    "Item Description",
                    "Custom Fields- SMT, TH, Process, Assembly,PCB, Mechanical",
                    "Preferred Part",
                    "Unit Cost",
                    "Default Units of Measure",
                    "Stock Unit-EA, ml, ounce, gram",
                    "per",
                    "Purchase Unit",
                    "Manufacturer Name",
                    "Manufacturer Part #",
                    "Supplier",
                    "Lead Time (WKS)",
                    "QC Required",
                    "TARIFF CODE (HTSUS)",
                    "Quote Number (Attach a Copy)",
                ],
            )
            npr_write_row = max(9, npr_header_row + 1)
            while npr_write_row <= npr_ws.max_row:
                if all(str(npr_ws.cell(row=npr_write_row, column=cc).value or "").strip() == "" for cc in range(1, 18)):
                    break
                npr_write_row += 1

            alts_header_row = _find_row_by_header(
                ws_alt,
                ["ELAN Inventory PN", "MfgNs under this PN", "Manufacturer", "link", "DESC", "BOM MfgN subbing for", "Designator", "NOTES"],
            )
            alts_write_row = max(5, alts_header_row + 1)
            while alts_write_row <= ws_alt.max_row:
                if all(str(ws_alt.cell(row=alts_write_row, column=cc).value or "").strip() == "" for cc in range(1, 9)):
                    break
                alts_write_row += 1

            def _normalize_text(v: Any) -> str:
                return str(v or "").strip()

            def _line_context_for_node(node: DecisionNode) -> dict:
                line_id = 0
                try:
                    _ws, _line = (node.id or "").split(":", 1)
                    line_id = int(_line)
                except Exception:
                    line_id = 0
                jr = by_line_id.get(line_id, {}) if line_id else {}

                qty = ""
                for k in ("input_qty", "qty", "input_quantity", "quantity"):
                    if k in jr and jr.get(k) not in (None, ""):
                        qty = jr.get(k)
                        break

                raw = jr.get("input_raw_json") or {}
                if isinstance(raw, str):
                    try:
                        raw = json.loads(raw)
                    except Exception:
                        raw = {}
                raw_fields = (raw.get("raw_fields") or {}) if isinstance(raw, dict) else {}

                refdes = ""
                for k in ("input_refdes", "refdes", "input_designator", "designator"):
                    if k in jr and jr.get(k) not in (None, ""):
                        refdes = jr.get(k)
                        break
                if not refdes:
                    refdes = raw_fields.get("designator", "") or ""

                mfgname = ""
                for k in ("input_mfgname", "mfgname", "manufacturer", "manufacturer_name"):
                    if k in jr and jr.get(k) not in (None, ""):
                        mfgname = jr.get(k)
                        break

                return {
                    "qty": qty,
                    "refdes": refdes,
                    "mfgname": mfgname,
                }

            def export_mfg_fields(node: DecisionNode) -> tuple[str, str]:
                ex = dict(getattr(node, "explain", {}) or {})
                preferred_inventory_mfgpn = _normalize_text(ex.get("preferred_inventory_mfgpn"))
                anchored_cpn = _normalize_text(getattr(node, "internal_part_number", "") or getattr(node, "assigned_part_number", ""))
                inv_alt = self._selected_inventory_alt(node)

                if anchored_cpn:
                    manufacturer = _normalize_text(getattr(inv_alt, "manufacturer", "") if inv_alt is not None else "")
                    mpn = (
                        preferred_inventory_mfgpn
                        or _normalize_text(getattr(node, "inventory_mpn", ""))
                        or _normalize_text(getattr(inv_alt, "manufacturer_part_number", "") if inv_alt is not None else "")
                        or _normalize_text(getattr(inv_alt, "_matched_mpn_ui", "") if inv_alt is not None else "")
                    )
                    if not mpn:
                        inv_obj = self._inv_by_itemnum.get(anchored_cpn)
                        if inv_obj is not None:
                            mpn = _normalize_text(getattr(inv_obj, "vendoritem", ""))
                            if not manufacturer:
                                manufacturer = _normalize_text(getattr(inv_obj, "mfgname", ""))
                    return manufacturer, mpn

                try:
                    for a in (node.selected_alternates() or []):
                        if (getattr(a, "source", "") or "").lower() == "inventory":
                            continue
                        if bool(getattr(a, "rejected", False)):
                            continue
                        return (_normalize_text(getattr(a, "manufacturer", "")), _normalize_text(getattr(a, "manufacturer_part_number", "")))
                except Exception:
                    pass

                return ("", _normalize_text(getattr(node, "bom_mpn", "")))

            def _selected_external_alts(node: DecisionNode) -> list[Alternate]:
                out = []
                try:
                    for a in (node.selected_alternates() or []):
                        if (getattr(a, "source", "") or "").lower() == "inventory":
                            continue
                        if bool(getattr(a, "rejected", False)):
                            continue
                        out.append(a)
                except Exception:
                    pass
                return out

            def _internal_selected_is_exact(node: DecisionNode, inv_alt: Optional[Alternate]) -> bool:
                if inv_alt is None:
                    return False
                bom_mpn = _normalize_text(getattr(node, "bom_mpn", ""))
                _mfg, selected_mpn = export_mfg_fields(node)
                selected_mpn = _normalize_text(selected_mpn)
                if bom_mpn and selected_mpn and bom_mpn.lower() == selected_mpn.lower():
                    return True
                rel = _normalize_text(getattr(inv_alt, "relationship", ""))
                if "exact" in rel.lower():
                    return True
                match_type = _normalize_text(getattr(node, "match_type", ""))
                if "EXACT" in match_type.upper():
                    return True
                return False

            def _append_npr_row(
                part_number: str,
                part_status: str,
                item_description: str,
                custom_field: str,
                preferred_part: str = "",
                manufacturer_name: str = "",
                manufacturer_part_number: str = "",
                supplier: str = "",
                unit_cost: Any = "",
                default_uom: str = "",
                stock_unit: str = "",
                per_value: str = "",
                purchase_unit: str = "",
                lead_time: Any = "",
                qc_required: Any = "",
                tariff_code: str = "",
                quote_number: str = "",
            ) -> None:
                nonlocal npr_write_row
                row_vals = [
                    part_number,
                    part_status,
                    item_description,
                    custom_field,
                    preferred_part,
                    unit_cost,
                    default_uom,
                    stock_unit,
                    per_value,
                    purchase_unit,
                    manufacturer_name,
                    manufacturer_part_number,
                    supplier,
                    lead_time,
                    qc_required,
                    tariff_code,
                    quote_number,
                ]
                for idx, val in enumerate(row_vals, start=1):
                    npr_ws.cell(row=npr_write_row, column=idx, value=val)
                npr_write_row += 1

            def _alts_optional_approved_col(ws) -> int | None:
                try:
                    for c in range(1, ws.max_column + 1):
                        hdr = _normalize_text(ws.cell(row=alts_header_row, column=c).value)
                        if hdr.lower().startswith("approved"):
                            return c
                except Exception:
                    pass
                return None

            def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
                from copy import copy as _cpy
                try:
                    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
                except Exception:
                    pass
                for cc in range(1, max_col + 1):
                    s = ws.cell(row=src_row, column=cc)
                    d = ws.cell(row=dst_row, column=cc)
                    try:
                        d._style = _cpy(s._style)
                        d.number_format = s.number_format
                        d.font = _cpy(s.font)
                        d.fill = _cpy(s.fill)
                        d.border = _cpy(s.border)
                        d.alignment = _cpy(s.alignment)
                        d.protection = _cpy(s.protection)
                    except Exception:
                        pass

            def _unmerge_overlapping(ws, row_start: int, row_end: int) -> None:
                try:
                    ranges = list(ws.merged_cells.ranges)
                except Exception:
                    ranges = []
                for mr in ranges:
                    try:
                        if not (mr.max_row < row_start or mr.min_row > row_end):
                            ws.unmerge_cells(str(mr))
                    except Exception:
                        pass

            def _set_link_cell(cell, value: str) -> None:
                value = _normalize_text(value)
                cell.value = value
                if value and (value.startswith("http://") or value.startswith("https://")):
                    try:
                        cell.hyperlink = value
                    except Exception:
                        pass

            def _render_alts_approval_sheet(internal_groups: list[dict], external_rows: list[dict]) -> None:
                nonlocal alts_header_row

                inv_start = alts_header_row + 1
                ext_label_row = None
                for rr in range(inv_start, ws_alt.max_row + 1):
                    v = _normalize_text(ws_alt.cell(row=rr, column=2).value)
                    if v.upper().startswith("EXTERNAL ALTERNATES"):
                        ext_label_row = rr
                        break
                if ext_label_row is None:
                    ext_label_row = max(inv_start + 1, ws_alt.max_row + 2)
                    ws_alt.cell(row=ext_label_row, column=2, value="EXTERNAL ALTERNATES")
                    ws_alt.cell(row=ext_label_row + 1, column=2, value="List of parts that which may be fit as an alternate of designated BOM Manufacturer Part Number (MfgN) not within ELAN INVENTORY")
                    ws_alt.cell(row=ext_label_row + 2, column=2, value="MfgNs")
                    ws_alt.cell(row=ext_label_row + 2, column=3, value="Manufacturer")
                    ws_alt.cell(row=ext_label_row + 2, column=4, value="link")
                    ws_alt.cell(row=ext_label_row + 2, column=5, value="DESC")
                    ws_alt.cell(row=ext_label_row + 2, column=6, value="BOM MfgN Alternate For")
                    ws_alt.cell(row=ext_label_row + 2, column=7, value="Designator")
                    ws_alt.cell(row=ext_label_row + 2, column=8, value="NOTES")

                ext_header_row_local = ext_label_row + 2
                ext_start = ext_header_row_local + 1
                inv_capacity = max(0, ext_label_row - inv_start)
                internal_needed = sum(max(1, len(g.get("rows", []) or [])) for g in (internal_groups or []))
                if internal_needed > inv_capacity:
                    need = internal_needed - inv_capacity
                    ws_alt.insert_rows(ext_label_row, amount=need)
                    for rr in range(ext_label_row, ext_label_row + need):
                        _copy_row_style(ws_alt, max(inv_start, ext_label_row - 1), rr, max(8, ws_alt.max_column))
                    ext_label_row += need
                    ext_header_row_local += need
                    ext_start += need

                ext_capacity = max(0, ws_alt.max_row - ext_start + 1)
                external_needed = len(external_rows or [])
                if external_needed > ext_capacity:
                    need = external_needed - ext_capacity
                    ws_alt.insert_rows(ws_alt.max_row + 1, amount=need)
                    for rr in range(ws_alt.max_row - need + 1, ws_alt.max_row + 1):
                        _copy_row_style(ws_alt, ext_start, rr, max(8, ws_alt.max_column))

                _unmerge_overlapping(ws_alt, inv_start, ws_alt.max_row)

                for rr in range(inv_start, ws_alt.max_row + 1):
                    for cc in range(1, max(8, ws_alt.max_column) + 1):
                        ws_alt.cell(row=rr, column=cc).value = None
                        try:
                            ws_alt.cell(row=rr, column=cc).hyperlink = None
                        except Exception:
                            pass

                approved_col = _alts_optional_approved_col(ws_alt)
                cur = inv_start
                for group in (internal_groups or []):
                    rows = list(group.get("rows", []) or [])
                    span = max(1, len(rows))
                    for off in range(span):
                        _copy_row_style(ws_alt, inv_start, cur + off, max(8, ws_alt.max_column))
                    if span > 1:
                        for col in (1, 5, 6, 7, 8):
                            ws_alt.merge_cells(start_row=cur, start_column=col, end_row=cur + span - 1, end_column=col)
                        if approved_col:
                            ws_alt.merge_cells(start_row=cur, start_column=approved_col, end_row=cur + span - 1, end_column=approved_col)
                    ws_alt.cell(row=cur, column=1, value=group.get("elan_pn", ""))
                    ws_alt.cell(row=cur, column=5, value=group.get("desc", ""))
                    ws_alt.cell(row=cur, column=6, value=group.get("bom_mpn", ""))
                    ws_alt.cell(row=cur, column=7, value=group.get("designator", ""))
                    ws_alt.cell(row=cur, column=8, value=group.get("notes", "") or "---")
                    if approved_col:
                        ws_alt.cell(row=cur, column=approved_col, value=group.get("approved", "Yes"))
                    for off, row in enumerate(rows):
                        rr = cur + off
                        ws_alt.cell(row=rr, column=2, value=row.get("mfgpn", ""))
                        ws_alt.cell(row=rr, column=3, value=row.get("manufacturer", ""))
                        _set_link_cell(ws_alt.cell(row=rr, column=4), row.get("link", ""))
                    cur += span

                cur = ext_start
                for row in (external_rows or []):
                    _copy_row_style(ws_alt, ext_start, cur, max(8, ws_alt.max_column))
                    ws_alt.cell(row=cur, column=2, value=row.get("mfgpn", ""))
                    ws_alt.cell(row=cur, column=3, value=row.get("manufacturer", ""))
                    _set_link_cell(ws_alt.cell(row=cur, column=4), row.get("link", ""))
                    ws_alt.cell(row=cur, column=5, value=row.get("desc", ""))
                    ws_alt.cell(row=cur, column=6, value=row.get("bom_mpn", ""))
                    ws_alt.cell(row=cur, column=7, value=row.get("designator", ""))
                    ws_alt.cell(row=cur, column=8, value=row.get("notes", "") or "---")
                    if approved_col:
                        ws_alt.cell(row=cur, column=approved_col, value=row.get("approved", "Yes"))
                    cur += 1

            def _iter_company_mpn_rows(company_pn: str) -> list[tuple[str, str, str]]:
                company_pn = _normalize_text(company_pn)
                if not company_pn:
                    return []
                inv_obj = self._inv_by_itemnum.get(company_pn)
                row_company = None
                try:
                    row_company = self.inv_company_repo.get(self.workspace_id, company_pn)
                except Exception:
                    row_company = None
                if not row_company:
                    try:
                        row_company = self.inv_company_repo.get(self._inventory_store_workspace_id, company_pn)
                    except Exception:
                        row_company = None

                out = []
                if row_company:
                    alts = row_company.get("alternates_json", []) or []
                    if isinstance(alts, str):
                        try:
                            alts = json.loads(alts)
                        except Exception:
                            alts = []
                    seen = set()
                    for a in (alts or []):
                        mpn = _normalize_text((a or {}).get("mpn") or (a or {}).get("mfgpn") or (a or {}).get("manufacturer_part_number"))
                        mfg = _normalize_text((a or {}).get("mfgname") or (a or {}).get("manufacturer"))
                        link = _normalize_text((a or {}).get("link") or (a or {}).get("url") or (a or {}).get("purchase_url") or (a or {}).get("datasheet") or "")
                        if not mpn:
                            continue
                        key = (mpn.lower(), mfg.lower(), link.lower())
                        if key in seen:
                            continue
                        seen.add(key)
                        out.append((mpn, mfg, link))
                elif inv_obj is not None:
                    mpns = []
                    rep = _normalize_text(getattr(inv_obj, "vendoritem", ""))
                    if rep:
                        mpns.append(rep)
                    mpns.extend([_normalize_text(x) for x in (getattr(inv_obj, "substitutes", []) or []) if _normalize_text(x)])
                    seen = set()
                    for mpn in mpns:
                        k = mpn.lower()
                        if k in seen:
                            continue
                        seen.add(k)
                        out.append((mpn, _normalize_text(getattr(inv_obj, "mfgname", "")), ""))
                return out

            internal_approval_groups = []
            external_approval_rows = []

            for node in ready_nodes:
                ctx = _line_context_for_node(node)
                desc = self._export_description_for_node(node)
                bom_mpn = _normalize_text(getattr(node, "bom_mpn", ""))
                company_pn = _normalize_text(getattr(node, "internal_part_number", "") or getattr(node, "assigned_part_number", "") or "NEW")
                bucket, type_val = section_bucket(self.get_node_bom_section(node.id))
                note_text = _normalize_text(getattr(node, "notes", ""))

                inv_alt = self._selected_inventory_alt(node)
                external_alts = _selected_external_alts(node)
                has_internal = bool(_normalize_text(getattr(node, "internal_part_number", "")))
                include_approval = bool(getattr(node, "needs_approval", False))

                ws_bomexist.append([
                    _normalize_text(getattr(node, "bom_uid", "")),
                    bom_mpn,
                    desc,
                    company_pn,
                    "EXISTING" if has_internal else "NEW",
                    note_text,
                ])

                if has_internal and not external_alts:
                    if include_approval and inv_alt is not None and not _internal_selected_is_exact(node, inv_alt):
                        internal_rows = []
                        for mpn, mfg, link in _iter_company_mpn_rows(company_pn):
                            internal_rows.append({"mfgpn": mpn, "manufacturer": mfg, "link": link})
                        if internal_rows:
                            internal_approval_groups.append({
                                "elan_pn": company_pn,
                                "rows": internal_rows,
                                "desc": desc,
                                "bom_mpn": bom_mpn,
                                "designator": _normalize_text(ctx.get("refdes")),
                                "notes": note_text,
                                "approved": "Yes",
                            })
                    continue

                if has_internal:
                    if include_approval and inv_alt is not None and (external_alts or not _internal_selected_is_exact(node, inv_alt)):
                        internal_rows = []
                        for mpn, mfg, link in _iter_company_mpn_rows(company_pn):
                            internal_rows.append({"mfgpn": mpn, "manufacturer": mfg, "link": link})
                        if internal_rows:
                            internal_approval_groups.append({
                                "elan_pn": company_pn,
                                "rows": internal_rows,
                                "desc": desc,
                                "bom_mpn": bom_mpn,
                                "designator": _normalize_text(ctx.get("refdes")),
                                "notes": note_text,
                                "approved": "Yes",
                            })
                else:
                    input_mfgname = _normalize_text(ctx.get("mfgname"))
                    _append_npr_row(
                        company_pn,
                        "NEW",
                        desc,
                        type_val,
                        preferred_part=bom_mpn,
                        manufacturer_name=input_mfgname,
                        manufacturer_part_number=bom_mpn,
                    )

                for alt in external_alts:
                    # IMPORTANT:
                    # External NPR rows should stay correlated to the committed header values
                    # for this BOM line. The alternate contributes the external sourcing fields
                    # (MFG / MFGPN / supplier / link / cost), but the exported ELAN-side identity
                    # stays on the node: company_pn + header description.
                    alt_desc = desc
                    alt_supplier = _normalize_text(getattr(alt, "supplier", "") or getattr(alt, "source", ""))
                    alt_link = _normalize_text(
                        getattr(alt, "link", "")
                        or getattr(alt, "url", "")
                        or getattr(alt, "purchase_url", "")
                        or getattr(alt, "datasheet", "")
                        or alt_supplier
                    )
                    alt_mfg = _normalize_text(getattr(alt, "manufacturer", ""))
                    alt_mpn = _normalize_text(getattr(alt, "manufacturer_part_number", ""))
                    _append_npr_row(
                        company_pn,
                        "NEW",
                        alt_desc,
                        type_val,
                        preferred_part=alt_mpn,
                        manufacturer_name=alt_mfg,
                        manufacturer_part_number=alt_mpn,
                        supplier=alt_supplier,
                        unit_cost=getattr(alt, "unit_cost", ""),
                        stock_unit=_normalize_text(getattr(alt, "stock_unit", "")),
                        lead_time=getattr(alt, "lead_time_weeks", ""),
                        qc_required=getattr(alt, "qc_required", ""),
                        tariff_code=_normalize_text(getattr(alt, "tariff_code", "")),
                    )
                    if include_approval:
                        external_approval_rows.append({
                            "mfgpn": alt_mpn,
                            "manufacturer": alt_mfg,
                            "link": alt_link,
                            "desc": alt_desc,
                            "bom_mpn": bom_mpn,
                            "designator": _normalize_text(ctx.get("refdes")),
                            "notes": note_text,
                            "approved": "Yes",
                        })

            _render_alts_approval_sheet(internal_approval_groups, external_approval_rows)
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = str(Path.cwd() / f"EXPORT_{timestamp}.xlsx")

            try:
                wb.save(output_path)
            except PermissionError:
                raise PermissionError(f"Cannot save. Close the Excel file first: {output_path}")
            except Exception as e:
                raise RuntimeError(f"Export failed to save workbook: {e}")

            return output_path

    # ----------------------------
    # Internal enforcement
    # ----------------------------

    def _find_alt(self, node: DecisionNode, alt_id: str) -> Alternate:
        for a in node.alternates:
            if a.id == alt_id:
                return a
        raise KeyError(f"No alternate with id={alt_id} on node={node.id}")

    def _ensure_unlocked(self, node: DecisionNode) -> None:
        if node.locked:
            raise PermissionError("Node is locked (READY_FOR_EXPORT). Unlocking not supported.")
        

    def _recompute_node_flags(self, node: DecisionNode) -> None:
        """
        Central place for status/approval heuristics.
        UI should just display these.
        """
        bt = node.base_type.upper()
        selected = len(node.selected_alternates())
        candidates = len(node.candidate_alternates())

        ex = dict(getattr(node, "explain", {}) or {})
        if "approval_export_selected" in ex:
            node.needs_approval = bool(ex.get("approval_export_selected"))
        else:
            node.needs_approval = False

        # Status heuristic unless already READY/locked
        if node.locked:
            node.status = DecisionStatus.READY_FOR_EXPORT
            return

        if getattr(node, "status", None) == DecisionStatus.READY_FOR_EXPORT and self._has_selected_external_alt(node):
            # External-alternate flow stays editable after Mark Ready so description / assigned PN
            # can still be adjusted without losing Ready-for-export state.
            node.status = DecisionStatus.READY_FOR_EXPORT
            return

        if bt == "NEW" and selected == 0:
            node.status = DecisionStatus.NEEDS_ALTERNATE
            return

        if bt == "EXISTS" and candidates > 0 and selected == 0:
            node.status = DecisionStatus.NEEDS_DECISION
            return

        # default: auto-ish state (still editable)
        node.status = DecisionStatus.FULL_MATCH



        ### input in a set_node_description which uses the passed in description to change the description in the Exported BOM and NPR.
        ### 